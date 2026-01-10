"""
Django management command to map Excel data to database entities.
Extracts master list from AR+NR+LR+LVR+CR+MR+SR row and maps to DB.
"""

import json
import os
import re
from pathlib import Path
from django.core.management.base import BaseCommand
from django.conf import settings
from difflib import SequenceMatcher


class Command(BaseCommand):
    help = 'Map Excel text values to DB entities using fuzzy matching'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel-file',
            type=str,
            default='/home/itpc6/Public/django/git-repo/7nov/topteenhtml/SMART_ALIGNED_CAREER_SHEET_FILLED.xlsx',
            help='Path to Excel file'
        )
        parser.add_argument(
            '--master-data-dir',
            type=str,
            default='static/data/combined_report_data',
            help='Directory containing master JSON files'
        )
        parser.add_argument(
            '--output-file',
            type=str,
            default='static/data/combined_report_data/excel_to_db_mapping.json',
            help='Output file for mapping JSON'
        )
        parser.add_argument(
            '--master-only',
            action='store_true',
            help='Process only the master row (AR+NR+LR+LVR+CR+MR+SR) for initial mapping'
        )

    def handle(self, *args, **options):
        excel_file = Path(options['excel_file'])
        master_data_dir = Path(options['master_data_dir'])
        output_file = Path(options['output_file'])
        
        # Ensure output directory exists
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        self.stdout.write(self.style.SUCCESS('Starting Excel to DB mapping...'))
        
        # Load master JSON files
        self.stdout.write('\nLoading master JSON files...')
        clusters_data = self.load_json(master_data_dir / 'master_career_clusters.json')
        roles_data = self.load_json(master_data_dir / 'master_career_roles.json')
        pathways_data = self.load_json(master_data_dir / 'master_educational_pathways.json')
        
        clusters = clusters_data.get('clusters', [])
        roles = roles_data.get('roles', [])
        pathways = pathways_data.get('pathways', [])
        
        self.stdout.write(f'  ✓ Loaded {len(clusters)} clusters, {len(roles)} roles, {len(pathways)} pathways')
        
        # Read Excel file
        self.stdout.write(f'\nReading Excel file: {excel_file}')
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Find the master row (AR+NR+LR+LVR+CR+MR+SR)
        master_row = None
        master_aptitude_code = "AR+NR+LR+LVR+CR+MR+SR"
        
        for row_idx in range(2, ws.max_row + 1):
            aptitude_code = ws.cell(row=row_idx, column=1).value
            if aptitude_code and str(aptitude_code).strip() == master_aptitude_code:
                master_row = row_idx
                break
        
        if not master_row:
            self.stdout.write(self.style.ERROR(f'  ✗ Master row with "{master_aptitude_code}" not found!'))
            return
        
        self.stdout.write(f'  ✓ Found master row at row {master_row}')
        
        if options.get('master_only'):
            self.stdout.write(self.style.WARNING('\n⚠ MASTER-ONLY MODE: Processing only master row for initial mapping'))
        
        # Extract data from master row
        cluster_text = str(ws.cell(row=master_row, column=3).value or '').strip()
        role_text = str(ws.cell(row=master_row, column=4).value or '').strip()
        pathway_text = str(ws.cell(row=master_row, column=5).value or '').strip()
        
        # Split by comma and clean
        excel_clusters = [c.strip() for c in cluster_text.split(',') if c.strip()]
        excel_roles = [r.strip() for r in role_text.split(',') if r.strip()]
        excel_pathways = [p.strip() for p in pathway_text.split(',') if p.strip()]
        
        self.stdout.write(f'\nExtracted from master row:')
        self.stdout.write(f'  - {len(excel_clusters)} clusters')
        self.stdout.write(f'  - {len(excel_roles)} roles')
        self.stdout.write(f'  - {len(excel_pathways)} pathways')
        
        # Create mappings
        self.stdout.write('\nCreating mappings...')
        cluster_mappings = self.map_clusters(excel_clusters, clusters)
        role_mappings = self.map_items(excel_roles, roles, 'role')
        pathway_mappings = self.map_items(excel_pathways, pathways, 'pathway')
        
        # Create output structure
        output_data = {
            "cluster_mappings": cluster_mappings,
            "role_mappings": role_mappings,
            "pathway_mappings": pathway_mappings,
            "statistics": {
                "excel_clusters_count": len(excel_clusters),
                "excel_roles_count": len(excel_roles),
                "excel_pathways_count": len(excel_pathways),
                "mapped_clusters_count": sum(1 for v in cluster_mappings.values() if v),
                "mapped_roles_count": sum(1 for v in role_mappings.values() if v),
                "mapped_pathways_count": sum(1 for v in pathway_mappings.values() if v),
            },
            "metadata": {
                "master_only_mode": options.get('master_only', False),
                "source_row": master_row,
                "aptitude_code": master_aptitude_code if options.get('master_only') else None
            }
        }
        
        # Save mapping
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)
        
        self.stdout.write(self.style.SUCCESS(f'\n✓ Mapping saved to {output_file}'))
        
        # Print statistics
        self.stdout.write('\nMapping Statistics:')
        self.stdout.write(f'  Clusters: {output_data["statistics"]["mapped_clusters_count"]}/{output_data["statistics"]["excel_clusters_count"]} mapped')
        self.stdout.write(f'  Roles: {output_data["statistics"]["mapped_roles_count"]}/{output_data["statistics"]["excel_roles_count"]} mapped')
        self.stdout.write(f'  Pathways: {output_data["statistics"]["mapped_pathways_count"]}/{output_data["statistics"]["excel_pathways_count"]} mapped')
        
        # Report unmatched items
        self.report_unmatched(cluster_mappings, role_mappings, pathway_mappings)

    def load_json(self, file_path):
        """Load JSON file"""
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def similarity(self, a, b):
        """Calculate similarity ratio between two strings"""
        return SequenceMatcher(None, a.lower(), b.lower()).ratio()

    def normalize_name(self, name):
        """Normalize name for comparison"""
        # Remove extra spaces, convert to lowercase
        name = re.sub(r'\s+', ' ', name.strip().lower())
        # Remove special characters for comparison
        name = re.sub(r'[^\w\s]', '', name)
        return name

    def map_clusters(self, excel_clusters, db_clusters):
        """
        Map Excel cluster names to DB clusters.
        Handles cases where Excel has combined names like "Architecture, Construction & Planning"
        """
        mappings = {}
        
        # Create lookup dictionaries
        db_cluster_by_name = {self.normalize_name(c['name']): c for c in db_clusters}
        db_cluster_names = list(db_cluster_by_name.keys())
        
        for excel_cluster in excel_clusters:
            excel_normalized = self.normalize_name(excel_cluster)
            matched_clusters = []
            
            # Strategy 1: Exact match
            if excel_normalized in db_cluster_by_name:
                matched_clusters.append(db_cluster_by_name[excel_normalized])
            else:
                # Strategy 2: Check if Excel cluster contains comma (might be multiple)
                if ',' in excel_cluster:
                    # Split and try to match each part
                    parts = [p.strip() for p in excel_cluster.split(',')]
                    for part in parts:
                        part_normalized = self.normalize_name(part)
                        # Try exact match first
                        if part_normalized in db_cluster_by_name:
                            matched_clusters.append(db_cluster_by_name[part_normalized])
                        else:
                            # Try fuzzy match
                            best_match = self.find_best_match(part_normalized, db_cluster_names, threshold=0.7)
                            if best_match:
                                matched_clusters.append(db_cluster_by_name[best_match])
                else:
                    # Strategy 3: Fuzzy match for single cluster name
                    best_match = self.find_best_match(excel_normalized, db_cluster_names, threshold=0.7)
                    if best_match:
                        matched_clusters.append(db_cluster_by_name[best_match])
                
                # Strategy 4: Check if any DB cluster name is contained in Excel cluster or vice versa
                if not matched_clusters:
                    for db_name in db_cluster_names:
                        # Check if Excel cluster is part of DB cluster name (e.g., "Education And Training" in "Arts, Humanities , Education & Training")
                        if excel_normalized in db_name or db_name in excel_normalized:
                            # Make sure it's a meaningful match (not too short)
                            if len(excel_normalized) >= 5 or len(db_name) >= 5:
                                matched_clusters.append(db_cluster_by_name[db_name])
                                break
                    
                    # Special case: "Education And Training" should map to "Arts, Humanities , Education & Training"
                    if not matched_clusters and 'education' in excel_normalized and 'training' in excel_normalized:
                        for db_name in db_cluster_names:
                            if 'education' in db_name and 'training' in db_name:
                                matched_clusters.append(db_cluster_by_name[db_name])
                                break
            
            # Remove duplicates while preserving order
            seen = set()
            unique_matched = []
            for cluster in matched_clusters:
                cluster_id = cluster['id']
                if cluster_id not in seen:
                    seen.add(cluster_id)
                    unique_matched.append(cluster)
            
            mappings[excel_cluster] = unique_matched if unique_matched else None
        
        return mappings

    def map_items(self, excel_items, db_items, item_type):
        """Map Excel items (roles or pathways) to DB items"""
        mappings = {}
        
        # Create lookup dictionaries
        db_item_by_name = {self.normalize_name(item['name']): item for item in db_items}
        db_item_names = list(db_item_by_name.keys())
        
        # Create manual override mappings for roles (to fix incorrect fuzzy matches)
        role_overrides = {
            'ux designer': 'ui and ux designer',
            'strategist': 'digital-content strategist',  # Keep current match, but could be improved
        }
        
        # Create abbreviation/alias mappings for pathways
        pathway_aliases = {
            'b.arch': ['bachelor of architecture', 'b.arch', 'bachelor architecture'],
            'b.com': ['bachelor of commerce', 'b.com', 'bachelor commerce'],
            'b.des': ['bachelor of design', 'b.des', 'bachelor design'],
            'b.ed': ['bachelor of education', 'b.ed', 'bachelor education'],
            'b.sc animation': ['bachelor of science in animation', 'bsc animation', 'b.sc in animation'],
            'b.sc logic': ['bachelor of science in logic', 'bsc logic', 'b.sc in logic'],
            'b.sc math': ['bachelor of science in mathematics', 'bsc mathematics', 'b.sc in math', 'bachelor of science mathematics'],
            'b.tech': ['bachelor of technology', 'btech', 'b.tech'],
            'b.tech (cs)': ['bachelor of technology in computer science', 'btech computer science', 'b.tech cse', 'bachelor of technology computer science engineering', 'bachelor of technology in computer science and engineering'],
            'b.tech (mech)': ['bachelor of technology in mechanical', 'btech mechanical', 'b.tech mechanical engineering', 'bachelor of technology mechanical engineering', 'bachelor of technology in mechanical engineering'],
            'ba english': ['bachelor of arts in english', 'ba in english', 'bachelor of arts english'],
            'bba': ['bachelor of business administration', 'bba', 'bachelor business administration'],
            'bfa': ['bachelor of fine arts', 'bfa', 'bachelor fine arts'],
            'llb': ['bachelor of law', 'llb', 'bachelor law', 'bachelor of laws'],
            'diploma in mech engg': ['diploma in mechanical engineering', 'diploma mechanical engineering'],
            'diploma in office management': ['diploma in office management', 'diploma office management'],
            'liberal arts': ['liberal arts', 'bachelor of arts', 'ba'],
        }
        
        for excel_item in excel_items:
            excel_normalized = self.normalize_name(excel_item)
            matched_item = None
            
            # Strategy 0: Manual overrides (for roles with known issues)
            if item_type == 'role' and excel_normalized in role_overrides:
                override_name = role_overrides[excel_normalized]
                if override_name in db_item_by_name:
                    matched_item = db_item_by_name[override_name]
                    mappings[excel_item] = matched_item
                    continue
            
            # Strategy 1: Exact match
            if excel_normalized in db_item_by_name:
                matched_item = db_item_by_name[excel_normalized]
            else:
                # Strategy 2: Check abbreviation/alias mappings (for pathways)
                if item_type == 'pathway':
                    for alias_key, alias_list in pathway_aliases.items():
                        if excel_normalized == alias_key or excel_normalized.startswith(alias_key):
                            # Try to find matching pathway using aliases
                            for alias in alias_list:
                                alias_normalized = self.normalize_name(alias)
                                # Try exact match first
                                if alias_normalized in db_item_by_name:
                                    matched_item = db_item_by_name[alias_normalized]
                                    break
                                # Try fuzzy match on aliases
                                best_match = self.find_best_match(alias_normalized, db_item_names, threshold=0.6)
                                if best_match:
                                    matched_item = db_item_by_name[best_match]
                                    break
                            if matched_item:
                                break
                    
                    # Also try direct search in DB names for abbreviations
                    if not matched_item:
                        # Remove parentheses and special chars for comparison
                        excel_clean = re.sub(r'[()]', '', excel_normalized)
                        for db_name in db_item_names:
                            db_clean = re.sub(r'[()]', '', db_name)
                            # Check if Excel abbreviation appears in DB name
                            if excel_clean in db_clean or db_clean.startswith(excel_clean):
                                # Make sure it's a reasonable match
                                if len(excel_clean) >= 2:
                                    matched_item = db_item_by_name[db_name]
                                    break
                
                # Strategy 3: Fuzzy match with improved logic
                if not matched_item:
                    threshold = 0.7 if item_type == 'pathway' else 0.8
                    
                    # For roles, check for key term matches first (e.g., "UX" should match "UI and UX")
                    if item_type == 'role':
                        # Extract key terms from Excel (remove common words)
                        excel_keywords = set(re.findall(r'\b\w+\b', excel_normalized))
                        common_words = {'and', 'the', 'a', 'an', 'of', 'in', 'for', 'with'}
                        excel_keywords = excel_keywords - common_words
                        
                        # Find matches where key terms appear
                        candidate_matches = []
                        for db_name in db_item_names:
                            db_keywords = set(re.findall(r'\b\w+\b', db_name))
                            overlap = excel_keywords & db_keywords
                            
                            # If significant overlap (especially for short terms like "UX")
                            if overlap:
                                # For short terms like "UX", require exact keyword match
                                if len(excel_keywords) <= 2:
                                    if excel_keywords.issubset(db_keywords):
                                        similarity = self.similarity(excel_normalized, db_name)
                                        candidate_matches.append((db_name, similarity))
                                else:
                                    # For longer terms, use similarity
                                    similarity = self.similarity(excel_normalized, db_name)
                                    if similarity >= threshold:
                                        candidate_matches.append((db_name, similarity))
                        
                        # Sort by similarity and pick best
                        if candidate_matches:
                            candidate_matches.sort(key=lambda x: x[1], reverse=True)
                            best_match_name = candidate_matches[0][0]
                            matched_item = db_item_by_name[best_match_name]
                    
                    # Fallback to standard fuzzy match
                    if not matched_item:
                        best_match = self.find_best_match(excel_normalized, db_item_names, threshold=threshold)
                        if best_match:
                            # Additional validation: for roles, avoid clearly wrong matches
                            if item_type == 'role':
                                # Check if match makes sense (e.g., "UX" shouldn't match "VFX")
                                excel_words = set(re.findall(r'\b\w+\b', excel_normalized.lower()))
                                match_words = set(re.findall(r'\b\w+\b', best_match.lower()))
                                
                                # Reject wrong matches
                                wrong_matches = [
                                    ('ux', 'vfx'),  # UX should not match VFX
                                    ('ux', 'ui'),  # UX should prefer matches with UX in them
                                ]
                                
                                should_reject = False
                                for excel_word, match_word in wrong_matches:
                                    if excel_word in excel_words and match_word in match_words and excel_word not in match_words:
                                        should_reject = True
                                        break
                                
                                if should_reject:
                                    # Try to find a better match
                                    # For UX, specifically search for "ux" in DB names
                                    if 'ux' in excel_words:
                                        for db_name in db_item_names:
                                            if 'ux' in db_name.lower() and 'vfx' not in db_name.lower():
                                                matched_item = db_item_by_name[db_name]
                                                break
                                
                                if not matched_item:
                                    matched_item = db_item_by_name[best_match]
                            else:
                                matched_item = db_item_by_name[best_match]
                
                # Strategy 4: Check if Excel item is contained in any DB item name or vice versa
                if not matched_item:
                    for db_name in db_item_names:
                        # For pathways, use more flexible matching
                        if item_type == 'pathway':
                            # Check if main terms match (e.g., "B.Tech (CS)" should match anything with "B.Tech" and "Computer" or "CS")
                            excel_terms = set(re.findall(r'\b\w+\b', excel_normalized))
                            db_terms = set(re.findall(r'\b\w+\b', db_name))
                            # Remove very common words
                            common_words = {'in', 'of', 'and', 'the', 'a', 'an', 'bachelor', 'technology', 'science', 'arts', 'diploma', 'engineering', 'degree'}
                            excel_terms = excel_terms - common_words
                            db_terms = db_terms - common_words
                            # Check for significant overlap
                            if excel_terms and db_terms:
                                overlap = excel_terms & db_terms
                                # If at least 2 meaningful terms match, or if it's a short abbreviation that matches
                                if len(overlap) >= 2 or (len(excel_normalized) <= 10 and len(overlap) >= 1):
                                    matched_item = db_item_by_name[db_name]
                                    break
                        else:
                            # For roles, use substring matching with better threshold
                            if excel_normalized in db_name or db_name in excel_normalized:
                                # Check if it's a reasonable match (not too short)
                                if len(excel_normalized) >= 3 and len(db_name) >= 3:
                                    matched_item = db_item_by_name[db_name]
                                    break
                            # Also try word-by-word matching for roles
                            excel_words = set(re.findall(r'\b\w+\b', excel_normalized))
                            db_words = set(re.findall(r'\b\w+\b', db_name))
                            if excel_words and db_words:
                                overlap = excel_words & db_words
                                # If most words match, consider it a match
                                if len(overlap) >= len(excel_words) * 0.6:  # At least 60% of words match
                                    matched_item = db_item_by_name[db_name]
                                    break
            
            mappings[excel_item] = matched_item
        
        return mappings

    def extract_keywords(self, text):
        """Extract key keywords from text (remove common words)"""
        # Common words to ignore
        stop_words = {'in', 'of', 'and', 'the', 'a', 'an', 'bachelor', 'technology', 'science', 'arts', 'diploma', 'engineering', 'degree'}
        words = re.findall(r'\b\w+\b', text.lower())
        keywords = [w for w in words if w not in stop_words and len(w) > 2]
        return keywords

    def find_best_match(self, search_term, candidates, threshold=0.8):
        """Find best matching candidate using similarity"""
        best_match = None
        best_score = 0
        
        for candidate in candidates:
            score = self.similarity(search_term, candidate)
            if score > best_score and score >= threshold:
                best_score = score
                best_match = candidate
        
        return best_match

    def report_unmatched(self, cluster_mappings, role_mappings, pathway_mappings):
        """Report items that couldn't be matched"""
        unmatched_clusters = [k for k, v in cluster_mappings.items() if not v]
        unmatched_roles = [k for k, v in role_mappings.items() if not v]
        unmatched_pathways = [k for k, v in pathway_mappings.items() if not v]
        
        if unmatched_clusters or unmatched_roles or unmatched_pathways:
            self.stdout.write(self.style.WARNING('\n⚠ Unmatched Items (require manual review):'))
            
            if unmatched_clusters:
                self.stdout.write(f'\n  Unmatched Clusters ({len(unmatched_clusters)}):')
                for item in unmatched_clusters[:10]:  # Show first 10
                    self.stdout.write(f'    - {item}')
                if len(unmatched_clusters) > 10:
                    self.stdout.write(f'    ... and {len(unmatched_clusters) - 10} more')
            
            if unmatched_roles:
                self.stdout.write(f'\n  Unmatched Roles ({len(unmatched_roles)}):')
                for item in unmatched_roles[:10]:
                    self.stdout.write(f'    - {item}')
                if len(unmatched_roles) > 10:
                    self.stdout.write(f'    ... and {len(unmatched_roles) - 10} more')
            
            if unmatched_pathways:
                self.stdout.write(f'\n  Unmatched Pathways ({len(unmatched_pathways)}):')
                for item in unmatched_pathways[:10]:
                    self.stdout.write(f'    - {item}')
                if len(unmatched_pathways) > 10:
                    self.stdout.write(f'    ... and {len(unmatched_pathways) - 10} more')
        else:
            self.stdout.write(self.style.SUCCESS('\n✓ All items matched successfully!'))
