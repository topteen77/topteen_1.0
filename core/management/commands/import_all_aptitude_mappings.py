"""
Django management command to import all 127 aptitude combination mappings from Excel.
Uses the corrected master mappings from database.
"""

import json
from pathlib import Path
from django.core.management.base import BaseCommand
from django.db import transaction
from app_post_matric.models import (
    ClusterMapping, RoleMapping, PathwayMapping, AptitudeCombinationMapping
)
from careers.models import CareerCluster, Career
from courses.models import Course


class Command(BaseCommand):
    help = 'Import all 127 aptitude combination mappings from Excel using corrected DB mappings'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel-file',
            type=str,
            default='/home/itpc6/Public/django/git-repo/7nov/topteenhtml/SMART_ALIGNED_CAREER_SHEET_FILLED.xlsx',
            help='Path to Excel file'
        )
        parser.add_argument(
            '--clear-existing',
            action='store_true',
            help='Clear existing aptitude combination mappings before importing'
        )

    def handle(self, *args, **options):
        excel_file = Path(options['excel_file'])
        clear_existing = options.get('clear_existing', False)
        
        self.stdout.write(self.style.SUCCESS('Starting full aptitude mapping import...'))
        
        # Clear existing if requested
        if clear_existing:
            self.stdout.write('\nClearing existing aptitude combination mappings...')
            AptitudeCombinationMapping.objects.all().delete()
            self.stdout.write('  ✓ Cleared existing mappings')
        
        # Read Excel file
        self.stdout.write(f'\nReading Excel file: {excel_file}')
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        total_rows = 0
        imported_count = 0
        error_count = 0
        
        # Process each row (starting from row 2, skipping header)
        self.stdout.write('\nProcessing rows...')
        for row_idx in range(2, ws.max_row + 1):
            aptitude_code = ws.cell(row=row_idx, column=1).value
            aptitude_area = ws.cell(row=row_idx, column=2).value
            cluster_text = ws.cell(row=row_idx, column=3).value
            role_text = ws.cell(row=row_idx, column=4).value
            pathway_text = ws.cell(row=row_idx, column=5).value
            
            # Skip empty rows
            if not aptitude_code:
                continue
            
            total_rows += 1
            aptitude_code = str(aptitude_code).strip()
            aptitude_area = str(aptitude_area or '').strip()
            cluster_text = str(cluster_text or '').strip()
            role_text = str(role_text or '').strip()
            pathway_text = str(pathway_text or '').strip()
            
            try:
                with transaction.atomic():
                    # Get or create aptitude combination
                    combination, created = AptitudeCombinationMapping.objects.get_or_create(
                        aptitude_code=aptitude_code,
                        defaults={'aptitude_areas': aptitude_area, 'is_complete': False}
                    )
                    
                    if not created:
                        combination.aptitude_areas = aptitude_area
                    
                    # Clear existing relationships
                    combination.clusters.clear()
                    combination.roles.clear()
                    combination.pathways.clear()
                    
                    # Map clusters
                    if cluster_text:
                        cluster_names = [c.strip() for c in cluster_text.split(',') if c.strip()]
                        for cluster_name in cluster_names:
                            try:
                                cluster_mapping = ClusterMapping.objects.get(excel_name=cluster_name)
                                if cluster_mapping.is_mapped and cluster_mapping.db_cluster:
                                    combination.clusters.add(cluster_mapping.db_cluster)
                            except ClusterMapping.DoesNotExist:
                                self.stdout.write(self.style.WARNING(f'  ⚠ Cluster mapping not found: {cluster_name}'))
                    
                    # Map roles
                    if role_text:
                        role_names = [r.strip() for r in role_text.split(',') if r.strip()]
                        for role_name in role_names:
                            try:
                                role_mapping = RoleMapping.objects.get(excel_name=role_name)
                                if role_mapping.is_mapped and role_mapping.db_role:
                                    combination.roles.add(role_mapping.db_role)
                            except RoleMapping.DoesNotExist:
                                self.stdout.write(self.style.WARNING(f'  ⚠ Role mapping not found: {role_name}'))
                    
                    # Map pathways
                    if pathway_text:
                        pathway_names = [p.strip() for p in pathway_text.split(',') if p.strip()]
                        for pathway_name in pathway_names:
                            try:
                                pathway_mapping = PathwayMapping.objects.get(excel_name=pathway_name)
                                if pathway_mapping.is_mapped and pathway_mapping.db_pathway:
                                    combination.pathways.add(pathway_mapping.db_pathway)
                            except PathwayMapping.DoesNotExist:
                                self.stdout.write(self.style.WARNING(f'  ⚠ Pathway mapping not found: {pathway_name}'))
                    
                    # Check if complete
                    has_clusters = combination.clusters.exists()
                    has_roles = combination.roles.exists()
                    has_pathways = combination.pathways.exists()
                    combination.is_complete = has_clusters and has_roles and has_pathways
                    combination.save()
                    
                    imported_count += 1
                    
                    if total_rows % 20 == 0:
                        self.stdout.write(f'  Processed {total_rows} rows...')
                        
            except Exception as e:
                error_count += 1
                self.stdout.write(self.style.ERROR(f'  ✗ Error processing {aptitude_code}: {str(e)}'))
        
        self.stdout.write(self.style.SUCCESS(f'\n✓ Import completed!'))
        self.stdout.write(f'  Total rows processed: {total_rows}')
        self.stdout.write(f'  Successfully imported: {imported_count}')
        self.stdout.write(f'  Errors: {error_count}')
        
        # Show statistics
        complete = AptitudeCombinationMapping.objects.filter(is_complete=True).count()
        incomplete = AptitudeCombinationMapping.objects.filter(is_complete=False).count()
        self.stdout.write(f'\nMapping Status:')
        self.stdout.write(f'  Complete: {complete}')
        self.stdout.write(f'  Incomplete: {incomplete}')
