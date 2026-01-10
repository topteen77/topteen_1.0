"""
Django management command to create master mapping for all aptitude combinations.
Reads Excel file and maps each aptitude combination to DB entities.
"""

import json
from pathlib import Path
from django.core.management.base import BaseCommand
from collections import defaultdict


class Command(BaseCommand):
    help = 'Create master mapping for all aptitude combinations (AR, AR+NR, etc.)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel-file',
            type=str,
            default='/home/itpc6/Public/django/git-repo/7nov/topteenhtml/SMART_ALIGNED_CAREER_SHEET_FILLED.xlsx',
            help='Path to Excel file'
        )
        parser.add_argument(
            '--mapping-file',
            type=str,
            default='static/data/combined_report_data/excel_to_db_mapping.json',
            help='Path to Excel to DB mapping file'
        )
        parser.add_argument(
            '--output-file',
            type=str,
            default='static/data/combined_report_data/aptitude_master_mapping.json',
            help='Output file for master mapping JSON'
        )

    def handle(self, *args, **options):
        excel_file = Path(options['excel_file'])
        mapping_file = Path(options['mapping_file'])
        output_file = Path(options['output_file'])
        
        # Ensure output directory exists
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        self.stdout.write(self.style.SUCCESS('Starting aptitude master mapping creation...'))
        
        # Load Excel to DB mapping
        self.stdout.write(f'\nLoading mapping file: {mapping_file}')
        with open(mapping_file, 'r', encoding='utf-8') as f:
            mapping_data = json.load(f)
        
        cluster_mappings = mapping_data.get('cluster_mappings', {})
        role_mappings = mapping_data.get('role_mappings', {})
        pathway_mappings = mapping_data.get('pathway_mappings', {})
        
        self.stdout.write('  ✓ Mapping data loaded')
        
        # Read Excel file
        self.stdout.write(f'\nReading Excel file: {excel_file}')
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        master_mapping = {}
        total_rows = 0
        mapped_rows = 0
        
        # Process each row (starting from row 2, skipping header)
        for row_idx in range(2, ws.max_row + 1):
            aptitude_code = ws.cell(row=row_idx, column=1).value
            aptitude_area = ws.cell(row=row_idx, column=2).value
            cluster_text = ws.cell(row=row_idx, column=3).value
            role_text = ws.cell(row=row_idx, column=4).value
            pathway_text = ws.cell(row=row_idx, column=5).value
            
            # Skip empty rows
            if not aptitude_code:
                continue
            
            total_rows += 1
            aptitude_code = str(aptitude_code).strip()
            aptitude_area = str(aptitude_area or '').strip()
            cluster_text = str(cluster_text or '').strip()
            role_text = str(role_text or '').strip()
            pathway_text = str(pathway_text or '').strip()
            
            # Map clusters
            mapped_clusters = self.map_clusters(cluster_text, cluster_mappings)
            
            # Map roles
            mapped_roles = self.map_roles(role_text, role_mappings)
            
            # Map pathways
            mapped_pathways = self.map_pathways(pathway_text, pathway_mappings)
            
            # Store in master mapping
            master_mapping[aptitude_code] = {
                'Areas': aptitude_area,
                'Career_Clusters': mapped_clusters,
                'Career_Roles': mapped_roles,
                'Educational_Pathways': mapped_pathways
            }
            
            if mapped_clusters or mapped_roles or mapped_pathways:
                mapped_rows += 1
        
        # Create output structure
        output_data = {
            'mappings': master_mapping,
            'statistics': {
                'total_aptitude_combinations': len(master_mapping),
                'total_rows_processed': total_rows,
                'mapped_rows': mapped_rows,
                'clusters_mapped': sum(1 for v in master_mapping.values() if v['Career_Clusters']),
                'roles_mapped': sum(1 for v in master_mapping.values() if v['Career_Roles']),
                'pathways_mapped': sum(1 for v in master_mapping.values() if v['Educational_Pathways']),
            }
        }
        
        # Save mapping
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)
        
        self.stdout.write(self.style.SUCCESS(f'\n✓ Master mapping saved to {output_file}'))
        
        # Print statistics
        stats = output_data['statistics']
        self.stdout.write('\nMapping Statistics:')
        self.stdout.write(f'  Total aptitude combinations: {stats["total_aptitude_combinations"]}')
        self.stdout.write(f'  Rows with mappings: {stats["mapped_rows"]}/{stats["total_rows_processed"]}')
        self.stdout.write(f'  Combinations with clusters: {stats["clusters_mapped"]}')
        self.stdout.write(f'  Combinations with roles: {stats["roles_mapped"]}')
        self.stdout.write(f'  Combinations with pathways: {stats["pathways_mapped"]}')

    def map_clusters(self, cluster_text, cluster_mappings):
        """Map cluster text to DB clusters"""
        if not cluster_text:
            return []
        
        # Split by comma
        cluster_names = [c.strip() for c in cluster_text.split(',') if c.strip()]
        mapped_clusters = []
        seen_ids = set()
        
        for cluster_name in cluster_names:
            # Look up in mapping
            if cluster_name in cluster_mappings:
                db_clusters = cluster_mappings[cluster_name]
                if db_clusters:  # Not None
                    for cluster in db_clusters:
                        cluster_id = cluster['id']
                        if cluster_id not in seen_ids:
                            seen_ids.add(cluster_id)
                            mapped_clusters.append({
                                'id': cluster['id'],
                                'name': cluster['name'],
                                'slug': cluster.get('slug', '')
                            })
        
        return mapped_clusters

    def map_roles(self, role_text, role_mappings):
        """Map role text to DB roles"""
        if not role_text:
            return []
        
        # Split by comma
        role_names = [r.strip() for r in role_text.split(',') if r.strip()]
        mapped_roles = []
        seen_ids = set()
        
        for role_name in role_names:
            # Look up in mapping
            if role_name in role_mappings:
                db_role = role_mappings[role_name]
                if db_role:  # Not None
                    role_id = db_role['id']
                    if role_id not in seen_ids:
                        seen_ids.add(role_id)
                        mapped_roles.append({
                            'id': db_role['id'],
                            'name': db_role['name'],
                            'slug': db_role.get('slug', '')
                        })
        
        return mapped_roles

    def map_pathways(self, pathway_text, pathway_mappings):
        """Map pathway text to DB pathways"""
        if not pathway_text:
            return []
        
        # Split by comma
        pathway_names = [p.strip() for p in pathway_text.split(',') if p.strip()]
        mapped_pathways = []
        seen_ids = set()
        
        for pathway_name in pathway_names:
            # Look up in mapping
            if pathway_name in pathway_mappings:
                db_pathway = pathway_mappings[pathway_name]
                if db_pathway:  # Not None
                    pathway_id = db_pathway['id']
                    if pathway_id not in seen_ids:
                        seen_ids.add(pathway_id)
                        mapped_pathways.append({
                            'id': db_pathway['id'],
                            'name': db_pathway['name'],
                            'slug': db_pathway.get('slug', '')
                        })
        
        return mapped_pathways
