"""
Django management command to validate and cross-check all mappings.
Adds flags (matched/error) and keeps Excel names for manual checking.
"""

import json
from pathlib import Path
from django.core.management.base import BaseCommand
import openpyxl


class Command(BaseCommand):
    help = 'Validate and cross-check all mappings with flags and error tracking'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel-file',
            type=str,
            default='/home/itpc6/Public/django/git-repo/7nov/topteenhtml/SMART_ALIGNED_CAREER_SHEET_FILLED.xlsx',
            help='Path to Excel file'
        )
        parser.add_argument(
            '--mapping-file',
            type=str,
            default='static/data/combined_report_data/excel_to_db_mapping.json',
            help='Path to Excel to DB mapping file'
        )
        parser.add_argument(
            '--master-mapping-file',
            type=str,
            default='static/data/combined_report_data/aptitude_master_mapping.json',
            help='Path to aptitude master mapping file'
        )
        parser.add_argument(
            '--output-file',
            type=str,
            default='static/data/combined_report_data/mapping_validation_report.json',
            help='Output file for validation report'
        )

    def handle(self, *args, **options):
        excel_file = Path(options['excel_file'])
        mapping_file = Path(options['mapping_file'])
        master_mapping_file = Path(options['master_mapping_file'])
        output_file = Path(options['output_file'])
        
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        self.stdout.write(self.style.SUCCESS('Starting mapping validation...'))
        
        # Load files
        self.stdout.write('\nLoading files...')
        with open(mapping_file, 'r', encoding='utf-8') as f:
            mapping_data = json.load(f)
        
        with open(master_mapping_file, 'r', encoding='utf-8') as f:
            master_mapping_data = json.load(f)
        
        # Read Excel file
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        self.stdout.write('  ✓ Files loaded')
        
        # Validate clusters
        self.stdout.write('\nValidating clusters...')
        cluster_validation = self.validate_clusters(ws, mapping_data['cluster_mappings'])
        
        # Validate roles
        self.stdout.write('Validating roles...')
        role_validation = self.validate_roles(ws, mapping_data['role_mappings'])
        
        # Validate pathways
        self.stdout.write('Validating pathways...')
        pathway_validation = self.validate_pathways(ws, mapping_data['pathway_mappings'])
        
        # Validate master mapping
        self.stdout.write('Validating master mapping...')
        master_validation = self.validate_master_mapping(ws, master_mapping_data['mappings'])
        
        # Create validation report
        validation_report = {
            'cluster_validation': cluster_validation,
            'role_validation': role_validation,
            'pathway_validation': pathway_validation,
            'master_mapping_validation': master_validation,
            'summary': {
                'total_clusters': len(cluster_validation),
                'matched_clusters': sum(1 for v in cluster_validation.values() if v['status'] == 'matched'),
                'error_clusters': sum(1 for v in cluster_validation.values() if v['status'] == 'error'),
                'total_roles': len(role_validation),
                'matched_roles': sum(1 for v in role_validation.values() if v['status'] == 'matched'),
                'error_roles': sum(1 for v in role_validation.values() if v['status'] == 'error'),
                'unmapped_roles': sum(1 for v in role_validation.values() if v['status'] == 'unmapped'),
                'total_pathways': len(pathway_validation),
                'matched_pathways': sum(1 for v in pathway_validation.values() if v['status'] == 'matched'),
                'error_pathways': sum(1 for v in pathway_validation.values() if v['status'] == 'error'),
            }
        }
        
        # Save report
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(validation_report, f, indent=2, ensure_ascii=False)
        
        self.stdout.write(self.style.SUCCESS(f'\n✓ Validation report saved to {output_file}'))
        
        # Print summary
        self.print_summary(validation_report)

    def validate_clusters(self, ws, cluster_mappings):
        """Validate cluster mappings"""
        validation = {}
        error_id = 1
        
        # Get master row
        master_row = self.find_master_row(ws)
        cluster_text = str(ws.cell(row=master_row, column=3).value or '').strip()
        excel_clusters = [c.strip() for c in cluster_text.split(',') if c.strip()]
        
        for excel_cluster in excel_clusters:
            entry = {
                'excel_name': excel_cluster,
                'status': 'unmapped',
                'mapped_to': None,
                'error_id': None,
                'error_message': None
            }
            
            if excel_cluster in cluster_mappings:
                cluster_data = cluster_mappings[excel_cluster]
                # Handle new structure with "matched" and "db_matches"
                if isinstance(cluster_data, dict) and 'db_matches' in cluster_data:
                    db_clusters = cluster_data['db_matches']
                    match_status = cluster_data.get('matched', 'no')
                else:
                    # Old structure - direct list
                    db_clusters = cluster_data
                    match_status = 'yes' if db_clusters else 'no'
                
                if db_clusters and len(db_clusters) > 0:
                    entry['status'] = 'matched' if match_status == 'yes' else 'error'
                    entry['mapped_to'] = [{'id': c['id'], 'name': c['name']} for c in db_clusters if isinstance(c, dict) and 'id' in c]
                    
                    # Check for potential issues
                    issues = self.check_cluster_issues(excel_cluster, db_clusters)
                    if issues:
                        entry['status'] = 'error'
                        entry['error_id'] = f'CLUSTER-{error_id:03d}'
                        entry['error_message'] = issues
                        error_id += 1
                else:
                    entry['status'] = 'error'
                    entry['error_id'] = f'CLUSTER-{error_id:03d}'
                    entry['error_message'] = 'No mapping found in database'
                    error_id += 1
            else:
                entry['status'] = 'error'
                entry['error_id'] = f'CLUSTER-{error_id:03d}'
                entry['error_message'] = 'Not found in mapping file'
                error_id += 1
            
            validation[excel_cluster] = entry
        
        return validation

    def validate_roles(self, ws, role_mappings):
        """Validate role mappings"""
        validation = {}
        error_id = 1
        
        # Get master row
        master_row = self.find_master_row(ws)
        role_text = str(ws.cell(row=master_row, column=4).value or '').strip()
        excel_roles = [r.strip() for r in role_text.split(',') if r.strip()]
        
        for excel_role in excel_roles:
            entry = {
                'excel_name': excel_role,
                'status': 'unmapped',
                'mapped_to': None,
                'error_id': None,
                'error_message': None
            }
            
            if excel_role in role_mappings:
                role_data = role_mappings[excel_role]
                # Handle new structure with "matched" and "db_match"
                if isinstance(role_data, dict):
                    if 'db_match' in role_data:
                        db_role = role_data['db_match']
                        match_status = role_data.get('matched', 'no')
                    elif 'id' in role_data:
                        # Old structure - direct dict
                        db_role = role_data
                        match_status = 'yes'
                    else:
                        db_role = None
                        match_status = 'no'
                else:
                    db_role = role_data
                    match_status = 'yes' if db_role else 'no'
                
                if db_role and isinstance(db_role, dict) and 'id' in db_role:
                    entry['status'] = 'matched' if match_status == 'yes' else 'error'
                    entry['mapped_to'] = {'id': db_role['id'], 'name': db_role['name']}
                    
                    # Check for potential issues
                    issues = self.check_role_issues(excel_role, db_role)
                    if issues:
                        entry['status'] = 'error'
                        entry['error_id'] = f'ROLE-{error_id:03d}'
                        entry['error_message'] = issues
                        error_id += 1
                else:
                    entry['status'] = 'unmapped'
                    entry['error_message'] = 'No mapping found in database (marked as pending)'
            else:
                entry['status'] = 'error'
                entry['error_id'] = f'ROLE-{error_id:03d}'
                entry['error_message'] = 'Not found in mapping file'
                error_id += 1
            
            validation[excel_role] = entry
        
        return validation

    def validate_pathways(self, ws, pathway_mappings):
        """Validate pathway mappings"""
        validation = {}
        error_id = 1
        
        # Get master row
        master_row = self.find_master_row(ws)
        pathway_text = str(ws.cell(row=master_row, column=5).value or '').strip()
        excel_pathways = [p.strip() for p in pathway_text.split(',') if p.strip()]
        
        for excel_pathway in excel_pathways:
            entry = {
                'excel_name': excel_pathway,
                'status': 'matched',
                'mapped_to': None,
                'error_id': None,
                'error_message': None
            }
            
            if excel_pathway in pathway_mappings:
                pathway_data = pathway_mappings[excel_pathway]
                # Handle new structure with "matched" and "db_match"
                if isinstance(pathway_data, dict):
                    if 'db_match' in pathway_data:
                        db_pathway = pathway_data['db_match']
                        match_status = pathway_data.get('matched', 'no')
                    elif 'id' in pathway_data:
                        # Old structure - direct dict
                        db_pathway = pathway_data
                        match_status = 'yes'
                    else:
                        db_pathway = None
                        match_status = 'no'
                else:
                    db_pathway = pathway_data
                    match_status = 'yes' if db_pathway else 'no'
                
                if db_pathway and isinstance(db_pathway, dict) and 'id' in db_pathway:
                    entry['status'] = 'matched' if match_status == 'yes' else 'error'
                    entry['mapped_to'] = {'id': db_pathway['id'], 'name': db_pathway['name']}
                    
                    # Check for potential issues (warnings, not errors)
                    issues = self.check_pathway_issues(excel_pathway, db_pathway)
                    if issues:
                        # These are warnings, not errors - generic names mapping to specific programs is acceptable
                        entry['status'] = 'matched'  # Keep as matched, but note the warning
                        entry['warning'] = issues
                        # Don't increment error_id for warnings
                else:
                    entry['status'] = 'error'
                    entry['error_id'] = f'PATHWAY-{error_id:03d}'
                    entry['error_message'] = 'No mapping found in database'
                    error_id += 1
            else:
                entry['status'] = 'error'
                entry['error_id'] = f'PATHWAY-{error_id:03d}'
                entry['error_message'] = 'Not found in mapping file'
                error_id += 1
            
            validation[excel_pathway] = entry
        
        return validation

    def validate_master_mapping(self, ws, master_mappings):
        """Validate master mapping against Excel"""
        validation = {}
        error_id = 1
        
        for row_idx in range(2, ws.max_row + 1):
            aptitude_code = ws.cell(row=row_idx, column=1).value
            if not aptitude_code:
                continue
            
            aptitude_code = str(aptitude_code).strip()
            
            if aptitude_code not in master_mappings:
                validation[aptitude_code] = {
                    'excel_code': aptitude_code,
                    'status': 'error',
                    'error_id': f'MASTER-{error_id:03d}',
                    'error_message': 'Not found in master mapping'
                }
                error_id += 1
                continue
            
            excel_areas = str(ws.cell(row=row_idx, column=2).value or '').strip()
            excel_clusters = str(ws.cell(row=row_idx, column=3).value or '').strip()
            excel_roles = str(ws.cell(row=row_idx, column=4).value or '').strip()
            excel_pathways = str(ws.cell(row=row_idx, column=5).value or '').strip()
            
            master_entry = master_mappings[aptitude_code]
            
            entry = {
                'excel_code': aptitude_code,
                'excel_areas': excel_areas,
                'excel_clusters': excel_clusters,
                'excel_roles': excel_roles,
                'excel_pathways': excel_pathways,
                'status': 'matched',
                'master_areas': master_entry['Areas'],
                'master_clusters_count': len(master_entry['Career_Clusters']),
                'master_roles_count': len(master_entry['Career_Roles']),
                'master_pathways_count': len(master_entry['Educational_Pathways']),
                'errors': []
            }
            
            # Check for issues
            if excel_areas != master_entry['Areas']:
                entry['errors'].append(f'AREAS-MISMATCH: Excel="{excel_areas}" vs Master="{master_entry["Areas"]}"')
            
            # Count expected items from Excel
            expected_clusters = len([c for c in excel_clusters.split(',') if c.strip()])
            expected_roles = len([r for r in excel_roles.split(',') if r.strip()])
            expected_pathways = len([p for p in excel_pathways.split(',') if p.strip()])
            
            # Note: Counts may differ due to deduplication (multiple Excel values mapping to same DB entity)
            # Only flag as error if master has 0 when Excel has values, or if count is significantly off
            if expected_clusters > 0 and entry['master_clusters_count'] == 0:
                entry['errors'].append(f'CLUSTER-COUNT-ERROR: Expected at least 1, got 0')
            elif expected_clusters > 0 and entry['master_clusters_count'] < expected_clusters * 0.5:
                # Flag if less than 50% of expected (might indicate mapping issues)
                entry['errors'].append(f'CLUSTER-COUNT-WARNING: Expected {expected_clusters}, got {entry["master_clusters_count"]} (may be due to deduplication)')
            
            if expected_roles > 0 and entry['master_roles_count'] == 0:
                entry['errors'].append(f'ROLE-COUNT-ERROR: Expected at least 1, got 0')
            elif expected_roles > 0 and entry['master_roles_count'] < expected_roles * 0.5:
                entry['errors'].append(f'ROLE-COUNT-WARNING: Expected {expected_roles}, got {entry["master_roles_count"]} (may be due to deduplication or unmapped roles)')
            
            if expected_pathways > 0 and entry['master_pathways_count'] == 0:
                entry['errors'].append(f'PATHWAY-COUNT-ERROR: Expected at least 1, got 0')
            elif expected_pathways > 0 and entry['master_pathways_count'] < expected_pathways * 0.5:
                entry['errors'].append(f'PATHWAY-COUNT-WARNING: Expected {expected_pathways}, got {entry["master_pathways_count"]} (may be due to deduplication)')
            
            if entry['errors']:
                entry['status'] = 'error'
                entry['error_id'] = f'MASTER-{error_id:03d}'
                error_id += 1
            
            validation[aptitude_code] = entry
        
        return validation

    def check_cluster_issues(self, excel_name, db_clusters):
        """Check for issues in cluster mapping"""
        issues = []
        
        # Check if Excel name is too generic but mapped to specific cluster
        if ',' not in excel_name and len(db_clusters) > 1:
            issues.append(f'Generic Excel name "{excel_name}" mapped to {len(db_clusters)} clusters')
        
        return '; '.join(issues) if issues else None

    def check_role_issues(self, excel_name, db_role):
        """Check for issues in role mapping"""
        issues = []
        
        excel_lower = excel_name.lower()
        db_name_lower = db_role['name'].lower()
        
        # Check for wrong matches
        if 'ux' in excel_lower and 'vfx' in db_name_lower and 'ux' not in db_name_lower:
            issues.append(f'ERROR: UX Designer matched to VFX Designer (wrong match)')
        
        if 'ux' in excel_lower and 'ui' not in db_name_lower and 'ux' not in db_name_lower:
            issues.append(f'POTENTIAL: UX-related role may not match correctly')
        
        # Check if names are very different
        excel_words = set(excel_name.lower().split())
        db_words = set(db_role['name'].lower().split())
        common_words = {'and', 'the', 'a', 'an', 'of', 'in', 'for', 'with'}
        excel_words = excel_words - common_words
        db_words = db_words - common_words
        
        if excel_words and db_words:
            overlap = excel_words & db_words
            if len(overlap) == 0 and len(excel_words) > 1:
                issues.append(f'POTENTIAL: Excel name "{excel_name}" and DB name "{db_role["name"]}" have no common words')
        
        return '; '.join(issues) if issues else None

    def check_pathway_issues(self, excel_name, db_pathway):
        """Check for issues in pathway mapping"""
        issues = []
        
        excel_lower = excel_name.lower()
        db_name_lower = db_pathway['name'].lower()
        
        # Check if Excel abbreviation is too generic but mapped to specific specialization
        if '(' in excel_name or excel_name in ['B.Des', 'B.Tech', 'B.Com', 'B.Sc Math']:
            # Check if DB name is too specific (has specialization)
            if 'in ' in db_name_lower or 'specialisation' in db_name_lower or 'specialization' in db_name_lower:
                issues.append(f'POTENTIAL: Generic Excel name "{excel_name}" mapped to specific specialization "{db_pathway["name"]}"')
        
        return '; '.join(issues) if issues else None

    def find_master_row(self, ws):
        """Find the master row (AR+NR+LR+LVR+CR+MR+SR)"""
        for row_idx in range(2, ws.max_row + 1):
            code = ws.cell(row=row_idx, column=1).value
            if code and str(code).strip() == "AR+NR+LR+LVR+CR+MR+SR":
                return row_idx
        return None

    def print_summary(self, report):
        """Print validation summary"""
        summary = report['summary']
        
        self.stdout.write('\n' + '=' * 60)
        self.stdout.write('VALIDATION SUMMARY')
        self.stdout.write('=' * 60)
        
        self.stdout.write(f'\nClusters:')
        self.stdout.write(f'  Total: {summary["total_clusters"]}')
        self.stdout.write(f'  Matched: {summary["matched_clusters"]}')
        self.stdout.write(f'  Errors: {summary["error_clusters"]}')
        
        self.stdout.write(f'\nRoles:')
        self.stdout.write(f'  Total: {summary["total_roles"]}')
        self.stdout.write(f'  Matched: {summary["matched_roles"]}')
        self.stdout.write(f'  Unmapped (pending): {summary["unmapped_roles"]}')
        self.stdout.write(f'  Errors: {summary["error_roles"]}')
        
        self.stdout.write(f'\nPathways:')
        self.stdout.write(f'  Total: {summary["total_pathways"]}')
        self.stdout.write(f'  Matched: {summary["matched_pathways"]}')
        self.stdout.write(f'  Errors: {summary["error_pathways"]}')
        
        # Count master mapping errors
        master_errors = sum(1 for v in report['master_mapping_validation'].values() if v.get('status') == 'error')
        self.stdout.write(f'\nMaster Mapping:')
        self.stdout.write(f'  Total combinations: {len(report["master_mapping_validation"])}')
        self.stdout.write(f'  Errors: {master_errors}')
        
        if summary["error_clusters"] > 0 or summary["error_roles"] > 0 or summary["error_pathways"] > 0 or master_errors > 0:
            self.stdout.write(self.style.WARNING(f'\n⚠ Found {summary["error_clusters"] + summary["error_roles"] + summary["error_pathways"] + master_errors} errors - check validation report'))
        else:
            self.stdout.write(self.style.SUCCESS('\n✓ All validations passed!'))
