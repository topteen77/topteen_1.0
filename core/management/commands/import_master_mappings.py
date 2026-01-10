"""
Django management command to import master row data from Excel into mapping tables.
This creates the initial mapping records that can then be edited in admin.
"""

import json
from pathlib import Path
from django.core.management.base import BaseCommand
from django.db import transaction
from app_post_matric.models import ClusterMapping, RoleMapping, PathwayMapping
from careers.models import CareerCluster, Career
from courses.models import Course


class Command(BaseCommand):
    help = 'Import master row data from Excel into mapping tables'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel-file',
            type=str,
            default='/home/itpc6/Public/django/git-repo/7nov/topteenhtml/SMART_ALIGNED_CAREER_SHEET_FILLED.xlsx',
            help='Path to Excel file'
        )
        parser.add_argument(
            '--mapping-file',
            type=str,
            default='static/data/combined_report_data/excel_to_db_mapping.json',
            help='Path to existing mapping JSON file (optional)'
        )
        parser.add_argument(
            '--clear-existing',
            action='store_true',
            help='Clear existing mappings before importing'
        )

    def handle(self, *args, **options):
        excel_file = Path(options['excel_file'])
        mapping_file = Path(options.get('mapping_file'))
        clear_existing = options.get('clear_existing', False)
        
        self.stdout.write(self.style.SUCCESS('Starting master mapping import...'))
        
        # Clear existing if requested
        if clear_existing:
            self.stdout.write('\nClearing existing mappings...')
            ClusterMapping.objects.all().delete()
            RoleMapping.objects.all().delete()
            PathwayMapping.objects.all().delete()
            self.stdout.write('  ✓ Cleared existing mappings')
        
        # Read Excel file
        self.stdout.write(f'\nReading Excel file: {excel_file}')
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Find master row
        master_row = None
        master_aptitude_code = "AR+NR+LR+LVR+CR+MR+SR"
        
        for row_idx in range(2, ws.max_row + 1):
            aptitude_code = ws.cell(row=row_idx, column=1).value
            if aptitude_code and str(aptitude_code).strip() == master_aptitude_code:
                master_row = row_idx
                break
        
        if not master_row:
            self.stdout.write(self.style.ERROR(f'  ✗ Master row with "{master_aptitude_code}" not found!'))
            return
        
        self.stdout.write(f'  ✓ Found master row at row {master_row}')
        
        # Extract data from master row
        cluster_text = str(ws.cell(row=master_row, column=3).value or '').strip()
        role_text = str(ws.cell(row=master_row, column=4).value or '').strip()
        pathway_text = str(ws.cell(row=master_row, column=5).value or '').strip()
        
        # Split by comma
        excel_clusters = [c.strip() for c in cluster_text.split(',') if c.strip()]
        excel_roles = [r.strip() for r in role_text.split(',') if r.strip()]
        excel_pathways = [p.strip() for p in pathway_text.split(',') if p.strip()]
        
        self.stdout.write(f'\nExtracted from master row:')
        self.stdout.write(f'  - {len(excel_clusters)} clusters')
        self.stdout.write(f'  - {len(excel_roles)} roles')
        self.stdout.write(f'  - {len(excel_pathways)} pathways')
        
        # Load existing mapping if available
        mapping_data = None
        if mapping_file and mapping_file.exists():
            self.stdout.write(f'\nLoading existing mapping file: {mapping_file}')
            with open(mapping_file, 'r', encoding='utf-8') as f:
                mapping_data = json.load(f)
            self.stdout.write('  ✓ Mapping file loaded')
        
        # Import clusters
        self.stdout.write('\nImporting clusters...')
        cluster_count = self.import_clusters(excel_clusters, mapping_data)
        self.stdout.write(f'  ✓ Imported {cluster_count} cluster mappings')
        
        # Import roles
        self.stdout.write('Importing roles...')
        role_count = self.import_roles(excel_roles, mapping_data)
        self.stdout.write(f'  ✓ Imported {role_count} role mappings')
        
        # Import pathways
        self.stdout.write('Importing pathways...')
        pathway_count = self.import_pathways(excel_pathways, mapping_data)
        self.stdout.write(f'  ✓ Imported {pathway_count} pathway mappings')
        
        self.stdout.write(self.style.SUCCESS('\n✓ Master mapping import completed!'))
        self.stdout.write('\nNext steps:')
        self.stdout.write('  1. Go to Django admin')
        self.stdout.write('  2. Review and fix mappings in:')
        self.stdout.write('     - Cluster Mappings')
        self.stdout.write('     - Role Mappings')
        self.stdout.write('     - Pathway Mappings')
        self.stdout.write('  3. Unmapped items are highlighted in red')

    @transaction.atomic
    def import_clusters(self, excel_clusters, mapping_data):
        """Import cluster mappings"""
        count = 0
        cluster_mappings = mapping_data.get('cluster_mappings', {}) if mapping_data else {}
        
        for excel_cluster in excel_clusters:
            mapping, created = ClusterMapping.objects.get_or_create(
                excel_name=excel_cluster,
                defaults={'is_mapped': False}
            )
            
            # Try to find DB cluster from mapping data
            if excel_cluster in cluster_mappings:
                db_clusters = cluster_mappings[excel_cluster]
                if db_clusters and len(db_clusters) > 0:
                    # Use first cluster (or handle multiple if needed)
                    db_cluster_data = db_clusters[0] if isinstance(db_clusters, list) else db_clusters
                    if isinstance(db_cluster_data, dict) and 'id' in db_cluster_data:
                        try:
                            db_cluster = CareerCluster.objects.get(id=db_cluster_data['id'])
                            mapping.db_cluster = db_cluster
                            mapping.is_mapped = True
                            mapping.save()
                        except CareerCluster.DoesNotExist:
                            pass
            
            count += 1
        
        return count

    @transaction.atomic
    def import_roles(self, excel_roles, mapping_data):
        """Import role mappings"""
        count = 0
        role_mappings = mapping_data.get('role_mappings', {}) if mapping_data else {}
        
        for excel_role in excel_roles:
            mapping, created = RoleMapping.objects.get_or_create(
                excel_name=excel_role,
                defaults={'is_mapped': False}
            )
            
            # Try to find DB role from mapping data
            if excel_role in role_mappings:
                db_role_data = role_mappings[excel_role]
                if db_role_data and isinstance(db_role_data, dict) and 'id' in db_role_data:
                    try:
                        db_role = Career.objects.get(id=db_role_data['id'])
                        mapping.db_role = db_role
                        mapping.is_mapped = True
                        mapping.save()
                    except Career.DoesNotExist:
                        pass
            
            count += 1
        
        return count

    @transaction.atomic
    def import_pathways(self, excel_pathways, mapping_data):
        """Import pathway mappings"""
        count = 0
        pathway_mappings = mapping_data.get('pathway_mappings', {}) if mapping_data else {}
        
        for excel_pathway in excel_pathways:
            mapping, created = PathwayMapping.objects.get_or_create(
                excel_name=excel_pathway,
                defaults={'is_mapped': False}
            )
            
            # Try to find DB pathway from mapping data
            if excel_pathway in pathway_mappings:
                db_pathway_data = pathway_mappings[excel_pathway]
                if db_pathway_data and isinstance(db_pathway_data, dict) and 'id' in db_pathway_data:
                    try:
                        db_pathway = Course.objects.get(id=db_pathway_data['id'])
                        mapping.db_pathway = db_pathway
                        mapping.is_mapped = True
                        mapping.save()
                    except Course.DoesNotExist:
                        pass
            
            count += 1
        
        return count
