from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode
from colleges.models import ActivityLog, UserExam,UserProject,\
    EssaySubmission,UserAdditionalCourse,UserAwardsHonours,UserGrade,UserGradeScore,ResumeScheduledExam,CollegeDates
from users.models import StudentProfile,User
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl import Workbook
from django.db.models import Sum
from datetime import datetime,timedelta
from django.utils import timezone
from activity_updates.models import Activity
from core import choices
from core.models import Configuration
from datetime import datetime
import openpyxl

def get_student_workbook(id,worksheet,request):

    BG_COL =  PatternFill(start_color="ffeb9c",
                                    end_color="ffeb9c",
                                    fill_type="solid")

    BG_HEADING_COL =  PatternFill(start_color="ffffcc",
                                    end_color="ffffcc",
                                    fill_type="solid")
    
    worksheet.column_dimensions['A'].width = 17
    worksheet.column_dimensions['B'].width = 34
    worksheet.column_dimensions['C'].width = 23
    worksheet.column_dimensions['D'].width = 23
    worksheet.column_dimensions['E'].width = 23

    row_num=2
    gen_date = str(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    gen_user = request.user.name
    worksheet.cell(row=1, column=1).value  = 'Genrated Date: '
    worksheet.cell(row=1, column=2).value  = gen_date
    worksheet.cell(row=1, column=1).fill = BG_COL

    worksheet.cell(row=1, column=3).value  = "Genrated By: "
    worksheet.cell(row=1, column=4).value  = gen_user
    worksheet.cell(row=1, column=3).fill = BG_COL

    img = openpyxl.drawing.image.Image('static/images/logo.jpg')
    img.width = 150
    img.height = 50
    # img.anchor(worksheet.cell(row=1, column=7))
    img.format.lower()
    worksheet.add_image(img,'G3')

    worksheet.cell(row=row_num, column=1).value  = 'Student Details'
    worksheet.cell(row=row_num, column=1).fill = BG_HEADING_COL
    print("=============================")
    print(id)
    print(User.objects.get(id=id).name)
    worksheet.title = User.objects.get(id=id).name
   
    # Define the titles for columns
    columns = ['Student Name','Student School', 'Current Grade']
    col_num = 2
    for row_num, column_title in enumerate(columns, 2):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        row_num+=1

    col_num += 1
    student =StudentProfile.objects.get(student=id)
    row=[student.student.name,student.school,student.grade]
   
    for row_num, cell_value in enumerate(row, 2):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        row_num +=1
    row_num+=1
    worksheet.cell(row=row_num, column=1).value  = 'Resume Builder'
    worksheet.cell(row=row_num, column=1).fill = BG_HEADING_COL

    h1 = ('Standardised Exams', 'Name', 'Total Score')
    h2 = ('Projects', 'Name', 'Completion Date')
    h3 = ('Awards & Honours ', 'Name', 'Rank')
    h4 = ('Courses', 'Name', 'Completion Date')
    h5 = ('Scheduled Exams', 'Name', 'Exam Date')
    h6 = ('Scores', 'Grade', 'Score')

    user_exams = UserExam.objects.filter(user=id)
    student_exam = []
    for user_exam in user_exams:
        # print(user_exam.exam.name)
        for score in user_exam.scores.all():
            # print(user_exam.exam.name,':',user_exam.marks_scored)
            a = (user_exam.exam.name,user_exam.marks_scored)
            student_exam.append(a)
            break;


    student_projects = UserProject.objects.filter(user=id)
    projects=[]
    if student_projects:
        for project in student_projects:
            a = (project.project_name,str(project.end_date))
            projects.append(a)

    student_award_honours = UserAwardsHonours.objects.filter(user=id)
    awards=[]
    if student_award_honours:
        for award in student_award_honours:
            a = (award.name,award.rank)
            awards.append(a)
    
    student_additional_course = UserAdditionalCourse.objects.filter(user = id)
    courses=[]
    if student_additional_course:
        for course in student_additional_course:
            a = (course.name,str(course.end_date))
            courses.append(a)
    
    student_scheduled_exams = ResumeScheduledExam.objects.filter(user = id)
    scheduled_exams=[]
    if student_scheduled_exams:
        for exam in student_scheduled_exams:
            a = (exam.exam.name,str(exam.scheduled_date))
            scheduled_exams.append(a)
    #9th
    user_grade_9th = UserGrade.objects.filter(user=id,grade=1).first()
    #10th
    user_grade_10th = UserGrade.objects.filter(user=id,grade=2).first()
    #11th
    user_grade_11th = UserGrade.objects.filter(user=id,grade=3).first()
    #12th
    user_grade_12th = UserGrade.objects.filter(user=id,grade=4).first()

    user_grades=[user_grade_9th, user_grade_10th, user_grade_11th, user_grade_12th ]
    grades=[]
    for grade in user_grades:
        if grade:
            user_grade_score = UserGradeScore.objects.filter(user_grade = grade)
            user_grade_score_count = user_grade_score.count()
            user_grade_score_total = user_grade_score.aggregate(Sum('score'))
            # score = '{}/{}'.format(user_grade_score_count,user_grade_score_total['score__sum'])
            if user_grade_score_total['score__sum']:
                score = user_grade_score_total['score__sum'] / user_grade_score_count
            else:
                score = 'Not Entered'
            a = ('Grade{}-{}'.format(grade.get_grade_display(),grade.get_board_display()),score)
            grades.append(a)


    headings = [h1,h2,h3,h4,h5,h6]
    obj_lists= [student_exam,projects,awards,courses,scheduled_exams,grades]
    

    row_num += 1
    for heading,obj_list in zip(headings,obj_lists):
        for col_num, column_title in enumerate(heading, 2):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = BG_COL
        row_num +=1
        
        if obj_list:
            for row_num, obj in enumerate(obj_list,row_num ):
                cell = worksheet.cell(row=row_num, column=3)
                cell.value = obj[0]
                cell = worksheet.cell(row=row_num, column=4)
                cell.value = obj[1]
        else:
            cell = worksheet.cell(row=row_num, column=3)
            cell.value = 'NA'
            cell = worksheet.cell(row=row_num, column=4)
            cell.value = 'NA'
        row_num +=2
    
    worksheet.cell(row=row_num, column=1).value = 'University'
    worksheet.cell(row=row_num, column=1).fill = BG_HEADING_COL
    row_num +=1
    uni_heading = ['Shortlisted List','Essay Name','Status','Application Deadline']
    for col_num, column_title in enumerate(uni_heading, 2):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = BG_COL
    row_num +=1
    
    shortlisted_colleges = student.student.college_shortlists.all()
    if shortlisted_colleges:
        for student_college in shortlisted_colleges:
            cell = worksheet.cell(row=row_num, column=2)
            cell.value = student_college.college.name
            cell = worksheet.cell(row=row_num, column=5)
            college_date = CollegeDates.objects.filter(college =  student_college.college.id)
            if college_date:
                date_list = []
                for m in college_date:
                    a = m.date_type
                    dateType = choices.CollegeDateType.CHOICES[a-1][1]
                    date_list.append(dateType)
                    date_list.append(" :")
                    date_list.append(m.date)
                    date_list.append(", ")
                convertList = ' '.join([str(e) for e in date_list])
                cell.value = convertList
            else:
                cell.value = "No Date"
            essays=student_college.college.essays.all()
            for row_num, essay in enumerate(essays,row_num ):
                cell = worksheet.cell(row=row_num, column=3)
                cell.value = essay.header
                cell = worksheet.cell(row=row_num, column=4)
                if EssaySubmission.objects.filter(user=student.student,essay = essay).exists():
                    status = EssaySubmission.objects.get(user=student.student,essay = essay)
                    cell.value = status.get_status_display()
                else:
                    cell.value = 'NA'
            row_num +=1
    else:
        for col_num in range(2,6):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = 'NA'

    return worksheet


def get_inactive_students(advisor_user_ids):
    threshold_date= timezone.now() - timedelta(days=int(Configuration.get("INACTIVE_USER_DAYS",30)))
    return advisor_user_ids.exclude(student__last_login__gte = threshold_date).values_list('student__name','student__last_login')
