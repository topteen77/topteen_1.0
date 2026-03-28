"""
Analytics Dashboard Views for Business Owner, Accounts, and Web Owner.
Provides comprehensive analytics reports and visualizations.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, Count, Avg, Q, F, OuterRef, Subquery, Max
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.utils.html import format_html
from django.middleware.csrf import get_token
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.conf import settings
from datetime import datetime, timedelta
from decimal import Decimal
from urllib.parse import unquote, urlencode
import os
import re
import json
import logging

logger = logging.getLogger(__name__)

from notifications.services import clear_service_monitor_tail_logs, get_runtime_service_status
from topteens.email_logging import format_ts_for_display, get_email_send_log_path, load_email_log_entries_newest_first

from django.contrib.contenttypes.models import ContentType
from user_analytics.models import (
    UserActivity,
    Lead,
    UserEvent,
    UserJourney,
    AnalyticsCache,
    EnquirySource,
    ChatbotPageRule,
)
# GA4Session imported conditionally in functions that need it
from user_analytics.ga4_service import GA4Service
from django.contrib import messages
from django.core.mail import EmailMultiAlternatives

from user_analytics.tasks import (
    track_page_view_sync,
    update_user_journey_sync,
    safe_track_user_event,
    send_daily_new_user_report,
)
from users.models import User
from payments.models import Payment
from psychometric_tests.models import PsychometricTestPayment
from skilllab.models import SkillLabCourse, SkilllabCoursePayment
from core import choices
from core.models import Configuration


def _normalize_daily_report_hhmm(value):
    """Return 'HH:MM' (24h) from Configuration string."""
    try:
        s = (value or '').strip()
        h_s, m_s = s.split(':', 1)
        h, m = int(h_s), int(m_s)
        if 0 <= h <= 23 and 0 <= m <= 59:
            return f'{h:02d}:{m:02d}'
    except Exception:
        pass
    return '15:00'


def _email_log_rows_for_template(raw_entries, max_rows=10, subject_max=120, error_max=300):
    """Build table rows for email JSONL audit (service monitor preview vs full page)."""
    if max_rows is not None:
        slice_list = raw_entries[:max_rows]
    else:
        slice_list = list(raw_entries)
    rows = []
    for e in slice_list:
        err = e.get('error') or ''
        rows.append(
            {
                'subject': (e.get('subject') or '')[:subject_max],
                'from_display': (e.get('from_email') or '').strip() or '—',
                'to_display': ', '.join(e.get('to') or []) or '—',
                'when_display': format_ts_for_display(e.get('ts')),
                'status': e.get('status') or '',
                'error': (
                    (err[:error_max] + ('…' if len(err) > error_max else ''))
                    if err
                    else '—'
                ),
            }
        )
    return rows


def is_staff_or_superuser(user):
    """Check if user is staff or superuser"""
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def is_superuser_or_accounts_staff(user):
    """
    Superuser or staff in Django group 'Accounts' or 'Accounts staff'.
    Used for per-row 'Update payment' on the business payments report (manual Razorpay completion).
    Create one of these groups in Django Admin and assign accounts team users.
    """
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    if not user.is_staff:
        return False
    return user.groups.filter(
        Q(name__iexact='Accounts') | Q(name__iexact='Accounts staff')
    ).exists()


def get_date_range_from_period(time_period, default_days=30):
    """
    Helper function to calculate date range from time period string.
    Returns (start_date, end_date) tuple. For 'alltime', returns (None, None).
    
    Args:
        time_period: One of 'today', 'yesterday', '7days', '30days', '90days', 'alltime'
        default_days: Default number of days if period is invalid (default: 30)
    
    Returns:
        tuple: (start_date, end_date) - both datetime objects, or (None, None) for 'alltime'
    """
    end_date = timezone.now()
    
    if time_period == 'today':
        start_date = end_date.replace(hour=0, minute=0, second=0, microsecond=0)
    elif time_period == 'yesterday':
        start_date = (end_date - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_date + timedelta(days=1)
    elif time_period == '7days':
        start_date = end_date - timedelta(days=7)
    elif time_period == '30days':
        start_date = end_date - timedelta(days=30)
    elif time_period == '90days':
        start_date = end_date - timedelta(days=90)
    elif time_period == 'alltime':
        # For all time, return None to indicate no date filtering
        return (None, None)
    else:
        # Default to specified number of days
        start_date = end_date - timedelta(days=default_days)
    
    return (start_date, end_date)


def get_date_range_from_request(request, period_param='period', date_from_param='date_from', date_to_param='date_to'):
    """
    Get (start_date, end_date) from request: use custom date_from/date_to if both provided,
    otherwise use period (e.g. 30days). Returns (start_date, end_date, time_period_used).
    """
    date_from_str = request.GET.get(date_from_param, '').strip()
    date_to_str = request.GET.get(date_to_param, '').strip()
    time_period = request.GET.get(period_param, '30days')
    if date_from_str and date_to_str:
        try:
            start_date = timezone.make_aware(datetime.strptime(date_from_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0))
            end_date = timezone.make_aware(datetime.strptime(date_to_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999))
            if start_date <= end_date:
                return (start_date, end_date, 'custom')
        except ValueError:
            pass
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    return (start_date, end_date, time_period)


def _payment_amount_rupees(payment):
    """Gateway Payment.amount is stored in whole rupees (BaseMoneyModel)."""
    return float(Decimal(payment.amount or 0))


def _payment_amount_rupees_from_event(ev, gateway_payment=None):
    """
    INR amount for display: prefer Payment row when resolved; else load the
    UserEvent's related purchase (PsychometricTestPayment / SkilllabCoursePayment / Payment).
    UserEvent.event_value can be stale or wrong (e.g. async task stored paise/decimal incorrectly);
    domain models use integer whole rupees (BaseMoneyModel).
    """
    if gateway_payment is not None:
        return _payment_amount_rupees(gateway_payment)
    oid = getattr(ev, 'object_id', None)
    ct = getattr(ev, 'content_type', None)
    if oid and ct:
        try:
            model = ct.model_class()
            if model is PsychometricTestPayment:
                row = PsychometricTestPayment.objects.filter(pk=oid).only('amount').first()
                if row is not None and row.amount is not None:
                    return float(row.amount)
            elif model is SkilllabCoursePayment:
                row = SkilllabCoursePayment.objects.filter(pk=oid).only('amount').first()
                if row is not None and row.amount is not None:
                    return float(row.amount)
            elif model is Payment:
                row = Payment.objects.filter(pk=oid).only('amount').first()
                if row is not None and row.amount is not None:
                    return float(row.amount)
        except Exception:
            pass
    return float(ev.event_value or 0)


def _order_amount_rupees_for_checkout_event(ev, gateway_payment=None):
    """
    Order/checkout amount (INR) for pending, failed, or cancelled flows.
    Tries linked models first, then metadata, then Payment lookup by gateway order id.
    """
    meta = getattr(ev, 'metadata', None) or {}
    v = _payment_amount_rupees_from_event(ev, gateway_payment)
    if v and float(v) > 0:
        return float(v)
    for key in ('order_amount_rupees', 'amount_rupees'):
        x = meta.get(key)
        if x is not None and str(x).strip() != '':
            try:
                return float(x)
            except (TypeError, ValueError):
                pass
    ap = meta.get('amount_paise')
    if ap is not None:
        try:
            return float(ap) / 100.0
        except (TypeError, ValueError):
            pass
    oid = (meta.get('gateway_order_id') or meta.get('order_id') or '').strip()
    if oid:
        p = Payment.objects.filter(gateway_order_id=oid).only('amount').first()
        if p is not None and p.amount is not None:
            return float(p.amount)
    return float(ev.event_value or 0)


def _payment_row_hide_manual_update_resolved_success(gp, gateway_order_id, payment_status_filter):
    """
    Stale analytics: UserEvent may still say failed while Payment was completed later (manual reconcile).
    If the Payment row is successful, hide manual update and show Success in the table.
    """
    if payment_status_filter not in ('fail', 'error', 'inprocess'):
        return False, None
    if gp and gp.is_success == choices.YesNoChoices.YES:
        return True, 'Success'
    oid = (gateway_order_id or '').strip()
    if oid:
        p = Payment.objects.filter(gateway_order_id=oid).only('id', 'is_success').first()
        if p and p.is_success == choices.YesNoChoices.YES:
            return True, 'Success'
    return False, None


def _resolve_gateway_payments_for_userevents(events):
    """
    Map UserEvent.id -> related Payment (with invoice prefetched) when the event
    points at Payment, PsychometricTestPayment, or SkilllabCoursePayment.
    """
    from skilllab.models import SkilllabCoursePayment

    if not events:
        return {}
    ct_payment = ContentType.objects.get_for_model(Payment)
    ct_psych = ContentType.objects.get_for_model(PsychometricTestPayment)
    ct_skill = ContentType.objects.get_for_model(SkilllabCoursePayment)

    direct_ids = []
    psych_obj_ids = []
    skill_obj_ids = []
    for ev in events:
        if ev.content_type_id == ct_payment.id and ev.object_id:
            direct_ids.append(ev.object_id)
        elif ev.content_type_id == ct_psych.id and ev.object_id:
            psych_obj_ids.append(ev.object_id)
        elif ev.content_type_id == ct_skill.id and ev.object_id:
            skill_obj_ids.append(ev.object_id)

    by_pk = {}
    if direct_ids:
        for p in Payment.objects.filter(pk__in=set(direct_ids)).select_related('invoice'):
            by_pk[p.id] = p

    psych_by_obj_id = {}
    if psych_obj_ids:
        for p in Payment.objects.filter(
            obj_type=choices.PaymentObjectType.PYSCHOMETRICTESTDETAIL,
            obj_id__in=set(psych_obj_ids),
        ).select_related('invoice'):
            psych_by_obj_id[p.obj_id] = p

    skill_by_obj_id = {}
    if skill_obj_ids:
        for p in Payment.objects.filter(
            obj_type=choices.PaymentObjectType.SKILLLABCOURSE,
            obj_id__in=set(skill_obj_ids),
        ).select_related('invoice'):
            skill_by_obj_id[p.obj_id] = p

    out = {}
    for ev in events:
        p = None
        if ev.content_type_id == ct_payment.id and ev.object_id:
            p = by_pk.get(ev.object_id)
        elif ev.content_type_id == ct_psych.id and ev.object_id:
            p = psych_by_obj_id.get(ev.object_id)
        elif ev.content_type_id == ct_skill.id and ev.object_id:
            p = skill_by_obj_id.get(ev.object_id)
        out[ev.id] = p
    return out


def is_superuser_only(user):
    """Check if user is superuser only"""
    return user.is_authenticated and user.is_superuser


@login_required
@user_passes_test(is_superuser_only)
def admin_dashboard(request):
    """
    Main Admin Dashboard - Connects all major admin areas.
    Only accessible to superusers.
    """
    time_period = request.GET.get('period', '30days')
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Quick Stats - Filtered by period
    total_users_query = User.objects.all()
    if start_date is not None:
        total_users_query = total_users_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    total_users = total_users_query.count()
    
    active_users_query = User.objects.filter(is_active=True)
    if start_date is not None:
        active_users_query = active_users_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    active_users = active_users_query.count()
    
    total_payments_query = Payment.objects.all()
    if start_date is not None:
        total_payments_query = total_payments_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    total_payments = total_payments_query.count()
    
    successful_payments_query = Payment.objects.filter(is_success=choices.YesNoChoices.YES)
    if start_date is not None:
        successful_payments_query = successful_payments_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    successful_payments = successful_payments_query.count()
    
    # Recent Activity - Filtered by period
    recent_registrations_query = User.objects.all()
    if start_date is not None:
        recent_registrations_query = recent_registrations_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    recent_registrations = recent_registrations_query.order_by('-created')[:10]
    
    recent_payments_query = Payment.objects.all()
    if start_date is not None:
        recent_payments_query = recent_payments_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    recent_payments = recent_payments_query.order_by('-created')[:10]
    
    # Analytics Summary
    total_page_views_query = UserActivity.objects.all()
    if start_date is not None:
        total_page_views_query = total_page_views_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    total_page_views = total_page_views_query.count()
    
    total_sessions_query = UserActivity.objects.all()
    if start_date is not None:
        total_sessions_query = total_sessions_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    total_sessions = total_sessions_query.values('session_id').distinct().count()
    
    total_revenue_query = UserEvent.objects.filter(event_type='payment_success')
    if start_date is not None:
        total_revenue_query = total_revenue_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    total_revenue = total_revenue_query.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    
    total_leads_query = Lead.objects.all()
    if start_date is not None:
        total_leads_query = total_leads_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    total_leads = total_leads_query.count()
    
    context = {
        'time_period': time_period,
        'start_date': start_date,  # Can be None for alltime
        'end_date': end_date,  # Can be None for alltime
        'total_users': total_users,
        'active_users': active_users,
        'total_payments': total_payments,
        'successful_payments': successful_payments,
        'total_page_views': total_page_views,
        'total_sessions': total_sessions,
        'total_revenue': total_revenue,
        'total_leads': total_leads,
        'recent_registrations': recent_registrations,
        'recent_payments': recent_payments,
    }
    
    return render(request, 'user_analytics/admin_dashboard.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def dashboard(request):
    """Main analytics dashboard - redirects to business dashboard"""
    return business_dashboard(request)


@login_required
@user_passes_test(is_staff_or_superuser)
def business_dashboard(request):
    """
    Business Owner Dashboard
    Shows revenue, payments, enrollments, and conversion metrics.
    """
    time_period = request.GET.get('period', '30days')
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Revenue Metrics - Check UserEvent first, then fallback to Payment model
    revenue_events = UserEvent.objects.filter(event_type='payment_success')
    if start_date is not None:
        revenue_events = revenue_events.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    total_revenue = revenue_events.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    successful_payments_from_events = revenue_events.count()
    
    # Fallback to Payment model if UserEvent has no data
    if total_revenue == 0 or successful_payments_from_events == 0:
        try:
            from payments.models import Payment
            from core import choices
            
            # Get successful payments from Payment model
            successful_payments_model = Payment.objects.filter(
                is_success=choices.YesNoChoices.YES
            )
            if start_date is not None:
                successful_payments_model = successful_payments_model.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            
            # Calculate revenue from Payment model
            payment_revenue = successful_payments_model.aggregate(total=Sum('amount'))['total'] or Decimal('0')
            
            # Use Payment model data if it has more revenue
            if payment_revenue > total_revenue:
                total_revenue = payment_revenue
                successful_payments = successful_payments_model.count()
                logger.info(f"Using Payment model data: {successful_payments} payments, {total_revenue} revenue")
            else:
                successful_payments = successful_payments_from_events
        except Exception as e:
            logger.warning(f"Error fetching Payment model data: {e}")
            successful_payments = successful_payments_from_events
    else:
        successful_payments = successful_payments_from_events
    
    # Failed Payments - Calculate count and attempted revenue
    failed_payment_events = UserEvent.objects.filter(event_type='payment_failed')
    if start_date is not None:
        failed_payment_events = failed_payment_events.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    failed_payments = failed_payment_events.count()
    failed_payments_revenue = failed_payment_events.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    
    # Fallback to Payment model for failed payments if UserEvent has no data
    if failed_payments == 0:
        try:
            from payments.models import Payment
            from core import choices
            
            failed_payments_model = Payment.objects.filter(
                is_success=choices.YesNoChoices.NO
            )
            if start_date is not None:
                failed_payments_model = failed_payments_model.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            failed_payments = failed_payments_model.count()
            failed_payments_revenue = failed_payments_model.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        except Exception:
            pass
    
    # Pending Payments - Calculate count and potential revenue
    pending_payment_events = UserEvent.objects.filter(event_type='payment_pending')
    if start_date is not None:
        pending_payment_events = pending_payment_events.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    pending_payments = pending_payment_events.count()
    pending_payments_revenue = pending_payment_events.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    
    # Enrollment Metrics
    enrollment_events = UserEvent.objects.filter(
        event_type__in=['course_enrolled', 'skilllab_enrolled', 'psychometric_test_completed', 'institute_student_registered']
    )
    if start_date is not None:
        enrollment_events = enrollment_events.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    total_enrollments = enrollment_events.count()
    
    psychometric_query = UserEvent.objects.filter(event_type='psychometric_test_completed')
    course_query = UserEvent.objects.filter(event_type__in=['course_enrolled', 'skilllab_enrolled'])
    
    if start_date is not None:
        psychometric_query = psychometric_query.filter(created__gte=start_date, created__lte=end_date)
        course_query = course_query.filter(created__gte=start_date, created__lte=end_date)
    
    psychometric_enrollments = psychometric_query.count()
    course_enrollments = course_query.count()
    
    # Psychometric Test Revenue Breakdown
    # Class 12 = Career Direction = ADVANCED test
    class12_query = UserEvent.objects.filter(event_type='payment_success').filter(
        Q(metadata__test_name='Career Direction') |
        Q(metadata__test_type='Advanced test') |
        Q(event_name__icontains='Career Direction')
    )
    if start_date is not None:
        class12_query = class12_query.filter(created__gte=start_date, created__lte=end_date)
    
    class12_psychometric_revenue = class12_query.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    class12_psychometric_count = class12_query.count()
    
    # Stream Sorter (Class 10-11) = BASIC test
    stream_sorter_query = UserEvent.objects.filter(event_type='payment_success').filter(
        Q(metadata__test_name='Stream Sorter') |
        Q(event_name__icontains='Stream Sorter')
    )
    if start_date is not None:
        stream_sorter_query = stream_sorter_query.filter(created__gte=start_date, created__lte=end_date)
    
    stream_sorter_revenue = stream_sorter_query.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    stream_sorter_count = stream_sorter_query.count()
    
    # Conversion Funnel
    visitors_query = UserActivity.objects.all()
    registrations_query = UserEvent.objects.filter(event_type='registration')
    
    if start_date is not None:
        visitors_query = visitors_query.filter(created__gte=start_date, created__lte=end_date)
        registrations_query = registrations_query.filter(created__gte=start_date, created__lte=end_date)
    
    total_visitors = visitors_query.values('session_id').distinct().count()
    total_registrations = registrations_query.count()
    
    from core import choices
    
    total_leads_query = Lead.objects.all()
    converted_leads_query = Lead.objects.filter(is_converted=True)
    
    if start_date is not None:
        total_leads_query = total_leads_query.filter(
            first_visit__gte=start_date,
            first_visit__lte=end_date
        )
        converted_leads_query = converted_leads_query.filter(
            converted_at__gte=start_date,
            converted_at__lte=end_date
        )
    
    total_leads = total_leads_query.count()
    converted_leads = converted_leads_query.count()
    
    # Revenue by Source - Query JSONField properly
    revenue_by_source_raw = UserEvent.objects.filter(event_type='payment_success').exclude(metadata__isnull=True)
    if start_date is not None:
        revenue_by_source_raw = revenue_by_source_raw.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    
    # Process revenue by source manually since JSONField queries can be tricky
    revenue_by_source_dict = {}
    for event in revenue_by_source_raw:
        source = event.metadata.get('source', 'Unknown') if event.metadata else 'Unknown'
        if source not in revenue_by_source_dict:
            revenue_by_source_dict[source] = {'revenue': Decimal('0'), 'count': 0}
        revenue_by_source_dict[source]['revenue'] += event.event_value or Decimal('0')
        revenue_by_source_dict[source]['count'] += 1
    
    # If no revenue by source from UserEvent, try Payment model by obj_type
    if not revenue_by_source_dict or sum(data['revenue'] for data in revenue_by_source_dict.values()) == 0:
        try:
            from payments.models import Payment
            from core import choices
            
            payment_revenue_query = Payment.objects.filter(is_success=choices.YesNoChoices.YES)
            if start_date is not None:
                payment_revenue_query = payment_revenue_query.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            
            # Group by obj_type (payment type)
            for payment in payment_revenue_query:
                obj_type_name = dict(choices.PaymentObjectType.CHOICES).get(payment.obj_type, f'Type {payment.obj_type}')
                if obj_type_name not in revenue_by_source_dict:
                    revenue_by_source_dict[obj_type_name] = {'revenue': Decimal('0'), 'count': 0}
                revenue_by_source_dict[obj_type_name]['revenue'] += payment.amount or Decimal('0')
                revenue_by_source_dict[obj_type_name]['count'] += 1
        except Exception as e:
            logger.warning(f"Error fetching revenue by source from Payment model: {e}")
    
    # Convert to list and sort by revenue
    revenue_by_source = [
        {'metadata__source': source, 'revenue': float(data['revenue']), 'count': data['count']}
        for source, data in sorted(revenue_by_source_dict.items(), key=lambda x: x[1]['revenue'], reverse=True)
    ][:10]
    
    # Daily Revenue Trend - Use TruncDate for better database compatibility
    # Try UserEvent first, fallback to Payment model
    daily_revenue_query = UserEvent.objects.filter(event_type='payment_success')
    if start_date is not None:
        daily_revenue_query = daily_revenue_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    daily_revenue = daily_revenue_query.annotate(
        day=TruncDate('created')
    ).values('day').annotate(
        revenue=Sum('event_value'),
        count=Count('id')
    ).order_by('day')
    
    # Fallback to Payment model if no UserEvent data
    if not daily_revenue.exists():
        try:
            from payments.models import Payment
            from core import choices
            
            payment_query = Payment.objects.filter(is_success=choices.YesNoChoices.YES)
            if start_date is not None:
                payment_query = payment_query.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            daily_revenue = payment_query.annotate(
                day=TruncDate('created')
            ).values('day').annotate(
                revenue=Sum('amount'),
                count=Count('id')
            ).order_by('day')
        except Exception as e:
            logger.warning(f"Error fetching daily revenue from Payment model: {e}")
    
    # Top Products/Services
    top_products_query = UserEvent.objects.filter(event_type='payment_success')
    if start_date is not None:
        top_products_query = top_products_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    top_products = top_products_query.values('event_name').annotate(
        revenue=Sum('event_value'),
        count=Count('id')
    ).order_by('-revenue')[:10]
    
    # Fallback to Payment model if no UserEvent data
    if not top_products.exists():
        try:
            from payments.models import Payment
            from core import choices
            
            # Get top products by payment object type
            payment_query = Payment.objects.filter(is_success=choices.YesNoChoices.YES)
            if start_date is not None:
                payment_query = payment_query.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            top_products = payment_query.values('obj_type').annotate(
                revenue=Sum('amount'),
                count=Count('id')
            ).order_by('-revenue')[:10]
            
            # Convert obj_type to readable names
            top_products_list = []
            for item in top_products:
                obj_type = item['obj_type']
                type_name = dict(choices.PaymentObjectType.CHOICES).get(obj_type, f'Type {obj_type}')
                top_products_list.append({
                    'event_name': type_name,
                    'revenue': float(item['revenue']),
                    'count': item['count']
                })
            top_products = top_products_list
        except Exception as e:
            logger.warning(f"Error fetching top products from Payment model: {e}")
    
    # Calculate summary metrics
    total_attempts = successful_payments + failed_payments + pending_payments
    success_rate = (successful_payments / total_attempts * 100) if total_attempts > 0 else 0
    
    # Calculate total attempted revenue - use Payment model if we're using it for successful/failed
    # This ensures consistency (all from Payment model or all from UserEvent)
    total_attempted_revenue = total_revenue + failed_payments_revenue + pending_payments_revenue
    
    # If we're using Payment model for successful/failed, also calculate total attempted from Payment model
    # to ensure consistency
    try:
        from payments.models import Payment
        from core import choices
        
        # Check if we should use Payment model for total attempted
        # (if successful or failed payments came from Payment model)
        payment_model_successful = Payment.objects.filter(is_success=choices.YesNoChoices.YES)
        payment_model_failed = Payment.objects.filter(is_success=choices.YesNoChoices.NO)
        if start_date is not None:
            payment_model_successful = payment_model_successful.filter(created__gte=start_date, created__lte=end_date)
            payment_model_failed = payment_model_failed.filter(created__gte=start_date, created__lte=end_date)
        
        payment_successful_revenue = payment_model_successful.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        payment_failed_revenue = payment_model_failed.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        payment_total_attempted = payment_successful_revenue + payment_failed_revenue
        
        # If Payment model has more data, use it for total attempted
        if payment_total_attempted > total_attempted_revenue:
            total_attempted_revenue = payment_total_attempted
            logger.info(f"Using Payment model for total attempted revenue: ₹{total_attempted_revenue}")
    except Exception as e:
        logger.warning(f"Error calculating total attempted revenue from Payment model: {e}")
    
    conversion_value_rate = (total_revenue / total_attempted_revenue * 100) if total_attempted_revenue > 0 else 0
    
    # Goal-Based Analytics
    goal_events_query = UserEvent.objects.exclude(
        event_type__in=['page_view', 'payment_failed', 'payment_pending']
    )
    if start_date is not None:
        goal_events_query = goal_events_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    
    # Group goals by type
    goal_analytics = {}
    goal_types = [
        ('registration', 'User Registration'),
        ('payment_success', 'Payment Success'),
        ('psychometric_test_started', 'Psychometric Test Started'),
        ('psychometric_test_completed', 'Psychometric Test Completed'),
        ('result_generated', 'Result Generated'),
        ('course_enrolled', 'Course Enrolled'),
        ('skilllab_enrolled', 'SkillLab Course Enrolled'),
        ('institute_student_registered', 'Institute Student Registered'),
    ]
    
    for goal_type, goal_name in goal_types:
        goal_count = goal_events_query.filter(event_type=goal_type).count()
        if goal_type == 'payment_success':
            goal_value = goal_events_query.filter(event_type=goal_type).aggregate(
                total=Sum('event_value')
            )['total'] or Decimal('0')
        else:
            goal_value = Decimal('0')
        
        if goal_count > 0:
            goal_analytics[goal_type] = {
                'name': goal_name,
                'count': goal_count,
                'value': float(goal_value),
            }
    
    # Calculate goal conversion rates
    if total_visitors > 0:
        for goal_type in goal_analytics:
            goal_analytics[goal_type]['conversion_rate'] = (
                goal_analytics[goal_type]['count'] / total_visitors * 100
            )
    else:
        for goal_type in goal_analytics:
            goal_analytics[goal_type]['conversion_rate'] = 0
    
    context = {
        'time_period': time_period,
        'start_date': start_date,
        'end_date': end_date,
        'total_revenue': total_revenue,
        'successful_payments': successful_payments,
        'failed_payments': failed_payments,
        'failed_payments_revenue': failed_payments_revenue,
        'pending_payments': pending_payments,
        'pending_payments_revenue': pending_payments_revenue,
        'total_attempted_revenue': total_attempted_revenue,
        'success_rate': success_rate,
        'conversion_value_rate': conversion_value_rate,
        'total_enrollments': total_enrollments,
        'psychometric_enrollments': psychometric_enrollments,
        'course_enrollments': course_enrollments,
        'class12_psychometric_revenue': class12_psychometric_revenue,
        'class12_psychometric_count': class12_psychometric_count,
        'stream_sorter_revenue': stream_sorter_revenue,
        'stream_sorter_count': stream_sorter_count,
        'total_visitors': total_visitors,
        'total_registrations': total_registrations,
        'total_leads': total_leads,
        'converted_leads': converted_leads,
        'revenue_by_source': list(revenue_by_source),
        'daily_revenue': list(daily_revenue),
        'top_products': list(top_products),
        'goal_analytics': goal_analytics,
    }
    
    return render(request, 'user_analytics/business_dashboard.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def accounts_dashboard(request):
    """
    Accounts Dashboard
    Shows user registration, payment status, prospects, and revenue breakdown.
    """
    time_period = request.GET.get('period', '30days')
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # User Registration Trends
    user_registrations = User.objects.all()
    if start_date is not None:
        user_registrations = user_registrations.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    total_registrations = user_registrations.count()
    
    daily_registrations = user_registrations.extra(
        select={'day': 'DATE(created)'}
    ).values('day').annotate(
        count=Count('id')
    ).order_by('day')
    
    # Payment Status Breakdown
    payment_success_query = UserEvent.objects.filter(event_type='payment_success')
    payment_failed_query = UserEvent.objects.filter(event_type='payment_failed')
    payment_pending_query = UserEvent.objects.filter(event_type='payment_pending')
    
    if start_date is not None:
        payment_success_query = payment_success_query.filter(created__gte=start_date, created__lte=end_date)
        payment_failed_query = payment_failed_query.filter(created__gte=start_date, created__lte=end_date)
        payment_pending_query = payment_pending_query.filter(created__gte=start_date, created__lte=end_date)
    
    payment_status = {
        'success': payment_success_query.count(),
        'failed': payment_failed_query.count(),
        'pending': payment_pending_query.count(),
    }
    
    # Revenue Metrics - Check UserEvent first, then fallback to Payment model
    total_revenue_query = UserEvent.objects.filter(event_type='payment_success')
    if start_date is not None:
        total_revenue_query = total_revenue_query.filter(created__gte=start_date, created__lte=end_date)
    total_revenue = total_revenue_query.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    successful_payments_from_events = payment_success_query.count()
    
    # Fallback to Payment model if UserEvent has no data
    if total_revenue == 0 or successful_payments_from_events == 0:
        try:
            from payments.models import Payment
            from core import choices
            
            # Get successful payments from Payment model
            successful_payments_model = Payment.objects.filter(
                is_success=choices.YesNoChoices.YES
            )
            if start_date is not None:
                successful_payments_model = successful_payments_model.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            
            # Calculate revenue from Payment model
            payment_revenue = successful_payments_model.aggregate(total=Sum('amount'))['total'] or Decimal('0')
            payment_success_count = successful_payments_model.count()
            
            # Use Payment model data if it has more revenue
            if payment_revenue > total_revenue:
                total_revenue = payment_revenue
                payment_status['success'] = payment_success_count
                logger.info(f"Using Payment model data for accounts dashboard: {payment_success_count} payments, ₹{total_revenue} revenue")
            
            # Also update failed payments count from Payment model
            failed_payments_model = Payment.objects.filter(is_success=choices.YesNoChoices.NO)
            if start_date is not None:
                failed_payments_model = failed_payments_model.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            payment_failed_count = failed_payments_model.count()
            if payment_failed_count > payment_status['failed']:
                payment_status['failed'] = payment_failed_count
        except Exception as e:
            logger.warning(f"Error fetching Payment model data for accounts dashboard: {e}")
    
    # Prospects (Leads)
    total_prospects_query = Lead.objects.all()
    converted_prospects_query = Lead.objects.filter(is_converted=True)
    pending_prospects_query = Lead.objects.filter(is_converted=False)
    
    if start_date is not None:
        total_prospects_query = total_prospects_query.filter(first_visit__gte=start_date, first_visit__lte=end_date)
        converted_prospects_query = converted_prospects_query.filter(converted_at__gte=start_date, converted_at__lte=end_date)
        pending_prospects_query = pending_prospects_query.filter(first_visit__gte=start_date, first_visit__lte=end_date)
    
    total_prospects = total_prospects_query.count()
    converted_prospects = converted_prospects_query.count()
    pending_prospects = pending_prospects_query.count()
    
    # Revenue by Source - Query JSONField properly
    revenue_by_source_raw = UserEvent.objects.filter(event_type='payment_success').exclude(metadata__isnull=True)
    if start_date is not None:
        revenue_by_source_raw = revenue_by_source_raw.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    
    # Process revenue by source manually since JSONField queries can be tricky
    revenue_by_source_dict = {}
    for event in revenue_by_source_raw:
        source = event.metadata.get('source', 'Unknown') if event.metadata else 'Unknown'
        if source not in revenue_by_source_dict:
            revenue_by_source_dict[source] = {'revenue': Decimal('0'), 'count': 0}
        revenue_by_source_dict[source]['revenue'] += event.event_value or Decimal('0')
        revenue_by_source_dict[source]['count'] += 1
    
    # Convert to list and sort by revenue
    revenue_by_source = [
        {'metadata__source': source, 'revenue': float(data['revenue']), 'count': data['count']}
        for source, data in sorted(revenue_by_source_dict.items(), key=lambda x: x[1]['revenue'], reverse=True)
    ]
    
    # Failed Payments Analysis
    failed_payments_query = UserEvent.objects.filter(event_type='payment_failed')
    if start_date is not None:
        failed_payments_query = failed_payments_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    failed_payments = failed_payments_query.values('event_name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Pending Payments
    pending_payments_query = UserEvent.objects.filter(event_type='payment_pending')
    if start_date is not None:
        pending_payments_query = pending_payments_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    pending_payments_list = pending_payments_query.select_related('user').order_by('-created')[:20]
    
    context = {
        'time_period': time_period,
        'start_date': start_date,
        'end_date': end_date,
        'total_registrations': total_registrations,
        'daily_registrations': list(daily_registrations),
        'payment_status': payment_status,
        'total_revenue': total_revenue,
        'total_prospects': total_prospects,
        'converted_prospects': converted_prospects,
        'pending_prospects': pending_prospects,
        'revenue_by_source': list(revenue_by_source),
        'failed_payments': list(failed_payments),
        'pending_payments_list': pending_payments_list,
    }
    
    return render(request, 'user_analytics/accounts_dashboard.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def web_owner_dashboard(request):
    """
    Web Owner Dashboard
    Shows traffic sources, user journey, engagement metrics, and page analytics.
    """
    time_period = request.GET.get('period', 'today')
    
    # Ensure default is today if not provided or invalid
    valid_periods = ['today', 'yesterday', '7days', '30days', '90days', 'alltime']
    if time_period not in valid_periods:
        time_period = 'today'
    
    logger.info("=" * 80)
    logger.info(f"WEB OWNER DASHBOARD - Time Period: {time_period}")
    logger.info("=" * 80)
    
    # Initialize GA4 Service
    ga4_service = GA4Service()
    logger.debug("GA4 Service initialized")
    
    # Get GA4 metrics - optimized for performance
    # Load critical data first, then optional data can be loaded via AJAX
    logger.debug("Fetching GA4 metrics (optimized for performance)...")
    try:
        # Critical data - load immediately (core metrics)
        user_metrics = ga4_service.get_user_metrics(time_period, use_cache=True)
        device_breakdown = ga4_service.get_device_breakdown(time_period, use_cache=True)
        top_pages = ga4_service.get_top_pages(time_period, limit=20, use_cache=True)
        traffic_sources = ga4_service.get_traffic_sources(time_period, limit=15, use_cache=True)
        
        # Real-time data - load only if period is 'today' (most relevant)
        # For other periods, skip real-time to improve performance
        if time_period == 'today':
            real_time_users = ga4_service.get_real_time_users()
            real_time_breakdown = ga4_service.get_real_time_users_breakdown()
            real_time_users_by_country = ga4_service.get_real_time_users_by_country()
        else:
            real_time_users = None
            real_time_breakdown = None
            real_time_users_by_country = []
        
        # Optional data - can be loaded via AJAX if needed (disabled for initial load)
        # These are less critical and can be loaded on-demand
        top_pages_with_trends = []  # Disabled - makes multiple API calls, very slow
        engagement = None  # Can be loaded via AJAX if needed
        users_by_country = []  # Can be loaded via AJAX if needed
        ga4_entry_pages = []  # Can be loaded via AJAX if needed
        ga4_exit_pages = []  # Can be loaded via AJAX if needed
    except Exception as e:
        logger.error(f"Error fetching GA4 data: {e}", exc_info=True)
        # Set defaults to None/empty lists on error
        user_metrics = None
        device_breakdown = None
        top_pages = []
        top_pages_with_trends = []
        traffic_sources = None
        engagement = None
        real_time_users = None
        real_time_breakdown = None
        real_time_users_by_country = []
        users_by_country = []
        ga4_entry_pages = []
        ga4_exit_pages = []
    
    # Calculate existing vs organic users from database (if available)
    # This is a fallback - GA4 doesn't directly tell us registered vs anonymous
    # We can infer from our database tracking
    # Only calculate if needed (optimize for performance)
    existing_users_count = 0
    organic_users_count = 0
    # Skip this expensive query for now - can be loaded via AJAX if needed
    # if UserActivity.objects.exists():
    #     # Count unique registered users in last hour
    #     recent_cutoff = timezone.now() - timedelta(hours=1)
    #     existing_users_count = UserActivity.objects.filter(
    #         created__gte=recent_cutoff,
    #         user__isnull=False
    #     ).values('user').distinct().count()
    #     
    #     # Count unique anonymous sessions in last hour
    #     organic_users_count = UserActivity.objects.filter(
    #         created__gte=recent_cutoff,
    #         user__isnull=True
    #     ).values('session_id').distinct().count()
    
    logger.info(f"GA4 Data Retrieved (Optimized):")
    logger.info(f"  - User Metrics: {'Available' if user_metrics else 'Not Available'}")
    logger.info(f"  - Device Breakdown: {'Available' if device_breakdown else 'Not Available'}")
    logger.info(f"  - Top Pages: {len(top_pages) if top_pages else 0} pages")
    logger.info(f"  - Traffic Sources: {'Available' if traffic_sources else 'Not Available'}")
    if time_period == 'today':
        logger.info(f"  - Real-time Users: {real_time_users or 0}")
        if real_time_breakdown:
            logger.info(f"    - New Users: {real_time_breakdown.get('new', 0)}")
            logger.info(f"    - Returning Users: {real_time_breakdown.get('returning', 0)}")
    else:
        logger.info(f"  - Real-time data: Skipped (not needed for {time_period})")
    logger.info(f"  - Optional data (Entry/Exit pages, Countries, Trends): Loaded via AJAX if needed")
    
    # Calculate date range for database queries
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    if start_date is not None:
        logger.info(f"Date Range: {start_date.strftime('%Y-%m-%d %H:%M:%S')} to {end_date.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"Duration: {(end_date - start_date).days} days")
    else:
        logger.info("Date Range: All Time (no date filtering)")
    
    # User Journey Metrics - Only calculate if needed (optimize for performance)
    # These queries can be expensive, so we'll use GA4 data primarily
    total_sessions = 0
    converted_sessions = 0
    avg_session_duration = 0
    
    # Only query database if GA4 sessions data is not available
    if not user_metrics or not user_metrics.get('sessions'):
        journey_query = UserJourney.objects.all()
        if start_date is not None:
            journey_query = journey_query.filter(
                start_time__gte=start_date,
                start_time__lte=end_date
            )
        
        total_sessions = journey_query.count()
        converted_sessions = journey_query.filter(converted=True).count()
        avg_session_duration = journey_query.aggregate(avg=Avg('total_time'))['avg'] or 0
    else:
        # Use GA4 data for sessions
        total_sessions = sum(user_metrics.get('sessions', [])) if user_metrics else 0
    
    logger.info(f"Database User Journey Metrics:")
    logger.info(f"  - Total Sessions: {total_sessions}")
    logger.info(f"  - Converted Sessions: {converted_sessions}")
    logger.info(f"  - Avg Session Duration: {avg_session_duration:.2f} seconds ({avg_session_duration/60:.2f} minutes)")
    
    # Use GA4 data for all sections (skip database)
    # Top Entry Pages - Use GA4
    top_entry_pages = ga4_entry_pages if ga4_entry_pages else []
    
    # Top Exit Pages - Use GA4
    top_exit_pages = ga4_exit_pages if ga4_exit_pages else []
    
    # Traffic Sources - Convert GA4 format to match template expectations
    ga4_traffic_sources_list = []
    if traffic_sources and 'sources' in traffic_sources and 'sessions' in traffic_sources:
        for i, source in enumerate(traffic_sources['sources']):
            ga4_traffic_sources_list.append({
                'utm_source': source or 'Direct',
                'sessions': traffic_sources['sessions'][i] if i < len(traffic_sources['sessions']) else 0,
                'pageviews': 0  # GA4 doesn't provide pageviews per source in this call
            })
    
    # Device Breakdown - Convert GA4 format to match template expectations
    ga4_device_breakdown_list = []
    if device_breakdown and 'devices' in device_breakdown and 'users' in device_breakdown:
        for i, device in enumerate(device_breakdown['devices']):
            ga4_device_breakdown_list.append({
                'device_type': device or 'Unknown',
                'count': device_breakdown['users'][i] if i < len(device_breakdown['users']) else 0
            })
    
    logger.info(f"GA4 Data Converted:")
    logger.info(f"  - Traffic Sources: {len(ga4_traffic_sources_list)} sources")
    logger.info(f"  - Device Breakdown: {len(ga4_device_breakdown_list)} device types")
    logger.info(f"  - Entry Pages: {len(top_entry_pages)} pages")
    logger.info(f"  - Exit Pages: {len(top_exit_pages)} pages")
    
    # Top Pages - Use database as primary source (faster than GA4 API)
    # Database data is more reliable and always up-to-date
    # Only query if we need it (if GA4 top_pages is empty)
    db_top_pages = []
    if not top_pages or len(top_pages) == 0:
        db_top_pages_query = UserActivity.objects.exclude(page_path__isnull=True).exclude(page_path='')
        if start_date is not None:
            db_top_pages_query = db_top_pages_query.filter(
                created__gte=start_date,
                created__lte=end_date
            )
        db_top_pages_raw = db_top_pages_query.values('page_path', 'page_title').annotate(
            pageviews=Count('id')
        ).order_by('-pageviews')[:20]
    
        # Convert to same format as GA4 top_pages
        db_top_pages = [
            {
                'path': item.get('page_path', 'N/A'),
                'title': item.get('page_title') or item.get('page_path', 'Unknown'),
                'pageviews': item.get('pageviews', 0)
            }
            for item in db_top_pages_raw
        ]
    
    # Use database as primary source (if available), fallback to GA4
    # Database data is more reliable and always up-to-date, but GA4 has broader coverage
    if db_top_pages:
        # Database has data - use it
        final_top_pages = db_top_pages
    elif top_pages:
        # No database data, but GA4 has data - use GA4
        final_top_pages = top_pages
    else:
        # No data from either source
        final_top_pages = []
    
    # Calculate total pageviews from database only if GA4 is unavailable (lazy evaluation)
    db_total_pageviews = 0
    db_total_users = 0
    if not user_metrics or not user_metrics.get('screenPageViews'):
        # Only query database if GA4 data is not available
        db_total_pageviews_query = UserActivity.objects.all()
        if start_date is not None:
            db_total_pageviews_query = db_total_pageviews_query.filter(
                created__gte=start_date,
                created__lte=end_date
            )
        db_total_pageviews = db_total_pageviews_query.count()
    
    if not user_metrics or not user_metrics.get('activeUsers'):
        # Only query database if GA4 data is not available
        db_total_users_query = User.objects.all()
        if start_date is not None:
            db_total_users_query = db_total_users_query.filter(
                created__gte=start_date,
                created__lte=end_date
            )
        db_total_users = db_total_users_query.count()
    
    logger.info(f"Database Summary Metrics:")
    logger.info(f"  - Total Users: {db_total_users}")
    logger.info(f"  - Total Pageviews: {db_total_pageviews}")
    logger.info(f"  - Top Pages (DB): {len(db_top_pages)} pages")
    
    # Summary metrics with proper fallbacks
    # Use GA4 data if available, otherwise fall back to database
    if user_metrics and 'activeUsers' in user_metrics and user_metrics['activeUsers']:
        total_users = sum(user_metrics['activeUsers'])
        users_source = "GA4"
    else:
        total_users = db_total_users
        users_source = "Database"
    
    if user_metrics and 'sessions' in user_metrics and user_metrics['sessions']:
        total_sessions_ga4 = sum(user_metrics['sessions'])
        sessions_source = "GA4"
    else:
        total_sessions_ga4 = total_sessions
        sessions_source = "Database"
    
    if user_metrics and 'screenPageViews' in user_metrics and user_metrics['screenPageViews']:
        total_pageviews = sum(user_metrics['screenPageViews'])
        pageviews_source = "GA4"
    else:
        total_pageviews = db_total_pageviews
        pageviews_source = "Database"
    
    if user_metrics and 'newUsers' in user_metrics and user_metrics['newUsers']:
        new_users = sum(user_metrics['newUsers'])
        new_users_source = "GA4"
    else:
        # For new users, we can use the same db_total_users as a fallback
        # since new users in the period would be all users created in that period
        new_users = db_total_users
        new_users_source = "Database"
    
    summary = {
        'totalUsers': total_users,
        'totalSessions': total_sessions_ga4,
        'totalPageviews': total_pageviews,
        'newUsers': new_users,
        'realTimeUsers': real_time_users or 0,
        'convertedSessions': converted_sessions,
        'avgSessionDuration': avg_session_duration or 0,
    }
    
    logger.info("=" * 80)
    logger.info("FINAL SUMMARY METRICS (What will be displayed):")
    logger.info("=" * 80)
    logger.info(f"  Total Users: {total_users} (Source: {users_source})")
    logger.info(f"  Total Sessions: {total_sessions_ga4} (Source: {sessions_source})")
    logger.info(f"  Total Pageviews: {total_pageviews} (Source: {pageviews_source})")
    logger.info(f"  New Users: {new_users} (Source: {new_users_source})")
    logger.info(f"  Real-time Users: {real_time_users or 0} (Source: GA4)")
    logger.info(f"  Converted Sessions: {converted_sessions} (Source: Database)")
    logger.info(f"  Avg Session Duration: {avg_session_duration/60:.2f} min (Source: Database)")
    logger.info(f"  Top Pages: {len(final_top_pages)} pages (Source: {'Database' if db_top_pages else 'GA4' if top_pages else 'None'})")
    
    if final_top_pages:
        logger.info("  Top 5 Pages:")
        for i, page in enumerate(final_top_pages[:5], 1):
            logger.info(f"    {i}. {page.get('title', 'Unknown')[:50]} - {page.get('pageviews', 0)} views")
    
    logger.info("=" * 80)
    
    # Check if database has any data for this time period (to determine if detail links should be clickable)
    # Only check if needed (optimize for performance)
    has_db_data = False
    if db_top_pages or total_sessions > 0:
        # If we already have database data, set flag to True
        has_db_data = True
    else:
        # Only query if we haven't already determined we have data
        has_db_data_query = UserActivity.objects.all()
        if start_date is not None:
            has_db_data_query = has_db_data_query.filter(
                created__gte=start_date,
                created__lte=end_date
            )
        has_db_data = has_db_data_query.exists()
    
    logger.info(f"Database has data for this period: {has_db_data}")
    
    context = {
        'time_period': time_period,
        'start_date': start_date,
        'end_date': end_date,
        'user_metrics': json.dumps(user_metrics) if user_metrics else None,
        'device_breakdown': json.dumps(device_breakdown) if device_breakdown else None,
        'db_device_breakdown': ga4_device_breakdown_list,  # Use GA4 data
        'top_pages': final_top_pages,
        'db_top_pages': db_top_pages,  # For template to know data source
        'traffic_sources': json.dumps(traffic_sources) if traffic_sources else None,
        'db_traffic_sources': ga4_traffic_sources_list,  # Use GA4 data
        'engagement': engagement,
        'summary': summary,
        'total_sessions': total_sessions,
        'converted_sessions': converted_sessions,
        'top_entry_pages': list(top_entry_pages) if top_entry_pages else [],
        'top_exit_pages': list(top_exit_pages) if top_exit_pages else [],
        'real_time_breakdown': real_time_breakdown,
        'real_time_users_by_country': real_time_users_by_country if real_time_users_by_country else [],
        'users_by_country': users_by_country if users_by_country else [],
        'total_country_users': sum(c.get('activeUsers', 0) for c in users_by_country) if users_by_country else 0,
        'top_pages_with_trends': top_pages_with_trends if top_pages_with_trends else [],
        'load_optional_data': False,  # Flag to indicate optional data should be loaded via AJAX
        'existing_users_count': existing_users_count,
        'organic_users_count': organic_users_count,
        'has_db_data': has_db_data,  # Flag to indicate if database has data for detail links
    }
    
    return render(request, 'user_analytics/web_owner_dashboard.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def web_owner_services_monitor(request):
    status = get_runtime_service_status()
    log_entries = load_email_log_entries_newest_first()
    email_log_preview = _email_log_rows_for_template(log_entries, max_rows=10)
    daily_report_time = _normalize_daily_report_hhmm(
        Configuration.get('DAILY_USER_REPORT_TIME', default='15:00', editable=True)
    )
    try:
        admin_configuration_url = reverse('admin:core_configuration_changelist')
    except Exception:
        admin_configuration_url = '/admin/core/configuration/'
    return render(
        request,
        'user_analytics/services_monitor.html',
        {
            'page_title': 'Service monitor',
            'status': status,
            'services': status.get('services', []),
            'logs': status.get('logs', {}),
            'all_required_ok': status.get('all_required_ok', False),
            'email_log_preview': email_log_preview,
            'email_send_log_path': get_email_send_log_path(),
            'email_backend': getattr(settings, 'EMAIL_BACKEND', ''),
            'daily_report_time': daily_report_time,
            'admin_configuration_url': admin_configuration_url,
        },
    )


@login_required
@user_passes_test(is_staff_or_superuser)
@require_POST
def web_owner_service_test_email(request):
    """Send a deliverability test to WEBADMINEMAIL, From TOPTEEN_FROM_EMAIL (same as other admin mail)."""
    web_admin = (getattr(settings, 'WEBADMINEMAIL', '') or '').strip()
    recipients = [e.strip() for e in web_admin.split(',') if e.strip()]
    if not recipients:
        messages.error(
            request,
            'WEBADMINEMAIL is not set or has no addresses. Set it in .env / settings to receive the test.',
        )
        return redirect('user_analytics:web_owner_services_monitor')

    from_email = (getattr(settings, 'TOPTEEN_FROM_EMAIL', '') or '').strip()
    if not from_email:
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', None) or 'noreply@example.com'
    subject = '[TopTeen] Test email – deliverability check'
    text_body = (
        'This is a test email from TopTeen (Service monitor).\n\n'
        'If you receive this in your inbox, SMTP is working.\n'
        'If it lands in spam, check SPF/DKIM/DMARC for your sending domain.'
    )
    html_body = (
        '<p>This is a test email from <strong>TopTeen</strong> (Service monitor).</p>'
        '<p>If you receive this in your <strong>inbox</strong>, SMTP is working.</p>'
        '<p>If it lands in <strong>spam</strong>, check SPF/DKIM/DMARC for your sending domain.</p>'
    )
    try:
        msg = EmailMultiAlternatives(
            subject=subject,
            body=text_body,
            from_email=from_email,
            to=recipients,
        )
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=False)
        messages.success(
            request,
            f'Test email sent to {", ".join(recipients)} (From: {from_email}). Check inbox and spam.',
        )
    except Exception as exc:
        logger.exception('Service monitor test email failed')
        messages.error(request, f'Failed to send test email: {exc}')
    return redirect('user_analytics:web_owner_services_monitor')


@login_required
@user_passes_test(is_staff_or_superuser)
@require_POST
def web_owner_send_daily_new_user_report(request):
    """
    Run the same logic as the scheduled daily new user report (force_send=True).
    Recipients come from WEBADMINEMAIL unless skipped.
    """
    try:
        result = send_daily_new_user_report(force_send=True, override_recipients=None)
    except Exception as exc:
        logger.exception('Manual daily new user report failed')
        messages.error(request, f'Daily new user report failed: {exc}')
        return redirect('user_analytics:web_owner_services_monitor')

    if result == 'skipped_missing_recipient':
        messages.warning(
            request,
            'WEBADMINEMAIL is not configured in settings/env. Set it to receive the daily report.',
        )
    elif result == 'skipped_invalid_recipient':
        messages.warning(
            request,
            'WEBADMINEMAIL has no valid recipient addresses after parsing.',
        )
    elif result == 'skipped_non_production':
        messages.info(request, 'Report skipped (non-production). This should not occur when forcing send.')
    elif isinstance(result, dict):
        messages.success(
            request,
            'Daily new user report sent. '
            f"Today: {result.get('today_new_users', '—')} new users; "
            f"week (Mon–today): {result.get('week_new_users', '—')}; "
            f"recipients: {result.get('recipient_count', '—')}.",
        )
    else:
        messages.info(request, f'Daily new user report result: {result!r}')
    return redirect('user_analytics:web_owner_services_monitor')


@login_required
@user_passes_test(is_staff_or_superuser)
@require_POST
def web_owner_daily_report_schedule(request):
    """Persist DAILY_USER_REPORT_TIME in core.Configuration (Celery beat reads at worker/beat startup)."""
    raw = (request.POST.get('daily_report_time') or '').strip()
    if not re.match(r'^\d{1,2}:\d{2}$', raw):
        messages.error(request, 'Use 24-hour time as HH:MM (e.g. 09:30 or 18:00).')
        return redirect('user_analytics:web_owner_services_monitor')
    try:
        parts = raw.split(':')
        h, m = int(parts[0]), int(parts[1])
    except ValueError:
        messages.error(request, 'Invalid time.')
        return redirect('user_analytics:web_owner_services_monitor')
    if not (0 <= h <= 23 and 0 <= m <= 59):
        messages.error(request, 'Hour must be 0–23 and minute 0–59.')
        return redirect('user_analytics:web_owner_services_monitor')
    normalized = f'{h:02d}:{m:02d}'
    obj, _ = Configuration.objects.get_or_create(
        key='DAILY_USER_REPORT_TIME',
        defaults={'value': normalized, 'editable': True},
    )
    obj.value = normalized
    obj.editable = True
    obj.save()
    messages.success(
        request,
        f'Daily new user report schedule saved as {normalized} IST (Asia/Kolkata). '
        'Restart the Celery beat process for the change to apply.',
    )
    return redirect('user_analytics:web_owner_services_monitor')


@login_required
@user_passes_test(is_staff_or_superuser)
@require_POST
def web_owner_clear_service_logs(request):
    """Truncate django/celery/gunicorn log files shown in Service monitor (paths under LOG_DIR or BASE_DIR/logs only)."""
    result = clear_service_monitor_tail_logs()
    cleared = result.get('cleared') or []
    errs = result.get('errors') or []
    skipped = result.get('skipped') or []
    if cleared:
        preview = ', '.join(cleared[:4])
        if len(cleared) > 4:
            preview += f' (+{len(cleared) - 4} more)'
        messages.success(
            request,
            f'Cleared {len(cleared)} service log file(s): {preview}',
        )
    elif not errs:
        messages.info(
            request,
            'No matching log files found under LOG_DIR, or paths point outside the project (e.g. /var/log). '
            + (skipped[0] if skipped else ''),
        )
    for err in errs[:8]:
        messages.warning(request, err)
    return redirect('user_analytics:web_owner_services_monitor')


@login_required
@user_passes_test(is_staff_or_superuser)
def web_owner_email_logs(request):
    """JSONL audit log from logging mail backends (logs/email_send.jsonl)."""
    entries = load_email_log_entries_newest_first()
    paginator = Paginator(entries, 50)
    page_obj = paginator.get_page(request.GET.get('page') or 1)
    rows = _email_log_rows_for_template(
        page_obj.object_list,
        max_rows=None,
        subject_max=2000,
        error_max=2000,
    )
    return render(
        request,
        'user_analytics/email_logs.html',
        {
            'page_title': 'Email logs',
            'page_obj': page_obj,
            'rows': rows,
            'log_path': get_email_send_log_path(),
        },
    )


@login_required
@user_passes_test(is_staff_or_superuser)
def user_journey_view(request, user_id=None):
    """
    User Journey Visualization
    Shows detailed journey for a specific user or all users.
    Supports filtering by user type (existing/registered vs organic/anonymous).
    """
    user_type_filter = request.GET.get('user_type', '')  # 'registered', 'organic', or ''
    goal_filter = request.GET.get('goal', '')  # 'registered', 'payment', 'test_started', 'test_completed', 'result_generated'
    enquiry_filter = (request.GET.get('enquiry_source') or '').strip()
    time_period = request.GET.get('period', '30days')
    search_query = request.GET.get('search', '')
    page_number = request.GET.get('page', 1)
    
    logger.info("=" * 80)
    logger.info(f"USER JOURNEY VIEW - User ID: {user_id}, User Type Filter: {user_type_filter}")
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Base queryset
    if user_id:
        journeys = UserJourney.objects.filter(user_id=user_id)
    else:
        journeys = UserJourney.objects.all()
    
    if start_date is not None:
        journeys = journeys.filter(
            start_time__gte=start_date,
            start_time__lte=end_date
        )
    journeys = journeys.select_related('user', 'enquiry_source').order_by('-start_time')
    
    # Apply user type filter
    if user_type_filter == 'registered':
        journeys = journeys.filter(user__isnull=False)
        logger.info("Filtering for registered users only")
    elif user_type_filter == 'organic':
        journeys = journeys.filter(user__isnull=True)
        logger.info("Filtering for organic/anonymous users only")
    
    # Apply goal filter
    if goal_filter:
        if goal_filter == 'registered':
            journeys = journeys.filter(is_registered=True)
            logger.info("Filtering for journeys with registration goal")
        elif goal_filter == 'payment':
            journeys = journeys.filter(has_payment=True)
            logger.info("Filtering for journeys with payment goal")
        elif goal_filter == 'test_started':
            journeys = journeys.filter(has_psychometric_test=True)
            logger.info("Filtering for journeys with psychometric test started goal")
        elif goal_filter == 'test_completed':
            journeys = journeys.filter(test_completed=True)
            logger.info("Filtering for journeys with test completed goal")
        elif goal_filter == 'result_generated':
            journeys = journeys.filter(result_generated=True)
            logger.info("Filtering for journeys with result generated goal")

    if enquiry_filter:
        journeys = journeys.filter(
            Q(enquiry_source__name__icontains=enquiry_filter) |
            Q(enquiry_source__token__icontains=enquiry_filter)
        )
    
    # Apply search filter
    if search_query:
        journeys = journeys.filter(
            Q(user__email__icontains=search_query) |
            Q(session_id__icontains=search_query) |
            Q(entry_page__icontains=search_query) |
            Q(exit_page__icontains=search_query)
        )
        logger.info(f"Applied search filter: {search_query}")
    
    # Calculate statistics
    total_journeys = journeys.count()
    registered_count = journeys.filter(user__isnull=False).count()
    organic_count = journeys.filter(user__isnull=True).count()
    
    logger.info(f"Journey Statistics:")
    logger.info(f"  - Total: {total_journeys}")
    logger.info(f"  - Registered Users: {registered_count}")
    logger.info(f"  - Organic Users: {organic_count}")
    
    # Pagination
    paginator = Paginator(journeys, 25)
    try:
        journeys_page = paginator.page(page_number)
    except PageNotAnInteger:
        journeys_page = paginator.page(1)
    except EmptyPage:
        journeys_page = paginator.page(paginator.num_pages)
    
    context = {
        'journeys': journeys_page,
        'user_id': user_id,
        'user_type_filter': user_type_filter,
        'goal_filter': goal_filter,
        'enquiry_filter': enquiry_filter,
        'time_period': time_period,
        'search_query': search_query,
        'total_count': total_journeys,
        'registered_count': registered_count,
        'organic_count': organic_count,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    logger.info("=" * 80)
    
    return render(request, 'user_analytics/user_journey.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def user_journey_detail_view(request, session_id):
    """
    Detailed view of a specific user journey showing all pages visited with timestamps.
    """
    try:
        # Get the journey
        journey = UserJourney.objects.select_related('user').get(session_id=session_id)
        
        # Get all activities for this session, ordered by time
        activities = UserActivity.objects.filter(session_id=session_id).order_by('created')
        
        # Define goals based on UserJourney fields
        goals_achieved = []
        
        # Goal 1: Registration
        if journey.is_registered and journey.registration_event:
            goals_achieved.append({
                'event': journey.registration_event,
                'goal': {'name': 'User Registration', 'color': 'success', 'icon': 'fa-user-plus'},
            })
        elif journey.is_registered:
            # If registered but no event, create a placeholder
            goals_achieved.append({
                'event': None,
                'event_time': journey.start_time,  # Use journey start as fallback
                'goal': {'name': 'User Registration', 'color': 'success', 'icon': 'fa-user-plus'},
            })
        
        # Goal 2: Payment
        if journey.has_payment and journey.payment_event:
            event_value = journey.payment_event.event_value or 0
            goals_achieved.append({
                'event': journey.payment_event,
                'goal': {'name': f'Payment Success - ₹{event_value}', 'color': 'success', 'icon': 'fa-check-circle'},
            })
        elif journey.has_payment:
            goals_achieved.append({
                'event': None,
                'event_time': journey.start_time,
                'goal': {'name': 'Payment Success', 'color': 'success', 'icon': 'fa-check-circle'},
            })
        
        # Goal 3: Psychometric Test Started
        if journey.has_psychometric_test and journey.psychometric_test_event:
            goals_achieved.append({
                'event': journey.psychometric_test_event,
                'goal': {'name': 'Psychometric Test Started', 'color': 'info', 'icon': 'fa-play-circle'},
            })
        elif journey.has_psychometric_test:
            goals_achieved.append({
                'event': None,
                'event_time': journey.start_time,
                'goal': {'name': 'Psychometric Test Started', 'color': 'info', 'icon': 'fa-play-circle'},
            })
        
        # Goal 4: Test Completed
        if journey.test_completed and journey.test_completion_event:
            goals_achieved.append({
                'event': journey.test_completion_event,
                'goal': {'name': 'Psychometric Test Completed', 'color': 'primary', 'icon': 'fa-check-circle'},
            })
        elif journey.test_completed:
            goals_achieved.append({
                'event': None,
                'event_time': journey.start_time,
                'goal': {'name': 'Psychometric Test Completed', 'color': 'primary', 'icon': 'fa-check-circle'},
            })
        
        # Goal 5: Result Generated
        if journey.result_generated and journey.result_generation_event:
            goals_achieved.append({
                'event': journey.result_generation_event,
                'goal': {'name': 'Result Generated', 'color': 'success', 'icon': 'fa-file-alt'},
            })
        elif journey.result_generated:
            goals_achieved.append({
                'event': None,
                'event_time': journey.start_time,
                'goal': {'name': 'Result Generated', 'color': 'success', 'icon': 'fa-file-alt'},
            })
        
        # Create a mapping of event timestamps to goals
        goal_map = {}
        for goal_data in goals_achieved:
            if goal_data['event']:
                event_time = goal_data['event'].created
            else:
                event_time = goal_data.get('event_time', journey.start_time)
            
            goal_map[event_time] = goal_data
        
        # Also check for any other events in the session that might not be linked
        events = UserEvent.objects.filter(session_id=session_id).order_by('created')
        for event in events:
            # Only add if not already in goal_map
            if event.created not in goal_map:
                goal_info = None
                if event.event_type == 'registration':
                    goal_info = {'name': 'User Registration', 'color': 'success', 'icon': 'fa-user-plus'}
                elif event.event_type == 'payment_success':
                    goal_info = {'name': f'Payment Success - ₹{event.event_value}', 'color': 'success', 'icon': 'fa-check-circle'}
                elif event.event_type == 'psychometric_test_started':
                    goal_info = {'name': 'Psychometric Test Started', 'color': 'info', 'icon': 'fa-play-circle'}
                elif event.event_type == 'psychometric_test_completed':
                    goal_info = {'name': 'Psychometric Test Completed', 'color': 'primary', 'icon': 'fa-check-circle'}
                elif event.event_type == 'result_generated':
                    goal_info = {'name': 'Result Generated', 'color': 'success', 'icon': 'fa-file-alt'}
                
                if goal_info:
                    goal_map[event.created] = {
                        'event': event,
                        'goal': goal_info,
                    }
        
        # Calculate time differences between pages and match goals
        activity_list = []
        prev_time = None
        
        for activity in activities:
            if prev_time is None:
                # First activity - calculate from journey start
                time_diff = (activity.created - journey.start_time).total_seconds()
            else:
                # Calculate from previous activity
                time_diff = (activity.created - prev_time).total_seconds()
            
            # Check if this activity corresponds to a goal (within 10 seconds tolerance)
            matched_goal = None
            matched_event_time = None
            for event_time, goal_data in goal_map.items():
                time_delta = abs((activity.created - event_time).total_seconds())
                if time_delta <= 10:  # 10 second tolerance
                    matched_goal = goal_data
                    matched_event_time = event_time
                    break
            
            # Remove matched goal from map
            if matched_event_time:
                del goal_map[matched_event_time]
            
            activity_list.append({
                'activity': activity,
                'time_since_previous': max(0, int(time_diff)),
                'timestamp': activity.created,
                'goal': matched_goal,
            })
            prev_time = activity.created
        
        # If there are unmatched goals, add them to the first activity or closest activity
        if goal_map:
            for event_time, goal_data in goal_map.items():
                # Find the closest activity to this event time
                closest_activity = None
                min_time_diff = float('inf')
                for item in activity_list:
                    time_diff = abs((item['timestamp'] - event_time).total_seconds())
                    if time_diff < min_time_diff:
                        min_time_diff = time_diff
                        closest_activity = item
                
                # If we found a close activity (within 60 seconds), add the goal
                if closest_activity and min_time_diff <= 60:
                    if not closest_activity['goal']:  # Only if no goal already assigned
                        closest_activity['goal'] = goal_data
        
        # Collect all goals achieved for summary
        all_goals = []
        if journey.is_registered:
            all_goals.append({'name': 'User Registration', 'color': 'success', 'icon': 'fa-user-plus'})
        if journey.has_payment:
            all_goals.append({'name': 'Payment Success', 'color': 'success', 'icon': 'fa-check-circle'})
        if journey.has_psychometric_test:
            all_goals.append({'name': 'Psychometric Test Started', 'color': 'info', 'icon': 'fa-play-circle'})
        if journey.test_completed:
            all_goals.append({'name': 'Psychometric Test Completed', 'color': 'primary', 'icon': 'fa-check-circle'})
        if journey.result_generated:
            all_goals.append({'name': 'Result Generated', 'color': 'success', 'icon': 'fa-file-alt'})
        
        context = {
            'journey': journey,
            'activities': activity_list,
            'total_activities': len(activity_list),
            'all_goals': all_goals,
        }
        
        return render(request, 'user_analytics/user_journey_detail.html', context)
    
    except UserJourney.DoesNotExist:
        from django.http import Http404
        raise Http404("Journey not found")


@csrf_exempt
@login_required
@user_passes_test(is_staff_or_superuser)
def api_dashboard_data(request):
    """API endpoint for AJAX dashboard data updates"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        dashboard_type = data.get('dashboard_type', 'business')
        time_period = data.get('period', '30days')
        
        # Calculate date range
        start_date, end_date = get_date_range_from_period(time_period, default_days=30)
        
        response_data = {}
        
        if dashboard_type == 'business':
            # Business metrics
            revenue_query = UserEvent.objects.filter(event_type='payment_success')
            if start_date is not None:
                revenue_query = revenue_query.filter(created__gte=start_date, created__lte=end_date)
            total_revenue = revenue_query.aggregate(total=Sum('event_value'))['total'] or 0
            
            payments_query = UserEvent.objects.filter(event_type='payment_success')
            enrollments_query = UserEvent.objects.filter(
                event_type__in=['course_enrolled', 'skilllab_enrolled', 'psychometric_test_completed']
            )
            if start_date is not None:
                payments_query = payments_query.filter(created__gte=start_date, created__lte=end_date)
                enrollments_query = enrollments_query.filter(created__gte=start_date, created__lte=end_date)
            
            response_data = {
                'total_revenue': float(total_revenue),
                'successful_payments': payments_query.count(),
                'total_enrollments': enrollments_query.count(),
            }
        
        elif dashboard_type == 'accounts':
            # Accounts metrics
            users_query = User.objects.all()
            revenue_query = UserEvent.objects.filter(event_type='payment_success')
            leads_query = Lead.objects.all()
            
            if start_date is not None:
                users_query = users_query.filter(created__gte=start_date, created__lte=end_date)
                revenue_query = revenue_query.filter(created__gte=start_date, created__lte=end_date)
                leads_query = leads_query.filter(first_visit__gte=start_date, first_visit__lte=end_date)
            
            response_data = {
                'total_registrations': users_query.count(),
                'total_revenue': float(revenue_query.aggregate(total=Sum('event_value'))['total'] or 0),
                'total_prospects': leads_query.count(),
            }
        
        elif dashboard_type == 'web_owner':
            # Web owner metrics
            ga4_service = GA4Service()
            user_metrics = ga4_service.get_user_metrics(time_period)
            real_time_users = ga4_service.get_real_time_users()
            
            response_data = {
                'totalUsers': sum(user_metrics['activeUsers']) if user_metrics else 0,
                'totalSessions': sum(user_metrics['sessions']) if user_metrics else 0,
                'totalPageviews': sum(user_metrics['screenPageViews']) if user_metrics else 0,
                'realTimeUsers': real_time_users or 0,
            }
        
        return JsonResponse(response_data)
    
    except Exception as e:
        import traceback
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)


@login_required
@user_passes_test(is_staff_or_superuser)
def successful_payments_detail(request):
    """Detail page for successful payments with filters (period or custom date range)"""
    # If AJAX request, return JSON data
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return successful_payments_api(request)
    
    search_query = request.GET.get('search', '')
    source_filter = request.GET.get('source', '').strip()
    payment_status_filter = request.GET.get('status', 'success').strip().lower()
    if payment_status_filter not in ('success', 'fail', 'error', 'inprocess'):
        payment_status_filter = 'success'
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    start_date, end_date, time_period = get_date_range_from_request(request)
    
    # Base queryset - Try UserEvent first, fallback to Payment model.
    # Source filter matches Enquiry Sources modal: session/ref attribution, not metadata.source only.
    if payment_status_filter == 'inprocess':
        pay_q = UserEvent.objects.filter(event_type='payment_pending').filter(
            Q(metadata__payment_stage='checkout_started') | Q(metadata__stage='started')
        )
        if start_date is not None:
            pay_q = pay_q.filter(created__gte=start_date, created__lte=end_date)
        if search_query:
            pay_q = pay_q.filter(
                Q(user__email__icontains=search_query) |
                Q(event_name__icontains=search_query) |
                Q(metadata__icontains=search_query)
            )
        if source_filter:
            pay_q = _apply_traffic_source_to_userevent_qs(pay_q, source_filter)
        term_q = _attribution_q_for_terminal_check(source_filter)
        candidates = list(pay_q.select_related('user', 'content_type').order_by('-created')[:2500])
        payments = [ev for ev in candidates if not _payment_has_terminal_outcome_after(term_q, ev)]
        total_count = len(payments)
        _gp_inproc = _resolve_gateway_payments_for_userevents(payments)
        total_revenue = sum(
            Decimal(str(_payment_amount_rupees_from_event(ev, _gp_inproc.get(ev.id))))
            for ev in payments
        )
    else:
        event_type = 'payment_failed' if payment_status_filter in ('fail', 'error') else 'payment_success'
        payments = UserEvent.objects.filter(event_type=event_type)
        if payment_status_filter == 'error':
            payments = payments.exclude(Q(metadata__stage='cancel') | Q(event_name__icontains='cancel'))
        if start_date is not None:
            payments = payments.filter(created__gte=start_date, created__lte=end_date)
        payments = payments.select_related('user', 'content_type').order_by('-created')
        if search_query:
            payments = payments.filter(
                Q(user__email__icontains=search_query) |
                Q(event_name__icontains=search_query) |
                Q(metadata__icontains=search_query)
            )
        if source_filter:
            payments = _apply_traffic_source_to_userevent_qs(payments, source_filter)
        if payment_status_filter in ('fail', 'error'):
            failed_candidates = list(payments.select_related('user', 'content_type').order_by('-created')[:2500])
            payments = [ev for ev in failed_candidates if not _payment_has_success_after(Q(), ev)]
            total_count = len(payments)
            _gp_fail = _resolve_gateway_payments_for_userevents(payments)
            total_revenue = sum(
                Decimal(str(_payment_amount_rupees_from_event(ev, _gp_fail.get(ev.id))))
                for ev in payments
            )
        else:
            ev_list = list(payments)
            _gp_succ = _resolve_gateway_payments_for_userevents(ev_list)
            total_count = len(ev_list)
            total_revenue = sum(
                Decimal(str(_payment_amount_rupees_from_event(ev, _gp_succ.get(ev.id))))
                for ev in ev_list
            )
    
    if total_count == 0:
        try:
            from payments.models import Payment
            from core import choices
            
            if payment_status_filter == 'inprocess':
                payment_query = Payment.objects.none()
            else:
                payment_query = Payment.objects.filter(
                    is_success=choices.YesNoChoices.NO if payment_status_filter in ('fail', 'error') else choices.YesNoChoices.YES
                )
            if start_date is not None:
                payment_query = payment_query.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            if search_query:
                payment_query = payment_query.filter(
                    Q(user__email__icontains=search_query) |
                    Q(user__name__icontains=search_query) |
                    Q(order_id__icontains=search_query)
                )
            if source_filter:
                latest_act = UserActivity.objects.filter(user_id=OuterRef('user_id')).order_by('-created')
                payment_query = payment_query.annotate(
                    _latest_src=Subquery(latest_act.values('utm_source')[:1]),
                    _latest_enq=Subquery(latest_act.values('enquiry_source__name')[:1])
                ).filter(Q(_latest_src=source_filter) | Q(_latest_enq=source_filter))
            if payment_status_filter == 'inprocess':
                total_count = total_count
                total_revenue = total_revenue
            else:
                total_count = payment_query.count()
                total_revenue = payment_query.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        except Exception as e:
            logger.warning(f"Error fetching Payment model data: {e}")
    
    # Create empty paginator for template (data loaded via AJAX)
    from django.core.paginator import Paginator, Page
    empty_list = []
    paginator = Paginator(empty_list, 25)
    payments_page = paginator.page(1)
    
    context = {
        'payments': payments_page,
        'time_period': time_period,
        'date_from': date_from,
        'date_to': date_to,
        'start_date': start_date,
        'end_date': end_date,
        'search_query': search_query,
        'source_filter': source_filter,
        'payment_status_filter': payment_status_filter,
        'total_revenue': total_revenue,
        'total_amount': total_revenue,  # For consistency
        'total_count': total_count,
        'page_title': (
            'Successful Payments' if payment_status_filter == 'success'
            else 'Failed Payments' if payment_status_filter == 'fail'
            else 'Payment Errors' if payment_status_filter == 'error'
            else 'Payments In Process'
        ),
        'payment_type': 'successful',
        'show_manual_payment_row_action': is_superuser_or_accounts_staff(request.user),
    }
    
    return render(request, 'user_analytics/payments_detail.html', context)


def _inprocess_explanation_from_metadata(meta, service_name):
    """
    Clarify ambiguous 'In process' rows: psychometric (no gateway) vs SkillLab/Razorpay (gateway pending).
    Uses metadata and service name when explicit error text is absent.
    """
    meta = meta or {}
    detail = _event_error_detail_from_metadata(meta)
    if detail:
        return detail
    test_name = (meta.get('test_name') or '').strip()
    if test_name or (service_name and 'Psychometric' in service_name):
        return (
            'Abandoned checkout — user cancelled or closed payment before completion; '
            'no gateway capture on our side yet.'
        )
    gw = (meta.get('gateway') or '').strip()
    if gw and gw.upper() != 'N/A':
        return (
            'Gateway return / verification pending — order may exist at Razorpay; '
            'confirm capture in the dashboard or use Manual complete after verification.'
        )
    return 'Checkout started but not completed (no success/fail recorded yet).'


def _payment_transaction_id_for_display(inv, gateway_payment, meta):
    """
    Invoice transaction id or gateway payment (capture) id only.
    Never use order id / gateway_order_id as a surrogate for transaction id.
    """
    meta = meta or {}
    if inv is not None:
        v = getattr(inv, 'transaction_id', None)
        if v is not None and str(v).strip():
            return str(v).strip()
    if gateway_payment is not None:
        v = getattr(gateway_payment, 'gateway_payment_id', None)
        if v is not None and str(v).strip():
            return str(v).strip()
    v = meta.get('gateway_payment_id')
    if v is not None and str(v).strip():
        return str(v).strip()
    return ''


@csrf_exempt
@login_required
@user_passes_test(is_staff_or_superuser)
def successful_payments_api(request):
    """API endpoint for payments data (AJAX). status via ?status=success|fail|error|inprocess."""
    search_query = request.GET.get('search', '').strip()
    source_filter = request.GET.get('source', '').strip()
    resolved_enquiry_source = _active_enquiry_source_by_name(source_filter) if source_filter else None
    page_number = request.GET.get('page', 1)
    start_date, end_date, _ = get_date_range_from_request(request)
    payment_status = request.GET.get('status', 'success').strip().lower()
    if payment_status not in ('success', 'fail', 'error', 'inprocess'):
        payment_status = 'success'
    status_label = {
        'success': 'Success',
        'fail': 'Fail',
        'error': 'Error',
        'inprocess': 'In Process',
    }.get(payment_status, 'Success')

    if payment_status == 'inprocess':
        pay_q = UserEvent.objects.filter(event_type='payment_pending').filter(
            Q(metadata__payment_stage='checkout_started') | Q(metadata__stage='started')
        )
        if start_date is not None:
            pay_q = pay_q.filter(created__gte=start_date, created__lte=end_date)
        if search_query:
            pay_q = pay_q.filter(
                Q(user__email__icontains=search_query) |
                Q(event_name__icontains=search_query) |
                Q(metadata__icontains=search_query)
            )
        if source_filter:
            pay_q = _apply_traffic_source_to_userevent_qs(pay_q, source_filter)
        term_q = _attribution_q_for_terminal_check(source_filter)
        candidates = list(pay_q.select_related('user', 'content_type').order_by('-created')[:2500])
        payments = [ev for ev in candidates if not _payment_has_terminal_outcome_after(term_q, ev)]
    else:
        event_type = 'payment_failed' if payment_status in ('fail', 'error') else 'payment_success'
        payments = UserEvent.objects.filter(event_type=event_type)
        if payment_status == 'error':
            payments = payments.exclude(Q(metadata__stage='cancel') | Q(event_name__icontains='cancel'))
        if start_date is not None:
            payments = payments.filter(
                created__gte=start_date,
                created__lte=end_date
            )
        payments = payments.select_related('user', 'content_type').order_by('-created')

        if search_query:
            payments = payments.filter(
                Q(user__email__icontains=search_query) |
                Q(event_name__icontains=search_query) |
                Q(metadata__icontains=search_query)
            )
        if source_filter:
            payments = _apply_traffic_source_to_userevent_qs(payments, source_filter)
        if payment_status in ('fail', 'error'):
            failed_candidates = list(payments.select_related('user', 'content_type').order_by('-created')[:2500])
            payments = [ev for ev in failed_candidates if not _payment_has_success_after(Q(), ev)]

    use_payment_model = False
    if not isinstance(payments, list) and payments.count() == 0:
        try:
            payment_query = Payment.objects.filter(
                is_success=choices.YesNoChoices.YES if payment_status == 'success' else choices.YesNoChoices.NO
            )
            if start_date is not None:
                payment_query = payment_query.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )

            if search_query:
                search_q = (
                    Q(user__email__icontains=search_query) |
                    Q(user__name__icontains=search_query) |
                    Q(gateway_order_id__icontains=search_query) |
                    Q(gateway_receipt__icontains=search_query)
                )
                if hasattr(choices, 'PaymentObjectType'):
                    for choice_value, choice_name in choices.PaymentObjectType.CHOICES:
                        if search_query.lower() in choice_name.lower() or choice_name.lower() in search_query.lower():
                            search_q |= Q(obj_type=choice_value)
                            break
                payment_query = payment_query.filter(search_q)

            latest_activity = UserActivity.objects.filter(user_id=OuterRef('user_id')).order_by('-created')
            payment_query = payment_query.annotate(
                latest_traffic_source=Subquery(latest_activity.values('utm_source')[:1]),
                latest_enquiry_source_name=Subquery(latest_activity.values('enquiry_source__name')[:1])
            )
            if source_filter:
                payment_query = payment_query.filter(
                    Q(latest_traffic_source=source_filter) | Q(latest_enquiry_source_name=source_filter)
                )
            payments = payment_query.select_related('user', 'invoice').order_by('-created')
            use_payment_model = True
        except Exception as e:
            logger.warning(f"Error fetching Payment model data: {e}")

    paginator = Paginator(payments, 25)
    try:
        payments_page = paginator.page(page_number)
    except PageNotAnInteger:
        payments_page = paginator.page(1)
    except EmptyPage:
        payments_page = paginator.page(paginator.num_pages)

    if use_payment_model:
        total_revenue = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_count = payments.count()
    else:
        ev_list = list(payments)
        gp_map = _resolve_gateway_payments_for_userevents(ev_list)
        total_revenue = Decimal('0')
        for ev in ev_list:
            gp = gp_map.get(ev.id)
            total_revenue += Decimal(str(_payment_amount_rupees_from_event(ev, gp)))
        total_count = len(ev_list)

    payments_data = []
    if use_payment_model:
        for payment in payments_page:
            obj_type_name = _payment_service_name_from_payment(payment) if payment else 'Payment'
            gateway_name = 'N/A'
            if hasattr(choices, 'GatewayChoices') and payment.gateway:
                gateway_name = dict(choices.GatewayChoices.CHOICES).get(payment.gateway, f'Gateway {payment.gateway}')
            inv = getattr(payment, 'invoice', None)
            source = (getattr(payment, 'latest_enquiry_source_name', None) or '').strip() or (getattr(payment, 'latest_traffic_source', None) or '').strip() or 'Direct'
            go = payment.gateway_order_id or ''
            gp = payment.gateway_payment_id or ''
            hide_manual_update = (
                payment_status in ('fail', 'error', 'inprocess')
                and payment.is_success == choices.YesNoChoices.YES
            )
            row_status_label = 'Success' if hide_manual_update else status_label
            payments_data.append({
                'id': payment.id,
                'payment_db_id': payment.id,
                'gateway_order_id': go,
                'gateway_payment_id': gp,
                'user_email': payment.user.email if payment.user else 'Anonymous',
                'invoice_number': getattr(inv, 'invoice_number', None) or 'N/A',
                'transaction_id': _payment_transaction_id_for_display(inv, payment, {}) or '',
                'student_id': payment.user_id or '',
                'event_name': obj_type_name,
                'amount': _payment_amount_rupees(payment),
                'date': payment.created.strftime('%b %d, %Y %H:%M') if payment.created else 'N/A',
                'source': source,
                'gateway': gateway_name,
                'order_id': go or 'N/A',
                'status_label': row_status_label,
                'hide_manual_update': hide_manual_update,
                'error_detail': '',
                'error_value': '',
            })
    else:
        gp_page = _resolve_gateway_payments_for_userevents(list(payments_page))
        for payment in payments_page:
            meta = payment.metadata or {}
            gp = gp_page.get(payment.id)
            inv = getattr(gp, 'invoice', None) if gp else None
            txn = _payment_transaction_id_for_display(inv, gp, meta)
            inv_no = inv.invoice_number if inv else 'N/A'
            amt = _payment_amount_rupees_from_event(payment, gp)
            order_id = meta.get('order_id') or meta.get('gateway_order_id')
            if gp:
                order_id = order_id or gp.gateway_order_id or 'N/A'
            else:
                order_id = order_id or 'N/A'
            go_row = (gp.gateway_order_id if gp else '') or (meta.get('gateway_order_id') or meta.get('order_id') or '')
            gp_id = (gp.gateway_payment_id if gp else '') or (meta.get('gateway_payment_id') or '')
            hide_manual_update, row_st_override = _payment_row_hide_manual_update_resolved_success(
                gp, go_row, payment_status
            )
            payment_db_id = gp.id if gp else None
            if hide_manual_update and payment_db_id is None and (go_row or '').strip():
                p_enrich = Payment.objects.filter(gateway_order_id=(go_row or '').strip()).only('id').first()
                if p_enrich:
                    payment_db_id = p_enrich.id
            row_status_label = row_st_override if row_st_override else status_label
            service_name = _payment_service_name_from_event(payment, gp)
            err_line = (
                _inprocess_explanation_from_metadata(meta, service_name)
                if payment_status == 'inprocess'
                else _event_error_detail_from_metadata(meta)
            )
            err_value = _event_error_detail_from_metadata(meta) or ''
            payments_data.append({
                'id': payment.id,
                'payment_db_id': payment_db_id,
                'gateway_order_id': go_row,
                'gateway_payment_id': gp_id,
                'user_email': payment.user.email if payment.user else 'Anonymous',
                'invoice_number': inv_no,
                'transaction_id': txn,
                'student_id': payment.user_id or '',
                'event_name': service_name,
                'amount': amt,
                'date': payment.created.strftime('%b %d, %Y %H:%M') if payment.created else 'N/A',
                'source': (meta.get('source') or '').strip() or (resolved_enquiry_source.name if resolved_enquiry_source else '') or 'Direct',
                'gateway': meta.get('gateway', 'N/A'),
                'order_id': order_id,
                'status_label': row_status_label,
                'hide_manual_update': hide_manual_update,
                'error_detail': err_line,
                'error_value': err_value,
            })

    return JsonResponse({
        'success': True,
        'payments': payments_data,
        'pagination': {
            'current_page': payments_page.number,
            'total_pages': paginator.num_pages,
            'has_previous': payments_page.has_previous(),
            'has_next': payments_page.has_next(),
            'previous_page': payments_page.previous_page_number() if payments_page.has_previous() else None,
            'next_page': payments_page.next_page_number() if payments_page.has_next() else None,
        },
        'totals': {
            'total_count': total_count,
            'total_revenue': float(total_revenue),
            'total_amount': float(total_revenue),
        }
    })


@login_required
@user_passes_test(is_staff_or_superuser)
def successful_payments_export_excel(request):
    """Export payments to Excel with same filters as the table (?status=success|fail|error|inprocess)."""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        return HttpResponse('Excel export requires openpyxl. pip install openpyxl', status=500)
    start_date, end_date, _ = get_date_range_from_request(request)
    search_query = request.GET.get('search', '').strip()
    source_filter = request.GET.get('source', '').strip()
    payment_status = request.GET.get('status', 'success').strip().lower()
    if payment_status not in ('success', 'fail', 'error', 'inprocess'):
        payment_status = 'success'
    status_label = {
        'success': 'Success',
        'fail': 'Fail',
        'error': 'Error',
        'inprocess': 'In Process',
    }.get(payment_status, 'Success')

    if payment_status == 'inprocess':
        pay_q = UserEvent.objects.filter(event_type='payment_pending').filter(
            Q(metadata__payment_stage='checkout_started') | Q(metadata__stage='started')
        )
        if start_date is not None:
            pay_q = pay_q.filter(created__gte=start_date, created__lte=end_date)
        if search_query:
            pay_q = pay_q.filter(
                Q(user__email__icontains=search_query) |
                Q(event_name__icontains=search_query) |
                Q(metadata__icontains=search_query)
            )
        if source_filter:
            pay_q = _apply_traffic_source_to_userevent_qs(pay_q, source_filter)
        term_q = _attribution_q_for_terminal_check(source_filter)
        candidates = list(pay_q.select_related('user', 'content_type').order_by('-created')[:2500])
        payments = [ev for ev in candidates if not _payment_has_terminal_outcome_after(term_q, ev)]
    else:
        event_type = 'payment_success' if payment_status == 'success' else 'payment_failed'
        payments = UserEvent.objects.filter(event_type=event_type)
        if payment_status == 'error':
            payments = payments.exclude(Q(metadata__stage='cancel') | Q(event_name__icontains='cancel'))
        if start_date is not None:
            payments = payments.filter(created__gte=start_date, created__lte=end_date)
        payments = payments.select_related('user', 'content_type').order_by('-created')
        if search_query:
            payments = payments.filter(
                Q(user__email__icontains=search_query) |
                Q(event_name__icontains=search_query) |
                Q(metadata__icontains=search_query)
            )
        if source_filter:
            payments = _apply_traffic_source_to_userevent_qs(payments, source_filter)
        if payment_status in ('fail', 'error'):
            failed_candidates = list(payments.select_related('user', 'content_type').order_by('-created')[:2500])
            payments = [ev for ev in failed_candidates if not _payment_has_success_after(Q(), ev)]
    use_payment_model = False
    if not isinstance(payments, list) and payments.count() == 0:
        try:
            payment_query = Payment.objects.filter(
                is_success=choices.YesNoChoices.YES if payment_status == 'success' else choices.YesNoChoices.NO
            )
            if start_date is not None:
                payment_query = payment_query.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            if search_query:
                search_q = (
                    Q(user__email__icontains=search_query) |
                    Q(user__name__icontains=search_query) |
                    Q(gateway_order_id__icontains=search_query) |
                    Q(gateway_receipt__icontains=search_query)
                )
                if hasattr(choices, 'PaymentObjectType'):
                    for choice_value, choice_name in choices.PaymentObjectType.CHOICES:
                        if search_query.lower() in choice_name.lower() or choice_name.lower() in search_query.lower():
                            search_q |= Q(obj_type=choice_value)
                            break
                payment_query = payment_query.filter(search_q)
            latest_activity_excel = UserActivity.objects.filter(user_id=OuterRef('user_id')).order_by('-created')
            payment_query = payment_query.annotate(
                latest_traffic_source=Subquery(latest_activity_excel.values('utm_source')[:1]),
                latest_enquiry_source_name=Subquery(latest_activity_excel.values('enquiry_source__name')[:1])
            )
            if source_filter:
                payment_query = payment_query.filter(
                    Q(latest_traffic_source=source_filter) | Q(latest_enquiry_source_name=source_filter)
                )
            payments = payment_query.select_related('user', 'invoice').order_by('-created')
            use_payment_model = True
        except Exception as e:
            logger.warning(f"Excel export Payment fallback error: {e}")

    ev_rows = list(payments) if not use_payment_model else []
    gp_excel = _resolve_gateway_payments_for_userevents(ev_rows) if ev_rows else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payments'
    headers = ['#', 'User', 'Invoice No', 'Transaction ID', 'Student ID', 'Service', 'Amount (₹)', 'Date', 'Status', 'Traffic source', 'Gateway', 'Order ID']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)
    for row_idx, payment in enumerate(payments, 2):
        if use_payment_model:
            inv = getattr(payment, 'invoice', None)
            obj_type_name = _payment_service_name_from_payment(payment) if payment else 'Payment'
            gateway_name = dict(choices.GatewayChoices.CHOICES).get(payment.gateway, str(payment.gateway)) if hasattr(choices, 'GatewayChoices') and payment.gateway else 'N/A'
            source = (getattr(payment, 'latest_enquiry_source_name', None) or '').strip() or (getattr(payment, 'latest_traffic_source', None) or '').strip() or 'Direct'
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=payment.user.email if payment.user else 'Anonymous')
            ws.cell(row=row_idx, column=3, value=getattr(inv, 'invoice_number', None) or 'N/A')
            ws.cell(row=row_idx, column=4, value=_payment_transaction_id_for_display(inv, payment, {}) or None)
            ws.cell(row=row_idx, column=5, value=payment.user_id or '')
            ws.cell(row=row_idx, column=6, value=obj_type_name)
            ws.cell(row=row_idx, column=7, value=_payment_amount_rupees(payment))
            ws.cell(row=row_idx, column=8, value=payment.created.strftime('%Y-%m-%d %H:%M') if payment.created else 'N/A')
            ws.cell(row=row_idx, column=9, value=status_label)
            ws.cell(row=row_idx, column=10, value=source)
            ws.cell(row=row_idx, column=11, value=gateway_name)
            ws.cell(row=row_idx, column=12, value=payment.gateway_order_id or 'N/A')
        else:
            gp = gp_excel.get(payment.id)
            inv = getattr(gp, 'invoice', None) if gp else None
            meta = payment.metadata or {}
            txn = _payment_transaction_id_for_display(inv, gp, meta) or None
            inv_no = inv.invoice_number if inv else 'N/A'
            amt = _payment_amount_rupees_from_event(payment, gp)
            order_id = meta.get('order_id') or meta.get('gateway_order_id')
            if gp:
                order_id = order_id or gp.gateway_order_id or 'N/A'
            else:
                order_id = order_id or 'N/A'
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=payment.user.email if payment.user else 'Anonymous')
            ws.cell(row=row_idx, column=3, value=inv_no)
            ws.cell(row=row_idx, column=4, value=txn)
            ws.cell(row=row_idx, column=5, value=payment.user_id or '')
            ws.cell(row=row_idx, column=6, value=_payment_service_name_from_event(payment, gp))
            ws.cell(row=row_idx, column=7, value=amt)
            ws.cell(row=row_idx, column=8, value=payment.created.strftime('%Y-%m-%d %H:%M') if payment.created else 'N/A')
            ws.cell(row=row_idx, column=9, value=status_label)
            ws.cell(row=row_idx, column=10, value=(meta.get('source') or '').strip() or 'Direct')
            ws.cell(row=row_idx, column=11, value=meta.get('gateway', 'N/A'))
            ws.cell(row=row_idx, column=12, value=order_id)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = 'payments_export.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _audit_manual_payment_reconciliation(request, payment, staff_note, verify_mode):
    """Staff audit trail (does not replace Payment/UserEvent success tracking from signals)."""
    content_type = ContentType.objects.get_for_model(payment)
    safe_track_user_event(
        event_type='form_submission',
        event_name='Staff manual payment completion',
        user_id=payment.user_id,
        event_value=float(payment.amount or 0),
        content_type_id=content_type.id,
        object_id=payment.id,
        session_id=getattr(request.session, 'session_key', None) or '',
        metadata={
            'reconciliation': 'staff_manual',
            'verify_mode': verify_mode,
            'staff_id': request.user.id,
            'staff_email': getattr(request.user, 'email', '') or '',
            'payment_db_id': payment.id,
            'gateway_order_id': payment.gateway_order_id or '',
            'gateway_payment_id': payment.gateway_payment_id or '',
            'note': (staff_note or '')[:500],
        },
    )


@login_required
@user_passes_test(is_staff_or_superuser)
def manual_payment_suggest_api(request):
    """AJAX typeahead for manual payment completion: search Payment by id, email, order id, pay id."""
    q = (request.GET.get('q') or '').strip()
    if len(q) < 2:
        return JsonResponse({'suggestions': []})
    filt = (
        Q(user__email__icontains=q)
        | Q(gateway_order_id__icontains=q)
        | Q(gateway_payment_id__icontains=q)
        | Q(gateway_receipt__icontains=q)
    )
    if q.isdigit():
        try:
            filt |= Q(pk=int(q))
        except ValueError:
            pass
    rows = list(
        Payment.objects.select_related('user')
        .filter(filt)
        .order_by('-created')[:15]
    )
    suggestions = []
    for p in rows:
        uid = p.user.email if p.user else str(p.user_id)
        label = '#{} — {} — ₹{} — {}'.format(p.id, uid, p.amount, p.gateway_order_id or '—')
        suggestions.append(
            {
                'payment_db_id': p.id,
                'gateway_order_id': p.gateway_order_id or '',
                'gateway_payment_id': p.gateway_payment_id or '',
                'label': label,
                'user_email': uid,
            }
        )
    return JsonResponse({'suggestions': suggestions})


@login_required
@user_passes_test(is_staff_or_superuser)
def manual_payment_reconciliation_view(request):
    """
    Staff tool: after verifying a captured payment in Razorpay Dashboard, complete our Payment row,
    allocate course / psychometric access (same as webhook + client success), and log an audit event.
    """
    from payments.models import Payment
    from payments.payment.razorpay import RazorpayService
    from payments.reconciliation import finalize_side_effects_after_gateway_success

    result_message = None
    result_level = None
    payment_preview = None

    if request.method == 'POST':
        confirm = request.POST.get('confirm_verified') == 'on'
        staff_note = (request.POST.get('staff_note') or '').strip()[:1000]
        payment_db_id = (request.POST.get('payment_db_id') or '').strip()
        gateway_order_id_in = (request.POST.get('gateway_order_id') or '').strip()
        rz_payment_id = (request.POST.get('razorpay_payment_id') or '').strip()
        rz_signature = (request.POST.get('razorpay_signature') or '').strip()
        verify_mode = (request.POST.get('verify_mode') or 'api').strip().lower()
        force_allocate = request.POST.get('force_allocate') == 'on' and request.user.is_superuser

        if not confirm:
            result_message = 'Check the confirmation box: you must have verified this payment in Razorpay (or other gateway) before completing here.'
            result_level = 'error'
        elif not payment_db_id and not gateway_order_id_in:
            result_message = 'Enter either internal Payment ID (database id) or Razorpay Order ID stored on the payment row.'
            result_level = 'error'
        elif not force_allocate and not rz_payment_id:
            result_message = 'Enter Razorpay Payment ID from the dashboard, or use superuser-only force completion (no gateway check).'
            result_level = 'error'
        else:
            payment = None
            if payment_db_id:
                try:
                    payment = Payment.objects.select_related('user').filter(pk=int(payment_db_id)).first()
                except (TypeError, ValueError):
                    payment = None
            if not payment and gateway_order_id_in:
                payment = Payment.objects.select_related('user').filter(gateway_order_id=gateway_order_id_in).first()

            if not payment:
                result_message = 'No Payment row found for that id or order id.'
                result_level = 'error'
            elif payment.is_success == choices.YesNoChoices.YES:
                result_message = 'This payment is already successful in our database. No changes made.'
                result_level = 'warning'
                payment_preview = payment
            elif force_allocate:
                if rz_payment_id:
                    payment.gateway_payment_id = rz_payment_id
                    payment.save(update_fields=['gateway_payment_id'])
                payment.is_success = choices.YesNoChoices.YES
                payment.save(update_fields=['is_success'])
                finalize_side_effects_after_gateway_success(payment)
                _audit_manual_payment_reconciliation(request, payment, staff_note, 'force_superuser')
                result_message = (
                    'Marked successful without live gateway verification. Side effects (course / test access, emails) were applied. '
                    'Use only when you have offline proof of payment.'
                )
                result_level = 'success'
                payment_preview = payment
            elif verify_mode == 'signature':
                order_for_sig = gateway_order_id_in or (payment.gateway_order_id or '')
                if not (rz_payment_id and rz_signature and order_for_sig):
                    result_message = 'Signature mode requires Razorpay Payment ID, Signature, and Order ID (form or already on the payment row).'
                    result_level = 'error'
                else:
                    ok = payment.update_payment(rz_payment_id, order_for_sig, rz_signature)
                    if ok:
                        finalize_side_effects_after_gateway_success(payment)
                        _audit_manual_payment_reconciliation(request, payment, staff_note, 'razorpay_signature')
                        result_message = 'Signature verified with Razorpay. Payment marked successful and allocation completed.'
                        result_level = 'success'
                        payment_preview = payment
                    else:
                        result_message = 'Razorpay rejected signature / payment verification. Check pasted values match the dashboard.'
                        result_level = 'error'
            else:
                payment.gateway_payment_id = rz_payment_id
                payment.save(update_fields=['gateway_payment_id'])
                rsvc = RazorpayService()
                if rsvc.verify_payment_amount_status_and_order(payment):
                    payment.is_success = choices.YesNoChoices.YES
                    payment.save(update_fields=['is_success'])
                    finalize_side_effects_after_gateway_success(payment)
                    _audit_manual_payment_reconciliation(request, payment, staff_note, 'razorpay_api')
                    result_message = 'Verified via Razorpay API (amount, status, order). Payment marked successful and allocation completed.'
                    result_level = 'success'
                    payment_preview = payment
                else:
                    result_message = (
                        'Razorpay API check failed (captured/authorized + amount + order id). '
                        'Try signature mode with values from the payment success callback in dashboard, or confirm order id on our Payment row.'
                    )
                    result_level = 'error'

    prefill = {
        'payment_db_id': (request.GET.get('payment_db_id') or '').strip(),
        'gateway_order_id': (request.GET.get('gateway_order_id') or '').strip(),
        'razorpay_payment_id': (request.GET.get('razorpay_payment_id') or '').strip(),
    }
    context = {
        'result_message': result_message,
        'result_level': result_level,
        'payment_preview': payment_preview,
        'page_title': 'Manual payment completion',
        'prefill': prefill,
        'manual_payment_suggest_url': reverse('user_analytics:manual_payment_suggest_api'),
        'csrf_input_html': format_html(
            '<input type="hidden" name="csrfmiddlewaretoken" value="{}">',
            get_token(request),
        ),
    }
    return render(request, 'user_analytics/manual_payment_reconciliation.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def failed_payments_detail(request):
    """Detail page for failed payments with filters"""
    # If AJAX request, return JSON data
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return failed_payments_api(request)
    
    time_period = request.GET.get('period', '30days')
    search_query = request.GET.get('search', '')
    page_number = request.GET.get('page', 1)
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Base queryset - Try UserEvent first, fallback to Payment model
    payments = UserEvent.objects.filter(event_type='payment_failed')
    if start_date is not None:
        payments = payments.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    payments = payments.select_related('user').order_by('-created')
    
    # Apply search filter
    if search_query:
        payments = payments.filter(
            Q(user__email__icontains=search_query) |
            Q(event_name__icontains=search_query) |
            Q(metadata__icontains=search_query)
        )
    
    # If no UserEvent data, fallback to Payment model (but API will handle the actual data)
    # For the template, we just need the count
    total_count = payments.count()
    total_amount = payments.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    
    # If no UserEvent data, check Payment model for count
    if total_count == 0:
        try:
            from payments.models import Payment
            from core import choices
            
            payment_query = Payment.objects.filter(is_success=choices.YesNoChoices.NO)
            if start_date is not None:
                payment_query = payment_query.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            if search_query:
                payment_query = payment_query.filter(
                    Q(user__email__icontains=search_query) |
                    Q(user__name__icontains=search_query) |
                    Q(gateway_order_id__icontains=search_query) |
                    Q(gateway_receipt__icontains=search_query)
                )
            total_count = payment_query.count()
            total_amount = payment_query.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        except Exception as e:
            logger.warning(f"Error fetching Payment model data: {e}")
    
    # Create empty paginator for template (data loaded via AJAX)
    from django.core.paginator import Paginator, Page
    empty_list = []
    paginator = Paginator(empty_list, 25)
    payments_page = paginator.page(1)
    
    context = {
        'payments': payments_page,
        'time_period': time_period,
        'start_date': start_date,
        'end_date': end_date,
        'search_query': search_query,
        'total_amount': total_amount,
        'total_count': total_count,
        'page_title': 'Failed Payments',
        'payment_type': 'failed',
        'show_manual_payment_row_action': False,
    }
    
    return render(request, 'user_analytics/payments_detail.html', context)


@csrf_exempt
@login_required
@user_passes_test(is_staff_or_superuser)
def failed_payments_api(request):
    """API endpoint for failed payments data (AJAX)"""
    time_period = request.GET.get('period', '30days')
    search_query = request.GET.get('search', '')
    page_number = request.GET.get('page', 1)
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Base queryset - Try UserEvent first, fallback to Payment model
    payments = UserEvent.objects.filter(event_type='payment_failed')
    if start_date is not None:
        payments = payments.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    payments = payments.select_related('user').order_by('-created')
    
    # Apply search filter
    if search_query:
        payments = payments.filter(
            Q(user__email__icontains=search_query) |
            Q(event_name__icontains=search_query) |
            Q(metadata__icontains=search_query)
        )
    
    # Check if we have data, if not fallback to Payment model
    use_payment_model = False
    if payments.count() == 0:
        try:
            from payments.models import Payment
            from core import choices
            
            payment_query = Payment.objects.filter(is_success=choices.YesNoChoices.NO)
            if start_date is not None:
                payment_query = payment_query.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            
            # Apply search filter to payments
            if search_query:
                # Build Q object for search
                search_q = Q(user__email__icontains=search_query) | \
                           Q(user__name__icontains=search_query) | \
                           Q(gateway_order_id__icontains=search_query) | \
                           Q(gateway_receipt__icontains=search_query)
                
                # Also search by PaymentObjectType display name
                if hasattr(choices, 'PaymentObjectType'):
                    # Check if search query matches any PaymentObjectType display name
                    for choice_value, choice_name in choices.PaymentObjectType.CHOICES:
                        if search_query.lower() in choice_name.lower() or choice_name.lower() in search_query.lower():
                            search_q |= Q(obj_type=choice_value)
                            break
                
                payment_query = payment_query.filter(search_q)
            
            payments = payment_query.select_related('user').order_by('-created')
            use_payment_model = True
        except Exception as e:
            logger.warning(f"Error fetching Payment model data: {e}")
    
    # Pagination
    paginator = Paginator(payments, 25)
    try:
        payments_page = paginator.page(page_number)
    except PageNotAnInteger:
        payments_page = paginator.page(1)
    except EmptyPage:
        payments_page = paginator.page(paginator.num_pages)
    
    # Calculate totals
    if use_payment_model:
        total_amount = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_count = payments.count()
    else:
        total_amount = payments.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
        total_count = payments.count()
    
    # Serialize data
    payments_data = []
    for payment in payments_page:
        if use_payment_model:
            # Payment model
            obj_type_name = 'Payment'
            if hasattr(choices, 'PaymentObjectType') and payment.obj_type:
                obj_type_name = dict(choices.PaymentObjectType.CHOICES).get(payment.obj_type, f'Type {payment.obj_type}')
            
            # Get gateway display name
            gateway_name = 'N/A'
            if hasattr(choices, 'GatewayChoices') and payment.gateway:
                gateway_name = dict(choices.GatewayChoices.CHOICES).get(payment.gateway, f'Gateway {payment.gateway}')
            
            payments_data.append({
                'id': payment.id,
                'user_email': payment.user.email if payment.user else 'Anonymous',
                'event_name': obj_type_name,
                'amount': float(payment.amount or 0),
                'date': payment.created.strftime('%b %d, %Y %H:%M') if payment.created else 'N/A',
                'source': 'Unknown',  # Payment model doesn't have source
                'gateway': gateway_name,
                'order_id': payment.gateway_order_id or 'N/A',
            })
        else:
            # UserEvent model
            payments_data.append({
                'id': payment.id,
                'user_email': payment.user.email if payment.user else 'Anonymous',
                'event_name': payment.event_name or 'N/A',
                'amount': float(payment.event_value or 0),
                'date': payment.created.strftime('%b %d, %Y %H:%M') if payment.created else 'N/A',
                'source': payment.metadata.get('source', 'Unknown') if payment.metadata else 'Unknown',
                'gateway': payment.metadata.get('gateway', 'N/A') if payment.metadata else 'N/A',
                'order_id': payment.metadata.get('order_id', 'N/A') if payment.metadata else 'N/A',
            })
    
    return JsonResponse({
        'success': True,
        'payments': payments_data,
        'pagination': {
            'current_page': payments_page.number,
            'total_pages': paginator.num_pages,
            'has_previous': payments_page.has_previous(),
            'has_next': payments_page.has_next(),
            'previous_page': payments_page.previous_page_number() if payments_page.has_previous() else None,
            'next_page': payments_page.next_page_number() if payments_page.has_next() else None,
        },
        'totals': {
            'total_count': total_count,
            'total_amount': float(total_amount),
            'total_revenue': float(total_amount),  # For consistency
        }
    })


@login_required
@user_passes_test(is_staff_or_superuser)
def pending_payments_detail(request):
    """Detail page for pending payments with filters"""
    # If AJAX request, return JSON data
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return pending_payments_api(request)
    
    time_period = request.GET.get('period', '30days')
    search_query = request.GET.get('search', '')
    page_number = request.GET.get('page', 1)
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Base queryset
    payments = UserEvent.objects.filter(event_type='payment_pending')
    if start_date is not None:
        payments = payments.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    payments = payments.select_related('user').order_by('-created')
    
    # Apply search filter
    if search_query:
        payments = payments.filter(
            Q(user__email__icontains=search_query) |
            Q(event_name__icontains=search_query) |
            Q(metadata__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(payments, 25)
    try:
        payments_page = paginator.page(page_number)
    except PageNotAnInteger:
        payments_page = paginator.page(1)
    except EmptyPage:
        payments_page = paginator.page(paginator.num_pages)
    
    # Calculate totals
    total_amount = payments.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    total_count = payments.count()
    
    context = {
        'payments': payments_page,
        'time_period': time_period,
        'start_date': start_date,
        'end_date': end_date,
        'search_query': search_query,
        'total_amount': total_amount,
        'total_count': total_count,
        'page_title': 'Pending Payments',
        'payment_type': 'pending',
        'show_manual_payment_row_action': False,
    }
    
    return render(request, 'user_analytics/payments_detail.html', context)


@csrf_exempt
@login_required
@user_passes_test(is_staff_or_superuser)
def pending_payments_api(request):
    """API endpoint for pending payments data (AJAX)"""
    time_period = request.GET.get('period', '30days')
    search_query = request.GET.get('search', '')
    page_number = request.GET.get('page', 1)
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Base queryset
    payments = UserEvent.objects.filter(event_type='payment_pending')
    if start_date is not None:
        payments = payments.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    payments = payments.select_related('user').order_by('-created')
    
    # Apply search filter
    if search_query:
        payments = payments.filter(
            Q(user__email__icontains=search_query) |
            Q(event_name__icontains=search_query) |
            Q(metadata__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(payments, 25)
    try:
        payments_page = paginator.page(page_number)
    except PageNotAnInteger:
        payments_page = paginator.page(1)
    except EmptyPage:
        payments_page = paginator.page(paginator.num_pages)
    
    # Calculate totals
    total_amount = payments.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    total_count = payments.count()
    
    # Serialize data
    payments_data = []
    for payment in payments_page:
        payments_data.append({
            'id': payment.id,
            'user_email': payment.user.email if payment.user else 'N/A',
            'event_name': payment.event_name or 'N/A',
            'amount': float(payment.event_value or 0),
            'date': payment.created.strftime('%b %d, %Y %H:%M') if payment.created else 'N/A',
            'order_id': payment.metadata.get('order_id', 'N/A') if payment.metadata else 'N/A',
        })
    
    return JsonResponse({
        'success': True,
        'payments': payments_data,
        'pagination': {
            'current_page': payments_page.number,
            'total_pages': paginator.num_pages,
            'has_previous': payments_page.has_previous(),
            'has_next': payments_page.has_next(),
            'previous_page': payments_page.previous_page_number() if payments_page.has_previous() else None,
            'next_page': payments_page.next_page_number() if payments_page.has_next() else None,
        },
        'totals': {
            'total_count': total_count,
            'total_amount': float(total_amount),
        }
    })


@login_required
@user_passes_test(is_staff_or_superuser)
def enrollments_detail(request):
    """Detail page for enrollments with filters"""
    # If AJAX request, return JSON data
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return enrollments_api(request)
    
    time_period = request.GET.get('period', '30days')
    event_type_filter = request.GET.get('event_type', '')
    search_query = request.GET.get('search', '')
    page_number = request.GET.get('page', 1)
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Base queryset
    enrollments = UserEvent.objects.filter(
        event_type__in=['course_enrolled', 'skilllab_enrolled', 'psychometric_test_completed', 'institute_student_registered']
    )
    if start_date is not None:
        enrollments = enrollments.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    enrollments = enrollments.select_related('user').order_by('-created')
    
    # Apply event type filter
    if event_type_filter:
        enrollments = enrollments.filter(event_type=event_type_filter)
    
    # Apply search filter
    if search_query:
        enrollments = enrollments.filter(
            Q(user__email__icontains=search_query) |
            Q(event_name__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(enrollments, 25)
    try:
        enrollments_page = paginator.page(page_number)
    except PageNotAnInteger:
        enrollments_page = paginator.page(1)
    except EmptyPage:
        enrollments_page = paginator.page(paginator.num_pages)
    
    # Calculate totals
    total_count = enrollments.count()
    
    # Get event type choices for filter dropdown
    event_type_choices = [
        ('', 'All Types'),
        ('course_enrolled', 'Course Enrolled'),
        ('skilllab_enrolled', 'SkillLab Enrolled'),
        ('psychometric_test_completed', 'Psychometric Test Completed'),
        ('institute_student_registered', 'Institute Student Registered'),
    ]
    
    context = {
        'enrollments': enrollments_page,
        'time_period': time_period,
        'start_date': start_date,
        'end_date': end_date,
        'search_query': search_query,
        'event_type_filter': event_type_filter,
        'event_type_choices': event_type_choices,
        'total_count': total_count,
        'page_title': 'Enrollments',
    }
    
    return render(request, 'user_analytics/enrollments_detail.html', context)


@csrf_exempt
@login_required
@user_passes_test(is_staff_or_superuser)
def business_metrics_api(request):
    """API endpoint for business dashboard key metrics (AJAX)"""
    time_period = request.GET.get('period', '30days')
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Calculate Total Revenue
    revenue_events = UserEvent.objects.filter(event_type='payment_success')
    if start_date is not None:
        revenue_events = revenue_events.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    total_revenue = revenue_events.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    successful_payments_from_events = revenue_events.count()
    
    # Fallback to Payment model if UserEvent has no data
    if total_revenue == 0 or successful_payments_from_events == 0:
        try:
            from payments.models import Payment
            from core import choices
            
            successful_payments_model = Payment.objects.filter(is_success=choices.YesNoChoices.YES)
            if start_date is not None:
                successful_payments_model = successful_payments_model.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            payment_revenue = successful_payments_model.aggregate(total=Sum('amount'))['total'] or Decimal('0')
            
            if payment_revenue > total_revenue:
                total_revenue = payment_revenue
                successful_payments = successful_payments_model.count()
            else:
                successful_payments = successful_payments_from_events
        except Exception as e:
            logger.warning(f"Error fetching Payment model data in API: {e}")
            successful_payments = successful_payments_from_events
    else:
        successful_payments = successful_payments_from_events
    
    # Calculate Total Enrollments
    enrollments_query = UserEvent.objects.filter(
        event_type__in=['course_enrolled', 'skilllab_enrolled', 'psychometric_test_completed', 'institute_student_registered']
    )
    if start_date is not None:
        enrollments_query = enrollments_query.filter(created__gte=start_date, created__lte=end_date)
    total_enrollments = enrollments_query.count()
    
    # Calculate other metrics
    failed_payment_events = UserEvent.objects.filter(event_type='payment_failed')
    if start_date is not None:
        failed_payment_events = failed_payment_events.filter(created__gte=start_date, created__lte=end_date)
    failed_payments = failed_payment_events.count()
    failed_payments_revenue = failed_payment_events.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    
    # Fallback to Payment model for failed payments
    if failed_payments == 0:
        try:
            from payments.models import Payment
            from core import choices
            
            failed_payments_model = Payment.objects.filter(is_success=choices.YesNoChoices.NO)
            if start_date is not None:
                failed_payments_model = failed_payments_model.filter(
                    created__gte=start_date,
                    created__lte=end_date
                )
            failed_payments = failed_payments_model.count()
            failed_payments_revenue = failed_payments_model.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        except Exception:
            pass
    
    pending_payment_events = UserEvent.objects.filter(event_type='payment_pending')
    if start_date is not None:
        pending_payment_events = pending_payment_events.filter(created__gte=start_date, created__lte=end_date)
    pending_payments = pending_payment_events.count()
    pending_payments_revenue = pending_payment_events.aggregate(total=Sum('event_value'))['total'] or Decimal('0')
    
    psychometric_query = UserEvent.objects.filter(event_type='psychometric_test_completed')
    course_query = UserEvent.objects.filter(event_type__in=['course_enrolled', 'skilllab_enrolled'])
    visitors_query = UserActivity.objects.all()
    registrations_query = UserEvent.objects.filter(event_type='registration')
    
    if start_date is not None:
        psychometric_query = psychometric_query.filter(created__gte=start_date, created__lte=end_date)
        course_query = course_query.filter(created__gte=start_date, created__lte=end_date)
        visitors_query = visitors_query.filter(created__gte=start_date, created__lte=end_date)
        registrations_query = registrations_query.filter(created__gte=start_date, created__lte=end_date)
    
    psychometric_enrollments = psychometric_query.count()
    course_enrollments = course_query.count()
    total_visitors = visitors_query.values('session_id').distinct().count()
    total_registrations = registrations_query.count()
    
    from core import choices
    
    total_leads_query = Lead.objects.all()
    converted_leads_query = Lead.objects.filter(is_converted=True)
    
    if start_date is not None:
        total_leads_query = total_leads_query.filter(first_visit__gte=start_date, first_visit__lte=end_date)
        converted_leads_query = converted_leads_query.filter(converted_at__gte=start_date, converted_at__lte=end_date)
    
    total_leads = total_leads_query.count()
    converted_leads = converted_leads_query.count()
    
    conversion_rate = (converted_leads / total_visitors * 100) if total_visitors > 0 else 0
    
    return JsonResponse({
        'success': True,
        'metrics': {
            'total_revenue': float(total_revenue),
            'successful_payments': successful_payments,
            'total_enrollments': total_enrollments,
            'failed_payments': failed_payments,
            'failed_payments_revenue': float(failed_payments_revenue),
            'pending_payments': pending_payments,
            'pending_payments_revenue': float(pending_payments_revenue),
            'psychometric_enrollments': psychometric_enrollments,
            'course_enrollments': course_enrollments,
            'total_visitors': total_visitors,
            'total_registrations': total_registrations,
            'total_leads': total_leads,
            'converted_leads': converted_leads,
            'conversion_rate': round(conversion_rate, 2),
        },
        'period': time_period,
        'start_date': start_date.isoformat() if start_date is not None else None,
        'end_date': end_date.isoformat() if end_date is not None else None,
    })


@csrf_exempt
@login_required
@user_passes_test(is_staff_or_superuser)
def enrollments_api(request):
    """API endpoint for enrollments data (AJAX)"""
    time_period = request.GET.get('period', '30days')
    event_type_filter = request.GET.get('event_type', '')
    search_query = request.GET.get('search', '')
    page_number = request.GET.get('page', 1)
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Base queryset
    enrollments = UserEvent.objects.filter(
        event_type__in=['course_enrolled', 'skilllab_enrolled', 'psychometric_test_completed', 'institute_student_registered']
    )
    if start_date is not None:
        enrollments = enrollments.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    enrollments = enrollments.select_related('user').order_by('-created')
    
    # Apply event type filter
    if event_type_filter:
        enrollments = enrollments.filter(event_type=event_type_filter)
    
    # Apply search filter
    if search_query:
        enrollments = enrollments.filter(
            Q(user__email__icontains=search_query) |
            Q(event_name__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(enrollments, 25)
    try:
        enrollments_page = paginator.page(page_number)
    except PageNotAnInteger:
        enrollments_page = paginator.page(1)
    except EmptyPage:
        enrollments_page = paginator.page(paginator.num_pages)
    
    # Calculate totals
    total_count = enrollments.count()
    
    # Serialize data
    enrollments_data = []
    for enrollment in enrollments_page:
        event_type_display = 'Unknown'
        if enrollment.event_type == 'course_enrolled':
            event_type_display = 'Course Enrolled'
        elif enrollment.event_type == 'skilllab_enrolled':
            event_type_display = 'SkillLab Enrolled'
        elif enrollment.event_type == 'psychometric_test_completed':
            event_type_display = 'Psychometric Test'
        elif enrollment.event_type == 'institute_student_registered':
            event_type_display = 'Institute Student'
        
        enrollments_data.append({
            'id': enrollment.id,
            'user_email': enrollment.user.email if enrollment.user else 'N/A',
            'event_type': enrollment.event_type,
            'event_type_display': event_type_display,
            'event_name': enrollment.event_name or 'N/A',
            'date': enrollment.created.strftime('%b %d, %Y %H:%M') if enrollment.created else 'N/A',
            'amount': float(enrollment.event_value or 0),
        })
    
    return JsonResponse({
        'success': True,
        'enrollments': enrollments_data,
        'pagination': {
            'current_page': enrollments_page.number,
            'total_pages': paginator.num_pages,
            'has_previous': enrollments_page.has_previous(),
            'has_next': enrollments_page.has_next(),
            'previous_page': enrollments_page.previous_page_number() if enrollments_page.has_previous() else None,
            'next_page': enrollments_page.next_page_number() if enrollments_page.has_next() else None,
        },
        'totals': {
            'total_count': total_count,
        }
    })


@login_required
@user_passes_test(is_staff_or_superuser)
def registrations_detail(request):
    """Detail page for user registrations with filters"""
    time_period = request.GET.get('period', '30days')
    search_query = request.GET.get('search', '')
    page_number = request.GET.get('page', 1)
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Base queryset
    users = User.objects.all()
    if start_date is not None:
        users = users.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    users = users.order_by('-created')
    
    # Apply search filter
    if search_query:
        users = users.filter(
            Q(email__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(mobile__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(users, 25)
    try:
        users_page = paginator.page(page_number)
    except PageNotAnInteger:
        users_page = paginator.page(1)
    except EmptyPage:
        users_page = paginator.page(paginator.num_pages)
    
    # Calculate totals
    total_count = users.count()
    
    context = {
        'users': users_page,
        'time_period': time_period,
        'start_date': start_date,
        'end_date': end_date,
        'search_query': search_query,
        'total_count': total_count,
        'page_title': 'User Registrations',
    }
    
    return render(request, 'user_analytics/registrations_detail.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def prospects_detail(request):
    """Detail page for prospects/leads with filters"""
    time_period = request.GET.get('period', '30days')
    search_query = request.GET.get('search', '')
    converted_filter = request.GET.get('converted', '')
    page_number = request.GET.get('page', 1)
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    # Base queryset
    leads = Lead.objects.all()
    if start_date is not None:
        leads = leads.filter(
            first_visit__gte=start_date,
            first_visit__lte=end_date
        )
    leads = leads.order_by('-first_visit')
    
    # Apply converted filter
    if converted_filter == 'yes':
        leads = leads.filter(is_converted=True)
    elif converted_filter == 'no':
        leads = leads.filter(is_converted=False)
    
    # Apply search filter
    if search_query:
        leads = leads.filter(
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(source__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(leads, 25)
    try:
        leads_page = paginator.page(page_number)
    except PageNotAnInteger:
        leads_page = paginator.page(1)
    except EmptyPage:
        leads_page = paginator.page(paginator.num_pages)
    
    # Calculate totals
    total_count = leads.count()
    
    context = {
        'leads': leads_page,
        'time_period': time_period,
        'start_date': start_date,
        'end_date': end_date,
        'search_query': search_query,
        'converted_filter': converted_filter,
        'total_count': total_count,
        'page_title': 'Prospects/Leads',
    }
    
    return render(request, 'user_analytics/prospects_detail.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def admin_user_analytics_view(request):
    """
    Admin User Analytics - User activities with filters, shown in the analytics dashboard.
    Filters: period, referrer source (Google, Facebook, iapply.io), device, country, search.
    """
    from user_analytics.utils import referrer_source_q
    from django.db.models import Count

    time_period = request.GET.get('period', '30days')
    source_filter = (request.GET.get('source') or '').strip()
    enquiry_filter = (request.GET.get('enquiry_source') or '').strip()
    device_filter = (request.GET.get('device') or '').strip()
    country_filter = (request.GET.get('country') or '').strip()
    traffic_category_filter = (request.GET.get('traffic_category') or '').strip()
    search_query = (request.GET.get('search') or '').strip()
    page_number = request.GET.get('page', 1)

    if country_filter:
        country_filter = unquote(country_filter)
    if source_filter:
        source_filter = unquote(source_filter)
    if enquiry_filter:
        enquiry_filter = unquote(enquiry_filter)
    if device_filter:
        device_filter = unquote(device_filter)

    start_date, end_date = get_date_range_from_period(time_period, default_days=30)

    qs = UserActivity.objects.all().select_related('user', 'enquiry_source')
    if start_date is not None:
        qs = qs.filter(created__gte=start_date, created__lte=end_date)

    if source_filter:
        if source_filter.lower() in ('(direct)', 'direct', '(not set)'):
            qs = qs.filter(
                Q(utm_source__iexact='(direct)') |
                Q(utm_source__iexact='direct') |
                Q(utm_source__iexact='(not set)') |
                Q(utm_source__isnull=True) |
                Q(utm_source='')
            )
        elif source_filter.lower() in ('google', 'facebook', 'iapply', 'iapply.io'):
            qs = qs.filter(referrer_source_q(source_filter))
        else:
            qs = qs.filter(Q(utm_source__iexact=source_filter) | Q(referrer__icontains=source_filter))
    if enquiry_filter:
        qs = qs.filter(
            Q(enquiry_source__name__icontains=enquiry_filter) |
            Q(enquiry_source__token__icontains=enquiry_filter)
        )

    if device_filter:
        qs = qs.filter(device_type__iexact=device_filter)
    if country_filter:
        qs = qs.filter(country__icontains=country_filter)
    if traffic_category_filter:
        qs = qs.filter(traffic_source_category__iexact=traffic_category_filter)

    if search_query:
        qs = qs.filter(
            Q(session_id__icontains=search_query) |
            Q(page_path__icontains=search_query) |
            Q(referrer__icontains=search_query) |
            Q(utm_source__icontains=search_query) |
            Q(enquiry_source__name__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )

    qs = qs.order_by('-created')

    # Summary stats (from same filtered queryset)
    total_activities = qs.count()
    unique_sessions = qs.values('session_id').distinct().count()

    # Counts by source (for summary cards)
    base_qs = UserActivity.objects.all()
    if start_date is not None:
        base_qs = base_qs.filter(created__gte=start_date, created__lte=end_date)
    count_google = base_qs.filter(referrer_source_q('google')).count()
    count_facebook = base_qs.filter(referrer_source_q('facebook')).count()
    count_iapply = base_qs.filter(referrer_source_q('iapply')).count()
    count_direct = base_qs.filter(
        Q(utm_source__isnull=True) | Q(utm_source='') |
        Q(utm_source__iexact='direct') | Q(utm_source__iexact='(direct)')
    ).count()

    # Counts by device
    by_device = base_qs.values('device_type').annotate(c=Count('id')).order_by('-c')
    device_counts = {r['device_type'] or 'unknown': r['c'] for r in by_device}

    paginator = Paginator(qs, 25)
    try:
        activities_page = paginator.page(page_number)
    except PageNotAnInteger:
        activities_page = paginator.page(1)
    except EmptyPage:
        activities_page = paginator.page(paginator.num_pages)

    # One-time cleanup output from POST redirect
    cleanup_output = request.session.pop('cleanup_output', None)

    context = {
        'activities': activities_page,
        'time_period': time_period,
        'source_filter': source_filter,
        'enquiry_filter': enquiry_filter,
        'device_filter': device_filter,
        'country_filter': country_filter,
        'traffic_category_filter': traffic_category_filter,
        'search_query': search_query,
        'total_activities': total_activities,
        'unique_sessions': unique_sessions,
        'count_google': count_google,
        'count_facebook': count_facebook,
        'count_iapply': count_iapply,
        'count_direct': count_direct,
        'device_counts': device_counts,
        'start_date': start_date,
        'end_date': end_date,
        'page_title': 'Admin User Analytics',
        'cleanup_output': cleanup_output,
        'cleanup_url': reverse('user_analytics:cleanup_analytics_data'),
        'csrf_input_html': format_html(
            '<input type="hidden" name="csrfmiddlewaretoken" value="{}">',
            get_token(request),
        ),
    }
    return render(request, 'user_analytics/admin_user_analytics.html', context)


def _cleanup_analytics_redirect(request):
    """After cleanup POST: stay on User Analytics page or return to Django admin app list."""
    return_to = (request.POST.get('return_to') or '').strip()
    if return_to == 'admin_app':
        return redirect('admin:app_list', app_label='user_analytics')
    return redirect('user_analytics:admin_user_analytics')


def _destructive_cleanup_enabled():
    """
    Safety lock for production: destructive cleanup must be explicitly enabled.
    Enable via Django setting ANALYTICS_CLEANUP_ALLOW_DESTRUCTIVE=True
    or env var ANALYTICS_CLEANUP_ALLOW_DESTRUCTIVE=1/true/yes/on.
    """
    flag = getattr(settings, 'ANALYTICS_CLEANUP_ALLOW_DESTRUCTIVE', False)
    if flag:
        return True
    env_val = (os.getenv('ANALYTICS_CLEANUP_ALLOW_DESTRUCTIVE') or '').strip().lower()
    return env_val in {'1', 'true', 'yes', 'on'}


@login_required
@user_passes_test(is_staff_or_superuser)
def cleanup_analytics_data_view(request):
    """
    POST-only: run analytics data cleanup (same logic as management command).
    Expects: days (int), dry_run (optional), confirm (required when not dry_run).
    For purge_all=1: deletes all UserActivity, UserJourney, UserEvent (optional dry_run).
    """
    from django.core.management import call_command
    from io import StringIO
    from django.contrib import messages

    if request.method != 'POST':
        return redirect('user_analytics:admin_user_analytics')

    dry_run = request.POST.get('dry_run') == 'on'
    confirm = request.POST.get('confirm') == 'on'
    purge_all = request.POST.get('purge_all') == '1'
    domain_cleanup = request.POST.get('domain_cleanup') == '1'
    domain = (request.POST.get('domain') or '').strip()
    destructive = not dry_run
    if destructive and not _destructive_cleanup_enabled():
        messages.error(
            request,
            'Destructive analytics cleanup is disabled. '
            'Set ANALYTICS_CLEANUP_ALLOW_DESTRUCTIVE=1 to allow this action.'
        )
        return _cleanup_analytics_redirect(request)

    if domain_cleanup:
        from user_analytics.domain_cleanup import run_domain_cleanup, VALID_DOMAINS

        if not confirm:
            messages.warning(
                request,
                'Check "I understand this will permanently delete data" to run domain cleanup.',
            )
            return _cleanup_analytics_redirect(request)
        if domain not in VALID_DOMAINS:
            messages.error(request, 'Invalid domain selection.')
            return _cleanup_analytics_redirect(request)
        out = StringIO()
        try:
            text, _ = run_domain_cleanup(domain, dry_run=dry_run)
            out.write(text)
            request.session['cleanup_output'] = out.getvalue()
            if dry_run:
                messages.info(request, 'Domain cleanup dry run completed. No data was deleted.')
            else:
                messages.success(request, 'Domain cleanup completed.')
            if request.POST.get('return_to') == 'admin_app':
                messages.info(
                    request,
                    format_html(
                        '<a href="{}">Open Admin User Analytics</a> to see the full command output.',
                        reverse('user_analytics:admin_user_analytics'),
                    ),
                )
        except Exception as e:
            logger.exception('Domain cleanup failed')
            messages.error(request, 'Domain cleanup failed: %s' % str(e))
        return _cleanup_analytics_redirect(request)

    if purge_all:
        if not confirm:
            messages.warning(
                request,
                'Check "I understand this will permanently delete data" to purge all activity, events, and journeys.',
            )
            return _cleanup_analytics_redirect(request)
        out = StringIO()
        try:
            call_command(
                'cleanup_analytics_data',
                purge_all=True,
                dry_run=dry_run,
                stdout=out,
            )
            output = out.getvalue()
            if dry_run:
                messages.info(request, 'Purge dry run completed. No data was deleted.')
            else:
                messages.success(request, 'Purge completed. All user activity, events, and journeys were removed.')
            request.session['cleanup_output'] = output
            if request.POST.get('return_to') == 'admin_app':
                messages.info(
                    request,
                    format_html(
                        '<a href="{}">Open Admin User Analytics</a> to see the full command output.',
                        reverse('user_analytics:admin_user_analytics'),
                    ),
                )
        except Exception as e:
            logger.exception('Purge analytics failed')
            messages.error(request, 'Purge failed: %s' % str(e))
        return _cleanup_analytics_redirect(request)

    days = request.POST.get('days')
    try:
        days = int(days) if days else 365
    except ValueError:
        days = 365

    if not dry_run and not confirm:
        messages.warning(request, 'Check "I understand this will permanently delete data" to run cleanup.')
        return _cleanup_analytics_redirect(request)

    out = StringIO()
    try:
        call_command(
            'cleanup_analytics_data',
            days=days,
            dry_run=dry_run,
            stdout=out,
        )
        output = out.getvalue()
        if dry_run:
            messages.info(request, 'Dry run completed. No data was deleted.')
        else:
            messages.success(request, 'Cleanup completed. See details below.')
        request.session['cleanup_output'] = output
    except Exception as e:
        logger.exception('Cleanup failed')
        messages.error(request, 'Cleanup failed: %s' % str(e))
    return _cleanup_analytics_redirect(request)


# ---------- Enquiry Sources (non-readable UTM links: ?ref=TOKEN) ----------


def _event_error_detail_from_metadata(meta):
    """Best-effort message for Razorpay/client errors, gateway failures, etc."""
    if not meta:
        return ''
    d = meta.get('detail') or meta.get('error_description') or meta.get('error_message') or meta.get('gateway_error')
    if d:
        return str(d).strip()
    err = meta.get('error')
    if isinstance(err, dict):
        parts = [err.get('description'), err.get('code'), err.get('message')]
        return ' — '.join(str(p) for p in parts if p) or ''
    if err:
        return str(err).strip()
    rc = meta.get('response_code')
    if rc:
        return 'Response code: %s' % rc
    return ''


def _user_event_is_payment_cancel_failure(ev):
    meta = getattr(ev, 'metadata', None) or {}
    st = (meta.get('stage') or '').strip().lower()
    if st == 'cancel':
        return True
    name = (getattr(ev, 'event_name', None) or '').lower()
    return 'cancel' in name


def _user_event_is_payment_error_failure(ev):
    """Gateway / Razorpay errors and server-side payment_failed (excludes user-dismissed checkout)."""
    if getattr(ev, 'event_type', None) != 'payment_failed':
        return False
    return not _user_event_is_payment_cancel_failure(ev)


def _payment_has_terminal_outcome_after(attribution_q, ev):
    """Later payment_success or payment_failed for the same payment row or order id."""
    q = UserEvent.objects.filter(attribution_q, created__gt=ev.created).filter(
        event_type__in=['payment_success', 'payment_failed']
    )
    if ev.content_type_id and ev.object_id:
        if q.filter(content_type_id=ev.content_type_id, object_id=ev.object_id).exists():
            return True
    meta = getattr(ev, 'metadata', None) or {}
    oid = meta.get('gateway_order_id') or meta.get('order_id')
    if oid:
        oid = str(oid).strip()
        if oid and q.filter(Q(metadata__gateway_order_id=oid) | Q(metadata__order_id=oid)).exists():
            return True
    if ev.session_id and not (ev.content_type_id and ev.object_id) and not oid:
        return q.filter(session_id=ev.session_id).exists()
    return False


def _payment_has_success_after(attribution_q, ev):
    """Later payment_success for the same payment row/order/session."""
    q = UserEvent.objects.filter(attribution_q, created__gt=ev.created, event_type='payment_success')
    if ev.content_type_id and ev.object_id:
        if q.filter(content_type_id=ev.content_type_id, object_id=ev.object_id).exists():
            return True
    meta = getattr(ev, 'metadata', None) or {}
    oid = meta.get('gateway_order_id') or meta.get('order_id')
    if oid:
        oid = str(oid).strip()
        if oid and q.filter(Q(metadata__gateway_order_id=oid) | Q(metadata__order_id=oid)).exists():
            return True
    # Retry payments often create a new gateway order/payment row for the same logical purchase.
    # Reconcile by domain identifiers (obj_id/test_name) and user before treating failure as unresolved.
    obj_id = meta.get('obj_id')
    obj_type = (meta.get('obj_type') or '').strip()
    if obj_id not in (None, ''):
        domain_q = Q(metadata__obj_id=obj_id)
        if obj_type:
            domain_q &= Q(metadata__obj_type=obj_type)
        if ev.user_id:
            domain_q &= Q(user_id=ev.user_id)
        if q.filter(domain_q).exists():
            return True
    test_name = (meta.get('test_name') or '').strip()
    if test_name:
        t_q = Q(metadata__test_name=test_name)
        if ev.user_id:
            t_q &= Q(user_id=ev.user_id)
        if q.filter(t_q).exists():
            return True
    if ev.session_id and ev.user_id and (test_name or 'psychometric' in obj_type.lower()):
        if q.filter(session_id=ev.session_id, user_id=ev.user_id).exists():
            return True
    if ev.session_id and not (ev.content_type_id and ev.object_id) and not oid:
        return q.filter(session_id=ev.session_id).exists()
    return False


def _payment_service_name_from_payment(payment_obj):
    obj_type = getattr(payment_obj, 'obj_type', None)
    obj_id = getattr(payment_obj, 'obj_id', None)
    try:
        if obj_type == choices.PaymentObjectType.PYSCHOMETRICTESTDETAIL and obj_id:
            ptest = PsychometricTestPayment.objects.filter(id=obj_id).only('test_type').first()
            if ptest:
                return f"Psychometric Test - {ptest.get_test_name()}"
        if obj_type == choices.PaymentObjectType.SKILLLABCOURSE and obj_id:
            course = SkillLabCourse.objects.filter(id=obj_id).only('name').first()
            if course and course.name:
                return f"SkillLab Course - {course.name}"
    except Exception:
        pass
    return dict(choices.PaymentObjectType.CHOICES).get(obj_type, f"Type {obj_type}")


def _payment_service_name_from_event(ev, gateway_payment=None):
    meta = (getattr(ev, 'metadata', None) or {})
    test_name = (meta.get('test_name') or '').strip()
    if test_name:
        return f"Psychometric Test - {test_name}"
    obj_type_text = (meta.get('obj_type') or '').strip()
    obj_id = meta.get('obj_id')
    try:
        if obj_id not in (None, '') and 'skilllab' in obj_type_text.lower():
            course = SkillLabCourse.objects.filter(id=obj_id).only('name').first()
            if course and course.name:
                return f"SkillLab Course - {course.name}"
    except Exception:
        pass
    if gateway_payment:
        return _payment_service_name_from_payment(gateway_payment)
    return ev.event_name or 'N/A'


def _enquiry_source_payment_in_process_count(attribution_q):
    """Checkout started (server or client) with no recorded success/fail yet (bounded scan)."""
    candidates = list(
        UserEvent.objects.filter(attribution_q, event_type='payment_pending')
        .filter(
            Q(metadata__payment_stage='checkout_started') | Q(metadata__stage='started')
        )
        .order_by('-created')[:2500]
    )
    # Terminal confirmation can be recorded without the same attribution payload;
    # remove those rows from "in process" as soon as success/fail exists anywhere.
    return sum(1 for ev in candidates if not _payment_has_terminal_outcome_after(Q(), ev))


def _enquiry_source_attribution_q(source):
    """Q object matching UserEvent rows attributable to this enquiry source, or None if no scope."""
    from django.db.models import Q

    source_activities = UserActivity.objects.filter(enquiry_source=source)
    activity_session_ids = list(
        source_activities.exclude(session_id__isnull=True).exclude(session_id='').values_list('session_id', flat=True).distinct()
    )
    journey_session_ids = list(
        UserJourney.objects.filter(enquiry_source=source).exclude(session_id__isnull=True).exclude(session_id='').values_list('session_id', flat=True).distinct()
    )
    session_ids = sorted(set(activity_session_ids) | set(journey_session_ids))
    source_user_ids = list(
        source_activities.exclude(user_id__isnull=True).values_list('user_id', flat=True).distinct()
    )
    if not session_ids and not source_user_ids:
        return None
    base_scope = Q()
    if session_ids:
        base_scope |= Q(session_id__in=session_ids)
    if source_user_ids:
        base_scope |= Q(user_id__in=source_user_ids)
    source_name_scope = Q(metadata__source=source.name)
    return base_scope | source_name_scope


def _enquiry_source_linked_session_id(source, user_id, event_session_id):
    """
    Prefer session_id on the event; else latest UserActivity/UserJourney for this user
    tied to this enquiry source (for modal links when analytics session was missing).
    """
    s = (event_session_id or '').strip()
    if s:
        return s
    if not user_id:
        return ''
    latest = (
        UserActivity.objects.filter(enquiry_source=source, user_id=user_id)
        .exclude(session_id__isnull=True).exclude(session_id='')
        .order_by('-created')
        .values_list('session_id', flat=True)
        .first()
    )
    if latest:
        return latest
    latest = (
        UserJourney.objects.filter(enquiry_source=source, user_id=user_id)
        .exclude(session_id__isnull=True).exclude(session_id='')
        .order_by('-created')
        .values_list('session_id', flat=True)
        .first()
    )
    return latest or ''


def _active_enquiry_source_by_name(name):
    """Active EnquirySource whose name matches (case-insensitive), or None."""
    n = (name or '').strip()
    if not n:
        return None
    return EnquirySource.objects.filter(
        name__iexact=n,
        object_status=choices.ObjectStatus.ACTIVE,
    ).first()


def _apply_traffic_source_to_userevent_qs(qs, source_filter):
    """
    Restrict UserEvent queryset to a traffic/enquiry source the same way Enquiry Sources
    drill-downs do: named EnquirySource → sessions/journeys/users from ?ref= plus
    metadata.source; otherwise plain metadata.source match.
    """
    sf = (source_filter or '').strip()
    if not sf:
        return qs
    enq = _active_enquiry_source_by_name(sf)
    if enq:
        aq = _enquiry_source_attribution_q(enq)
        if aq is not None:
            return qs.filter(aq)
        return qs.filter(metadata__source=enq.name)
    return qs.filter(metadata__source__iexact=sf)


def _attribution_q_for_terminal_check(source_filter):
    """
    When deciding if checkout is still 'in process', limit 'later success/fail' events
    to the same attribution scope as the enquiry source modal (when filter is a source name).
    """
    sf = (source_filter or '').strip()
    if not sf:
        return Q()
    enq = _active_enquiry_source_by_name(sf)
    if not enq:
        return Q()
    aq = _enquiry_source_attribution_q(enq)
    if aq is not None:
        return aq
    return Q(metadata__source=enq.name)


def _enquiry_source_stats(source):
    """Return dict of visit and conversion counts for an EnquirySource.

    Attribution strategy:
    - Session-based: events from sessions that touched this enquiry source.
    - User-based fallback: events by users who visited via this source (covers cases where session_id is absent).

    Checkout "start" counts only server-side payment_pending (Payment / PsychometricTestPayment create).
    Client payment-status API used to duplicate with metadata.stage='started'; that row is excluded here.
    """
    # Base activity for this source (most reliable signal for ref tracking).
    source_activities = UserActivity.objects.filter(enquiry_source=source)
    page_views = source_activities.count()

    activity_session_ids = list(
        source_activities.exclude(session_id__isnull=True).exclude(session_id='').values_list('session_id', flat=True).distinct()
    )
    journey_session_ids = list(
        UserJourney.objects.filter(enquiry_source=source).exclude(session_id__isnull=True).exclude(session_id='').values_list('session_id', flat=True).distinct()
    )
    session_ids = sorted(set(activity_session_ids) | set(journey_session_ids))
    visit_count = len(session_ids)

    source_user_ids = list(
        source_activities.exclude(user_id__isnull=True).values_list('user_id', flat=True).distinct()
    )

    payments_attributed = UserEvent.objects.filter(
        event_type='payment_success',
        metadata__source=source.name
    ).count()
    attribution_q = _enquiry_source_attribution_q(source)
    strict_payment_scope_q = Q(metadata__source__iexact=source.name)
    if session_ids:
        strict_payment_scope_q |= Q(session_id__in=session_ids)
    recent_source_user_ids = list(
        source_activities.filter(created__gte=timezone.now() - timedelta(days=2))
        .exclude(user_id__isnull=True)
        .values_list('user_id', flat=True)
        .distinct()
    )
    unattributed_recent_user_fallback_q = Q()
    if recent_source_user_ids:
        unattributed_recent_user_fallback_q = (
            Q(user_id__in=recent_source_user_ids)
            & (Q(session_id__isnull=True) | Q(session_id=''))
            & (
                Q(metadata__source__isnull=True)
                | Q(metadata__source='')
                | Q(metadata__source__iexact='direct')
                | Q(metadata__source__iexact='internal')
            )
        )
    if attribution_q is None:
        return {
            'page_views': page_views,
            'visit_count': 0,
            'registrations': 0,
            'payment_success': 0,
            'payments_attributed': payments_attributed,
            'course_enrolled': 0,
            'converted_sessions': 0,
            'payment_started': 0,
            'payment_failed': 0,
            'payment_errors': 0,
            'payment_cancelled': 0,
            'payment_in_process': 0,
        }

    reg = UserEvent.objects.filter(attribution_q, event_type='registration').distinct().count()
    reg_by_users = 0
    if reg == 0:
        if session_ids:
            reg_by_users = UserJourney.objects.filter(
                session_id__in=session_ids,
                user__isnull=False,
            ).exclude(user_id__isnull=True).values('user_id').distinct().count()
        if source_user_ids:
            reg_by_users = max(
                reg_by_users,
                User.objects.filter(id__in=source_user_ids).values('id').distinct().count(),
            )
    pay_qs = UserEvent.objects.filter(
        (strict_payment_scope_q | unattributed_recent_user_fallback_q),
        event_type='payment_success'
    )
    pay = pay_qs.distinct().count()
    payment_model_success = 0
    payment_model_failed = 0
    payment_model_enrolled = 0
    payment_started = (
        UserEvent.objects.filter(strict_payment_scope_q, event_type='payment_pending')
        .filter(Q(metadata__stage__isnull=True) | ~Q(metadata__stage='started'))
        .distinct()
        .count()
    )
    payment_failed = UserEvent.objects.filter(strict_payment_scope_q, event_type='payment_failed').distinct().count()
    failed_base = UserEvent.objects.filter(strict_payment_scope_q, event_type='payment_failed')
    payment_cancelled = failed_base.filter(
        Q(metadata__stage='cancel') | Q(event_name__icontains='Cancelled')
    ).distinct().count()
    payment_errors = failed_base.exclude(
        Q(metadata__stage='cancel') | Q(event_name__icontains='Cancelled')
    ).distinct().count()
    payment_in_process = _enquiry_source_payment_in_process_count(strict_payment_scope_q)
    # "Enrolled" = purchased course access. Primary signal is successful payment for course-like objects.
    # This keeps Enrolled aligned with "student bought" even if an explicit enrollment event is missing.
    payment_type_map = dict(choices.PaymentObjectType.CHOICES)
    enrolled_obj_types = [
        payment_type_map.get(choices.PaymentObjectType.SKILLLABCOURSE),
        payment_type_map.get(choices.PaymentObjectType.COUNSELOR),
    ]
    enrolled_by_payment = UserEvent.objects.filter(
        (strict_payment_scope_q | unattributed_recent_user_fallback_q),
        event_type='payment_success',
        metadata__obj_type__in=[x for x in enrolled_obj_types if x],
    ).distinct().count()
    # Fallback: explicit enrollment events (kept for legacy flows).
    enrolled_by_event = UserEvent.objects.filter(
        (strict_payment_scope_q | unattributed_recent_user_fallback_q),
        event_type__in=['course_enrolled', 'skilllab_enrolled', 'counselor_course_enrolled'],
    ).distinct().count()
    course = max(enrolled_by_payment, enrolled_by_event, payment_model_enrolled)
    need_payment_fallback = (pay == 0 or payment_failed == 0 or course == 0)
    if need_payment_fallback:
        try:
            # Fast fallback: use users who touched this enquiry source.
            # Avoid heavy per-row subqueries that can make the list page time out.
            p_base = Payment.objects.none()
            if source_user_ids:
                p_base = Payment.objects.filter(user_id__in=source_user_ids)
            if pay == 0:
                payment_model_success = p_base.filter(is_success=choices.YesNoChoices.YES).count()
            if payment_failed == 0:
                payment_model_failed = p_base.filter(is_success=choices.YesNoChoices.NO).count()
            if course == 0:
                payment_model_enrolled = p_base.filter(
                    is_success=choices.YesNoChoices.YES,
                    obj_type__in=[
                        choices.PaymentObjectType.SKILLLABCOURSE,
                        choices.PaymentObjectType.COUNSELOR,
                    ],
                ).count()
        except Exception:
            payment_model_success = 0
            payment_model_failed = 0
            payment_model_enrolled = 0
    course = max(course, payment_model_enrolled)

    paid_out = max(pay, payment_model_success)
    converted = (
        UserJourney.objects.filter(session_id__in=session_ids, converted=True)
        .exclude(session_id__isnull=True)
        .exclude(session_id='')
        .values('session_id')
        .distinct()
        .count()
        if session_ids
        else 0
    )
    # Same heuristic as before: when journey.converted lags payment capture, cap by visits vs paid.
    converted_out = max(converted, min(visit_count, paid_out))
    attributed_out = max(payments_attributed, payment_model_success)
    reg_out = max(reg, reg_by_users)
    failed_out = max(payment_failed, payment_model_failed)
    return {
        'page_views': page_views,
        'visit_count': visit_count,
        'registrations': reg_out,
        'payment_success': paid_out,
        'payments_attributed': attributed_out,
        'course_enrolled': course,
        'converted_sessions': converted_out,
        'payment_started': payment_started,
        'payment_failed': failed_out,
        'payment_errors': payment_errors,
        'payment_cancelled': payment_cancelled,
        'payment_in_process': payment_in_process,
    }


@login_required
@user_passes_test(is_staff_or_superuser)
def enquiry_sources_list_view(request):
    """List enquiry sources with copy URL, QR, download QR, and stats. Optional filter by agency/event."""
    from django.conf import settings
    base_url = getattr(settings, 'ENQUIRY_SOURCE_BASE_URL', '')
    sources = EnquirySource.objects.filter(object_status=choices.ObjectStatus.ACTIVE).order_by('-created')
    agency_filter = (request.GET.get('agency') or '').strip()
    event_filter = (request.GET.get('event') or '').strip()
    if agency_filter:
        sources = sources.filter(agency_name__icontains=agency_filter)
    if event_filter:
        sources = sources.filter(event__icontains=event_filter)
    source_stats = []
    for s in sources:
        full_url = s.get_full_url()
        stats = _enquiry_source_stats(s)
        source_stats.append({'source': s, 'full_url': full_url, 'stats': stats})
    # Distinct values for filter dropdowns
    all_sources = EnquirySource.objects.filter(object_status=choices.ObjectStatus.ACTIVE)
    agencies = sorted({x for x in all_sources.values_list('agency_name', flat=True).distinct() if x})
    events = sorted({x for x in all_sources.values_list('event', flat=True).distinct() if x})
    context = {
        'source_stats': source_stats,
        'enquiry_base_url': base_url,
        'agency_filter': agency_filter,
        'event_filter': event_filter,
        'agencies': agencies,
        'events': events,
        'page_title': 'Enquiry Sources (UTM Links)',
        'csrf_input_html': format_html(
            '<input type="hidden" name="csrfmiddlewaretoken" value="{}">',
            get_token(request),
        ),
    }
    return render(request, 'user_analytics/enquiry_sources_list.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
@require_GET
def enquiry_source_events_api(request):
    """JSON list of metric rows for an enquiry source (staff). Used by enquiry sources table modal."""
    try:
        source_id = int(request.GET.get('source_id', ''))
    except (TypeError, ValueError):
        return JsonResponse({'ok': False, 'error': 'Invalid source_id'}, status=400)
    kind = (request.GET.get('kind') or '').strip()
    if kind not in {
        'page_views',
        'sessions',
        'registration',
        'payment_success',
        'payment_started',
        'payment_failed',
        'payment_in_process',
        'course_enrolled',
        'converted_sessions',
    }:
        return JsonResponse({'ok': False, 'error': 'Invalid kind'}, status=400)
    source = EnquirySource.objects.filter(pk=source_id, object_status=choices.ObjectStatus.ACTIVE).first()
    if not source:
        return JsonResponse({'ok': False, 'error': 'Not found'}, status=404)
    limit = 100
    session_user_cache = {}

    def _resolve_user_email(session_id='', explicit_user_email=None):
        if explicit_user_email:
            return explicit_user_email
        sid = (session_id or '').strip()
        if not sid:
            return None
        if sid in session_user_cache:
            return session_user_cache[sid]
        email = (
            UserActivity.objects.filter(session_id=sid, user_id__isnull=False)
            .order_by('-created')
            .values_list('user__email', flat=True)
            .first()
        )
        if not email:
            email = (
                UserJourney.objects.filter(session_id=sid, user_id__isnull=False)
                .order_by('-created')
                .values_list('user__email', flat=True)
                .first()
            )
        if not email:
            email = (
                UserEvent.objects.filter(session_id=sid, user_id__isnull=False)
                .order_by('-created')
                .values_list('user__email', flat=True)
                .first()
            )
        session_user_cache[sid] = email or None
        return session_user_cache[sid]

    if kind == 'page_views':
        page_qs = UserActivity.objects.filter(enquiry_source=source).order_by('-created')
        total = page_qs.count()
        rows = list(page_qs[: limit + 1])
        truncated = len(rows) > limit
        rows = rows[:limit]
        data_rows = [{
            'id': r.id,
            'created': timezone.localtime(r.created).strftime('%Y-%m-%d %H:%M:%S'),
            'event_type': 'page_view',
            'event_name': r.page_title or r.page_path or 'Page View',
            'user_email': _resolve_user_email(r.session_id or '', r.user.email if r.user_id else None),
            'session_id': r.session_id or '',
            'metadata': {
                'page_path': r.page_path or '',
                'referrer': r.referrer or '',
                'utm_source': r.utm_source or '',
                'utm_medium': r.utm_medium or '',
            },
            'status_override': {'text': 'Viewed', 'variant': 'secondary'},
        } for r in rows]
        return JsonResponse({'ok': True, 'events': data_rows, 'total': total, 'truncated': truncated})

    if kind == 'sessions':
        activity_sessions = list(
            UserActivity.objects.filter(enquiry_source=source)
            .exclude(session_id__isnull=True).exclude(session_id='')
            .values('session_id')
            .annotate(last_seen=Max('created'))
        )
        journey_sessions = list(
            UserJourney.objects.filter(enquiry_source=source)
            .exclude(session_id__isnull=True).exclude(session_id='')
            .values('session_id')
            .annotate(last_seen=Max('created'))
        )
        by_session = {}
        for row in activity_sessions + journey_sessions:
            sid = row['session_id']
            seen = row['last_seen']
            if sid not in by_session or seen > by_session[sid]:
                by_session[sid] = seen
        ordered = sorted(by_session.items(), key=lambda item: item[1], reverse=True)
        total = len(ordered)
        truncated = total > limit
        selected = ordered[:limit]
        data_rows = [{
            'id': sid,
            'created': timezone.localtime(last_seen).strftime('%Y-%m-%d %H:%M:%S'),
            'event_type': 'session',
            'event_name': 'Session Visit',
            'user_email': _resolve_user_email(sid, None),
            'session_id': sid,
            'metadata': {},
            'status_override': {'text': 'Visited', 'variant': 'info'},
        } for sid, last_seen in selected]
        return JsonResponse({'ok': True, 'events': data_rows, 'total': total, 'truncated': truncated})

    attribution_q = _enquiry_source_attribution_q(source)
    if attribution_q is None:
        return JsonResponse({'ok': True, 'events': [], 'truncated': False, 'total': 0})
    source_user_ids = list(
        UserActivity.objects.filter(enquiry_source=source)
        .exclude(user_id__isnull=True)
        .values_list('user_id', flat=True)
        .distinct()
    )

    if kind == 'payment_in_process':
        candidates = list(
            UserEvent.objects.filter(attribution_q, event_type='payment_pending')
            .filter(
                Q(metadata__payment_stage='checkout_started') | Q(metadata__stage='started')
            )
            .select_related('user', 'content_type')
            .order_by('-created')[:2500]
        )
        filtered = [ev for ev in candidates if not _payment_has_terminal_outcome_after(Q(), ev)]
        total = len(filtered)
        scan_capped = len(candidates) >= 2500
        rows = filtered[: limit + 1]
        truncated = len(rows) > limit or scan_capped
        rows = rows[:limit]
        gp_inproc = _resolve_gateway_payments_for_userevents(rows)
        data_rows = []
        for ev in rows:
            sid = _enquiry_source_linked_session_id(source, ev.user_id, ev.session_id or '')
            data_rows.append({
                'id': ev.id,
                'created': timezone.localtime(ev.created).strftime('%Y-%m-%d %H:%M:%S'),
                'event_type': ev.event_type,
                'event_name': ev.event_name,
                'user_email': _resolve_user_email(sid, ev.user.email if ev.user_id else None),
                'session_id': sid,
                'metadata': ev.metadata or {},
                'effective_status': None,
                'error_detail': '',
                'amount_rupees': round(_order_amount_rupees_for_checkout_event(ev, gp_inproc.get(ev.id)), 2),
                'status_override': {'text': 'In process', 'variant': 'info'},
            })
        return JsonResponse({
            'ok': True,
            'events': data_rows,
            'total': total,
            'truncated': truncated,
            'failure_subset': None,
            'supports_failure_filter': False,
        })

    if kind == 'converted_sessions':
        # Total must match Enquiry Sources column: max(converted, min(visit_count, paid_out)).
        # Rows list concrete sessions: journey.converted OR payment_success in attributed session
        # (so the popup is not empty when the heuristic count is from paid vs visits).
        stats = _enquiry_source_stats(source)
        total = int(stats.get('converted_sessions') or 0)
        activity_session_ids = list(
            UserActivity.objects.filter(enquiry_source=source)
            .exclude(session_id__isnull=True).exclude(session_id='')
            .values_list('session_id', flat=True).distinct()
        )
        journey_session_ids = list(
            UserJourney.objects.filter(enquiry_source=source)
            .exclude(session_id__isnull=True).exclude(session_id='')
            .values_list('session_id', flat=True).distinct()
        )
        session_ids = sorted(set(activity_session_ids) | set(journey_session_ids))
        attribution_q = _enquiry_source_attribution_q(source)
        if attribution_q is None:
            return JsonResponse({'ok': True, 'events': [], 'total': total, 'truncated': False})
        by_sid = {}
        if session_ids:
            for row in (
                UserJourney.objects.filter(session_id__in=session_ids, converted=True)
                .exclude(session_id__isnull=True).exclude(session_id='')
                .values('session_id')
                .annotate(last_seen=Max('created'))
            ):
                sid = row['session_id']
                ls = row['last_seen']
                by_sid[sid] = max(by_sid.get(sid, ls), ls)
            for row in (
                UserEvent.objects.filter(attribution_q, event_type='payment_success')
                .exclude(session_id__isnull=True).exclude(session_id='')
                .filter(session_id__in=session_ids)
                .values('session_id')
                .annotate(last_seen=Max('created'))
            ):
                sid = row['session_id']
                ls = row['last_seen']
                if sid not in by_sid or ls > by_sid[sid]:
                    by_sid[sid] = ls
        ordered = sorted(by_sid.items(), key=lambda item: item[1], reverse=True)
        n_listed = len(ordered)
        truncated = n_listed > limit
        selected = ordered[:limit]
        data_rows = [{
            'id': sid,
            'created': timezone.localtime(last_seen).strftime('%Y-%m-%d %H:%M:%S'),
            'event_type': 'converted_session',
            'event_name': 'Converted Session',
            'user_email': _resolve_user_email(sid, None),
            'session_id': sid,
            'metadata': {'converted': True},
            'amount_rupees': None,
            'status_override': {'text': 'Converted', 'variant': 'secondary'},
        } for sid, last_seen in selected]
        if not data_rows and total > 0:
            fb = list(
                UserEvent.objects.filter(attribution_q, event_type='payment_success')
                .select_related('user', 'content_type')
                .order_by('-created')[: limit + 1]
            )
            truncated = len(fb) > limit or truncated
            fb = fb[:limit]
            gp_fb = _resolve_gateway_payments_for_userevents(fb)
            data_rows = []
            for ev in fb:
                sid = _enquiry_source_linked_session_id(source, ev.user_id, ev.session_id or '')
                data_rows.append({
                    'id': ev.id,
                    'created': timezone.localtime(ev.created).strftime('%Y-%m-%d %H:%M:%S'),
                    'event_type': 'converted_session',
                    'event_name': 'Converted (payment success)',
                    'user_email': _resolve_user_email(sid, ev.user.email if ev.user_id else None),
                    'session_id': sid,
                    'metadata': {'converted': True, 'fallback': 'payment_success_attributed'},
                    'amount_rupees': round(_payment_amount_rupees_from_event(ev, gp_fb.get(ev.id)), 2),
                    'status_override': {'text': 'Converted', 'variant': 'secondary'},
                })
        # Stats paid_out can come from Payment rows when UserEvents are missing; same as PAID modal fallback.
        if not data_rows and total > 0 and source_user_ids:
            p_qs = (
                Payment.objects.filter(
                    user_id__in=source_user_ids,
                    is_success=choices.YesNoChoices.YES,
                )
                .select_related('user')
                .order_by('-created')
            )
            p_list = list(p_qs[: limit + 1])
            truncated = len(p_list) > limit or truncated
            p_list = p_list[:limit]
            for p in p_list:
                psid = _enquiry_source_linked_session_id(source, p.user_id, '')
                data_rows.append({
                    'id': f'payment-{p.id}',
                    'created': timezone.localtime(p.created).strftime('%Y-%m-%d %H:%M:%S') if p.created else '',
                    'event_type': 'converted_session',
                    'event_name': 'Converted (payment record)',
                    'user_email': p.user.email if p.user_id and p.user else None,
                    'session_id': psid,
                    'metadata': {'converted': True, 'fallback': 'payment_model'},
                    'amount_rupees': round(_payment_amount_rupees(p), 2),
                    'status_override': {'text': 'Converted', 'variant': 'secondary'},
                })
        return JsonResponse({'ok': True, 'events': data_rows, 'total': total, 'truncated': truncated})

    qs = UserEvent.objects.filter(attribution_q)
    if kind == 'registration':
        qs = qs.filter(event_type='registration')
    elif kind == 'payment_success':
        qs = qs.filter(event_type='payment_success')
    elif kind == 'payment_failed':
        qs = qs.filter(event_type='payment_failed')
        failure_subset = (request.GET.get('failure_subset') or 'all').strip().lower()
        if failure_subset not in {'all', 'error', 'cancel'}:
            failure_subset = 'all'
        if failure_subset == 'error':
            qs = qs.exclude(Q(metadata__stage='cancel') | Q(event_name__icontains='Cancelled'))
        elif failure_subset == 'cancel':
            qs = qs.filter(Q(metadata__stage='cancel') | Q(event_name__icontains='Cancelled'))
    elif kind == 'payment_started':
        qs = qs.filter(event_type='payment_pending').filter(Q(metadata__stage__isnull=True) | ~Q(metadata__stage='started'))
    else:  # course_enrolled
        # Align with Enquiry Sources "Enrolled" column: prefer successful course-like payments,
        # with fallback to explicit enrollment events.
        payment_type_map = dict(choices.PaymentObjectType.CHOICES)
        enrolled_obj_types = [
            payment_type_map.get(choices.PaymentObjectType.SKILLLABCOURSE),
            payment_type_map.get(choices.PaymentObjectType.COUNSELOR),
        ]
        qs_payment = UserEvent.objects.filter(
            attribution_q,
            event_type='payment_success',
            metadata__obj_type__in=[x for x in enrolled_obj_types if x],
        ).select_related('user', 'content_type')
        qs_event = qs.filter(event_type__in=['course_enrolled', 'skilllab_enrolled', 'counselor_course_enrolled']).select_related('user', 'content_type')
        qs = (qs_payment | qs_event)
    if kind != 'course_enrolled':
        qs = qs.select_related('user', 'content_type')
    total = qs.distinct().count()
    rows = list(qs.distinct().order_by('-created')[: limit + 1])
    truncated = len(rows) > limit
    rows = rows[:limit]
    gp_modal = {}
    if kind in ('payment_success', 'payment_failed', 'payment_started'):
        gp_modal = _resolve_gateway_payments_for_userevents(rows)
    elif kind == 'course_enrolled':
        gp_modal = _resolve_gateway_payments_for_userevents(
            [ev for ev in rows if getattr(ev, 'event_type', None) == 'payment_success']
        )
    if kind == 'payment_success' and total == 0 and source_user_ids:
        # Keep modal aligned with paid counter fallback when event rows are absent/soft-deleted.
        p_qs = Payment.objects.filter(
            user_id__in=source_user_ids,
            is_success=choices.YesNoChoices.YES,
        ).select_related('user').order_by('-created')
        p_total = p_qs.count()
        p_rows = list(p_qs[: limit + 1])
        p_truncated = len(p_rows) > limit
        p_rows = p_rows[:limit]
        data_rows = []
        for p in p_rows:
            psid = _enquiry_source_linked_session_id(source, p.user_id, '')
            data_rows.append({
                'id': f'payment-{p.id}',
                'created': timezone.localtime(p.created).strftime('%Y-%m-%d %H:%M:%S') if p.created else '',
                'event_type': 'payment_success',
                'event_name': _payment_service_name_from_payment(p),
                'user_email': p.user.email if p.user_id and p.user else None,
                'session_id': psid,
                'metadata': {
                    'source': source.name,
                    'gateway_order_id': p.gateway_order_id or '',
                    'gateway_payment_id': p.gateway_payment_id or '',
                    'obj_type': dict(choices.PaymentObjectType.CHOICES).get(p.obj_type, str(p.obj_type)),
                    'obj_id': p.obj_id,
                    'fallback': 'payment_model',
                },
                'effective_status': 'success',
                'error_detail': '',
                'amount_rupees': round(_payment_amount_rupees(p), 2),
                'status_override': {'text': 'Success', 'variant': 'secondary'},
            })
        return JsonResponse({'ok': True, 'events': data_rows, 'total': p_total, 'truncated': p_truncated})
    if kind == 'course_enrolled' and total == 0 and source_user_ids:
        # Same Payment fallback as Enrolled column in _enquiry_source_stats (course-like success rows).
        p_qs = Payment.objects.filter(
            user_id__in=source_user_ids,
            is_success=choices.YesNoChoices.YES,
            obj_type__in=[
                choices.PaymentObjectType.SKILLLABCOURSE,
                choices.PaymentObjectType.COUNSELOR,
            ],
        ).select_related('user').order_by('-created')
        p_total = p_qs.count()
        p_rows = list(p_qs[: limit + 1])
        p_truncated = len(p_rows) > limit
        p_rows = p_rows[:limit]
        data_rows = []
        for p in p_rows:
            psid = _enquiry_source_linked_session_id(source, p.user_id, '')
            data_rows.append({
                'id': f'payment-{p.id}',
                'created': timezone.localtime(p.created).strftime('%Y-%m-%d %H:%M:%S') if p.created else '',
                'event_type': 'payment_success',
                'event_name': _payment_service_name_from_payment(p),
                'user_email': p.user.email if p.user_id and p.user else None,
                'session_id': psid,
                'metadata': {
                    'source': source.name,
                    'gateway_order_id': p.gateway_order_id or '',
                    'gateway_payment_id': p.gateway_payment_id or '',
                    'obj_type': dict(choices.PaymentObjectType.CHOICES).get(p.obj_type, str(p.obj_type)),
                    'obj_id': p.obj_id,
                    'fallback': 'payment_model_course_enrolled',
                },
                'effective_status': 'success',
                'error_detail': '',
                'amount_rupees': round(_payment_amount_rupees(p), 2),
                'status_override': {'text': 'Enrolled', 'variant': 'info'},
            })
        return JsonResponse({'ok': True, 'events': data_rows, 'total': p_total, 'truncated': p_truncated})
    # For payment rows, compute an "effective" status based on the latest payment event for the same payment/order.
    # This avoids showing "Pending/Cancelled" for checkouts that later became Paid, etc.
    # Do not apply this when listing payment_failed: each row is an explicit failure and must stay "Fail"
    # even if the same order/object later succeeded (retry).
    effective_by_id = {}
    if kind in {'payment_started', 'payment_success'}:
        payment_rows = [ev for ev in rows if ev.event_type in {'payment_pending', 'payment_failed', 'payment_success'}]
        if payment_rows:
            keys = []
            order_ids = set()
            for ev in payment_rows:
                if ev.content_type_id and ev.object_id:
                    keys.append(('obj', ev.content_type_id, ev.object_id))
                meta = ev.metadata or {}
                oid = meta.get('gateway_order_id') or meta.get('order_id')
                if oid:
                    order_ids.add(str(oid))

            match_q = Q()
            # Build safe OR-of-pairs (avoid cross-product from __in lists).
            for _, ct_id, obj_id in [k for k in keys if k[0] == 'obj']:
                match_q |= Q(content_type_id=ct_id, object_id=obj_id)
            if order_ids:
                match_q |= Q(metadata__gateway_order_id__in=list(order_ids)) | Q(metadata__order_id__in=list(order_ids))

            latest_by_key = {}
            if match_q:
                related = UserEvent.objects.filter(
                    attribution_q,
                    match_q,
                    event_type__in=['payment_success', 'payment_failed', 'payment_pending'],
                ).order_by('-created')
                for ev in related:
                    key = None
                    if ev.content_type_id and ev.object_id:
                        key = ('obj', ev.content_type_id, ev.object_id)
                    else:
                        meta = ev.metadata or {}
                        oid = meta.get('gateway_order_id') or meta.get('order_id')
                        if oid:
                            key = ('order', str(oid))
                    if key and key not in latest_by_key:
                        latest_by_key[key] = ev

            def _effective_from_latest(latest_ev):
                if not latest_ev:
                    return None
                if latest_ev.event_type == 'payment_success':
                    return 'success'
                if latest_ev.event_type == 'payment_pending':
                    return 'pending'
                # payment_failed (cancel/error/other) -> fail
                return 'fail'

            for ev in payment_rows:
                k = None
                if ev.content_type_id and ev.object_id:
                    k = ('obj', ev.content_type_id, ev.object_id)
                else:
                    meta = ev.metadata or {}
                    oid = meta.get('gateway_order_id') or meta.get('order_id')
                    if oid:
                        k = ('order', str(oid))
                latest = latest_by_key.get(k) if k else None
                eff = _effective_from_latest(latest)
                if not eff:
                    # Fallback to the row itself.
                    eff = _effective_from_latest(ev)
                effective_by_id[ev.id] = eff

    failure_subset_out = None
    if kind == 'payment_failed':
        failure_subset_out = (request.GET.get('failure_subset') or 'all').strip().lower()
        if failure_subset_out not in {'all', 'error', 'cancel'}:
            failure_subset_out = 'all'

    data_rows = []
    for ev in rows:
        meta = ev.metadata or {}
        err_detail = _event_error_detail_from_metadata(meta)
        if kind == 'payment_failed' and not err_detail and _user_event_is_payment_error_failure(ev):
            ps = (meta.get('payment_stage') or '').strip().lower()
            if ps == 'gateway_error':
                err_detail = 'Gateway verification failed'
        sid = _enquiry_source_linked_session_id(source, ev.user_id, ev.session_id or '')
        row = {
            'id': ev.id,
            'created': timezone.localtime(ev.created).strftime('%Y-%m-%d %H:%M:%S'),
            'event_type': ev.event_type,
            'event_name': ev.event_name,
            'user_email': _resolve_user_email(sid, ev.user.email if ev.user_id else None),
            'session_id': sid,
            'metadata': meta,
            'effective_status': effective_by_id.get(ev.id),
            'error_detail': err_detail,
            'status_override': None,
            'amount_rupees': None,
        }
        if kind in ('payment_success', 'payment_failed', 'payment_started'):
            gp = gp_modal.get(ev.id)
            if kind == 'payment_failed':
                row['amount_rupees'] = round(_order_amount_rupees_for_checkout_event(ev, gp), 2)
            else:
                row['amount_rupees'] = round(_payment_amount_rupees_from_event(ev, gp), 2)
        elif kind == 'course_enrolled':
            if ev.event_type == 'payment_success':
                gp = gp_modal.get(ev.id)
                row['amount_rupees'] = round(_payment_amount_rupees_from_event(ev, gp), 2)
            else:
                row['amount_rupees'] = round(float(ev.event_value or 0), 2)
        if kind == 'registration':
            row['status_override'] = {'text': 'Registered', 'variant': 'info'}
        elif kind == 'course_enrolled':
            row['status_override'] = {'text': 'Enrolled', 'variant': 'info'}
        data_rows.append(row)

    payload = {'ok': True, 'events': data_rows, 'total': total, 'truncated': truncated}
    if kind == 'payment_failed':
        payload['failure_subset'] = failure_subset_out
        payload['supports_failure_filter'] = True
    return JsonResponse(payload)


@login_required
@user_passes_test(is_staff_or_superuser)
def _enquiry_source_form_data(request):
    """Extract name, agency_name, user_name, event, base_url from POST. Returns (name, agency_name, user_name, event, base_url)."""
    def strip(s):
        return (s or '').strip() or None
    return (
        strip(request.POST.get('name')),
        strip(request.POST.get('agency_name')),
        strip(request.POST.get('user_name')),
        strip(request.POST.get('event')),
        strip(request.POST.get('base_url')),
    )


def enquiry_source_create_view(request):
    """Create a new enquiry source (name + optional agency, user, event, base_url). Token is auto-generated."""
    from django.contrib import messages
    from django.conf import settings
    if request.method == 'POST':
        name, agency_name, user_name, event, base_url = _enquiry_source_form_data(request)
        if not name:
            messages.error(request, 'Name is required.')
            return redirect('user_analytics:enquiry_source_create')
        try:
            EnquirySource.objects.create(
                name=name,
                agency_name=agency_name,
                user_name=user_name,
                event=event,
                base_url=base_url,
            )
            messages.success(request, 'Enquiry source created. Use the link with ?ref= token only (non-readable).')
            return redirect('user_analytics:enquiry_sources_list')
        except Exception as e:
            messages.error(request, str(e))
            return redirect('user_analytics:enquiry_sources_list')
    enquiry_base_url = getattr(settings, 'ENQUIRY_SOURCE_BASE_URL', '') or request.build_absolute_uri('/').rstrip('/')
    context = {
        'source': None,
        'page_title': 'Add Enquiry Source',
        'enquiry_base_url': enquiry_base_url,
        'csrf_input_html': format_html(
            '<input type="hidden" name="csrfmiddlewaretoken" value="{}">',
            get_token(request),
        ),
    }
    return render(request, 'user_analytics/enquiry_source_form.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def enquiry_source_edit_view(request, pk):
    """Edit enquiry source name, agency, user, event, base_url. Token is not changed."""
    from django.contrib import messages
    from django.http import HttpResponseNotFound
    from django.conf import settings
    try:
        source = EnquirySource.objects.get(pk=pk, object_status=choices.ObjectStatus.ACTIVE)
    except EnquirySource.DoesNotExist:
        return HttpResponseNotFound('Enquiry source not found.')
    if request.method == 'POST':
        name, agency_name, user_name, event, base_url = _enquiry_source_form_data(request)
        if not name:
            messages.error(request, 'Name is required.')
            return redirect('user_analytics:enquiry_source_edit', pk=pk)
        try:
            source.name = name
            source.agency_name = agency_name
            source.user_name = user_name
            source.event = event
            source.base_url = base_url
            source.save()
            messages.success(request, 'Enquiry source updated.')
            return redirect('user_analytics:enquiry_sources_list')
        except Exception as e:
            messages.error(request, str(e))
    enquiry_base_url = getattr(settings, 'ENQUIRY_SOURCE_BASE_URL', '') or request.build_absolute_uri('/').rstrip('/')
    context = {
        'source': source,
        'page_title': 'Edit Enquiry Source',
        'enquiry_base_url': enquiry_base_url,
        'csrf_input_html': format_html(
            '<input type="hidden" name="csrfmiddlewaretoken" value="{}">',
            get_token(request),
        ),
    }
    return render(request, 'user_analytics/enquiry_source_form.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def enquiry_source_delete_view(request, pk):
    """Soft-delete an enquiry source."""
    from django.contrib import messages
    from django.http import HttpResponseNotFound
    if request.method != 'POST':
        return redirect('user_analytics:enquiry_sources_list')
    try:
        source = EnquirySource.objects.get(pk=pk)
    except EnquirySource.DoesNotExist:
        return HttpResponseNotFound('Enquiry source not found.')
    source.delete(hard_delete=False)
    messages.success(request, 'Enquiry source deactivated.')
    return redirect('user_analytics:enquiry_sources_list')


@login_required
@user_passes_test(is_staff_or_superuser)
def enquiry_source_test_ref_view(request):
    """
    Manual test helper: GET ?ref=TOKEN returns JSON with lookup result and current counts.
    Use to verify the token is recognized before testing a full page visit.
    """
    from django.http import JsonResponse
    ref_token = (request.GET.get('ref') or '').strip()
    if not ref_token:
        return JsonResponse({
            'ok': False,
            'error': 'Missing ref parameter. Use ?ref=YOUR_TOKEN',
            'hint': 'Copy the ref value from your enquiry link (e.g. ...?ref=U14fSkYzV50)',
        }, status=400)
    try:
        es = EnquirySource.objects.filter(
            token=ref_token,
            is_active=True,
            object_status=choices.ObjectStatus.ACTIVE,
        ).first()
        if not es:
            return JsonResponse({
                'ok': True,
                'ref': ref_token[:12] + '...',
                'found': False,
                'message': 'No active EnquirySource with this token. Check token or create the source.',
            })
        page_views = UserActivity.objects.filter(enquiry_source=es).count()
        sessions = UserJourney.objects.filter(enquiry_source=es).count()
        ref_landing_url = request.build_absolute_uri('/ref-landing/') + '?ref=' + es.token
        return JsonResponse({
            'ok': True,
            'ref': ref_token[:12] + '...',
            'found': True,
            'enquiry_source_id': es.id,
            'source_name': es.name,
            'page_views': page_views,
            'sessions': sessions,
            'message': f'Token is valid. To record a visit, open the URL below in a new incognito tab (page must return 200).',
            'ref_landing_url': ref_landing_url,
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@require_GET
def enquiry_ref_hit_api(request):
    """
    Public lightweight endpoint to record enquiry-source hit from frontend on page load.
    Helps when landing pages strip query params after client-side navigation.
    """
    ref_token = (request.GET.get('ref') or '').strip()
    page_path = (request.GET.get('path') or request.path or '/').strip()
    page_title = (request.GET.get('title') or '').strip()[:500]
    if not ref_token:
        return JsonResponse({'ok': False, 'error': 'Missing ref parameter'}, status=400)

    if not page_path.startswith('/'):
        page_path = '/' + page_path

    try:
        source = EnquirySource.objects.filter(
            token=ref_token,
            is_active=True,
            object_status=choices.ObjectStatus.ACTIVE,
        ).first()
        if not source:
            return JsonResponse({'ok': False, 'error': 'Invalid or inactive ref token'}, status=404)

        session_id = request.session.get('analytics_session_id')
        if not session_id:
            import uuid
            session_id = str(uuid.uuid4())
            request.session['analytics_session_id'] = session_id
        # Persist source in session for attribution across subsequent requests.
        request.session['enquiry_source_id'] = source.id
        request.session['enquiry_ref_token'] = source.token

        # Deduplicate rapid duplicate hits from retries/reloads.
        recent_exists = UserActivity.objects.filter(
            session_id=session_id,
            page_path=page_path,
            enquiry_source=source,
            created__gte=timezone.now() - timedelta(seconds=5),
        ).exists()
        if recent_exists:
            return JsonResponse({'ok': True, 'deduped': True})

        ga4_client_id = request.META.get('HTTP_X_GA4_CLIENT_ID') or request.session.get('ga4_client_id')
        user_id = request.user.id if request.user.is_authenticated else None
        referrer = request.META.get('HTTP_REFERER', '')
        user_agent = request.META.get('HTTP_USER_AGENT', '')
        ip_address = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip() or request.META.get('REMOTE_ADDR')

        track_page_view_sync(
            session_id=session_id,
            user_id=user_id,
            ga4_client_id=ga4_client_id,
            page_path=page_path,
            page_url=request.build_absolute_uri(page_path),
            page_title=page_title,
            referrer=referrer,
            user_agent=user_agent,
            ip_address=ip_address,
            utm_source='',
            utm_medium='',
            utm_campaign='',
            utm_term='',
            utm_content='',
            enquiry_source_id=source.id,
        )
        update_user_journey_sync(
            session_id=session_id,
            user_id=user_id,
            ga4_client_id=ga4_client_id,
            page_path=page_path,
            referrer=referrer,
            enquiry_source_id=source.id,
        )
        return JsonResponse({'ok': True, 'recorded': True, 'source_id': source.id})
    except Exception as e:
        logger.exception('enquiry_ref_hit_api failed')
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def payment_status_capture_api(request):
    """
    Capture client-side payment lifecycle statuses (start/cancel/error).
    """
    try:
        payload = json.loads(request.body or '{}')
    except Exception:
        payload = {}

    status_value = (payload.get('status') or '').strip().lower()
    payment_id = payload.get('payment_id')
    order_id = (payload.get('order_id') or '').strip()
    gateway = (payload.get('gateway') or 'razorpay').strip().lower()
    detail = (payload.get('detail') or '').strip()

    if status_value not in {'started', 'cancel', 'error'}:
        return JsonResponse({'ok': False, 'error': 'Invalid status'}, status=400)

    payment = None
    if payment_id:
        payment = Payment.objects.filter(id=payment_id, user=request.user).first()
    if not payment and order_id:
        payment = Payment.objects.filter(gateway_order_id=order_id, user=request.user).first()

    metadata = {
        'gateway': gateway,
        'stage': status_value,
        'order_id': order_id or (payment.gateway_order_id if payment else ''),
        'detail': detail,
    }
    content_type_id = None
    object_id = None
    if payment:
        content_type = ContentType.objects.get_for_model(payment)
        content_type_id = content_type.id
        object_id = payment.id
        metadata['obj_type'] = payment.get_obj_type_display() if hasattr(payment, 'get_obj_type_display') else str(payment.obj_type)
        metadata['gateway_receipt'] = payment.gateway_receipt or ''

    event_type = 'payment_pending' if status_value == 'started' else 'payment_failed'
    event_name_map = {
        'started': 'Payment Checkout Started',
        'cancel': 'Payment Checkout Cancelled',
        'error': 'Payment Checkout Error',
    }
    safe_track_user_event(
        event_type=event_type,
        event_name=event_name_map[status_value],
        user_id=request.user.id,
        event_value=0,
        content_type_id=content_type_id,
        object_id=object_id,
        metadata=metadata,
        session_id=request.session.get('analytics_session_id'),
    )
    return JsonResponse({'ok': True})


@login_required
@user_passes_test(is_staff_or_superuser)
def enquiry_source_qr_view(request, pk):
    """Serve QR code PNG for the enquiry source link (for display and download)."""
    from django.http import HttpResponse, HttpResponseNotFound
    try:
        source = EnquirySource.objects.get(pk=pk, object_status=choices.ObjectStatus.ACTIVE)
    except EnquirySource.DoesNotExist:
        return HttpResponseNotFound('Not found.')
    fallback_base = request.build_absolute_uri('/').rstrip('/')
    full_url = source.get_full_url(fallback_base)
    if not full_url:
        return HttpResponseNotFound('No base URL configured.')
    try:
        import qrcode
        import io
        qr = qrcode.QRCode(version=1, box_size=6, border=2)
        qr.add_data(full_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color='black', back_color='white')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        response = HttpResponse(buf.getvalue(), content_type='image/png')
        if request.GET.get('download'):
            response['Content-Disposition'] = 'attachment; filename="qr-%s.png"' % source.token
        else:
            response['Content-Disposition'] = 'inline; filename="qr-%s.png"' % source.token
        return response
    except Exception:
        logger.exception('QR generation failed')
        return HttpResponseNotFound('QR generation failed.')


def _redirect_chatbot_rules(request, q=None):
    """Preserve ?q= URL search after POST."""
    if q is None:
        q = (request.POST.get('redirect_q') or request.GET.get('q') or '').strip()
    else:
        q = (q or '').strip()
    if q:
        return redirect(f"{reverse('user_analytics:chatbot_rules')}?{urlencode({'q': q})}")
    return redirect('user_analytics:chatbot_rules')


@login_required
@user_passes_test(is_staff_or_superuser)
def chatbot_rules_search_api(request):
    """JSON list of rules filtered by page_url (instant search on admin page)."""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    q = (request.GET.get('q') or '').strip()
    qs = ChatbotPageRule.objects.all().order_by('priority', '-modified')
    if q:
        qs = qs.filter(page_url__icontains=q)
    rules = []
    for r in qs:
        if r.bot_name == 'chat_this_page':
            bot_label = 'Chat this page'
        elif r.bot_name == 'career_counsellor':
            bot_label = 'Career counsellor'
        else:
            bot_label = r.bot_name
        rules.append(
            {
                'id': r.id,
                'page_url': r.page_url,
                'bot_name': r.bot_name,
                'bot_label': bot_label,
                'is_visible': r.is_visible,
                'include_subpages': r.include_subpages,
                'position': r.position,
                'priority': r.priority,
            }
        )
    return JsonResponse({'rules': rules, 'count': len(rules), 'q': q})


@login_required
@user_passes_test(is_staff_or_superuser)
def chatbot_rules_view(request):
    """
    Manage chatbot visibility rules from User Analytics dashboard.
    Fields: page_url, bot_name, show/hide, include_subpages, position.
    """
    from django.contrib import messages

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        if action == 'toggle_bots':
            page_chat_enabled = request.POST.get('chat_this_page_engine') == 'on'
            legacy_chatbot_enabled = request.POST.get('legacy_chatbot_engine') == 'on'

            Configuration.objects.complete().update_or_create(
                key='chat_this_page_engine',
                defaults={
                    'value': 'true' if page_chat_enabled else 'false',
                    'editable': True,
                    'object_status': choices.ObjectStatus.ACTIVE,
                },
            )
            Configuration.objects.complete().update_or_create(
                key='legacy_chatbot_engine',
                defaults={
                    'value': 'true' if legacy_chatbot_enabled else 'false',
                    'editable': True,
                    'object_status': choices.ObjectStatus.ACTIVE,
                },
            )
            messages.success(request, 'Bot enable/disable settings updated.')
            return _redirect_chatbot_rules(request)

        if action == 'update':
            rule_id = request.POST.get('rule_id')
            if not rule_id:
                messages.error(request, 'Rule id is required.')
                return _redirect_chatbot_rules(request)
            rule = get_object_or_404(ChatbotPageRule, pk=rule_id)
            chat_this_page_enabled = str(
                Configuration.get('chat_this_page_engine', 'true', editable=True)
            ).strip().lower() in ('true', '1', 'yes', 'on')
            legacy_chatbot_enabled = str(
                Configuration.get('legacy_chatbot_engine', 'false', editable=True)
            ).strip().lower() in ('true', '1', 'yes', 'on')
            page_url = (request.POST.get('page_url') or '').strip()
            if not page_url:
                messages.error(request, 'Page URL is required.')
                return _redirect_chatbot_rules(request)
            if not page_url.startswith('/'):
                page_url = '/' + page_url
            bot_name = (request.POST.get('bot_name') or '').strip()
            if bot_name not in ('chat_this_page', 'career_counsellor'):
                messages.error(request, 'Invalid bot name.')
                return _redirect_chatbot_rules(request)
            if bot_name == 'chat_this_page' and not chat_this_page_enabled:
                messages.error(request, '"Chat this page" is globally disabled. Enable it first.')
                return _redirect_chatbot_rules(request)
            if bot_name == 'career_counsellor' and not legacy_chatbot_enabled:
                messages.error(request, '"Career Counsellor (cb-root)" is globally disabled. Enable it first.')
                return _redirect_chatbot_rules(request)
            visibility = (request.POST.get('is_visible') or 'show').strip()
            is_visible = visibility == 'show'
            include_subpages = request.POST.get('include_subpages') == 'yes'
            position = (request.POST.get('position') or 'right').strip().lower()
            if position not in ('left', 'right'):
                position = 'right'
            try:
                priority = int(request.POST.get('priority') or 1)
            except Exception:
                priority = 1
            duplicate_qs = ChatbotPageRule.objects.filter(
                page_url=page_url,
                bot_name=bot_name,
                is_visible=is_visible,
                include_subpages=include_subpages,
                position=position,
                object_status=choices.ObjectStatus.ACTIVE,
            ).exclude(pk=rule.pk)
            if duplicate_qs.exists():
                messages.warning(
                    request,
                    'Another rule already exists with the same URL, bot, visibility, nesting, and position.',
                )
                return _redirect_chatbot_rules(request)
            rule.page_url = page_url
            rule.bot_name = bot_name
            rule.is_visible = is_visible
            rule.include_subpages = include_subpages
            rule.position = position
            rule.priority = priority
            rule.save()
            messages.success(request, 'Chatbot rule updated.')
            return _redirect_chatbot_rules(request)

        if action == 'create':
            chat_this_page_enabled = str(
                Configuration.get('chat_this_page_engine', 'true', editable=True)
            ).strip().lower() in ('true', '1', 'yes', 'on')
            legacy_chatbot_enabled = str(
                Configuration.get('legacy_chatbot_engine', 'false', editable=True)
            ).strip().lower() in ('true', '1', 'yes', 'on')
            page_url = (request.POST.get('page_url') or '').strip()
            if not page_url:
                messages.error(request, 'Page URL is required.')
                return _redirect_chatbot_rules(request)
            if not page_url.startswith('/'):
                page_url = '/' + page_url
            bot_name = (request.POST.get('bot_name') or '').strip()
            if bot_name not in ('chat_this_page', 'career_counsellor'):
                messages.error(request, 'Invalid bot name.')
                return _redirect_chatbot_rules(request)
            if bot_name == 'chat_this_page' and not chat_this_page_enabled:
                messages.error(request, '"Chat this page" is globally disabled. Enable it first to add rules.')
                return _redirect_chatbot_rules(request)
            if bot_name == 'career_counsellor' and not legacy_chatbot_enabled:
                messages.error(request, '"Career Counsellor (cb-root)" is globally disabled. Enable it first to add rules.')
                return _redirect_chatbot_rules(request)
            visibility = (request.POST.get('is_visible') or 'show').strip()
            is_visible = (visibility == 'show')
            include_subpages = (request.POST.get('include_subpages') == 'yes')
            position = (request.POST.get('position') or 'right').strip().lower()
            if position not in ('left', 'right'):
                position = 'right'
            try:
                priority = int(request.POST.get('priority') or 1)
            except Exception:
                priority = 1
            duplicate_qs = ChatbotPageRule.objects.filter(
                page_url=page_url,
                bot_name=bot_name,
                is_visible=is_visible,
                include_subpages=include_subpages,
                position=position,
                object_status=choices.ObjectStatus.ACTIVE,
            )
            if duplicate_qs.exists():
                messages.warning(
                    request,
                    'Duplicate rule already exists for same URL, bot, visibility, nesting, and position.',
                )
                return _redirect_chatbot_rules(request)
            ChatbotPageRule.objects.create(
                page_url=page_url,
                bot_name=bot_name,
                is_visible=is_visible,
                include_subpages=include_subpages,
                position=position,
                priority=priority,
            )
            messages.success(request, 'Chatbot rule added.')
            return _redirect_chatbot_rules(request)

        if action == 'delete':
            ids = request.POST.getlist('ids')
            if not ids:
                messages.warning(request, 'No rules selected.')
                return _redirect_chatbot_rules(request)
            delete_result = ChatbotPageRule.objects.filter(id__in=ids).delete()
            deleted = delete_result[0] if isinstance(delete_result, tuple) else int(delete_result or 0)
            messages.success(request, f'{deleted} rule(s) deleted.')
            return _redirect_chatbot_rules(request)

        messages.error(request, 'Invalid action.')
        return _redirect_chatbot_rules(request)

    rules = ChatbotPageRule.objects.all().order_by('priority', '-modified')
    chat_this_page_enabled = str(
        Configuration.get('chat_this_page_engine', 'true', editable=True)
    ).strip().lower() in ('true', '1', 'yes', 'on')
    legacy_chatbot_enabled = str(
        Configuration.get('legacy_chatbot_engine', 'false', editable=True)
    ).strip().lower() in ('true', '1', 'yes', 'on')
    context = {
        'rules': rules,
        'chatbot_rules_search_url': reverse('user_analytics:chatbot_rules_search'),
        'chat_this_page_enabled': chat_this_page_enabled,
        'legacy_chatbot_enabled': legacy_chatbot_enabled,
        'page_title': 'Chatbot Rules',
        'csrf_input_html': format_html(
            '<input type="hidden" name="csrfmiddlewaretoken" value="{}">',
            get_token(request),
        ),
    }
    return render(request, 'user_analytics/chatbot_rules.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def visitors_detail(request):
    """Detail page for visitors/sessions with filters"""
    from user_analytics.tasks import sync_ga4_sessions_task
    from django.utils import timezone as tz
    from django.db import connection
    
    # Import GA4Session only if table exists
    GA4Session = None
    try:
        from user_analytics.models import GA4Session as GA4SessionModel
        # Check if table exists
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*) 
                FROM information_schema.tables 
                WHERE table_schema = DATABASE() 
                AND table_name = 'user_analytics_ga4session'
            """)
            if cursor.fetchone()[0] > 0:
                GA4Session = GA4SessionModel
    except Exception:
        pass
    
    time_period = request.GET.get('period', 'today')
    search_query = request.GET.get('search', '')
    source_filter = request.GET.get('source', '')
    device_filter = request.GET.get('device', '')
    entry_page_filter = request.GET.get('entry_page', '')
    exit_page_filter = request.GET.get('exit_page', '')
    country_filter = request.GET.get('country', '')
    user_type_filter = request.GET.get('user_type', 'all')  # all, registered, new
    page_number = request.GET.get('page', 1)
    
    logger.info("=" * 80)
    logger.info(f"VISITORS DETAIL - Time Period: {time_period}")
    logger.info(f"Source Filter (raw): {source_filter}")
    logger.info(f"Device Filter (raw): {device_filter}")
    logger.info(f"Entry Page Filter (raw): {entry_page_filter}")
    logger.info(f"Exit Page Filter (raw): {exit_page_filter}")
    logger.info(f"Country Filter (raw): {country_filter}")
    logger.info(f"Search Query: {search_query}")
    
    # Decode all URL-encoded filters
    if source_filter:
        source_filter = unquote(source_filter)
        logger.info(f"Source Filter (decoded): {source_filter}")
    if device_filter:
        device_filter = unquote(device_filter)
        logger.info(f"Device Filter (decoded): {device_filter}")
    if entry_page_filter:
        entry_page_filter = unquote(entry_page_filter)
        logger.info(f"Entry Page Filter (decoded): {entry_page_filter}")
    if exit_page_filter:
        exit_page_filter = unquote(exit_page_filter)
        logger.info(f"Exit Page Filter (decoded): {exit_page_filter}")
    if country_filter:
        country_filter = unquote(country_filter)
        logger.info(f"Country Filter (decoded): {country_filter}")
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    if start_date is not None:
        logger.info(f"Date Range: {start_date.strftime('%Y-%m-%d %H:%M:%S')} to {end_date.strftime('%Y-%m-%d %H:%M:%S')}")
    else:
        logger.info("Date Range: All Time (no date filtering)")
    
    # Check if GA4 data needs to be synced
    # Check if we have recent GA4 data in DB for this period
    ga4_needs_sync = False
    if GA4Session is not None and start_date is not None:
        try:
            latest_ga4_sync = GA4Session.objects.filter(
                date__gte=start_date.date(),
                date__lte=end_date.date()
            ).order_by('-synced_at').first()
            
            if not latest_ga4_sync:
                # No GA4 data in DB, trigger sync
                ga4_needs_sync = True
                logger.info("No GA4 data found in DB, triggering sync")
            else:
                # Check if sync is older than 1 hour
                sync_age = tz.now() - latest_ga4_sync.synced_at
                if sync_age.total_seconds() > 3600:  # 1 hour
                    ga4_needs_sync = True
                    logger.info(f"GA4 data is {sync_age.total_seconds()/60:.1f} minutes old, triggering sync")
        except Exception as e:
            logger.warning(f"Error checking GA4 sync status: {e}")
    else:
        logger.info("GA4Session table does not exist yet - migration needed")
        ga4_needs_sync = False  # Don't try to sync if table doesn't exist
    
    # Trigger sync if needed (in background)
    if ga4_needs_sync:
        try:
            sync_ga4_sessions_task.delay(time_period=time_period, link_users=True)
            logger.info("GA4 sync task triggered in background")
        except Exception as e:
            logger.error(f"Error triggering GA4 sync: {e}")
    
    # Start with UserJourney to apply entry_page/exit_page filters early
    # This is more efficient since UserJourney has these fields indexed
    journey_query = UserJourney.objects.all()
    if start_date is not None:
        journey_query = journey_query.filter(
            start_time__gte=start_date,
            start_time__lte=end_date
        )
    
    # Apply entry_page filter early
    if entry_page_filter:
        # Try exact match first, then contains for flexibility
        # This handles cases where entry_page might have trailing slashes or query params
        journey_query = journey_query.filter(
            Q(entry_page__iexact=entry_page_filter) |
            Q(entry_page__icontains=entry_page_filter)
        )
        logger.info(f"Filtering for entry_page: {entry_page_filter}")
    
    # Apply exit_page filter early
    if exit_page_filter:
        journey_query = journey_query.filter(
            Q(exit_page__iexact=exit_page_filter) |
            Q(exit_page__icontains=exit_page_filter)
        )
        logger.info(f"Filtering for exit_page: {exit_page_filter}")
    
    # Apply device filter on UserJourney if available
    if device_filter:
        journey_query = journey_query.filter(device_type__iexact=device_filter)
        logger.info(f"Filtering for device: {device_filter}")
    
    # Apply country filter on UserJourney if available
    if country_filter:
        journey_query = journey_query.filter(country__icontains=country_filter)
        logger.info(f"Filtering for country: {country_filter}")
    
    # Apply user type filter
    if user_type_filter == 'registered':
        journey_query = journey_query.filter(user__isnull=False)
        logger.info("Filtering for registered users only")
    elif user_type_filter == 'new':
        # New users: sessions without a linked user (anonymous/new visitors)
        journey_query = journey_query.filter(user__isnull=True)
        logger.info("Filtering for new users only")
    # 'all' means no user filter
    
    # Get session IDs from UserJourney that match the filters above
    # Only get journey session IDs if we have entry/exit/device/country filters
    has_journey_filters = entry_page_filter or exit_page_filter or device_filter or country_filter
    journey_session_ids = None
    
    if has_journey_filters:
        journey_session_ids = set(journey_query.values_list('session_id', flat=True).distinct())
        logger.info(f"Found {len(journey_session_ids)} session IDs from UserJourney matching entry/exit/device/country filters")
    else:
        logger.info("No journey filters applied, will use UserActivity as primary source")
    
    # Now filter by source from UserActivity (since source is in UserActivity)
    session_query = UserActivity.objects.all()
    if start_date is not None:
        session_query = session_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    
    # If we have journey session IDs from filters, restrict UserActivity query to those sessions
    if journey_session_ids is not None:
        if len(journey_session_ids) == 0:
            # No matching journeys, set empty session_ids and skip further filtering
            session_ids = []
            logger.info("No matching journeys found, returning empty result")
        else:
            session_query = session_query.filter(session_id__in=journey_session_ids)
    
    # Apply source filter (always check UserActivity for source)
    if source_filter and (journey_session_ids is None or len(journey_session_ids) > 0):
        # Handle variations of "direct" traffic
        # GA4 might store it as "(direct)", "direct", "(not set)", or empty/null
        if source_filter.lower() in ['(direct)', 'direct', '(not set)']:
            # Match direct traffic - could be stored as various forms
            session_query = session_query.filter(
                Q(utm_source__iexact='(direct)') |
                Q(utm_source__iexact='direct') |
                Q(utm_source__iexact='(not set)') |
                Q(utm_source__isnull=True) |
                Q(utm_source='')
            )
            logger.info("Filtering for direct traffic (handling variations)")
        else:
            # Use referrer_source_q for Google, Facebook, iapply.io (matches utm_source + referrer)
            from user_analytics.utils import referrer_source_q
            if source_filter.lower() in ('google', 'facebook', 'iapply', 'iapply.io'):
                session_query = session_query.filter(referrer_source_q(source_filter))
                logger.info(f"Filtering for referrer source: {source_filter}")
            else:
                session_query = session_query.filter(utm_source__iexact=source_filter)
                logger.info(f"Filtering for source: {source_filter}")
    
    # If device filter wasn't applied to UserJourney (no journey filters case), apply to UserActivity
    if device_filter and not has_journey_filters:
        session_query = session_query.filter(device_type__iexact=device_filter)
        logger.info(f"Filtering for device (UserActivity): {device_filter}")
    
    # Get final session IDs (only if we haven't already set it to empty)
    if journey_session_ids is None or (journey_session_ids is not None and len(journey_session_ids) > 0):
        session_ids = list(session_query.values_list('session_id', flat=True).distinct())
        logger.info(f"Found {len(session_ids)} unique session IDs matching all filters")
    
    # Check if we have any data at all in the date range
    total_activities_query = UserActivity.objects.all()
    if start_date is not None:
        total_activities_query = total_activities_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    total_activities = total_activities_query.count()
    logger.info(f"Total UserActivity records in date range: {total_activities}")
    
    # Check unique sources in database for debugging
    unique_sources_query = UserActivity.objects.exclude(utm_source__isnull=True).exclude(utm_source='')
    if start_date is not None:
        unique_sources_query = unique_sources_query.filter(
            created__gte=start_date,
            created__lte=end_date
        )
    unique_sources = unique_sources_query.values_list('utm_source', flat=True).distinct()[:10]
    logger.info(f"Sample unique sources in database: {list(unique_sources)}")
    
    # Also include GA4Session data from DB if available
    ga4_sessions_from_db = []
    if GA4Session is not None:
        try:
            ga4_query = GA4Session.objects.all()
            if start_date is not None:
                ga4_query = ga4_query.filter(
                    date__gte=start_date.date(),
                    date__lte=end_date.date()
                )
            
            # Apply filters to GA4Session
            if source_filter:
                if source_filter.lower() in ['(direct)', 'direct', '(not set)']:
                    ga4_query = ga4_query.filter(
                        Q(source__iexact='(direct)') |
                        Q(source__iexact='direct') |
                        Q(source__iexact='(not set)') |
                        Q(source__isnull=True) |
                        Q(source='')
                    )
                else:
                    ga4_query = ga4_query.filter(source__iexact=source_filter)
            
            if device_filter:
                ga4_query = ga4_query.filter(device__iexact=device_filter)
            
            if country_filter:
                ga4_query = ga4_query.filter(country__icontains=country_filter)
            
            if entry_page_filter:
                ga4_query = ga4_query.filter(
                    Q(entry_page__iexact=entry_page_filter) |
                    Q(entry_page__icontains=entry_page_filter)
                )
            
            if exit_page_filter:
                ga4_query = ga4_query.filter(
                    Q(exit_page__iexact=exit_page_filter) |
                    Q(exit_page__icontains=exit_page_filter)
                )
            
            # Apply user type filter
            if user_type_filter == 'registered':
                ga4_query = ga4_query.filter(user__isnull=False)
            elif user_type_filter == 'new':
                ga4_query = ga4_query.filter(user__isnull=True)
            
            # Get GA4 sessions
            for ga4_session in ga4_query[:1000]:  # Limit to prevent memory issues
                ga4_sessions_from_db.append({
                    'session_id': ga4_session.django_session_id or f"ga4-{ga4_session.ga4_client_id[:10]}",
                    'user': ga4_session.user,
                    'first_visit': tz.make_aware(datetime.combine(ga4_session.date, datetime.min.time())),
                    'page_views': ga4_session.pageviews,
                    'device_type': ga4_session.device or 'Unknown',
                    'utm_source': ga4_session.source or 'Direct',
                    'country': ga4_session.country,
                    'entry_page': ga4_session.entry_page,
                    'is_ga4_data': True,
                    'sessions_count': ga4_session.sessions_count,
                })
            
            logger.info(f"Found {len(ga4_sessions_from_db)} GA4 sessions from DB")
        except Exception as e:
            logger.error(f"Error fetching GA4 sessions from DB: {e}", exc_info=True)
    else:
        logger.info("GA4Session table does not exist - skipping GA4 data query")
    
    # If no session IDs found, try to get GA4 aggregated data from API
    use_ga4_fallback = len(session_ids) == 0 and total_activities == 0 and len(ga4_sessions_from_db) == 0
    ga4_sessions_data = None
    
    if use_ga4_fallback and (source_filter or country_filter or device_filter or entry_page_filter or exit_page_filter):
        logger.info("No database records found. Fetching aggregated GA4 data from API...")
        try:
            ga4_service = GA4Service()
            ga4_sessions_data = ga4_service.get_sessions_by_filters(
                time_period=time_period,
                source=source_filter,
                country=country_filter,
                device=device_filter,
                entry_page=entry_page_filter,
                exit_page=exit_page_filter,
                limit=1000,
                use_cache=False
            )
            if ga4_sessions_data:
                logger.info(f"Found {len(ga4_sessions_data)} aggregated session records from GA4 API")
        except Exception as e:
            logger.error(f"Error fetching GA4 sessions data: {e}", exc_info=True)
    
    # Get first activity for each session
    session_details = []
    for session_id in session_ids:
        first_activity = UserActivity.objects.filter(
            session_id=session_id
        ).order_by('created').first()
        
        if first_activity:
            # Entry/exit page, device, and country filters are already applied at the query level
            # No need to check them again here if they were applied to journey_query
            
            # However, if country filter wasn't applied earlier (no journey filters case), check it here
            if country_filter and not has_journey_filters:
                # Country filter wasn't applied to journey_query, check here
                country_match = False
                if first_activity.country and country_filter.lower() in first_activity.country.lower():
                    country_match = True
                else:
                    journey = UserJourney.objects.filter(session_id=session_id).first()
                    if journey and journey.country and country_filter.lower() in journey.country.lower():
                        country_match = True
                if not country_match:
                    continue
            
            # Apply search filter early
            if search_query:
                user_match = first_activity.user and first_activity.user.email and search_query.lower() in first_activity.user.email.lower()
                session_match = search_query.lower() in session_id.lower()
                if not (user_match or session_match):
                    continue
            
            session_details.append({
                'session_id': session_id,
                'user': first_activity.user,
                'first_visit': first_activity.created,
                'page_views': UserActivity.objects.filter(session_id=session_id).count(),
                'device_type': first_activity.device_type,
                'utm_source': first_activity.utm_source or 'Direct',
            })
    
    # If we have GA4 data but no database data, convert GA4 data to session_details format
    if ga4_sessions_data and len(session_details) == 0:
        logger.info("Converting GA4 aggregated data to session details format...")
        for ga4_session in ga4_sessions_data:
            # Create a pseudo-session detail from GA4 aggregated data
            try:
                from datetime import datetime
                session_date = datetime.strptime(ga4_session.get('date', ''), '%Y%m%d')
            except:
                session_date = timezone.now()
            
            session_details.append({
                'session_id': f"ga4-{ga4_session.get('date', 'unknown')}-{ga4_session.get('source', 'unknown')[:10]}",
                'user': None,  # GA4 doesn't provide user info
                'first_visit': session_date,
                'page_views': ga4_session.get('pageviews', 0),
                'device_type': ga4_session.get('device', 'Unknown'),
                'utm_source': ga4_session.get('source', 'Direct'),
                'country': ga4_session.get('country', 'Unknown'),
                'entry_page': ga4_session.get('entry_page', 'N/A'),
                'sessions_count': ga4_session.get('sessions', 0),
                'is_ga4_data': True,  # Flag to indicate this is GA4 data
            })
        logger.info(f"Converted {len(session_details)} GA4 records to session details")
    
    # Merge GA4 sessions from DB with regular session details
    session_details.extend(ga4_sessions_from_db)
    
    # Sort by first visit
    session_details.sort(key=lambda x: x['first_visit'], reverse=True)
    
    # Calculate totals before pagination
    total_count = len(session_details)
    
    logger.info(f"Total sessions found: {total_count}")
    logger.info("=" * 80)
    
    # Manual pagination for list
    items_per_page = 25
    try:
        page_number = int(page_number)
    except (ValueError, TypeError):
        page_number = 1
    
    start_index = (page_number - 1) * items_per_page
    end_index = start_index + items_per_page
    sessions_page_data = session_details[start_index:end_index]
    
    # Create a paginator-like object for template compatibility
    from django.core.paginator import Page, Paginator
    paginator = Paginator(range(total_count), items_per_page)
    try:
        paginator_page = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        paginator_page = paginator.page(1)
    
    # Create a custom page object with our data
    class CustomPage:
        def __init__(self, data, paginator_page):
            self.object_list = data
            self.number = paginator_page.number
            self.paginator = paginator_page.paginator
            self.has_previous = paginator_page.has_previous()
            self.has_next = paginator_page.has_next()
            self.previous_page_number = paginator_page.previous_page_number() if self.has_previous else None
            self.next_page_number = paginator_page.next_page_number() if self.has_next else None
            self.start_index = start_index + 1
        
        @property
        def has_other_pages(self):
            return self.paginator.num_pages > 1
    
    sessions_page = CustomPage(sessions_page_data, paginator_page)
    
    # Build filter params for pagination links
    filter_params = {
        'period': time_period,
    }
    if source_filter:
        filter_params['source'] = source_filter
    if device_filter:
        filter_params['device'] = device_filter
    if entry_page_filter:
        filter_params['entry_page'] = entry_page_filter
    if exit_page_filter:
        filter_params['exit_page'] = exit_page_filter
    if country_filter:
        filter_params['country'] = country_filter
    if search_query:
        filter_params['search'] = search_query
    if user_type_filter and user_type_filter != 'all':
        filter_params['user_type'] = user_type_filter
    
    context = {
        'sessions': sessions_page,
        'time_period': time_period,
        'start_date': start_date,
        'end_date': end_date,
        'search_query': search_query,
        'source_filter': source_filter,
        'device_filter': device_filter,
        'entry_page_filter': entry_page_filter,
        'exit_page_filter': exit_page_filter,
        'country_filter': country_filter,
        'user_type_filter': user_type_filter,
        'filter_params': filter_params,
        'total_count': total_count,
        'total_activities_in_range': total_activities,
        'use_ga4_fallback': use_ga4_fallback,
        'has_database_data': total_activities > 0,  # Flag to indicate if database has any data
        'ga4_sessions_data': ga4_sessions_data,  # GA4 aggregated data if available
        'ga4_needs_sync': ga4_needs_sync,  # Flag to show sync status
        'page_title': 'Visitors/Sessions',
    }
    
    logger.info("=" * 80)
    
    return render(request, 'user_analytics/visitors_detail.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def visitors_filter_options_api(request):
    """
    API endpoint to return filter options for autocomplete/select dropdowns.
    Returns unique values for sources, devices, countries, and entry pages.
    """
    from user_analytics.models import UserActivity, UserJourney
    from django.db.models import Q
    from django.db import connection
    
    # Import GA4Session only if table exists
    GA4Session = None
    try:
        from user_analytics.models import GA4Session as GA4SessionModel
        # Check if table exists
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*) 
                FROM information_schema.tables 
                WHERE table_schema = DATABASE() 
                AND table_name = 'user_analytics_ga4session'
            """)
            if cursor.fetchone()[0] > 0:
                GA4Session = GA4SessionModel
    except Exception:
        pass
    
    filter_type = request.GET.get('filter_type', '')  # source, device, country, entry_page
    time_period = request.GET.get('period', '30days')
    search_query = request.GET.get('q', '')  # Search term for filtering options
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    options = set()
    
    try:
        if filter_type == 'source':
            # Preferred referrer sources (always show in dropdown for quick filter)
            preferred_sources = ['google', 'facebook', 'iapply']
            options.update(preferred_sources)
            
            # Get sources from UserActivity, UserJourney, and GA4Session
            sources_ua = UserActivity.objects.all()
            sources_uj = UserJourney.objects.all()
            
            if start_date is not None:
                sources_ua = sources_ua.filter(created__gte=start_date, created__lte=end_date)
                sources_uj = sources_uj.filter(start_time__gte=start_date, start_time__lte=end_date)
            
            sources_ua = sources_ua.exclude(utm_source__isnull=True).exclude(utm_source='').values_list('utm_source', flat=True).distinct()
            sources_uj = sources_uj.exclude(utm_source__isnull=True).exclude(utm_source='').values_list('utm_source', flat=True).distinct()
            
            # Only query GA4Session if table exists
            try:
                sources_ga4 = GA4Session.objects.filter(
                ).exclude(source__isnull=True).exclude(source='')
                if start_date is not None:
                    sources_ga4 = sources_ga4.filter(
                        date__gte=start_date.date(),
                        date__lte=end_date.date()
                    )
                sources_ga4 = sources_ga4.values_list('source', flat=True).distinct()
                options.update(sources_ga4)
            except Exception:
                # Table doesn't exist, skip
                pass
            
            options.update(sources_ua)
            options.update(sources_uj)
            
        elif filter_type == 'device':
            # Get devices from UserActivity, UserJourney, and GA4Session
            devices_ua = UserActivity.objects.all()
            devices_uj = UserJourney.objects.all()
            
            if start_date is not None:
                devices_ua = devices_ua.filter(created__gte=start_date, created__lte=end_date)
                devices_uj = devices_uj.filter(start_time__gte=start_date, start_time__lte=end_date)
            
            devices_ua = devices_ua.exclude(device_type__isnull=True).exclude(device_type='').values_list('device_type', flat=True).distinct()
            devices_uj = devices_uj.exclude(device_type__isnull=True).exclude(device_type='').values_list('device_type', flat=True).distinct()
            
            # Only query GA4Session if table exists
            try:
                devices_ga4 = GA4Session.objects.all()
                if start_date is not None:
                    devices_ga4 = devices_ga4.filter(
                        date__gte=start_date.date(),
                        date__lte=end_date.date()
                    )
                devices_ga4 = devices_ga4.exclude(device__isnull=True).exclude(device='').values_list('device', flat=True).distinct()
                options.update(devices_ga4)
            except Exception:
                # Table doesn't exist, skip
                pass
            
            options.update(devices_ua)
            options.update(devices_uj)
            
        elif filter_type == 'country':
            # Get countries from UserActivity, UserJourney, and GA4Session
            countries_ua = UserActivity.objects.all()
            countries_uj = UserJourney.objects.all()
            
            if start_date is not None:
                countries_ua = countries_ua.filter(created__gte=start_date, created__lte=end_date)
                countries_uj = countries_uj.filter(start_time__gte=start_date, start_time__lte=end_date)
            
            countries_ua = countries_ua.exclude(country__isnull=True).exclude(country='').values_list('country', flat=True).distinct()
            countries_uj = countries_uj.exclude(country__isnull=True).exclude(country='').values_list('country', flat=True).distinct()
            
            # Only query GA4Session if table exists
            try:
                countries_ga4 = GA4Session.objects.all()
                if start_date is not None:
                    countries_ga4 = countries_ga4.filter(
                        date__gte=start_date.date(),
                        date__lte=end_date.date()
                    )
                countries_ga4 = countries_ga4.exclude(country__isnull=True).exclude(country='').values_list('country', flat=True).distinct()
                options.update(countries_ga4)
            except Exception:
                # Table doesn't exist, skip
                pass
            
            options.update(countries_ua)
            options.update(countries_uj)
            
        elif filter_type == 'entry_page':
            # Get entry pages from UserJourney and GA4Session
            entry_pages_uj = UserJourney.objects.all()
            if start_date is not None:
                entry_pages_uj = entry_pages_uj.filter(start_time__gte=start_date, start_time__lte=end_date)
            entry_pages_uj = entry_pages_uj.exclude(entry_page__isnull=True).exclude(entry_page='').values_list('entry_page', flat=True).distinct()
            
            # Only query GA4Session if table exists
            try:
                entry_pages_ga4 = GA4Session.objects.all()
                if start_date is not None:
                    entry_pages_ga4 = entry_pages_ga4.filter(
                        date__gte=start_date.date(),
                        date__lte=end_date.date()
                    )
                entry_pages_ga4 = entry_pages_ga4.exclude(entry_page__isnull=True).exclude(entry_page='').values_list('entry_page', flat=True).distinct()
                options.update(entry_pages_ga4)
            except Exception:
                # Table doesn't exist, skip
                pass
            
            options.update(entry_pages_uj)
        
        # Filter by search query if provided
        if search_query:
            search_lower = search_query.lower()
            options = {opt for opt in options if opt and search_lower in str(opt).lower()}
        
        # Convert to sorted list and limit
        options_list = sorted([str(opt) for opt in options if opt])[:100]  # Limit to 100 options
        
        # For source filter, include display labels for Google, Facebook, iapply.io
        response_data = {
            'success': True,
            'options': options_list,
            'count': len(options_list),
            'filter_type': filter_type,
        }
        if filter_type == 'source':
            response_data['source_labels'] = {
                'google': 'Google',
                'facebook': 'Facebook',
                'iapply': 'iapply.io',
            }
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.error(f"Error fetching filter options: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e),
            'options': [],
        }, status=500)


@login_required
@user_passes_test(is_staff_or_superuser)
def pageviews_detail(request):
    """Detail page for pageviews with filters"""
    time_period = request.GET.get('period', '30days')
    search_query = request.GET.get('search', '')
    path_filter = request.GET.get('path', '')
    source_filter = request.GET.get('source', '')
    page_number = request.GET.get('page', 1)
    
    logger.info("=" * 80)
    logger.info(f"PAGEVIEWS DETAIL - Time Period: {time_period}")
    logger.info(f"Path Filter (raw): {path_filter}")
    logger.info(f"Source Filter: {source_filter}")
    logger.info(f"Search Query: {search_query}")
    
    # Calculate date range
    start_date, end_date = get_date_range_from_period(time_period, default_days=30)
    
    if start_date is not None:
        logger.info(f"Date Range: {start_date.strftime('%Y-%m-%d %H:%M:%S')} to {end_date.strftime('%Y-%m-%d %H:%M:%S')}")
    else:
        logger.info("Date Range: All Time (no date filtering)")
    
    # Use GA4 data directly (skip database for now)
    ga4_pageviews = None
    data_source = 'ga4'
    pageviews_page = None
    total_count = 0
    
    # If path filter is set, get data from GA4
    if path_filter:
        # Decode URL-encoded path
        path_filter_decoded = unquote(path_filter)
        # Store decoded version for display
        path_filter = path_filter_decoded
        
        logger.info(f"Path Filter (decoded): {path_filter_decoded}")
        logger.info("Fetching data from GA4...")
        
        ga4_service = GA4Service()
        ga4_pageviews = ga4_service.get_pageviews_by_path(
            path_filter_decoded, 
            time_period=time_period, 
            limit=1000, 
            use_cache=False
        )
        
        if ga4_pageviews:
            logger.info(f"Found {len(ga4_pageviews)} pageview records from GA4")
            
            # Apply source filter if provided
            if source_filter:
                source_filter_decoded = unquote(source_filter)
                ga4_pageviews = [p for p in ga4_pageviews if p.get('source', '').lower() == source_filter_decoded.lower()]
                logger.info(f"After source filter: {len(ga4_pageviews)} records")
            
            # Apply search filter if provided
            if search_query:
                search_lower = search_query.lower()
                ga4_pageviews = [
                    p for p in ga4_pageviews 
                    if search_lower in p.get('page_title', '').lower() 
                    or search_lower in p.get('page_path', '').lower()
                    or search_lower in p.get('source', '').lower()
                ]
                logger.info(f"After search filter: {len(ga4_pageviews)} records")
            
            # Paginate GA4 data
            if ga4_pageviews:
                total_count = len(ga4_pageviews)
                items_per_page = 25
                try:
                    page_number = int(page_number)
                except (ValueError, TypeError):
                    page_number = 1
                
                start_index = (page_number - 1) * items_per_page
                end_index = start_index + items_per_page
                ga4_pageviews_page = ga4_pageviews[start_index:end_index]
                
                # Create paginator-like object for GA4 data
                paginator = Paginator(range(total_count), items_per_page)
                try:
                    paginator_page = paginator.page(page_number)
                except (PageNotAnInteger, EmptyPage):
                    paginator_page = paginator.page(1)
                
                # Create custom page object
                class GA4Page:
                    def __init__(self, data, paginator_page, total_count):
                        self.object_list = data
                        self.number = paginator_page.number
                        self.paginator = paginator_page.paginator
                        self.has_previous = paginator_page.has_previous()
                        self.has_next = paginator_page.has_next()
                        self.previous_page_number = paginator_page.previous_page_number() if self.has_previous else None
                        self.next_page_number = paginator_page.next_page_number() if self.has_next else None
                        self.start_index = (paginator_page.number - 1) * items_per_page + 1
                    
                    @property
                    def has_other_pages(self):
                        return self.paginator.num_pages > 1
                
                pageviews_page = GA4Page(ga4_pageviews_page, paginator_page, total_count)
                logger.info(f"Created GA4Page with {len(ga4_pageviews_page)} items on page {page_number}")
            else:
                # No data after filtering
                pageviews_page = None
                logger.info("No data after applying filters")
        else:
            logger.info("No GA4 data found for this path")
            pageviews_page = None
    else:
        # No path filter - show message that path filter is required
        logger.info("No path filter provided. GA4 requires a path filter to show pageview details.")
        pageviews_page = None
    
    logger.info("=" * 80)
    logger.info(f"FINAL RESULTS:")
    logger.info(f"  Data Source: {data_source}")
    logger.info(f"  Total pageviews matching filters: {total_count}")
    logger.info(f"  Page number: {page_number}")
    if pageviews_page and hasattr(pageviews_page, 'object_list'):
        logger.info(f"  Pageviews on this page: {len(pageviews_page.object_list)}")
        if pageviews_page.object_list:
            logger.info(f"  Sample data (first item): {pageviews_page.object_list[0]}")
    else:
        logger.info(f"  Pageviews on this page: None (no data)")
    logger.info("=" * 80)
    
    context = {
        'pageviews': pageviews_page,
        'time_period': time_period,
        'start_date': start_date,
        'end_date': end_date,
        'search_query': search_query,
        'path_filter': path_filter,
        'source_filter': source_filter,
        'total_count': total_count,
        'page_title': 'Pageviews',
        'data_source': data_source,  # 'database' or 'ga4'
    }
    
    return render(request, 'user_analytics/pageviews_detail.html', context)


@login_required
@user_passes_test(is_staff_or_superuser)
def pageviews_paths_api(request):
    """API endpoint to get all available page paths for autocomplete"""
    time_period = request.GET.get('period', '30days')
    
    logger.info("=" * 80)
    logger.info(f"PAGEVIEWS PATHS API - Time Period: {time_period}")
    
    ga4_service = GA4Service()
    paths = ga4_service.get_all_page_paths(time_period=time_period, limit=1000, use_cache=False)
    
    logger.info(f"Found {len(paths)} unique paths")
    if paths:
        logger.info(f"Sample paths (first 10): {paths[:10]}")
    logger.info("=" * 80)
    
    return JsonResponse({
        'success': True,
        'paths': paths,
        'count': len(paths)
    })


@csrf_exempt
@login_required
@user_passes_test(is_staff_or_superuser)
def web_owner_optional_data_api(request):
    """API endpoint for optional web owner dashboard data (loaded via AJAX)"""
    time_period = request.GET.get('period', 'today')
    data_type = request.GET.get('type', 'all')  # 'entry_pages', 'exit_pages', 'users_by_country', 'engagement', 'trends', 'all'
    
    try:
        ga4_service = GA4Service()
        response_data = {
            'success': True,
            'time_period': time_period,
        }
        
        # Load requested data types
        if data_type in ['all', 'entry_pages']:
            try:
                entry_pages = ga4_service.get_entry_pages(time_period, limit=10, use_cache=True)
                response_data['entry_pages'] = entry_pages if entry_pages else []
            except Exception as e:
                logger.warning(f"Error fetching entry pages: {e}")
                response_data['entry_pages'] = []
        
        if data_type in ['all', 'exit_pages']:
            try:
                exit_pages = ga4_service.get_exit_pages(time_period, limit=10, use_cache=True)
                response_data['exit_pages'] = exit_pages if exit_pages else []
            except Exception as e:
                logger.warning(f"Error fetching exit pages: {e}")
                response_data['exit_pages'] = []
        
        if data_type in ['all', 'users_by_country']:
            try:
                users_by_country = ga4_service.get_users_by_country(time_period, limit=10, use_cache=True)
                response_data['users_by_country'] = users_by_country if users_by_country else []
            except Exception as e:
                logger.warning(f"Error fetching users by country: {e}")
                response_data['users_by_country'] = []
        
        if data_type in ['all', 'engagement']:
            try:
                engagement = ga4_service.get_user_engagement(time_period, use_cache=True)
                response_data['engagement'] = engagement if engagement else None
            except Exception as e:
                logger.warning(f"Error fetching engagement: {e}")
                response_data['engagement'] = None
        
        if data_type in ['all', 'trends']:
            try:
                top_pages_with_trends = ga4_service.get_top_pages_with_trends(time_period, limit=20, use_cache=True)
                response_data['top_pages_with_trends'] = top_pages_with_trends if top_pages_with_trends else []
            except Exception as e:
                logger.warning(f"Error fetching top pages with trends: {e}")
                response_data['top_pages_with_trends'] = []
        
        return JsonResponse(response_data)
    
    except Exception as e:
        logger.error(f"Error in web_owner_optional_data_api: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
