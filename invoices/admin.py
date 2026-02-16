"""
Accounts section: Invoice list with date filters, download PDF, Excel export, resend, callback health alert.
"""
import io
import json
import traceback
from django.contrib import admin
from django.utils.html import format_html
from django.urls import path, reverse
from django.shortcuts import redirect
from django.http import HttpResponse
from django.utils import timezone
from django.db import connection
from .models import Invoice, InvoiceConfiguration, InvoiceEmailLog, PaymentGatewayHealth
from .services import resend_invoice_email
from core import choices

# #region agent log
def _debug_log(message, hypothesis_id, data=None):
    try:
        payload = {"id": "inv_admin", "timestamp": timezone.now().timestamp() * 1000, "location": "invoices.admin", "message": message, "hypothesisId": hypothesis_id, "runId": "changelist"}
        if data is not None:
            payload["data"] = data
        with open("/home/itpc6/Public/django/git-repo/7nov/git/new_template-demo-topteens/topteen_1.0/.cursor/debug.log", "a") as f:
            f.write(json.dumps(payload) + "\n")
    except Exception:
        pass
# #endregion

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class InvoiceEmailLogInline(admin.TabularInline):
    model = InvoiceEmailLog
    extra = 0
    readonly_fields = ('recipient_type', 'recipient_email', 'success', 'error_message', 'created')
    can_delete = False
    max_num = 20


@admin.register(Invoice)
class InvoiceAdmin(admin.ModelAdmin):
    change_list_template = 'admin/invoices/invoice/change_list.html'
    list_display = [
        'invoice_number',
        'transaction_id',
        'created',
        'customer_name',
        'service',
        'amount_display',
        'gateway_display',
        'download_pdf_link',
    ]
    list_filter = ('service_obj_type', 'created')
    # date_hierarchy disabled: triggers "invalid datetime" on MySQL when tz tables not loaded (production).
    # Use list_filter "Created" for date filtering.
    search_fields = ('invoice_number', 'transaction_id', 'customer_name', 'customer_email')
    readonly_fields = (
        'payment', 'invoice_number', 'transaction_id', 'service', 'service_obj_type',
        'amount', 'currency', 'gst_rate', 'taxable_value', 'cgst', 'sgst',
        'customer_name', 'customer_email', 'customer_address', 'invoice_pdf', 'created', 'modified',
    )
    inlines = [InvoiceEmailLogInline]
    ordering = ('-created',)

    def amount_display(self, obj):
        return '₹ {}'.format(obj.amount)
    amount_display.short_description = 'Amount'

    def gateway_display(self, obj):
        return obj.payment.get_gateway_display() if obj.payment_id else '-'
    gateway_display.short_description = 'Gateway'

    def download_pdf_link(self, obj):
        if not obj.invoice_pdf:
            return '-'
        url = reverse('admin:invoices_invoice_download_pdf', args=[obj.pk])
        return format_html('<a href="{}">Download PDF</a>', url)
    download_pdf_link.short_description = 'PDF'

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('<int:pk>/download-pdf/', self.admin_site.admin_view(self.download_pdf_view), name='invoices_invoice_download_pdf'),
            path('export-excel/', self.admin_site.admin_view(self.export_excel_view), name='invoices_invoice_export_excel'),
        ]
        return custom + urls

    def download_pdf_view(self, request, pk):
        from django.shortcuts import get_object_or_404
        invoice = get_object_or_404(Invoice, pk=pk)
        if not invoice.invoice_pdf:
            return HttpResponse('PDF not generated yet.', status=404)
        response = HttpResponse(invoice.invoice_pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="invoice_{}.pdf"'.format(
            invoice.invoice_number.replace('/', '-')
        )
        return response

    def export_excel_view(self, request):
        if not HAS_OPENPYXL:
            from django.contrib import messages
            messages.error(request, 'openpyxl is required for Excel export. pip install openpyxl')
            return redirect('admin:invoices_invoice_changelist')
        qs = self.get_queryset(request)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Invoices'
        headers = [
            'Invoice Number', 'Transaction ID', 'Date', 'Customer Name', 'Customer Email',
            'Service', 'Amount (₹)', 'Taxable Value', 'CGST', 'SGST', 'Gateway', 'Created',
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h, font=Font(bold=True))
        for row, inv in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=inv.invoice_number)
            ws.cell(row=row, column=2, value=inv.transaction_id)
            ws.cell(row=row, column=3, value=inv.created.strftime('%d-%m-%Y') if inv.created else '')
            ws.cell(row=row, column=4, value=inv.customer_name)
            ws.cell(row=row, column=5, value=inv.customer_email)
            ws.cell(row=row, column=6, value=inv.service)
            ws.cell(row=row, column=7, value=inv.amount)
            ws.cell(row=row, column=8, value=str(inv.taxable_value))
            ws.cell(row=row, column=9, value=str(inv.cgst))
            ws.cell(row=row, column=10, value=str(inv.sgst))
            ws.cell(row=row, column=11, value=inv.payment.get_gateway_display() if inv.payment_id else '')
            ws.cell(row=row, column=12, value=inv.created.isoformat() if inv.created else '')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="invoices_export.xlsx"'
        return response

    def changelist_view(self, request, extra_context=None):
        # #region agent log
        _debug_log("changelist_view entry", "H2")
        try:
            with connection.cursor() as cur:
                cur.execute("SELECT @@session.time_zone")
                tz_row = cur.fetchone()
            _debug_log("session time_zone", "H1", {"time_zone": str(tz_row[0]) if tz_row else None})
        except Exception as e:
            _debug_log("session time_zone failed", "H1", {"error": str(e)})
        try:
            sample = list(Invoice.objects.values_list("id", "created").order_by("-id")[:5])
            _debug_log("sample Invoice created", "H3", {"sample": [[i, str(c)] for i, c in sample]})
        except Exception as e:
            _debug_log("sample Invoice failed", "H3", {"error": str(e)})
        # #endregion
        extra_context = extra_context or {}
        broken = list(PaymentGatewayHealth.objects.filter(
            object_status=choices.ObjectStatus.ACTIVE
        ).exclude(last_callback_success=True))
        extra_context['callback_health_broken'] = [h for h in broken if not h.is_working]
        # #region agent log
        try:
            out = super().changelist_view(request, extra_context)
            _debug_log("changelist_view success", "H2")
            return out
        except ValueError as e:
            _debug_log("changelist_view ValueError", "H2,H4,H5", {"message": str(e), "traceback": traceback.format_exc()})
            raise
        # #endregion

    actions = ['resend_to_admin', 'resend_to_customer', 'export_selected_excel']

    def resend_to_admin(self, request, queryset):
        for inv in queryset:
            try:
                resend_invoice_email(inv, InvoiceEmailLog.RECIPIENT_ADMIN)
                self.message_user(request, 'Resent to admin for {}'.format(inv.invoice_number), level=25)
            except Exception as e:
                self.message_user(request, 'Failed {}: {}'.format(inv.invoice_number, e), level=40)
    resend_to_admin.short_description = 'Resend invoice email to Admin'

    def resend_to_customer(self, request, queryset):
        for inv in queryset:
            if not inv.customer_email:
                self.message_user(request, 'No customer email for {}'.format(inv.invoice_number), level=40)
                continue
            try:
                resend_invoice_email(inv, InvoiceEmailLog.RECIPIENT_CUSTOMER)
                self.message_user(request, 'Resent to customer for {}'.format(inv.invoice_number), level=25)
            except Exception as e:
                self.message_user(request, 'Failed {}: {}'.format(inv.invoice_number, e), level=40)
    resend_to_customer.short_description = 'Resend invoice email to Customer'

    def export_selected_excel(self, request, queryset):
        if not HAS_OPENPYXL:
            self.message_user(request, 'openpyxl required.', level=40)
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Invoices'
        headers = [
            'Invoice Number', 'Transaction ID', 'Date', 'Customer Name', 'Customer Email',
            'Service', 'Amount (₹)', 'Taxable Value', 'CGST', 'SGST', 'Gateway', 'Created',
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h, font=Font(bold=True))
        for row, inv in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=inv.invoice_number)
            ws.cell(row=row, column=2, value=inv.transaction_id)
            ws.cell(row=row, column=3, value=inv.created.strftime('%d-%m-%Y') if inv.created else '')
            ws.cell(row=row, column=4, value=inv.customer_name)
            ws.cell(row=row, column=5, value=inv.customer_email)
            ws.cell(row=row, column=6, value=inv.service)
            ws.cell(row=row, column=7, value=inv.amount)
            ws.cell(row=row, column=8, value=str(inv.taxable_value))
            ws.cell(row=row, column=9, value=str(inv.cgst))
            ws.cell(row=row, column=10, value=str(inv.sgst))
            ws.cell(row=row, column=11, value=inv.payment.get_gateway_display() if inv.payment_id else '')
            ws.cell(row=row, column=12, value=inv.created.isoformat() if inv.created else '')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="invoices_export.xlsx"'
        return response
    export_selected_excel.short_description = 'Export selected invoices to Excel'


@admin.register(InvoiceConfiguration)
class InvoiceConfigurationAdmin(admin.ModelAdmin):
    list_display = ('company_name', 'accounts_email', 'auto_send_invoice_to_customer', 'default_gst_rate', 'invoice_prefix')
    list_editable = ('auto_send_invoice_to_customer',)

    def has_add_permission(self, request):
        return not InvoiceConfiguration.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(InvoiceEmailLog)
class InvoiceEmailLogAdmin(admin.ModelAdmin):
    list_display = (
        'invoice',
        'recipient_type',
        'recipient_email',
        'status_display',
        'sent_at_display',
        'error_message',
    )
    list_display_links = ('invoice',)
    list_filter = ('recipient_type', 'success', 'created')
    readonly_fields = ('invoice', 'recipient_type', 'recipient_email', 'success', 'error_message', 'created', 'modified')
    search_fields = ('recipient_email', 'invoice__invoice_number')
    ordering = ('-created',)
    list_per_page = 25
    list_empty_value_display = '-'

    def status_display(self, obj):
        if obj.success:
            return format_html('<span style="color: #155724; font-weight: 600;">Sent</span>')
        return format_html('<span style="color: #721c24; font-weight: 600;">Failed</span>')
    status_display.short_description = 'Status'

    def sent_at_display(self, obj):
        if not obj.created:
            return '-'
        return obj.created.strftime('%d %b %Y, %H:%M')
    sent_at_display.short_description = 'Sent date & time'


@admin.register(PaymentGatewayHealth)
class PaymentGatewayHealthAdmin(admin.ModelAdmin):
    list_display = ('gateway', 'last_callback_at', 'last_callback_success', 'is_working_display', 'callback_url')
    readonly_fields = ('last_callback_at', 'last_callback_success', 'last_error_message')

    def is_working_display(self, obj):
        if obj.is_working:
            return format_html('<span style="color: green;">OK</span>')
        return format_html('<span style="color: red;">Not working</span>')
    is_working_display.short_description = 'Status'
