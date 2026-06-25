"""Parse and validate Class 11–12 aptitude consolidated report Excel data."""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

VALID_CODES = frozenset({'AR', 'NR', 'LR', 'LVR', 'CR', 'MR', 'SR'})
EXPECTED_ROW_COUNT = 127
EXPECTED_SINGLE_CODES = VALID_CODES

EXCEL_COLUMNS = (
    'Reasoning Combination',
    'Aptitude Description',
    'Interpretation Narrative',
    'Career Clusters',
    'Career Pathways',
    'Degree Pathways',
    'Degree Pathways ',  # workbook header has trailing space
)

MIN_DESCRIPTION_LEN = 20
MIN_NARRATIVE_LEN = 40

# Consolidated Excel labels that must match CareerCluster.name in the careers DB.
CAREER_CLUSTER_DB_ALIASES: dict[str, str] = {
    'Agriculture, Natural Resources & Allied Sciences': 'Agriculture , Natural Resources & Allied Sciences',
    'Arts, Humanities, Education and Training': 'Arts, Humanities , Education & Training',
    'Vocational Studies': 'Vocational Careers',
}


def normalize_career_cluster_names(names: list[str] | None) -> list[str]:
    return [CAREER_CLUSTER_DB_ALIASES.get(name, name) for name in (names or [])]

_HTML_TAG_RE = re.compile(r'<[^>]+>')


def html_to_plain(text: str) -> str:
    """Strip HTML tags for validation and plain-text list splitting."""
    return _HTML_TAG_RE.sub('', str(text or '')).strip()


def rich_to_plain(value: Any) -> str:
    """Extract plain text from a cell value (no HTML escaping)."""
    from openpyxl.cell.rich_text import CellRichText

    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, CellRichText):
        parts: list[str] = []
        for block in value:
            if isinstance(block, str):
                parts.append(block)
            else:
                parts.append(block.text or '')
        return ''.join(parts).strip()
    return str(value).strip()


def rich_to_html(value: Any) -> str:
    """Convert an openpyxl cell value (plain or rich text) to HTML with <strong> for bold."""
    from html import escape

    from openpyxl.cell.rich_text import CellRichText

    if value is None:
        return ''
    if isinstance(value, str):
        return escape(value).strip()
    if isinstance(value, CellRichText):
        parts: list[str] = []
        for block in value:
            if isinstance(block, str):
                if block:
                    parts.append(escape(block))
                continue
            text = block.text or ''
            if not text:
                continue
            esc = escape(text)
            bold = getattr(getattr(block, 'font', None), 'bold', False)
            parts.append(f'<strong>{esc}</strong>' if bold else esc)
        html = ''.join(parts)
    else:
        html = escape(str(value))
    return html.replace('<strong></strong>', '').strip()


def split_html_list(cell_value: Any, delimiter: str) -> list[str]:
    """
    Split a cell into list items.

    Uses plain text for normal cells. When inline bold is present (e.g. Degree
    Pathways with "Global:"), returns one HTML string preserving <strong> tags.
    """
    html = rich_to_html(cell_value).replace('<strong></strong>', '').strip()
    if '<strong>' in html:
        return [html] if html else []
    plain = rich_to_plain(cell_value)
    if delimiter == ';':
        return split_semicolon_list(plain)
    return split_comma_list(plain)


@dataclass
class ImportResult:
    ok: bool
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    rows: list[dict[str, Any]] = field(default_factory=list)
    created: int = 0
    updated: int = 0
    skipped: int = 0
    source: str = ''


def normalize_combination_key(raw: str) -> str:
    """Canonical key: sorted codes joined with ' + ', e.g. 'AR + CR'."""
    parts = sorted(p.strip() for p in str(raw).replace(' ', '').split('+') if p.strip())
    return ' + '.join(parts)


def split_semicolon_list(text: str) -> list[str]:
    return [x.strip() for x in str(text or '').split(';') if x.strip()]


def split_comma_list(text: str) -> list[str]:
    return [x.strip() for x in str(text or '').split(',') if x.strip()]



def parse_excel_file(excel_path: Path) -> list[dict[str, Any]]:
    """Read Sheet1 and return normalized row dicts (HTML preserves bold formatting)."""
    import openpyxl

    if not excel_path.is_file():
        raise FileNotFoundError(f'Excel file not found: {excel_path}')

    wb = openpyxl.load_workbook(excel_path, data_only=False, rich_text=True)
    try:
        if 'Sheet1' not in wb.sheetnames:
            raise ValueError(f'Expected sheet "Sheet1"; found: {wb.sheetnames}')
        ws = wb['Sheet1']

        header_cols: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            header_val = ws.cell(1, col).value
            if header_val is None:
                continue
            header_cols[str(header_val).strip()] = col

        def col_for(*names: str) -> int | None:
            for name in names:
                key = name.strip()
                if key in header_cols:
                    return header_cols[key]
            return None

        def cell_html(row_num: int, *names: str) -> str:
            col = col_for(*names)
            if not col:
                return ''
            return rich_to_html(ws.cell(row_num, col).value)

        def cell_value(row_num: int, *names: str):
            col = col_for(*names)
            if not col:
                return None
            return ws.cell(row_num, col).value

        parsed: list[dict[str, Any]] = []

        for row_num in range(2, ws.max_row + 1):
            combination_raw = rich_to_plain(cell_value(row_num, 'Reasoning Combination'))
            if not combination_raw:
                if all(
                    rich_to_plain(cell_value(row_num, h)) == ''
                    for h in (
                        'Aptitude Description',
                        'Interpretation Narrative',
                        'Career Clusters',
                        'Career Pathways',
                        'Degree Pathways',
                    )
                ):
                    continue
                continue

            key = normalize_combination_key(combination_raw)
            codes = key.split(' + ') if key else []

            aptitude_description = rich_to_plain(cell_value(row_num, 'Aptitude Description'))
            interpretation_narrative = cell_html(row_num, 'Interpretation Narrative')

            parsed.append({
                'row_num': row_num,
                'reasoning_combination_raw': combination_raw,
                'reasoning_combination': key,
                'codes': codes,
                'aptitude_description': aptitude_description,
                'interpretation_narrative': interpretation_narrative,
                'career_clusters': normalize_career_cluster_names(
                    split_html_list(cell_value(row_num, 'Career Clusters'), ';')
                ),
                'career_pathways': split_html_list(
                    cell_value(row_num, 'Career Pathways'), ','
                ),
                'degree_pathways': split_html_list(
                    cell_value(row_num, 'Degree Pathways', 'Degree Pathways '), ','
                ),
            })
        return parsed
    finally:
        wb.close()


def validate_rows(
    rows: list[dict[str, Any]],
    *,
    source: str = '',
    expected_count: int | None = EXPECTED_ROW_COUNT,
) -> ImportResult:
    """Validate parsed rows; populate errors and warnings."""
    result = ImportResult(ok=True, rows=rows, source=source)

    if expected_count is not None and len(rows) != expected_count:
        result.errors.append(
            f'Expected {expected_count} data rows, found {len(rows)}.'
        )

    seen_keys: dict[str, int] = {}
    found_single_codes: set[str] = set()

    for row in rows:
        row_num = row.get('row_num', '?')
        key = row.get('reasoning_combination', '')
        raw_key = row.get('reasoning_combination_raw', '')

        if not key:
            result.errors.append(f'Row {row_num}: empty Reasoning Combination.')
            continue

        if raw_key != raw_key.strip():
            result.warnings.append(
                f'Row {row_num} ({key}): Reasoning Combination had leading/trailing spaces '
                f'(normalized from {raw_key!r}).'
            )

        if key in seen_keys:
            result.errors.append(
                f'Row {row_num}: duplicate key {key!r} (first seen on row {seen_keys[key]}).'
            )
        else:
            seen_keys[key] = row_num

        codes = row.get('codes') or []
        invalid_codes = [c for c in codes if c not in VALID_CODES]
        if invalid_codes:
            result.errors.append(
                f'Row {row_num} ({key}): invalid code(s): {", ".join(invalid_codes)}.'
            )

        if len(codes) == 1:
            found_single_codes.add(codes[0])

        if not row.get('aptitude_description'):
            result.errors.append(f'Row {row_num} ({key}): Aptitude Description is empty.')
        elif len(html_to_plain(row['aptitude_description'])) < MIN_DESCRIPTION_LEN:
            result.warnings.append(
                f'Row {row_num} ({key}): Aptitude Description is short '
                f'({len(html_to_plain(row["aptitude_description"]))} chars).'
            )

        if not row.get('interpretation_narrative'):
            result.errors.append(f'Row {row_num} ({key}): Interpretation Narrative is empty.')
        elif len(html_to_plain(row['interpretation_narrative'])) < MIN_NARRATIVE_LEN:
            result.warnings.append(
                f'Row {row_num} ({key}): Interpretation Narrative is short '
                f'({len(html_to_plain(row["interpretation_narrative"]))} chars).'
            )

        if not row.get('career_clusters'):
            result.errors.append(f'Row {row_num} ({key}): Career Clusters is empty.')
        if not row.get('career_pathways'):
            result.errors.append(f'Row {row_num} ({key}): Career Pathways is empty.')
        if not row.get('degree_pathways'):
            result.errors.append(f'Row {row_num} ({key}): Degree Pathways is empty.')

    missing_singles = sorted(EXPECTED_SINGLE_CODES - found_single_codes)
    if missing_singles:
        result.errors.append(
            f'Missing single-code rows: {", ".join(missing_singles)}.'
        )

    extra_singles = sorted(found_single_codes - EXPECTED_SINGLE_CODES)
    if extra_singles:
        result.errors.append(
            f'Unexpected single-code rows: {", ".join(extra_singles)}.'
        )

    result.ok = len(result.errors) == 0
    return result


def load_and_validate(excel_path: Path) -> ImportResult:
    """Parse Excel and validate in one step."""
    rows = parse_excel_file(excel_path)
    return validate_rows(rows, source=excel_path.name)


DEFAULT_JSON_PATH = Path(__file__).resolve().parent / 'data' / 'class12_aptitude_consolidated_report.json'


def load_json_payload(json_path: Path | None = None) -> dict[str, Any]:
    """Load generated consolidated report JSON from disk."""
    import json

    path = json_path or DEFAULT_JSON_PATH
    if not path.is_file():
        raise FileNotFoundError(f'Consolidated report JSON not found: {path}')
    with path.open(encoding='utf-8') as handle:
        payload = json.load(handle)
    combinations = payload.get('combinations') or {}
    if not isinstance(combinations, dict):
        raise ValueError('Invalid JSON: "combinations" must be an object.')
    return payload


def lookup_combination(
    combination_key: str,
    *,
    json_path: Path | None = None,
) -> dict[str, Any] | None:
    """Look up a row by normalized reasoning combination key (DB first, then JSON)."""
    key = normalize_combination_key(combination_key)
    db_row = _lookup_combination_from_db(key)
    if db_row:
        return db_row
    payload = load_json_payload(json_path)
    combinations: dict[str, Any] = payload.get('combinations') or {}
    return combinations.get(key)


def _lookup_combination_from_db(key: str) -> dict[str, Any] | None:
    try:
        from app.models import Class12AptitudeConsolidatedReport

        obj = Class12AptitudeConsolidatedReport.objects.filter(
            reasoning_combination=key,
            is_active=True,
        ).first()
        if obj:
            return obj.to_row_dict()
    except Exception:
        return None
    return None


def import_rows_to_db(
    *,
    rows: list[dict[str, Any]] | None = None,
    source: str = 'json',
    replace: bool = False,
    json_path: Path | None = None,
) -> dict[str, Any]:
    """
    Upsert consolidated rows into Django admin model.

    source: 'json' loads from generated JSON file; 'rows' uses provided parsed rows.
    """
    from app.class12_aptitude_report_utils import clear_consolidated_lookup_cache
    from app.models import Class12AptitudeConsolidatedReport

    try:
        if rows is None:
            if source == 'json':
                payload = load_json_payload(json_path)
                combinations = payload.get('combinations') or {}
                rows = list(combinations.values())
            else:
                return {'ok': False, 'error': 'No rows provided for import.'}

        if replace:
            Class12AptitudeConsolidatedReport.objects.all().delete()

        count = 0
        for row in rows:
            key = normalize_combination_key(row.get('reasoning_combination', ''))
            if not key:
                continue
            Class12AptitudeConsolidatedReport.objects.update_or_create(
                reasoning_combination=key,
                defaults={
                    'codes': list(row.get('codes') or []),
                    'aptitude_description': row.get('aptitude_description') or '',
                    'interpretation_narrative': row.get('interpretation_narrative') or '',
                    'career_clusters': normalize_career_cluster_names(
                        list(row.get('career_clusters') or [])
                    ),
                    'career_pathways': list(row.get('career_pathways') or []),
                    'degree_pathways': list(row.get('degree_pathways') or []),
                    'real_life_sign_ids': list(row.get('real_life_sign_ids') or []),
                    'daily_life_impact_ids': list(row.get('daily_life_impact_ids') or []),
                    'is_active': True,
                },
            )
            count += 1

        clear_consolidated_lookup_cache()
        return {'ok': True, 'count': count}
    except Exception as exc:
        return {'ok': False, 'error': str(exc)}


def build_json_payload(rows: list[dict[str, Any]], *, source: str) -> dict[str, Any]:
    """Build exportable JSON structure (used in Phase 2)."""
    combinations = {}
    for row in rows:
        key = row['reasoning_combination']
        combinations[key] = {
            'reasoning_combination': key,
            'codes': row['codes'],
            'aptitude_description': row['aptitude_description'],
            'interpretation_narrative': row['interpretation_narrative'],
            'real_life_sign_ids': list(row.get('real_life_sign_ids') or []),
            'daily_life_impact_ids': list(row.get('daily_life_impact_ids') or []),
            'career_clusters': normalize_career_cluster_names(list(row.get('career_clusters') or [])),
            'career_pathways': row['career_pathways'],
            'degree_pathways': row['degree_pathways'],
        }
    return {
        'version': 1,
        'source': source,
        'imported_at': datetime.now(timezone.utc).isoformat(),
        'combinations': combinations,
    }


CODE_TO_INTERPRETATION_AREA: dict[str, str] = {
    'AR': 'Abstract Reasoning',
    'NR': 'Numerical Reasoning',
    'LR': 'Logical Reasoning',
    'LVR': 'Language and Verbal Reasoning',
    'CR': 'Clerical speed & Accuracy',
    'MR': 'Mechanical Reasoning',
    'SR': 'Spatial Reasoning',
}


def load_aptitude_interpretation_data() -> dict[str, Any]:
    """Load legacy per-area interpretation JSON used for Real-life signs / Daily life impact."""
    import json

    from django.conf import settings

    path = Path(settings.BASE_DIR) / 'static' / 'data' / 'aptitute interpretation.json'
    if not path.is_file():
        return {}
    with path.open(encoding='utf-8') as handle:
        return json.load(handle)


def _merge_unique_text_items(*lists: list[Any] | None) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for items in lists:
        for item in items or []:
            text = str(item).strip()
            if not text or text in seen:
                continue
            seen.add(text)
            merged.append(text)
    return merged


def merge_legacy_signs_impact_for_codes(
    codes: list[str],
    interpretation_data: dict[str, Any] | None,
) -> tuple[list[str], list[str]]:
    """Merge Real-life signs and Daily life impact from legacy JSON for reasoning codes."""
    if not interpretation_data or not codes:
        return [], []

    by_area = {
        item.get('Area'): item
        for item in interpretation_data.get('Aptitude_Interpretations', [])
        if item.get('Area')
    }
    sign_lists: list[list[str]] = []
    impact_lists: list[list[str]] = []
    for code in codes:
        area = CODE_TO_INTERPRETATION_AREA.get(code)
        if not area:
            continue
        interp = by_area.get(area)
        if not interp:
            continue
        sign_lists.append(list(interp.get('Real life signs') or []))
        impact_lists.append(list(interp.get('Daily life impact') or []))
    return (
        _merge_unique_text_items(*sign_lists),
        _merge_unique_text_items(*impact_lists),
    )
