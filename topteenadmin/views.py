from careers.models import (Career, CareerFAQ, CareerMedia, CareerPath, CareerPathStep, CareerTags, Profession,
                            ProspectiveEmploymentArea, ProspectiveRecruiter,
                            Skill,CareerCluster,Videos,VideoCategory)
from core.utils import build_breadcrumb, build_html_head
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import (authenticate, get_user_model, login, logout,
                                 update_session_auth_hash)
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import (AuthenticationForm, PasswordChangeForm,
                                       PasswordResetForm, SetPasswordForm)
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import BadHeaderError, EmailMultiAlternatives, send_mail
from django.forms.models import inlineformset_factory
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils.decorators import method_decorator
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.views.generic import TemplateView, View
from users.models import User
from skilllab.models import SkillLabCourse,SkillLabCourseActivity,SkillLabCourseChapter

from blog.models import Blog,BlogCategory,BlogTag
from topteenadmin.filters import (CareerFAQFilter, CareerFilter, CareerMediaFilter, CareerRelatedCareersFilter,
                                  CareerPathFilter, CareerPathStepFilter, CommonFAQFilter, ProfessionFilter,
                                  ProspectiveEmploymentAreaFilter,
                                  ProspectiveRecruiterFilter, ReviewFilter, SkillFilter,CollegeFilter,CollegeImagesFilter,CollegeFlatTextFilter,CollegeTextFilter,CollegeFactsFilter,
                                  RecruitingCompaniesFilter,CollegeRecruitingCompaniesFilter,FacilityFilter,CollegeFacilityFilter,CollegeMoneyValueFilter,
                                  CountryFilter, SkillLabCourseActivityFilter,StateFilter,CityFilter,
                                  ProspectiveRecruiterFilter, SkillFilter,ProfessionFilter,StreamFilter,CourseFilter,CourseFactsFilter,CourseIntakeFilter,
                                  CourseTextFilter,CourseMoneyValueFilter,CourseEnglighRequirementsFilter,SkillLabCourseChapterFilter,
                                  EntranceExamFilter,BlogFilter,BlogCategoryFilter,BlogTagFilter,CareerClusterFilter,SkillLabCourseFilter,SkillLabCourseActivity,ExamTagsFilter,VideoCategoryFilter,VideosFilter,HobbiesFilter,SubjectFilter,UserFigureOutFilter,StoriesFilter,ApilogFilter,LeadFilter,PsychometricFaqFilter,StudentFilter,VocationalCourseCategoryFilter,VocationalCourseFilter,ExtracurricularActivityCategoryFilter,ExtracurricularActivityFilter,EntranceTestPrepCategoryFilter,EntranceTestPrepExamFilter)

from .base_views import (BaseCreateView, BaseDeleteView, BaseDetailView,
                         BaseListView, BaseUpdateView)
from .utils import build_admin_breadcrumb
from .forms import ( CareerFAQModelForm, CareerMediaModelForm,
                    CareerModelForm, CareerRelatedCareersForm, CareerPathModelForm,CareerPathStepModelForm, CommonFAQModelForm,
                    ProspectiveEmploymentAreaModelForm,
                    ProspectiveRecruiterModelForm, SkillModelForm,CollegeModelForm,CollegeImagesModelForm,
                    CollegeFlatTextModelForm,CollegeTextModelForm,CollegeFactsModelForm,
                    RecruitingCompaniesModelForm,CollegeRecruitingCompaniesModelForm,
                    FacilityModelForm,CollegeFacilityModelForm,CollegeMoneyValueModelForm,CountryModelForm
                    ,CityModelForm,StateModelForm,SkillLabCourseChapterModelForm,
                    ProspectiveRecruiterModelForm, SkillModelForm,ProfessionModelForm,StreamModelForm,
                    CourseModelForm,CourseFactsModelForm,CourseTextModelForm,CourseMoneyValueModelForm,
                    CourseIntakeModelForm,CourseEnglighRequirementsModelForm,
                    EntranceExamModelForm,CareerTagsForm,BlogModelForm,
                    BlogCategoryModelForm,TagModelForm,CareerClusterModelForm,ReviewModelForm,SkillLabCourseModelForm,SkillLabCourseActivityModelForm,
                    ExamTagsModelForm,VideosForm,VideoCategoryForm,HobbiesModelForm,SubjectModelForm,UserFigureOutModelForm,StoriesModelForm,LeadModelForm,ApilogModelForm,PsychometricFaqModelForm,VocationalCourseCategoryModelForm,VocationalCourseModelForm,ExtracurricularActivityCategoryModelForm,ExtracurricularActivityModelForm,EntranceTestPrepCategoryModelForm,EntranceTestPrepExamModelForm)

from colleges.models import (College, CollegeFacts,CollegeImages,CollegeFlatText,RecruitingCompanies,CollegeFacility,
                            CollegeText,RecruitingCompanies,Facility,CollegeRecruitingCompanies,CollegeMoneyValue)
from core.models import CommonFAQ, Country,City,State,Review,Hobbies,Subject,UserFigureOut,Stories,APILog,VocationalCourseCategory,VocationalCourse,ExtracurricularActivityCategory,ExtracurricularActivity,EntranceTestPrepCategory,EntranceTestPrepExam
from courses.models import (Stream,Course,CourseFacts,CourseIntake,CourseText,CourseMoneyValue,CourseEnglighRequirements)
from entrance_exams.models import EntranceExam,ExamTags
from crm.models import Lead
from psychometric_tests.models import PsychometricFAQ
from app_post_matric.models import Test, TestSession, TestResult, Question, Sections, UserResponse, SectionSession, TestTopCategories, Answer
from app.models import TestCompletion
from users.models import UserProfile
from institute.models import StudentManagement
from django.db.models import Q, Count, Case, When, IntegerField, Prefetch, Exists, OuterRef
from django.http import JsonResponse
from django.core.paginator import Paginator
# Create your views here.
@method_decorator(login_required,name='dispatch')
class TopteensDashboard(TemplateView):
    template_name = "topteenadmin/home.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['meta_title'] = 'Dashboard'
        ctx['active_tab'] = 'dashboard'
        ctx['html_head'] = {'title': 'Dashboard', 'description': ''}
        ctx['breadcrumb'] = build_admin_breadcrumb([
            {'title': 'Dashboard', 'text': 'Dashboard', 'url': '#'},
        ])
        return ctx
    
class CareerListView(BaseListView):
    template_name = "topteenadmin/career_list.html"
    title = "Career"
    active_tab = "career"
    model = Career
    filterset_class = CareerFilter
    context_object_name = "career_list"

    def get_queryset(self):
        return super().get_queryset().prefetch_related('career_cluster', 'related_careers')


def _career_form_extra_context(career):
    """Frontend preview URL and career reference for description/accordion templates."""
    if not career or not getattr(career, 'pk', None):
        return {'career': career, 'career_frontend_preview_url': None}
    ctx = {'career': career, 'career_frontend_preview_url': None}
    slug = getattr(career, 'slug', None)
    if slug:
        ctx['career_frontend_preview_url'] = reverse(
            'careers:careerdetail', args=[slug, career.pk]
        )
    return ctx


class CreateCareer(BaseCreateView):
    template_name = "topteenadmin/career_form.html"
    model=Career
    form_class=CareerModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerlist')
    title="Career"
    active_tab="career"
    success_message="Cereer created successfully."

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_career_form_extra_context(self.object))
        return ctx
    
class CareerUpdateView(BaseUpdateView):
    template_name = "topteenadmin/career_form.html"
    model=Career
    title="Career"
    form_class=CareerModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerlist')
    active_tab="career"
    success_message="Cereer updated successfully."

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_career_form_extra_context(self.object))
        return ctx

class CareerDeleteView(BaseDeleteView):
    model = Career
    active_tab="career"
    success_url = reverse_lazy('topteenadminmanaged:careerlist')
    success_message="Career deleted successfully."
    
class CareerDetailView(BaseDetailView):
    template_name = "topteenadmin/career_detail.html"
    model=Career
    title="Career"
    form_class=CareerModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerlist')
    active_tab="career"

    def get_queryset(self):
        return super().get_queryset().prefetch_related('related_careers', 'career_cluster')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_career_form_extra_context(self.object))
        return ctx


class CareerRelatedCareersListView(BaseListView):
    template_name = 'topteenadmin/related_career_list.html'
    title = 'Related Careers'
    active_tab = 'career_related'
    model = Career
    filterset_class = CareerRelatedCareersFilter
    context_object_name = 'career_list'

    def get_queryset(self):
        return super().get_queryset().prefetch_related(
            'career_cluster',
            'related_careers',
        ).order_by('name', 'id')

    def _breadcrumb(self):
        return build_admin_breadcrumb([
            {'title': self.title, 'text': self.title,
             'url': reverse('topteenadminmanaged:relatedcareerlist')},
        ])


class CareerRelatedCareersEditView(BaseUpdateView):
    template_name = 'topteenadmin/related_career_edit.html'
    model = Career
    form_class = CareerRelatedCareersForm
    title = 'Edit Related Careers'
    active_tab = 'career_related'
    success_message = 'Related careers updated successfully.'

    def get_success_url(self):
        return reverse_lazy('topteenadminmanaged:relatedcareerlist')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_career_form_extra_context(self.object))
        ctx['career'] = self.object
        ctx['autocomplete_url'] = reverse('careers:admin_autocomplete_careers')
        ctx['related_careers_initial'] = [
            {'id': c.id, 'text': c.name}
            for c in self.object.related_careers.all().order_by('name')
        ]
        clusters = self.object.career_cluster.all()
        ctx['cluster_names'] = ', '.join(c.name for c in clusters if c.name) or '—'
        ctx['breadcrumb'] = build_admin_breadcrumb([
            {'title': 'Related Careers', 'text': 'Related Careers',
             'url': reverse('topteenadminmanaged:relatedcareerlist')},
            {'title': self.object.name or 'Edit', 'text': self.object.name or 'Edit', 'url': '#'},
        ])
        return ctx


class SkillListView(BaseListView):
    template_name = "topteenadmin/skill_list.html"
    title="Skill"
    active_tab="skill"
    model = Skill
    filterset_class = SkillFilter
    
class CreateSkill(BaseCreateView):
    template_name = "topteenadmin/skill_form.html"
    model=Skill
    form_class=SkillModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllist')
    title="Skill"
    active_tab="skill"
    success_message="Skill created successfully."
    
class SkillUpdateView(BaseUpdateView):
    template_name = "topteenadmin/skill_form.html"
    model=Skill
    title="Skill"
    form_class=SkillModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllist')
    active_tab="skill"
    success_message="Skill updated successfully."

class SkillDeleteView(BaseDeleteView):
    model = Skill
    active_tab="skill"
    success_url = reverse_lazy('topteenadminmanaged:skilllist')
    success_message="Skill deleted successfully."
    
class SkillDetailView(BaseDetailView):
    template_name = "topteenadmin/skill_detail.html"
    model=Skill
    title="Skill"
    form_class=SkillModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllist')
    active_tab="skill"
    
class ProspectiveEmploymentAreaListView(BaseListView):
    template_name = "topteenadmin/prospectiveemploymentarea_list.html"
    title="Prospective Employment Area"
    active_tab="prospectiveemploymentarea"
    model = ProspectiveEmploymentArea
    filterset_class = ProspectiveEmploymentAreaFilter
    
class CreateProspectiveEmploymentArea(BaseCreateView):
    template_name = "topteenadmin/prospectiveemploymentarea_form.html"
    model=ProspectiveEmploymentArea
    form_class=ProspectiveEmploymentAreaModelForm
    success_url=reverse_lazy('topteenadminmanaged:prospectiveemploymentarealist')
    title="Prospective Employment Area"
    active_tab="prospectiveemploymentarea"
    success_message="Prospectiveemploymentarea created successfully."
    
class ProspectiveEmploymentAreaUpdateView(BaseUpdateView):
    template_name = "topteenadmin/prospectiveemploymentarea_form.html"
    model=ProspectiveEmploymentArea
    title="Prospective Employment Area"
    form_class=ProspectiveEmploymentAreaModelForm
    success_url=reverse_lazy('topteenadminmanaged:prospectiveemploymentarealist')
    active_tab="prospectiveemploymentarea"
    success_message="Prospectiveemploymentarea updated successfully."

class ProspectiveEmploymentAreaDeleteView(BaseDeleteView):
    model = ProspectiveEmploymentArea
    active_tab="prospectiveemploymentarea"
    success_url = reverse_lazy('topteenadminmanaged:prospectiveemploymentarealist')
    success_message="Prospectiveemploymentarea deleted successfully."
    
class ProspectiveEmploymentAreaDetailView(BaseDetailView):
    template_name = "topteenadmin/prospectiveemploymentarea_detail.html"
    model=ProspectiveEmploymentArea
    title="Prospective Employment Area"
    form_class=ProspectiveEmploymentAreaModelForm
    success_url=reverse_lazy('topteenadminmanaged:prospectiveemploymentarealist')
    active_tab="prospectiveemploymentarea"
    
class ProspectiveRecruiterListView(BaseListView):
    template_name = "topteenadmin/prospectiverecruiter_list.html"
    title="Prospective Recruiter"
    active_tab="prospectiverecruiter"
    model = ProspectiveRecruiter
    filterset_class = ProspectiveRecruiterFilter
    
class CreateProspectiveRecruiter(BaseCreateView):
    template_name = "topteenadmin/prospectiverecruiter_form.html"
    model=ProspectiveRecruiter
    form_class=ProspectiveRecruiterModelForm
    success_url=reverse_lazy('topteenadminmanaged:prospectiverecruiterlist')
    title="Prospective Recruiter"
    active_tab="prospectiverecruiter"
    success_message="Prospectiverecruiter created successfully."
    
class ProspectiveRecruiterUpdateView(BaseUpdateView):
    template_name = "topteenadmin/prospectiverecruiter_form.html"
    model=ProspectiveRecruiter
    title="Prospective Recruiter"
    form_class=ProspectiveRecruiterModelForm
    success_url=reverse_lazy('topteenadminmanaged:prospectiverecruiterlist')
    active_tab="prospectiverecruiter"
    success_message="Prospectiverecruiter updated successfully."

class ProspectiveRecruiterDeleteView(BaseDeleteView):
    model = ProspectiveRecruiter
    active_tab="prospectiverecruiter"
    success_url = reverse_lazy('topteenadminmanaged:prospectiverecruiterlist')
    success_message="Prospectiverecruiter deleted successfully."
    
class ProspectiveRecruiterDetailView(BaseDetailView):
    template_name = "topteenadmin/prospectiverecruiter_detail.html"
    model=ProspectiveRecruiter
    title="Prospective Recruiter"
    form_class=ProspectiveRecruiterModelForm
    success_url=reverse_lazy('topteenadminmanaged:prospectiverecruiterlist')
    active_tab="prospectiverecruiter"

class CareerMediaListView(BaseListView):
    template_name = "topteenadmin/careermedia_list.html"
    title="CareerMedia"
    active_tab="careermedia"
    model = CareerMedia
    filterset_class = CareerMediaFilter
    
class CreateCareerMedia(BaseCreateView):
    template_name = "topteenadmin/careermedia_form.html"
    model=CareerMedia
    form_class=CareerMediaModelForm
    success_url=reverse_lazy('topteenadminmanaged:careermedialist')
    title="CareerMedia"
    active_tab="careermedia"
    success_message="CareerMedia created successfully."
    
class CareerMediaUpdateView(BaseUpdateView):
    template_name = "topteenadmin/careermedia_form.html"
    model=CareerMedia
    title="CareerMedia"
    form_class=CareerMediaModelForm
    success_url=reverse_lazy('topteenadminmanaged:careermedialist')
    active_tab="careermedia"
    success_message="CareerPath updated successfully."

class CareerMediaDeleteView(BaseDeleteView):
    model = CareerMedia
    active_tab="careermedia"
    success_url = reverse_lazy('topteenadminmanaged:careermedialist')
    success_message="CareerMedia deleted successfully."
    
class CareerMediaDetailView(BaseDetailView):
    template_name = "topteenadmin/careermedia_detail.html"
    model=CareerMedia
    title="CareerMedia"
    form_class=CareerMediaModelForm
    success_url=reverse_lazy('topteenadminmanaged:careermedialist')
    active_tab="careermedia"
    
class CareerFAQListView(BaseListView):
    template_name = "topteenadmin/careerfaq_list.html"
    title="CareerFAQ"
    active_tab="careerFAQ"
    model = CareerFAQ
    filterset_class = CareerFAQFilter

class CreateCareerFAQ(BaseCreateView):
    template_name = "topteenadmin/careerfaq_form.html"
    model=CareerFAQ
    form_class=CareerFAQModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerfaqlist')
    title="CareerFAQ"
    active_tab="careerFAQ"
    success_message="CareerFAQ created successfully."      


class CareerFAQUpdateView(BaseUpdateView):
    template_name = "topteenadmin/careerfaq_form.html"
    model=CareerFAQ
    title="CareerFAQ"
    form_class=CareerFAQModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerfaqlist')
    active_tab="careerFAQ"
    success_message="CareerFAQ updated successfully."    

class CareerFAQDeleteView(BaseDeleteView):
    model = CareerFAQ
    active_tab="careerFAQ"
    success_url = reverse_lazy('topteenadminmanaged:careerfaqlist')
    success_message="CareerFAQ deleted successfully."

class CareerFAQDetailView(BaseDetailView):
    template_name = "topteenadmin/careerfaq_detail.html"
    model=CareerFAQ
    title="CareerFAQ"
    form_class=CareerFAQModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerfaqlist')
    active_tab="careerFAQ"    
        
class CareerPathStepListView(BaseListView):
    template_name = "topteenadmin/careerpathstep_list.html"
    title="CareerPathStep"
    active_tab="careerpathstep"
    model = CareerPathStep
    filterset_class = CareerPathStepFilter
    
class CreateCareerPathStep(BaseCreateView):
    template_name = "topteenadmin/careerpathstep_form.html"
    model=CareerPathStep
    form_class=CareerPathStepModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerpathsteplist')
    title="CareerPathStep"
    active_tab="careerpathstep"
    success_message="CareerPathStep created successfully."    

class CareerPathStepUpdateView(BaseUpdateView):
    template_name = "topteenadmin/careerpathstep_form.html"
    model=CareerPathStep
    title="CareerPathStep"
    form_class=CareerPathStepModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerpathsteplist')
    active_tab="careerpathstep"
    success_message="CareerPathStep updated successfully."    

class CareerPathStepDeleteView(BaseDeleteView):
    model = CareerPathStep
    active_tab="careerpathstep"
    success_url = reverse_lazy('topteenadminmanaged:careerpathsteplist')
    success_message="CareerPathStep deleted successfully."

class CareerPathStepDetailView(BaseDetailView):
    template_name = "topteenadmin/careerpathstep_detail.html"
    model=CareerPathStep
    title="CareerPathStep"
    form_class=CareerPathStepModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerpathsteplist')
    active_tab="careerpathstep"    
        
class CareerPathListView(BaseListView):
    template_name = "topteenadmin/careerpath_list.html"
    title="CareerPath"
    active_tab="careerpath"
    model = CareerPath
    filterset_class = CareerPathFilter
    
class CreateCareerPath(BaseCreateView):
    template_name = "topteenadmin/careerpath_form.html"
    model=CareerPath
    form_class=CareerPathModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerpathlist')
    title="CareerPath"
    active_tab="careerpath"
    success_message="CareerPath created successfully."

class CareerPathUpdateView(BaseUpdateView):
    template_name = "topteenadmin/careerpath_form.html"
    model=CareerPath
    title="CareerPath"
    form_class=CareerPathModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerpathlist')
    active_tab="careerpath"
    success_message="CareerPath updated successfully."

class CareerPathDeleteView(BaseDeleteView):
    model = CareerPath
    active_tab="careerpath"
    success_url = reverse_lazy('topteenadminmanaged:careerpathlist')
    success_message="CareerPath deleted successfully."
    
class CareerPathDetailView(BaseDetailView):
    template_name = "topteenadmin/careerpath_detail.html"
    model=CareerPath
    title="CareerPath"
    form_class=CareerPathModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerpathlist')
    active_tab="careerpath"

# Vocational Courses (core)
class VocationalCourseCategoryListView(BaseListView):
    template_name = "topteenadmin/vocationalcoursecategory_list.html"
    title = "Vocational Course Category"
    active_tab = "vocationalcoursecategory"
    model = VocationalCourseCategory
    filterset_class = VocationalCourseCategoryFilter
    context_object_name = "vocationalcoursecategory_list"

class CreateVocationalCourseCategory(BaseCreateView):
    template_name = "topteenadmin/vocationalcoursecategory_form.html"
    model = VocationalCourseCategory
    form_class = VocationalCourseCategoryModelForm
    success_url = reverse_lazy('topteenadminmanaged:vocationalcoursecategorylist')
    title = "Vocational Course Category"
    active_tab = "vocationalcoursecategory"
    success_message = "Vocational course category created successfully."

class VocationalCourseCategoryUpdateView(BaseUpdateView):
    template_name = "topteenadmin/vocationalcoursecategory_form.html"
    model = VocationalCourseCategory
    form_class = VocationalCourseCategoryModelForm
    success_url = reverse_lazy('topteenadminmanaged:vocationalcoursecategorylist')
    title = "Vocational Course Category"
    active_tab = "vocationalcoursecategory"
    success_message = "Vocational course category updated successfully."

class VocationalCourseCategoryDeleteView(BaseDeleteView):
    model = VocationalCourseCategory
    active_tab = "vocationalcoursecategory"
    success_url = reverse_lazy('topteenadminmanaged:vocationalcoursecategorylist')
    success_message = "Vocational course category deleted successfully."

class VocationalCourseCategoryDetailView(BaseDetailView):
    template_name = "topteenadmin/vocationalcoursecategory_detail.html"
    model = VocationalCourseCategory
    title = "Vocational Course Category"
    form_class = VocationalCourseCategoryModelForm
    success_url = reverse_lazy('topteenadminmanaged:vocationalcoursecategorylist')
    active_tab = "vocationalcoursecategory"

class VocationalCourseListView(BaseListView):
    template_name = "topteenadmin/vocationalcourse_list.html"
    title = "Vocational Course"
    active_tab = "vocationalcourse"
    model = VocationalCourse
    filterset_class = VocationalCourseFilter
    context_object_name = "vocationalcourse_list"

class CreateVocationalCourse(BaseCreateView):
    template_name = "topteenadmin/vocationalcourse_form.html"
    model = VocationalCourse
    form_class = VocationalCourseModelForm
    success_url = reverse_lazy('topteenadminmanaged:vocationalcourselist')
    title = "Vocational Course"
    active_tab = "vocationalcourse"
    success_message = "Vocational course created successfully."

class VocationalCourseUpdateView(BaseUpdateView):
    template_name = "topteenadmin/vocationalcourse_form.html"
    model = VocationalCourse
    form_class = VocationalCourseModelForm
    success_url = reverse_lazy('topteenadminmanaged:vocationalcourselist')
    title = "Vocational Course"
    active_tab = "vocationalcourse"
    success_message = "Vocational course updated successfully."

class VocationalCourseDeleteView(BaseDeleteView):
    model = VocationalCourse
    active_tab = "vocationalcourse"
    success_url = reverse_lazy('topteenadminmanaged:vocationalcourselist')
    success_message = "Vocational course deleted successfully."

class VocationalCourseDetailView(BaseDetailView):
    template_name = "topteenadmin/vocationalcourse_detail.html"
    model = VocationalCourse
    title = "Vocational Course"
    form_class = VocationalCourseModelForm
    success_url = reverse_lazy('topteenadminmanaged:vocationalcourselist')
    active_tab = "vocationalcourse"

# Extracurricular Activities (core)
class ExtracurricularActivityCategoryListView(BaseListView):
    template_name = "topteenadmin/extracurricularactivitycategory_list.html"
    title = "Extracurricular Category"
    active_tab = "extracurricularactivitycategory"
    model = ExtracurricularActivityCategory
    filterset_class = ExtracurricularActivityCategoryFilter
    context_object_name = "extracurricularactivitycategory_list"

class CreateExtracurricularActivityCategory(BaseCreateView):
    template_name = "topteenadmin/extracurricularactivitycategory_form.html"
    model = ExtracurricularActivityCategory
    form_class = ExtracurricularActivityCategoryModelForm
    success_url = reverse_lazy('topteenadminmanaged:extracurricularactivitycategorylist')
    title = "Extracurricular Category"
    active_tab = "extracurricularactivitycategory"
    success_message = "Extracurricular category created successfully."

class ExtracurricularActivityCategoryUpdateView(BaseUpdateView):
    template_name = "topteenadmin/extracurricularactivitycategory_form.html"
    model = ExtracurricularActivityCategory
    form_class = ExtracurricularActivityCategoryModelForm
    success_url = reverse_lazy('topteenadminmanaged:extracurricularactivitycategorylist')
    title = "Extracurricular Category"
    active_tab = "extracurricularactivitycategory"
    success_message = "Extracurricular category updated successfully."

class ExtracurricularActivityCategoryDeleteView(BaseDeleteView):
    model = ExtracurricularActivityCategory
    active_tab = "extracurricularactivitycategory"
    success_url = reverse_lazy('topteenadminmanaged:extracurricularactivitycategorylist')
    success_message = "Extracurricular category deleted successfully."

class ExtracurricularActivityCategoryDetailView(BaseDetailView):
    template_name = "topteenadmin/extracurricularactivitycategory_detail.html"
    model = ExtracurricularActivityCategory
    title = "Extracurricular Category"
    form_class = ExtracurricularActivityCategoryModelForm
    success_url = reverse_lazy('topteenadminmanaged:extracurricularactivitycategorylist')
    active_tab = "extracurricularactivitycategory"

class ExtracurricularActivityListView(BaseListView):
    template_name = "topteenadmin/extracurricularactivity_list.html"
    title = "Extracurricular Activity"
    active_tab = "extracurricularactivity"
    model = ExtracurricularActivity
    filterset_class = ExtracurricularActivityFilter
    context_object_name = "extracurricularactivity_list"

class CreateExtracurricularActivity(BaseCreateView):
    template_name = "topteenadmin/extracurricularactivity_form.html"
    model = ExtracurricularActivity
    form_class = ExtracurricularActivityModelForm
    success_url = reverse_lazy('topteenadminmanaged:extracurricularactivitylist')
    title = "Extracurricular Activity"
    active_tab = "extracurricularactivity"
    success_message = "Extracurricular activity created successfully."

class ExtracurricularActivityUpdateView(BaseUpdateView):
    template_name = "topteenadmin/extracurricularactivity_form.html"
    model = ExtracurricularActivity
    form_class = ExtracurricularActivityModelForm
    success_url = reverse_lazy('topteenadminmanaged:extracurricularactivitylist')
    title = "Extracurricular Activity"
    active_tab = "extracurricularactivity"
    success_message = "Extracurricular activity updated successfully."

class ExtracurricularActivityDeleteView(BaseDeleteView):
    model = ExtracurricularActivity
    active_tab = "extracurricularactivity"
    success_url = reverse_lazy('topteenadminmanaged:extracurricularactivitylist')
    success_message = "Extracurricular activity deleted successfully."

class ExtracurricularActivityDetailView(BaseDetailView):
    template_name = "topteenadmin/extracurricularactivity_detail.html"
    model = ExtracurricularActivity
    title = "Extracurricular Activity"
    form_class = ExtracurricularActivityModelForm
    success_url = reverse_lazy('topteenadminmanaged:extracurricularactivitylist')
    active_tab = "extracurricularactivity"

# Entrance Test Prep (core)
class EntranceTestPrepCategoryListView(BaseListView):
    template_name = "topteenadmin/entrancetestprepcategory_list.html"
    title = "Entrance Test Prep Category"
    active_tab = "entrancetestprepcategory"
    model = EntranceTestPrepCategory
    filterset_class = EntranceTestPrepCategoryFilter
    context_object_name = "entrancetestprepcategory_list"

class CreateEntranceTestPrepCategory(BaseCreateView):
    template_name = "topteenadmin/entrancetestprepcategory_form.html"
    model = EntranceTestPrepCategory
    form_class = EntranceTestPrepCategoryModelForm
    success_url = reverse_lazy('topteenadminmanaged:entrancetestprepcategorylist')
    title = "Entrance Test Prep Category"
    active_tab = "entrancetestprepcategory"
    success_message = "Category created successfully."

class EntranceTestPrepCategoryUpdateView(BaseUpdateView):
    template_name = "topteenadmin/entrancetestprepcategory_form.html"
    model = EntranceTestPrepCategory
    form_class = EntranceTestPrepCategoryModelForm
    success_url = reverse_lazy('topteenadminmanaged:entrancetestprepcategorylist')
    title = "Entrance Test Prep Category"
    active_tab = "entrancetestprepcategory"
    success_message = "Category updated successfully."

class EntranceTestPrepCategoryDeleteView(BaseDeleteView):
    model = EntranceTestPrepCategory
    active_tab = "entrancetestprepcategory"
    success_url = reverse_lazy('topteenadminmanaged:entrancetestprepcategorylist')
    success_message = "Category deleted successfully."

class EntranceTestPrepCategoryDetailView(BaseDetailView):
    template_name = "topteenadmin/entrancetestprepcategory_detail.html"
    model = EntranceTestPrepCategory
    title = "Entrance Test Prep Category"
    form_class = EntranceTestPrepCategoryModelForm
    success_url = reverse_lazy('topteenadminmanaged:entrancetestprepcategorylist')
    active_tab = "entrancetestprepcategory"

class EntranceTestPrepExamListView(BaseListView):
    template_name = "topteenadmin/entrancetestprepexam_list.html"
    title = "Entrance Test Prep Exam"
    active_tab = "entrancetestprepexam"
    model = EntranceTestPrepExam
    filterset_class = EntranceTestPrepExamFilter
    context_object_name = "entrancetestprepexam_list"

class CreateEntranceTestPrepExam(BaseCreateView):
    template_name = "topteenadmin/entrancetestprepexam_form.html"
    model = EntranceTestPrepExam
    form_class = EntranceTestPrepExamModelForm
    success_url = reverse_lazy('topteenadminmanaged:entrancetestprepexamlist')
    title = "Entrance Test Prep Exam"
    active_tab = "entrancetestprepexam"
    success_message = "Exam created successfully."

class EntranceTestPrepExamUpdateView(BaseUpdateView):
    template_name = "topteenadmin/entrancetestprepexam_form.html"
    model = EntranceTestPrepExam
    form_class = EntranceTestPrepExamModelForm
    success_url = reverse_lazy('topteenadminmanaged:entrancetestprepexamlist')
    title = "Entrance Test Prep Exam"
    active_tab = "entrancetestprepexam"
    success_message = "Exam updated successfully."

class EntranceTestPrepExamDeleteView(BaseDeleteView):
    model = EntranceTestPrepExam
    active_tab = "entrancetestprepexam"
    success_url = reverse_lazy('topteenadminmanaged:entrancetestprepexamlist')
    success_message = "Exam deleted successfully."

class EntranceTestPrepExamDetailView(BaseDetailView):
    template_name = "topteenadmin/entrancetestprepexam_detail.html"
    model = EntranceTestPrepExam
    title = "Entrance Test Prep Exam"
    form_class = EntranceTestPrepExamModelForm
    success_url = reverse_lazy('topteenadminmanaged:entrancetestprepexamlist')
    active_tab = "entrancetestprepexam"

class ProfessionListView(BaseListView):
    template_name = "topteenadmin/profession_list.html"
    title="Profession"
    active_tab="profession"
    model = Profession
    filterset_class = ProfessionFilter

class CreateProfession(BaseCreateView):
    template_name = "topteenadmin/profession_form.html"
    model=Profession
    form_class=ProfessionModelForm
    success_url=reverse_lazy('topteenadminmanaged:professionlist')
    title="Profession"
    active_tab="profession"
    success_message="Profession created successfully."
    
class ProfessionUpdateView(BaseUpdateView):
    template_name = "topteenadmin/profession_form.html"
    model=Profession
    title="Profession"
    form_class=ProfessionModelForm
    success_url=reverse_lazy('topteenadminmanaged:professionlist')
    active_tab="profession"
    success_message="Profession updated successfully."

class ProfessionDeleteView(BaseDeleteView):
    model = Profession
    active_tab="profession"
    success_url = reverse_lazy('topteenadminmanaged:professionlist')
    success_message="Profession deleted successfully."
    
class ProfessionDetailView(BaseDetailView):
    template_name = "topteenadmin/profession_detail.html"
    model=Profession
    title="Profession"
    form_class=ProfessionModelForm
    success_url=reverse_lazy('topteenadminmanaged:professionlist')
    active_tab="Profession"




class LoginView(TemplateView):
    template_name = "topteenadmin/user/login.html"
    
    def post(self,request,*args,**kwargs):
        data={}
        email= request.POST.get('email')
        password = request.POST.get('password')
        if email and password:
            user = authenticate(request, username=email, password=password)
            if user and user.is_staff:
                login(request, user, backend='users.backends.CustomUserBackend')
                return redirect(reverse('topteenadmin:topteendashboard'))
            else:
                messages.info(request, "Login failed. Invalid username/password..")
        return render(request, self.template_name,data)


def password_reset_request(request, *args, **kwargs):
    template_name="template20/admin/forgot_password.html"
    password_reset_form=None
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        data = request.POST.get('email')
        if password_reset_form.is_valid():
            user = User.objects.filter(email=data).first()
            if user:
                c = { 
                "email":user.email,
                'domain':request.META['HTTP_HOST'],
                'site_name': 'topteen',
                "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                "user": user,
                'token': default_token_generator.make_token(user),
                'protocol': 'http',
                }
                url=c['protocol']+"://"+c['domain']+"/topteenadmin/changepassword/"+c['uid']+"/"+c['token']
                from communication.email_templates import render_transactional_email
                subject, text_content, html_content = render_transactional_email(
                    'password_reset',
                    format_context={'email': user.email, 'url': url},
                    django_context={'url': url},
                )
                try:
                    msg = EmailMultiAlternatives(
                        subject,
                        text_content,
                        settings.DEFAULT_FROM_EMAIL,
                        [user.email],
                    )
                    msg.attach_alternative(html_content, 'text/html')
                    msg.send(fail_silently=False)
                except BadHeaderError:
                    return HttpResponse('Invalid header found.')
                messages.info(request, "Password reset instructions have been sent to the email address entered.")
                return redirect(reverse('topteenadmin:forgotpassword'))
            else:
                msg={}
                msg['message']="Invalid email address"
                messages.error(request, 'An invalid email has been entered.')
                return render(request, template_name,msg)

    password_reset_form = PasswordResetForm()
    return render(request, template_name, {"password_reset_form":password_reset_form})


class ChangePasswordView(TemplateView):
    template_name = 'template20/admin/change_password.html'

    def html_head(self):
        s='Edit My Profile'
        return build_html_head(title=s, description=s)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['uidb64'] = kwargs.get('uidb64', '')
        ctx['token'] = kwargs.get('token', '')
        return ctx

    def post(self,request, uidb64,token,*args, **kwargs):
        ctx={}
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
        password=request.POST.get('password',False)
        password2=request.POST.get('password2',False)
        status,message=False,""
        if password == password2:
            user.set_password(password)
            update_session_auth_hash(request,user)
            user.save()
            messages.info(request, "Password updated successfully. Login again to continue.")
            return redirect(reverse('topteenadmin:login'))
        else:
            messages.error(request, "Entered passwords do not match, please try again.")

        ctx['status']=status
        ctx['message']=message
        return render(request, self.template_name, ctx)


def profileupdate(request, *args, **kwargs):
    ctx = {}
    template_name = 'topteenadmin/profile_update.html'
    user = User.objects.get(pk=request.user.id)
    ctx['profile_user'] = user  # avoid clash with auth context processor's 'user' (request.user)
    ctx['meta_title'] = 'Profile'
    ctx['breadcrumb'] = build_admin_breadcrumb([{'title': 'Profile', 'text': 'Profile', 'url': reverse('topteenadmin:UpdateProfile')}])

    if request.method == "POST":
        if "name" in request.POST or request.FILES.get('image') or request.POST.get('image_remove') or "oldpassword" in request.POST:
            name = request.POST.get('name')
            image = request.FILES.get('image')
            if name:
                user.name = name
            if request.POST.get('image_remove'):
                if user.image:
                    user.image.delete(save=False)
                user.image = None
            elif image:
                user.image = image
            oldpassword = request.POST.get('oldpassword')
            if oldpassword and user.check_password(oldpassword):
                password = request.POST.get('newpassword')
                password2 = request.POST.get('confirmpassword')
                if password == password2:
                    user.set_password(password)
                    messages.success(request, 'Password updated successfully.')
                else:
                    ctx['msg'] = "New Password and confirm Password is Not Same"
            elif oldpassword:
                ctx['message'] = "Your old password does not match with existing password"
        user.save()
        if not ctx.get('msg') and not ctx.get('message'):
            messages.success(request, 'Profile updated successfully.')

    return render(request, template_name, ctx)


def custom_logout(request):
    print('Loggin out {}'.format(request.user))
    logout(request)
    return redirect(reverse('topteenadmin:login'))

class CountryListView(BaseListView):
    template_name = "topteenadmin/country_list.html"
    title="Country"
    active_tab="country"
    model = Country
    filterset_class = CountryFilter
    
class CreateCountry(BaseCreateView):
    template_name = "topteenadmin/country_form.html"
    model=Country
    form_class=CountryModelForm
    success_url=reverse_lazy('topteenadminmanaged:countrylist')
    title="Country"
    active_tab="college"
    success_message="Country created successfully."
    
class CountryUpdateView(BaseUpdateView):
    template_name = "topteenadmin/country_form.html"
    model=Country
    title="Country"
    form_class=CountryModelForm
    success_url=reverse_lazy('topteenadminmanaged:countrylist')
    active_tab="country"
    success_message="Country updated successfully."

class CountryDeleteView(BaseDeleteView):
    model = Country
    active_tab="country"
    success_url = reverse_lazy('topteenadminmanaged:countrylist')
    success_message="Country deleted successfully."
    
class CountryDetailView(BaseDetailView):
    template_name = "topteenadmin/country_detail.html"
    model=Country
    title="Country"
    form_class=CountryModelForm
    success_url=reverse_lazy('topteenadminmanaged:countrylist')
    active_tab="country"
    
class StateListView(BaseListView):
    template_name = "topteenadmin/state_list.html"
    title="State"
    active_tab="state"
    model = State
    filterset_class = StateFilter
    
class CreateState(BaseCreateView):
    template_name = "topteenadmin/state_form.html"
    model=State
    form_class=StateModelForm
    success_url=reverse_lazy('topteenadminmanaged:statelist')
    title="State"
    active_tab="state"
    success_message="State created successfully."
    
class StateUpdateView(BaseUpdateView):
    template_name = "topteenadmin/state_form.html"
    model=State
    title="State"
    form_class=StateModelForm
    success_url=reverse_lazy('topteenadminmanaged:statelist')
    active_tab="state"
    success_message="State updated successfully."

class StateDeleteView(BaseDeleteView):
    model = State
    active_tab="state"
    success_url = reverse_lazy('topteenadminmanaged:statelist')
    success_message="State deleted successfully."
    
class StateDetailView(BaseDetailView):
    template_name = "topteenadmin/state_detail.html"
    model=State
    title="State"
    form_class=StateModelForm
    success_url=reverse_lazy('topteenadminmanaged:statelist')
    active_tab="state"
    
class CityListView(BaseListView):
    template_name = "topteenadmin/city_list.html"
    title="City"
    active_tab="City"
    model = City
    filterset_class = CityFilter
    
class CreateCity(BaseCreateView):
    template_name = "topteenadmin/city_form.html"
    model=City
    form_class=CityModelForm
    success_url=reverse_lazy('topteenadminmanaged:citylist')
    title="City"
    active_tab="city"
    success_message="City created successfully."
    
class CityUpdateView(BaseUpdateView):
    template_name = "topteenadmin/city_form.html"
    model=City
    title="City"
    form_class=CityModelForm
    success_url=reverse_lazy('topteenadminmanaged:citylist')
    active_tab="state"
    success_message="City updated successfully."

class CityDeleteView(BaseDeleteView):
    model = City
    active_tab="city"
    success_url = reverse_lazy('topteenadminmanaged:citylist')
    success_message="City deleted successfully."
    
class CityDetailView(BaseDetailView):
    template_name = "topteenadmin/city_detail.html"
    model=City
    title="City"
    form_class=CityModelForm
    success_url=reverse_lazy('topteenadminmanaged:citylist')
    active_tab="city"

class AjaxCollegeFilter(BaseListView):
    template_name ="topteenadmin/includes/ajaxcollegefilter.html"

    def get_context(self,request,*args, **kwargs):
        ctx={}
        print("this is the search",request.GET.get('search'))
         
        clgfilter=College.objects.filter(name__contains=request.GET.get('search'))
        print("This is filter result",clgfilter)
        ctx['college_list']=clgfilter
        return ctx

    def get(self, request,*args, **kwargs):
        html = render_to_string(self.template_name,self.get_context(request, *args, **kwargs))
        return HttpResponse(html)

class CollegeListView(BaseListView):
    template_name = "topteenadmin/college_list.html"
    title="College"
    active_tab="college"
    model = College
    filterset_class = CollegeFilter
    
class CreateCollege(BaseCreateView):
    template_name = "topteenadmin/college_form.html"
    model=College
    form_class=CollegeModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegelist')
    title="College"
    active_tab="college"
    success_message="College created successfully."
    
class CollegeUpdateView(BaseUpdateView):
    template_name = "topteenadmin/college_form.html"
    model=College
    title="College"
    form_class=CollegeModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegelist')
    active_tab="college"
    success_message="College updated successfully."

class CollegeDeleteView(BaseDeleteView):
    model = College
    active_tab="college"
    success_url = reverse_lazy('topteenadminmanaged:collegelist')
    success_message="College deleted successfully."
    
class CollegeDetailView(BaseDetailView):
    template_name = "topteenadmin/college_detail.html"
    model=College
    title="College"
    form_class=CollegeModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegelist')
    active_tab="college"
    
class CollegeImagesListView(BaseListView):
    template_name = "topteenadmin/collegeimages_list.html"
    title="CollegeImages"
    active_tab="collegeimages"
    model = CollegeImages
    filterset_class = CollegeImagesFilter
    
class CreateCollegeImages(BaseCreateView):
    template_name = "topteenadmin/collegeimages_form.html"
    model=CollegeImages
    form_class=CollegeImagesModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegeimageslist')
    title="CollegeImages"
    active_tab="collegeimages"
    success_message="CollegeImages created successfully."
    
class CollegeImagesUpdateView(BaseUpdateView):
    template_name = "topteenadmin/collegeimages_form.html"
    model=CollegeImages
    title="CollegeImages"
    form_class=CollegeImagesModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegeimageslist')
    active_tab="collegeimages"
    success_message="CollegeImages updated successfully."

class CollegeImagesDeleteView(BaseDeleteView):
    model = CollegeImages
    active_tab="collegeimages"
    success_url = reverse_lazy('topteenadminmanaged:collegeimageslist')
    success_message="CollegeImages deleted successfully."
    
class CollegeImagesDetailView(BaseDetailView):
    template_name = "topteenadmin/collegeimages_detail.html"
    model=CollegeImages
    title="CollegeImages"
    form_class=CollegeImagesModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegeimageslist')
    active_tab="collegeimages"
    
class CollegeFlatTextListView(BaseListView):
    template_name = "topteenadmin/collegeflattext_list.html"
    title="CollegeFlatText"
    active_tab="collegeflattext"
    model = CollegeFlatText
    filterset_class = CollegeFlatTextFilter
    
class CreateCollegeFlatText(BaseCreateView):
    template_name = "topteenadmin/collegeflattext_form.html"
    model=CollegeFlatText
    form_class=CollegeFlatTextModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegeflattextlist')
    title="CollegeFlatText"
    active_tab="collegeflatext"
    success_message="CollegeFlatText created successfully."
    
class CollegeFlatTextUpdateView(BaseUpdateView):
    template_name = "topteenadmin/collegeflattext_form.html"
    model=CollegeFlatText
    title="CollegeFlatText"
    form_class=CollegeFlatTextModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegeflattextlist')
    active_tab="collegeflattext"
    success_message="CollegeFlatText updated successfully."

class CollegeFlatTextDeleteView(BaseDeleteView):
    model = CollegeFlatText
    active_tab="collegeflattext"
    success_url = reverse_lazy('topteenadminmanaged:collegeflattextlist')
    success_message="CollegeFlatText deleted successfully."
    
class CollegeFlatTextDetailView(BaseDetailView):
    template_name = "topteenadmin/collegeflattext_detail.html"
    model=CollegeFlatText
    title="CollegeFlatText"
    form_class=CollegeFlatTextModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegeflattextlist')
    active_tab="collegeflattext"
    
class CollegeTextListView(BaseListView):
    template_name = "topteenadmin/collegetext_list.html"
    title="CollegText"
    active_tab="collegetext"
    model = CollegeText
    filterset_class = CollegeTextFilter
    
class CreateCollegeText(BaseCreateView):
    template_name = "topteenadmin/collegetext_form.html"
    model=CollegeText
    form_class=CollegeTextModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegetextlist')
    title="CollegeText"
    active_tab="collegetext"
    success_message="CollegeText created successfully."
    
class CollegeTextUpdateView(BaseUpdateView):
    template_name = "topteenadmin/collegetext_form.html"
    model=CollegeText
    title="Collegetext"
    form_class=CollegeTextModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegetextlist')
    active_tab="collegetext"
    success_message="CollegeText updated successfully."

class CollegeTextDeleteView(BaseDeleteView):
    model = CollegeText
    active_tab="collegetext"
    success_url = reverse_lazy('topteenadminmanaged:collegetextlist')
    success_message="CollegeText deleted successfully."
    
class CollegeTextDetailView(BaseDetailView):
    template_name = "topteenadmin/collegetext_detail.html"
    model=CollegeText
    title="CollegeText"
    form_class=CollegeTextModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegetextlist')
    active_tab="colleFacts"
    
class CollegeFactsListView(BaseListView):
    template_name = "topteenadmin/collegefacts_list.html"
    title="CollegeFacts"
    active_tab="collegefacts"
    model = CollegeFacts
    filterset_class = CollegeFactsFilter
    
class CreateCollegeFacts(BaseCreateView):
    template_name = "topteenadmin/collegefacts_form.html"
    model=CollegeFacts
    form_class=CollegeFactsModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegefactslist')
    title="CollegeFacts"
    active_tab="collegefacts"
    success_message="CollegeFacts created successfully."
    
class CollegeFactsUpdateView(BaseUpdateView):
    template_name = "topteenadmin/collegefacts_form.html"
    model=CollegeFacts
    title="CollegeFacts"
    form_class=CollegeFactsModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegefactslist')
    active_tab="collegefacts"
    success_message="CollegeFacts updated successfully."

class CollegeFactsDeleteView(BaseDeleteView):
    model = CollegeFacts
    active_tab="collegefacts"
    success_url = reverse_lazy('topteenadminmanaged:collegefactslist')
    success_message="CollegeFacts deleted successfully."
    
class CollegeFactsDetailView(BaseDetailView):
    template_name = "topteenadmin/collegefacts_detail.html"
    model=CollegeFacts
    title="CollegeFacts"
    form_class=CollegeFactsModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegefactslist')
    active_tab="collegefacts"
    
class RecruitingCompaniesListView(BaseListView):
    template_name = "topteenadmin/recruitingcompanies_list.html"
    title="RecruitingCompanies"
    active_tab="recruitingcompanies"
    model = RecruitingCompanies
    filterset_class = RecruitingCompaniesFilter
    
class CreateRecruitingCompanies(BaseCreateView):
    template_name = "topteenadmin/recruitingcompanies_form.html"
    model=RecruitingCompanies
    form_class=RecruitingCompaniesModelForm
    success_url=reverse_lazy('topteenadminmanaged:recruitingcompanieslist')
    title="RecruitingCompanies"
    active_tab="recruitingcompanies"
    success_message="RecruitingCompanies created successfully."
    
class RecruitingCompaniesUpdateView(BaseUpdateView):
    template_name = "topteenadmin/recruitingcompanies_form.html"
    model=RecruitingCompanies
    title="RecruitingCompanies"
    form_class=RecruitingCompaniesModelForm
    success_url=reverse_lazy('topteenadminmanaged:recruitingcompanieslist')
    active_tab="recruitingcompanies"
    success_message="RecruitingCompanies updated successfully."

class RecruitingCompaniesDeleteView(BaseDeleteView):
    model = RecruitingCompanies
    active_tab="recruitingcompanies"
    success_url = reverse_lazy('topteenadminmanaged:recruitingcompanieslist')
    success_message="RecruitingCompanies deleted successfully."
    
class RecruitingCompaniesDetailView(BaseDetailView):
    template_name = "topteenadmin/recruitingcompanies_detail.html"
    model=RecruitingCompanies
    title="RecruitingCompanies"
    form_class=RecruitingCompaniesModelForm
    success_url=reverse_lazy('topteenadminmanaged:recruitingcompanieslist')
    active_tab="recruitingcompanies"
    
class CollegeRecruitingCompaniesListView(BaseListView):
    template_name = "topteenadmin/collegerecruitingcompanies_list.html"
    title="CollegeRecruitingCompanies"
    active_tab="collegerecruitingcompanies"
    model = CollegeRecruitingCompanies
    filterset_class = CollegeRecruitingCompaniesFilter
    
class CreateCollegeRecruitingCompanies(BaseCreateView):
    template_name = "topteenadmin/collegerecruitingcompanies_form.html"
    model=CollegeRecruitingCompanies
    form_class=CollegeRecruitingCompaniesModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegerecruitingcompanieslist')
    title="CollegeRecruitingCompanies"
    active_tab="collegerecruitingcompanies"
    success_message="CollegeRecruitingCompanies created successfully."
    
class CollegeRecruitingCompaniesUpdateView(BaseUpdateView):
    template_name = "topteenadmin/collegerecruitingcompanies_form.html"
    model=CollegeRecruitingCompanies
    title="CollegeRecruitingCompanies"
    form_class=CollegeRecruitingCompaniesModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegerecruitingcompanieslist')
    active_tab="collegerecruitingcompanies"
    success_message="CollegeRecruitingCompanies updated successfully."

class CollegeRecruitingCompaniesDeleteView(BaseDeleteView):
    model = CollegeRecruitingCompanies
    active_tab="collegerecruitingcompanies"
    success_url = reverse_lazy('topteenadminmanaged:collegerecruitingcompanieslist')
    success_message="CollegeRecruitingCompanies deleted successfully."
    
class CollegeRecruitingCompaniesDetailView(BaseDetailView):
    template_name = "topteenadmin/collegerecruitingcompanies_detail.html"
    model=CollegeRecruitingCompanies
    title="CollegeRecruitingCompanies"
    form_class=CollegeRecruitingCompaniesModelForm
    success_url=reverse_lazy('topteenadminmanaged:collgerecruitingcompanieslist')
    active_tab="collegerecruitingcompanies"
    
class FacilityListView(BaseListView):
    template_name = "topteenadmin/facility_list.html"
    title="Facility"
    active_tab="facility"
    model = Facility
    filterset_class = FacilityFilter
    
class CreateFacility(BaseCreateView):
    template_name = "topteenadmin/facility_form.html"
    model=Facility
    form_class=FacilityModelForm
    success_url=reverse_lazy('topteenadminmanaged:facilitylist')
    title="Facility"
    active_tab="facility"
    success_message="Facility created successfully."
    
class FacilityUpdateView(BaseUpdateView):
    template_name = "topteenadmin/facility_form.html"
    model=Facility
    title="Facility"
    form_class=FacilityModelForm
    success_url=reverse_lazy('topteenadminmanaged:facilitylist')
    active_tab="recruitingcompanies"
    success_message="Facility updated successfully."

class FacilityDeleteView(BaseDeleteView):
    model = Facility
    active_tab="facility"
    success_url = reverse_lazy('topteenadminmanaged:facilitylist')
    success_message="Facility deleted successfully."
    
class FacilityDetailView(BaseDetailView):
    template_name = "topteenadmin/facility_detail.html"
    model=Facility
    title="Facility"
    form_class=FacilityModelForm
    success_url=reverse_lazy('topteenadminmanaged:facilitylist')
    active_tab="facility"
    
class CollegeFacilityListView(BaseListView):
    template_name = "topteenadmin/collegefacility_list.html"
    title="CollegeFacility"
    active_tab="collegefacility"
    model = CollegeFacility
    filterset_class = CollegeFacilityFilter
    
class CreateCollegeFacility(BaseCreateView):
    template_name = "topteenadmin/collegefacility_form.html"
    model=CollegeFacility
    form_class=CollegeFacilityModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegefacilitylist')
    title="CollegeFacility"
    active_tab="collegefacility"
    success_message="CollegeFacility created successfully."
    
class CollegeFacilityUpdateView(BaseUpdateView):
    template_name = "topteenadmin/collegefacility_form.html"
    model=CollegeFacility
    title="CollegeFacility"
    form_class=CollegeFacilityModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegefacilitylist')
    active_tab="collegefacility"
    success_message="CollegeFacility updated successfully."

class CollegeFacilityDeleteView(BaseDeleteView):
    model = CollegeFacility
    active_tab="collegefacility"
    success_url = reverse_lazy('topteenadminmanaged:collegefacilitylist')
    success_message="CollegeFacility deleted successfully."
    
class CollegeFacilityDetailView(BaseDetailView):
    template_name = "topteenadmin/collegefacility_detail.html"
    model=CollegeFacility
    title="CollegeFacility"
    form_class=CollegeFacilityModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegefacilitylist')
    active_tab="collegefacility"
    
    
class CollegeMoneyValueListView(BaseListView):
    template_name = "topteenadmin/collegemoneyvalue_list.html"
    title="CollegeMoneyValue"
    active_tab="collegemoneyvalue"
    model = CollegeMoneyValue
    filterset_class = CollegeMoneyValueFilter
    
class CreateCollegeMoneyValue(BaseCreateView):
    template_name = "topteenadmin/collegemoneyvalue_form.html"
    model=CollegeMoneyValue
    form_class=CollegeMoneyValueModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegemoneyvaluelist')
    title="CollegeMoneyValue"
    active_tab="collegemoneyvalue"
    success_message="CollegeMoneyValue created successfully."
    
class CollegeMoneyValueUpdateView(BaseUpdateView):
    template_name = "topteenadmin/collegemoneyvalue_form.html"
    model=CollegeMoneyValue
    title="CollegeMoneyValue"
    form_class=CollegeMoneyValueModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegemoneyvaluelist')
    active_tab="collegemoneyvalue"
    success_message="CollegeMoneyValue updated successfully."

class CollegeMoneyValueDeleteView(BaseDeleteView):
    model = CollegeMoneyValue
    active_tab="collegemoneyvalue"
    success_url = reverse_lazy('topteenadminmanaged:collegemoneyvaluelist')
    success_message="CollegeMoneyValue deleted successfully."
    
class CollegeMoneyValueDetailView(BaseDetailView):
    template_name = "topteenadmin/collegemoneyvalue_detail.html"
    model=CollegeMoneyValue
    title="CollegeMoneyValue"
    form_class=CollegeMoneyValueModelForm
    success_url=reverse_lazy('topteenadminmanaged:collegemoneyvaluelist')
    active_tab="collegemoneyvalue"
    
class StreamListView(BaseListView):
    template_name = "topteenadmin/stream_list.html"
    title="Stream"
    active_tab="stream"
    model = Stream
    filterset_class = StreamFilter
    
class CreateStream(BaseCreateView):
    template_name = "topteenadmin/stream_form.html"
    model=Stream
    form_class=StreamModelForm
    success_url=reverse_lazy('topteenadminmanaged:streamlist')
    title="Stream"
    active_tab="stream"
    success_message="course created successfully."
    
class StreamUpdateView(BaseUpdateView):
    template_name = "topteenadmin/stream_form.html"
    model=Stream
    title="Stream"
    form_class=StreamModelForm
    success_url=reverse_lazy('topteenadminmanaged:streamlist')
    active_tab="stream"
    success_message="Stream updated successfully."

class StreamDeleteView(BaseDeleteView):
    model = Stream
    active_tab="stream"
    success_url = reverse_lazy('topteenadminmanaged:streamlist')
    success_message="Stream deleted successfully."
    
class StreamDetailView(BaseDetailView):
    template_name = "topteenadmin/stream_detail.html"
    model=Stream
    title="Stream"
    form_class=StreamModelForm
    success_url=reverse_lazy('topteenadminmanaged:streamlist')
    active_tab="stream"
     
class CourseListView(BaseListView):
    template_name = "topteenadmin/course_list.html"
    title="Course"
    active_tab="course"
    model = Course
    filterset_class = CourseFilter
    
class CreateCourse(BaseCreateView):
    template_name = "topteenadmin/course_form.html"
    model=Course
    form_class=CourseModelForm
    success_url=reverse_lazy('topteenadminmanaged:courselist')
    title="Course"
    active_tab="course"
    success_message="Course created successfully."
    
class CourseUpdateView(BaseUpdateView):
    template_name = "topteenadmin/course_form.html"
    model=Course
    title="Course"
    form_class=CourseModelForm
    success_url=reverse_lazy('topteenadminmanaged:courselist')
    active_tab="course"
    success_message="Course updated successfully."

class CourseDeleteView(BaseDeleteView):
    model = Course
    active_tab="course"
    success_url = reverse_lazy('topteenadminmanaged:courselist')
    success_message="Course deleted successfully."
    
class CourseDetailView(BaseDetailView):
    template_name = "topteenadmin/course_detail.html"
    model=Course
    title="Course"
    form_class=CourseModelForm
    success_url=reverse_lazy('topteenadminmanaged:courselist')
    active_tab="course"
    
class CourseFactsListView(BaseListView):
    template_name = "topteenadmin/coursefacts_list.html"
    title="CourseFacts"
    active_tab="coursefacts"
    model = CourseFacts
    filterset_class = CourseFactsFilter
    
class CreateCourseFacts(BaseCreateView):
    template_name = "topteenadmin/coursefacts_form.html"
    model=CourseFacts
    form_class=CourseFactsModelForm
    success_url=reverse_lazy('topteenadminmanaged:coursefactslist')
    title="CourseFacts"
    active_tab="coursefacts"
    success_message="CourseFacts created successfully."
    
class CourseFactsUpdateView(BaseUpdateView):
    template_name = "topteenadmin/coursefacts_form.html"
    model=CourseFacts
    title="CourseFacts"
    form_class=CourseFactsModelForm
    success_url=reverse_lazy('topteenadminmanaged:coursefactslist')
    active_tab="coursefacts"
    success_message="CourseFacts updated successfully."

class CourseFactsDeleteView(BaseDeleteView):
    model = CourseFacts
    active_tab="coursefacts"
    success_url = reverse_lazy('topteenadminmanaged:coursefactslist')
    success_message="CourseFacts deleted successfully."
    
class CourseFactsDetailView(BaseDetailView):
    template_name = "topteenadmin/course_detail.html"
    model=CourseFacts
    title="CourseFacts"
    form_class=CourseFactsModelForm
    success_url=reverse_lazy('topteenadminmanaged:coursefactslist')
    active_tab="coursefacts"
    
class CourseTextListView(BaseListView):
    template_name = "topteenadmin/coursetext_list.html"
    title="CourseText"
    active_tab="coursetext"
    model = CourseText
    filterset_class = CourseTextFilter
    
class CreateCourseText(BaseCreateView):
    template_name = "topteenadmin/coursetext_form.html"
    model=CourseText
    form_class=CourseTextModelForm
    success_url=reverse_lazy('topteenadminmanaged:coursetextlist')
    title="CourseText"
    active_tab="coursetext"
    success_message="CourseText created successfully."
    
class CourseTextUpdateView(BaseUpdateView):
    template_name = "topteenadmin/coursetext_form.html"
    model=CourseText
    title="CourseText"
    form_class=CourseTextModelForm
    success_url=reverse_lazy('topteenadminmanaged:coursetextlist')
    active_tab="coursetext"
    success_message="CourseText updated successfully."

class CourseTextDeleteView(BaseDeleteView):
    model = CourseText
    active_tab="coursetext"
    success_url = reverse_lazy('topteenadminmanaged:coursetextlist')
    success_message="CourseText deleted successfully."
    
class CourseTextDetailView(BaseDetailView):
    template_name = "topteenadmin/coursetext_detail.html"
    model=CourseText
    title="CourseText"
    form_class=CourseTextModelForm
    success_url=reverse_lazy('topteenadminmanaged:coursetextlist')
    active_tab="coursetext"
    
class CourseMoneyValueListView(BaseListView):
    template_name = "topteenadmin/coursemoneyvalue_list.html"
    title="CourseMoneyValue"
    active_tab="coursemoneyvalue"
    model = CourseMoneyValue
    filterset_class = CourseMoneyValueFilter
    
class CreateCourseMoneyValue(BaseCreateView):
    template_name = "topteenadmin/coursemoneyvalue_form.html"
    model=CourseMoneyValue
    form_class=CourseMoneyValueModelForm
    success_url=reverse_lazy('topteenadminmanaged:coursemoneyvaluelist')
    title="CourseMoneyValue"
    active_tab="coursemoneyvalue"
    success_message="CourseMoneyValue created successfully."
    
class CourseMoneyValueUpdateView(BaseUpdateView):
    template_name = "topteenadmin/coursemoneyvalue_form.html"
    model=CourseMoneyValue
    title="CourseMoneyValue"
    form_class=CourseMoneyValueModelForm
    success_url=reverse_lazy('topteenadminmanaged:coursemoneyvaluelist')
    active_tab="coursemoneyvalue"
    success_message="CourseMoneyValue updated successfully."

class CourseMoneyValueDeleteView(BaseDeleteView):
    model = CourseMoneyValue
    active_tab="coursemoneyvalue"
    success_url = reverse_lazy('topteenadminmanaged:coursemoneyvaluelist')
    success_message="CourseMoneyValue deleted successfully."
    
class CourseMoneyValueDetailView(BaseDetailView):
    template_name = "topteenadmin/coursemoneyvalue_detail.html"
    model=CourseMoneyValue
    title="CourseMoneyValue"
    form_class=CourseMoneyValueModelForm
    success_url=reverse_lazy('topteenadminmanaged:coursemoneyvaluelist')
    active_tab="coursemoneyvalue"
    
class CourseIntakeListView(BaseListView):
    template_name = "topteenadmin/courseintake_list.html"
    title="CourseIntake"
    active_tab="courseintake"
    model = CourseIntake
    filterset_class = CourseIntakeFilter
    
class CreateCourseIntake(BaseCreateView):
    template_name = "topteenadmin/courseintake_form.html"
    model=CourseIntake
    form_class=CourseIntakeModelForm
    success_url=reverse_lazy('topteenadminmanaged:courseintakelist')
    title="CourseIntake"
    active_tab="courseintake"
    success_message="CourseIntake created successfully."
    
class CourseIntakeUpdateView(BaseUpdateView):
    template_name = "topteenadmin/courseintake_form.html"
    model=CourseIntake
    title="CourseIntake"
    form_class=CourseIntakeModelForm
    success_url=reverse_lazy('topteenadminmanaged:courseintakelist')
    active_tab="courseintake"
    success_message="CourseIntake updated successfully."

class CourseIntakeDeleteView(BaseDeleteView):
    model = CourseIntake
    active_tab="courseintake"
    success_url = reverse_lazy('topteenadminmanaged:courseintakelist')
    success_message="CourseIntake deleted successfully."
    
class CourseIntakeDetailView(BaseDetailView):
    template_name = "topteenadmin/courseintake_detail.html"
    model=CourseIntake
    title="CourseIntake"
    form_class=CourseIntakeModelForm
    success_url=reverse_lazy('topteenadminmanaged:courseintakelist')
    active_tab="courseintake"
    
    
class CourseEnglighRequirementsListView(BaseListView):
    template_name = "topteenadmin/courseenglighrequirements_list.html"
    title="CourseEnglighRequirements"
    active_tab="courseenglighrequirements"
    model = CourseEnglighRequirements
    filterset_class = CourseEnglighRequirementsFilter
    
class CreateCourseEnglighRequirements(BaseCreateView):
    template_name = "topteenadmin/courseenglighrequirements_form.html"
    model=CourseEnglighRequirements
    form_class=CourseEnglighRequirementsModelForm
    success_url=reverse_lazy('topteenadminmanaged:courseenglighrequirementslist')
    title="CourseEnglighRequirements"
    active_tab="courseenglighrequirements"
    success_message="CourseEnglighRequirements created successfully."
    
class CourseEnglighRequirementsUpdateView(BaseUpdateView):
    template_name = "topteenadmin/courseenglighrequirements_form.html"
    model=CourseEnglighRequirements
    title="CourseEnglighRequirements"
    form_class=CourseEnglighRequirementsModelForm
    success_url=reverse_lazy('topteenadminmanaged:courseenglighrequirementslist')
    active_tab="courseenglighrequirements"
    success_message="CourseEnglighRequirements updated successfully."

class CourseEnglighRequirementsDeleteView(BaseDeleteView):
    model = CourseEnglighRequirements
    active_tab="courseenglighrequirements"
    success_url = reverse_lazy('topteenadminmanaged:courseenglighrequirementslist')
    success_message="CourseEnglighRequirements deleted successfully."
    
class CourseEnglighRequirementsDetailView(BaseDetailView):
    template_name = "topteenadmin/courseenglighrequirements_detail.html"
    model=CourseEnglighRequirements
    title="CourseEnglighRequirements"
    form_class=CourseEnglighRequirementsModelForm
    success_url=reverse_lazy('topteenadminmanaged:courseenglighrequirementslist')
    active_tab="courseenglighrequirements"
    
class EntranceExamListView(BaseListView):
    template_name = "topteenadmin/entranceexam_list.html"
    title="EntranceExam"
    active_tab="entranceexam"
    model = EntranceExam
    filterset_class = EntranceExamFilter
    
class CreateEntranceExam(BaseCreateView):
    template_name = "topteenadmin/entranceexam_form.html"
    model=EntranceExam
    form_class=EntranceExamModelForm
    success_url=reverse_lazy('topteenadminmanaged:entranceexamlist')
    title="EntranceExam"
    active_tab="entranceexam"
    success_message="EntranceExam created successfully."
    
class EntranceExamUpdateView(BaseUpdateView):
    template_name = "topteenadmin/entranceexam_form.html"
    model=EntranceExam
    title="EntranceExam"
    form_class=EntranceExamModelForm
    success_url=reverse_lazy('topteenadminmanaged:entranceexamlist')
    active_tab="entranceexam"
    success_message="EntranceExam updated successfully."

class EntranceExamDeleteView(BaseDeleteView):
    model = EntranceExam
    active_tab="entranceexam"
    success_url = reverse_lazy('topteenadminmanaged:entranceexamlist')
    success_message="EntranceExam deleted successfully."
    
class EntranceExamDetailView(BaseDetailView):
    template_name = "topteenadmin/entranceexam_detail.html"
    model=EntranceExam
    title="EntranceExam"
    form_class=EntranceExamModelForm
    success_url=reverse_lazy('topteenadminmanaged:entranceexamlist')
    active_tab="entranceexam"

class ExamTagsListView(BaseListView):
    template_name = "topteenadmin/examtags_list.html"
    title="ExamTags"
    active_tab="examtags"
    model = ExamTags
    filterset_class = ExamTagsFilter
    
class CreateExamTags(BaseCreateView):
    template_name = "topteenadmin/examtags_form.html"
    model=ExamTags
    form_class=ExamTagsModelForm
    success_url=reverse_lazy('topteenadminmanaged:examtagslist')
    title="ExamTags"
    active_tab="examtags"
    success_message="ExamTag created successfully."
    
class ExamTagsUpdateView(BaseUpdateView):
    template_name = "topteenadmin/examtags_form.html"
    model=ExamTags
    title="ExamTags"
    form_class=ExamTagsModelForm
    success_url=reverse_lazy('topteenadminmanaged:examtagslist')
    active_tab="examtags"
    success_message="ExamTag updated successfully."

class ExamTagsDeleteView(BaseDeleteView):
    model = ExamTags
    active_tab="examtags"
    success_url = reverse_lazy('topteenadminmanaged:examtagslist')
    success_message="ExamTag deleted successfully."
    
class ExamTagsDetailView(BaseDetailView):
    template_name = "topteenadmin/examtags_details.html"
    model=ExamTags
    title="ExamTags"
    form_class=ExamTagsModelForm
    success_url=reverse_lazy('topteenadminmanaged:examtagslist')
    active_tab="examtags"


class CareerTagsView(BaseListView):
    template_name = "topteenadmin/careertags_list.html"
    title="CareerTags"
    active_tab="careertags"
    model = CareerTags
    filterset_class = CareerFilter
    
class CreateCareerTags(BaseCreateView):
    template_name = "topteenadmin/careertags_form.html"
    model=CareerTags
    form_class=CareerTagsForm
    success_url=reverse_lazy('topteenadminmanaged:careertagslist')
    title="CareerTags"
    active_tab="careertags"
    success_message="Cereer created successfully."
    
class CareerTagsUpdateView(BaseUpdateView):
    template_name = "topteenadmin/careertags_form.html"
    model=CareerTags
    title="CareerTags"
    form_class=CareerTagsForm
    success_url=reverse_lazy('topteenadminmanaged:careertagslist')
    active_tab="careertags"
    success_message="Cereer updated successfully."

class CareerTagsDeleteView(BaseDeleteView):
    model = CareerTags
    active_tab="careertags"
    success_url = reverse_lazy('topteenadminmanaged:careertagslist')
    success_message="Career deleted successfully."
    
class CareerTagsDetailView(BaseDetailView):
    template_name = "topteenadmin/careertags_details.html"
    model=CareerTags
    title="CareerTags"
    form_class=CareerTagsForm
    success_url=reverse_lazy('topteenadminmanaged:careertagslist')
    active_tab="careertags"
    
class BlogListView(BaseListView):
    template_name = "topteenadmin/blog_list.html"
    title="Blog"
    active_tab="Blog"
    model =Blog
    filterset_class = BlogFilter
    
class BlogCreate(BaseCreateView):
    template_name = "topteenadmin/blog_form.html"
    model=Blog
    form_class=BlogModelForm
    success_url=reverse_lazy('topteenadminmanaged:bloglist')
    title="Blog"
    active_tab="blog"
    success_message="Blog created successfully."
    
class BlogUpdateView(BaseUpdateView):
    template_name = "topteenadmin/blog_form.html"
    model=Blog
    title="Blog"
    form_class=BlogModelForm
    success_url=reverse_lazy('topteenadminmanaged:bloglist')
    active_tab="blog"
    success_message="Blog updated successfully."

class BlogDeleteView(BaseDeleteView):
    model = Blog
    active_tab="blog"
    success_url = reverse_lazy('topteenadminmanaged:bloglist')
    success_message="Blog deleted successfully."
    
class BlogDetailView(BaseDetailView):
    template_name = "topteenadmin/blog_detail.html"
    model = Blog
    title="Blog"
    form_class=BlogModelForm
    success_url=reverse_lazy('topteenadminmanaged:bloglist')
    active_tab="blog"


class BlogCategoryListView(BaseListView):
    template_name = "topteenadmin/blog_category_list.html"
    title="Blog Category"
    active_tab="blogcategory"
    model = BlogCategory
    filterset_class = BlogCategoryFilter
    
class BlogCategoryCreate(BaseCreateView):
    template_name = "topteenadmin/blog_category_form.html"
    model=BlogCategory
    form_class=BlogCategoryModelForm
    success_url=reverse_lazy('topteenadminmanaged:blogcategorylist')
    title="Blog Category"
    active_tab="blogcategory"
    success_message="Blog Category created successfully."
    
class BlogCategoryUpdateView(BaseUpdateView):
    template_name = "topteenadmin/blog_category_form.html"
    model=BlogCategory
    title="Blog Category"
    form_class=BlogCategoryModelForm
    success_url=reverse_lazy('topteenadminmanaged:blogcategorylist')
    active_tab="blogcategory"
    success_message="BlogCategory updated successfully."

class BlogCategoryDeleteView(BaseDeleteView):
    model = BlogCategory
    active_tab="blogcategory"
    success_url = reverse_lazy('topteenadminmanaged:blogcategorylist')
    success_message="Blog Category deleted successfully."
    
class BlogCategoryDetailView(BaseDetailView):
    template_name = "topteenadmin/category_detail.html"
    model = BlogCategory
    title="Blog Category"
    form_class=BlogCategoryModelForm
    success_url=reverse_lazy('topteenadminmanaged:blogcategorylist')
    active_tab="blogcategory"



class BlogTagListView(BaseListView):
    template_name = "topteenadmin/blogtag_list.html"
    title="Blog Tag"
    active_tab="blogtag"
    model=BlogTag
    filterset_class = BlogTagFilter
    
class BlogTagCreate(BaseCreateView):
    template_name = "topteenadmin/blogtag_form.html"
    model=BlogTag
    form_class=TagModelForm
    success_url=reverse_lazy('topteenadminmanaged:blogtaglist')
    title="Blog Tag"
    active_tab="blogtag"
    success_message="Blog Tag created successfully."
    
class BlogTagUpdateView(BaseUpdateView):
    template_name = "topteenadmin/blogtag_form.html"
    model=BlogTag
    title="Blog Tag"
    form_class=TagModelForm
    success_url=reverse_lazy('topteenadminmanaged:blogtaglist')
    active_tab="blogtag"
    success_message="Blog Tag updated successfully."

class BlogTagDeleteView(BaseDeleteView):
    model=BlogTag
    active_tab="blogtag"
    success_url = reverse_lazy('topteenadminmanaged:blogtaglist')
    success_message="Blog Tag deleted successfully."
    
class BlogTagDetailView(BaseDetailView):
    template_name = "topteenadmin/blogtag_detail.html"
    model=BlogTag
    title="Blog Tag"
    form_class=TagModelForm
    success_url=reverse_lazy('topteenadminmanaged:blogtaglist')
    active_tab="blogtag"


class CareerClusterListView(BaseListView):
    template_name = "topteenadmin/career_cluster_list.html"
    title="CareerCluster"
    active_tab="careercluster"
    model = CareerCluster
    filterset_class = CareerClusterFilter
    
class CreateCareerCluster(BaseCreateView):
    template_name = "topteenadmin/career_cluster_form.html"
    model=CareerCluster
    form_class=CareerClusterModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerclusterlist')
    title="CareerCluster"
    active_tab="careercluster"
    success_message="CareerCluster created successfully."
    
class CareerClusterUpdateView(BaseUpdateView):
    template_name = "topteenadmin/career_cluster_form.html"
    model=CareerCluster
    title="CareerCluster"
    form_class=CareerClusterModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerclusterlist')
    active_tab="careercluster"
    success_message="CareerCluster updated successfully."

class CareerClusterDeleteView(BaseDeleteView):
    model = CareerCluster
    active_tab="careercluster"
    success_url = reverse_lazy('topteenadminmanaged:careerclusterlist')
    success_message="CareerCluster deleted successfully."
    
class CareerClusterDetailView(BaseDetailView):
    template_name = "topteenadmin/career_cluster_detail.html"
    model=CareerCluster
    title="CareerCluster"
    form_class=CareerClusterModelForm
    success_url=reverse_lazy('topteenadminmanaged:careerclusterlist')
    active_tab="careercluster"


class ReviewListView(BaseListView):
    template_name = "topteenadmin/review_list.html"
    title="Review"
    active_tab="review"
    model = Review
    filterset_class = ReviewFilter

class ReviewCreateView(BaseCreateView):
    template_name = "topteenadmin/review_form.html"
    model=Review
    form_class=ReviewModelForm
    success_url=reverse_lazy('topteenadminmanaged:reviewlist')
    title="Review"
    active_tab="review"
    success_message="review created successfully."

class ReviewUpdateView(BaseUpdateView):
    template_name = "topteenadmin/review_form.html"
    model=Review
    title="Review"
    form_class=ReviewModelForm
    success_url=reverse_lazy('topteenadminmanaged:reviewlist')
    active_tab="review"
    success_message="review updated successfully."

class ReviewDeleteView(BaseDeleteView):
    model = Review
    active_tab="review"
    success_url = reverse_lazy('topteenadminmanaged:reviewlist')
    success_message="review deleted successfully."    

class ReviewDetailView(BaseDetailView):
    template_name = "topteenadmin/review_detail.html"
    model=Review
    title="Review"
    form_class=ReviewModelForm
    success_url=reverse_lazy('topteenadminmanaged:reviewlist')
    active_tab="review"


class CommonFAQListView(BaseListView):
    template_name = "topteenadmin/commonfaq_list.html"
    title="CommonFAQ"
    active_tab="commonfaq"
    model = CommonFAQ
    filterset_class = CommonFAQFilter

class CommonFAQCreateView(BaseCreateView):
    template_name = "topteenadmin/commonfaq_form.html"
    model=CommonFAQ
    form_class=CommonFAQModelForm
    success_url=reverse_lazy('topteenadminmanaged:commonfaqlist')
    title="CommonFAQ"
    active_tab="commonfaq"
    success_message="CommonFAQ created successfully."

class CommonFAQUpdateView(BaseUpdateView):
    template_name = "topteenadmin/commonfaq_form.html"
    model=CommonFAQ
    title="CommonFAQ"
    form_class=CommonFAQModelForm
    success_url=reverse_lazy('topteenadminmanaged:commonfaqlist')
    active_tab="commonfaq"
    success_message="CommonFAQ updated successfully."

class CommonFAQDeleteView(BaseDeleteView):
    model = CommonFAQ
    active_tab="commonfaq"
    success_url = reverse_lazy('topteenadminmanaged:commonfaqlist')
    success_message="CommonFAQ deleted successfully."        

class CommonFAQDetailView(BaseDetailView):
    template_name = "topteenadmin/commonfaq_detail.html"
    model=CommonFAQ
    title="CommonFAQ"
    form_class=CommonFAQModelForm
    success_url=reverse_lazy('topteenadminmanaged:commonfaqlist')
    active_tab="commonfaq"    

class SkillLabCourseListView(BaseListView):
    template_name = "topteenadmin/skilllabcourse_list.html"
    title="SkillLabCourse"
    active_tab="skilllabcourse"
    model = SkillLabCourse
    filterset_class = SkillLabCourseFilter

class CreateSkillLabCourse(BaseCreateView):
    template_name = "topteenadmin/skilllabcourse_form.html"
    model=SkillLabCourse
    form_class=SkillLabCourseModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllabcourselist')
    title="SkillLabCourse"
    active_tab="skilllabcourse"
    success_message="SkillLabCourse created successfully."
    
class SkillLabCourseUpdateView(BaseUpdateView):
    template_name = "topteenadmin/skilllabcourse_form.html"
    model=SkillLabCourse
    title="SkillLabCourse"
    form_class=SkillLabCourseModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllabcourselist')
    active_tab="skilllabcourse"
    success_message="SkillLabCourse updated successfully."

class SkillLabCourseDeleteView(BaseDeleteView):
    model = SkillLabCourse
    active_tab="skilllabcourse"
    success_url = reverse_lazy('topteenadminmanaged:skilllabcourselist')
    success_message="SkillLabCourse deleted successfully."
    
class SkillLabCourseDetailView(BaseDetailView):
    template_name = "topteenadmin/skilllabcourse_detail.html"
    model=SkillLabCourse
    title="SkillLabCourse"
    form_class=SkillLabCourseModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllabcourselist')
    active_tab="skilllabcourse"




class SkillLabCourseChapterListView(BaseListView):
    template_name = "topteenadmin/skilllabcoursechapter_list.html"
    title="SkillLabCourseChapter"
    active_tab="skilllabcoursechapter"
    model = SkillLabCourseChapter
    filterset_class = SkillLabCourseChapterFilter

class CreateSkillLabCourseChapter(BaseCreateView):
    template_name = "topteenadmin/skilllabcoursechapter_form.html"
    model=SkillLabCourseChapter
    form_class=SkillLabCourseChapterModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllabcoursechapterlist')
    title="SkillLabCourseChapter"
    active_tab="skilllabcoursechapter"
    success_message="SkillLabCourse Chapter created successfully."
    
class SkillLabCourseChapterUpdateView(BaseUpdateView):
    template_name = "topteenadmin/skilllabcoursechapter_form.html"
    model=SkillLabCourseChapter
    title="SkillLabCourse Chapter"
    form_class=SkillLabCourseChapterModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllabcoursechapterlist')
    active_tab="skilllabcoursechapter"
    success_message="SkillLabCourse Chapter updated successfully."

class SkillLabCourseChapterDeleteView(BaseDeleteView):
    model = SkillLabCourseChapter
    active_tab="skilllabcoursechapter"
    success_url = reverse_lazy('topteenadminmanaged:skilllabcoursechapterlist')
    success_message="SkillLabCourse chapter deleted successfully."
    
class SkillLabCourseChapterDetailView(BaseDetailView):
    template_name = "topteenadmin/skilllabcoursechapter_detail.html"
    model=SkillLabCourseChapter
    title="SkillLabCourse Chapter"
    form_class=SkillLabCourseChapterModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllabcoursechapterlist')
    active_tab="skilllabcoursechapter"


class SkillLabCourseActivityListView(BaseListView):
    template_name = "topteenadmin/skilllabcourseactivity_list.html"
    title="SkillLabCourseActivity"
    active_tab="skilllabcourseactivity"
    model = SkillLabCourseActivity
    filterset_class = SkillLabCourseActivityFilter

class CreateSkillLabCourseActivity(BaseCreateView):
    template_name = "topteenadmin/skilllabcourseactivity_form.html"
    model=SkillLabCourseActivity
    form_class=SkillLabCourseActivityModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllabcourseactivitylist')
    title="SkillLabCourse"
    active_tab="skilllcourseabactivity"
    success_message="SkillLabCourse created successfully."
    
class SkillLabCourseActivityUpdateView(BaseUpdateView):
    template_name = "topteenadmin/skilllabcourseactivity_form.html"
    model=SkillLabCourseActivity
    title="SkillLabCourse"
    form_class=SkillLabCourseActivityModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllabcourseactivitylist')
    active_tab="skilllabcourseactivity"
    success_message="SkillLabCourse updated successfully."

class SkillLabCourseActivityDeleteView(BaseDeleteView):
    model = SkillLabCourseActivity
    active_tab="skilllabcourseactivity"
    success_url = reverse_lazy('topteenadminmanaged:skilllabcourseactivitylist')
    success_message="SkillLabCourse Activity deleted successfully."
    
class SkillLabCourseActivityDetailView(BaseDetailView):
    template_name = "topteenadmin/skilllabcourseactivity_detail.html"
    model=SkillLabCourseActivity
    title="SkillLabCourseActivity"
    form_class=SkillLabCourseActivityModelForm
    success_url=reverse_lazy('topteenadminmanaged:skilllabcourseactivitylist')
    active_tab="skilllabcourseactivity"

class CategoryListView(BaseListView):
    template_name = "topteenadmin/video_category_list.html"
    title="Video Category"
    active_tab="videocategory"
    model = VideoCategory
    filterset_class = VideoCategoryFilter
    
class CategoryCreate(BaseCreateView):
    template_name = "topteenadmin/video_category_form.html"
    model=VideoCategory
    form_class=VideoCategoryForm
    success_url=reverse_lazy('topteenadminmanaged:videocategorylist')
    title="Video Category"
    active_tab="videocategory"
    success_message="VideoCategory created successfully."
    
class CategoryUpdateView(BaseUpdateView):
    template_name = "topteenadmin/video_category_form.html"
    model=VideoCategory
    title="Video Category"
    form_class=VideoCategoryForm
    success_url=reverse_lazy('topteenadminmanaged:videocategorylist')
    active_tab="video category"
    success_message="VideoCategory updated successfully."

class CategoryDeleteView(BaseDeleteView):
    model = VideoCategory
    active_tab="videocategory"
    success_url = reverse_lazy('topteenadminmanaged:videocategorylist')
    success_message="VideoCategory deleted successfully."
    
class CategoryDetailView(BaseDetailView):
    template_name = "topteenadmin/video_category_detail.html"
    model = VideoCategory
    title="Video Category"
    form_class=VideoCategoryForm
    success_url=reverse_lazy('topteenadminmanaged:videocategorylist')
    active_tab="videocategory"


class VideosListView(BaseListView):
    template_name = "topteenadmin/videos_list.html"
    title="Videos"
    active_tab="videos"
    model = Videos
    filterset_class = VideosFilter
    
class VideosCreate(BaseCreateView):
    template_name = "topteenadmin/videos_form.html"
    model=Videos
    form_class=VideosForm
    success_url=reverse_lazy('topteenadminmanaged:videoslist')
    title="Videos"
    active_tab="videos"
    success_message="Videos created successfully."
    
class VideosUpdateView(BaseUpdateView):
    template_name = "topteenadmin/videos_form.html"
    model=Videos
    title="Videos"
    form_class=VideosForm
    success_url=reverse_lazy('topteenadminmanaged:videoslist')
    active_tab="videos"
    success_message="Videos updated successfully."

class VideosDeleteView(BaseDeleteView):
    model = Videos
    active_tab="videos"
    success_url = reverse_lazy('topteenadminmanaged:videoslist')
    success_message="Videos deleted successfully."
    
class VideosDetailView(BaseDetailView):
    template_name = "topteenadmin/videos_detail.html"
    model = Videos
    title="Videos"
    form_class=VideosForm
    success_url=reverse_lazy('topteenadminmanaged:videoslist')
    active_tab="videos"

class HobbiesListView(BaseListView):
    template_name = "topteenadmin/hobbies_list.html"
    title="Hobbies"
    active_tab="Hobbies"
    model = Hobbies
    filterset_class = HobbiesFilter
    
class HobbiesCreate(BaseCreateView):
    template_name = "topteenadmin/hobbies_form.html"
    model=Hobbies
    form_class=HobbiesModelForm
    success_url=reverse_lazy('topteenadminmanaged:hobbieslist')
    title="Hobbies"
    active_tab="hobbies"
    success_message="hobbies created successfully."
    
class HobbiesUpdateView(BaseUpdateView):
    template_name = "topteenadmin/hobbies_form.html"
    model=Hobbies
    title="Hobbies"
    form_class=HobbiesModelForm
    success_url=reverse_lazy('topteenadminmanaged:hobbieslist')
    active_tab="hobbies"
    success_message="Hobbies updated successfully."

class HobbiesDeleteView(BaseDeleteView):
    model = Hobbies
    active_tab="hobbies"
    success_url = reverse_lazy('topteenadminmanaged:hobbieslist')
    success_message="hobbies deleted successfully."
    
class HobbiesDetailView(BaseDetailView):
    template_name = "topteenadmin/hobbies_detail.html"
    model = Hobbies
    title="Hobbies"
    form_class=HobbiesModelForm
    success_url=reverse_lazy('topteenadminmanaged:hobbieslist')
    active_tab="hobbies"

class SubjectListView(BaseListView):
    template_name = "topteenadmin/subject_list.html"
    title="Subject"
    active_tab="subject"
    model = Subject
    filterset_class = SubjectFilter
    
class SubjectCreate(BaseCreateView):
    template_name = "topteenadmin/subject_form.html"
    model=Subject
    form_class=SubjectModelForm
    success_url=reverse_lazy('topteenadminmanaged:subjectlist')
    title="Subject"
    active_tab="subject"
    success_message="Subject created successfully."
    
class SubjectUpdateView(BaseUpdateView):
    template_name = "topteenadmin/subject_form.html"
    model=Subject
    title="Subject"
    form_class=SubjectModelForm
    success_url=reverse_lazy('topteenadminmanaged:subjectlist')
    active_tab="subject"
    success_message="Subject updated successfully."

class SubjectDeleteView(BaseDeleteView):
    model = Subject
    active_tab="subject"
    success_url = reverse_lazy('topteenadminmanaged:subjectlist')
    success_message="Subject deleted successfully."
    
class SubjectDetailView(BaseDetailView):
    template_name = "topteenadmin/subject_detail.html"
    model = Subject
    title="Subject"
    form_class=SubjectModelForm
    success_url=reverse_lazy('topteenadminmanaged:subjectlist')
    active_tab="subject"

class UserFigureOutListView(BaseListView):
    template_name = "topteenadmin/userfigureout_list.html"
    title="Userfigureout"
    active_tab="userfigureout"
    model = UserFigureOut
    filterset_class = UserFigureOutFilter
    
class UserFigureOutCreate(BaseCreateView):
    template_name = "topteenadmin/userfigureout_form.html"
    model=UserFigureOut
    form_class=UserFigureOutModelForm
    success_url=reverse_lazy('topteenadminmanaged:userfigureoutlist')
    title="Userfigureout"
    active_tab="userfigureout"
    success_message="User figure out created successfully."
    
class UserFigureOutUpdateView(BaseUpdateView):
    template_name = "topteenadmin/userfigureout_form.html"
    model=UserFigureOut
    title="Userfigureout"
    form_class=UserFigureOutModelForm
    success_url=reverse_lazy('topteenadminmanaged:userfigureoutlist')
    active_tab="userfigureout"
    success_message="User figure out updated successfully."

class UserFigureOutDeleteView(BaseDeleteView):
    model = UserFigureOut
    active_tab="userfigureout"
    success_url = reverse_lazy('topteenadminmanaged:userfigureoutlist')
    success_message="User figure out deleted successfully."
    
class UserFigureOutDetailView(BaseDetailView):
    template_name = "topteenadmin/userfigureout_detail.html"
    model = UserFigureOut
    title="UserFigureOut"
    form_class=UserFigureOutModelForm
    success_url=reverse_lazy('topteenadminmanaged:userfigureoutlist')
    active_tab="userfigureout"

class StoriesListView(BaseListView):
    template_name = "topteenadmin/stories_list.html"
    title="Stories"
    active_tab="stories"
    model = Stories
    filterset_class = StoriesFilter
    
class StoriesCreate(BaseCreateView):
    template_name = "topteenadmin/stories_form.html"
    model=Stories
    form_class=StoriesModelForm
    success_url=reverse_lazy('topteenadminmanaged:storieslist')
    title="Stories"
    active_tab="stories"
    success_message="Stories created successfully."
    
class StoriesUpdateView(BaseUpdateView):
    template_name = "topteenadmin/stories_form.html"
    model=Stories
    title="Stories"
    form_class=StoriesModelForm
    success_url=reverse_lazy('topteenadminmanaged:storieslist')
    active_tab="stories"
    success_message="Stories updated successfully."

class StoriesDeleteView(BaseDeleteView):
    model = Stories
    active_tab="stories"
    success_url = reverse_lazy('topteenadminmanaged:storieslist')
    success_message="Stories deleted successfully."
    
class StoriesDetailView(BaseDetailView):
    template_name = "topteenadmin/stories_detail.html"
    model = Stories
    title="Stories"
    form_class=StoriesModelForm
    success_url=reverse_lazy('topteenadminmanaged:storieslist')
    active_tab="stories"
    
class APILogListView(BaseListView):
    template_name = "topteenadmin/apilog_list.html"
    title="Apilog"
    active_tab="apilog"
    model = APILog
    filterset_class = ApilogFilter

class APILogDetailView(BaseDetailView):
    template_name = "topteenadmin/apilog_detail.html"
    model = APILog
    title="Apilog"
    form_class=ApilogModelForm
    success_url=reverse_lazy('topteenadminmanaged:apiloglist')
    active_tab="apilog"
    
class LeadListView(BaseListView):
    template_name = "topteenadmin/lead_list.html"
    title="Lead"
    active_tab="lead"
    model = Lead
    filterset_class = LeadFilter
    
class LeadDetailView(BaseDetailView):
    template_name = "topteenadmin/lead_detail.html"
    model = Lead
    title="Lead"
    form_class=LeadModelForm
    success_url=reverse_lazy('topteenadminmanaged:leadlist')
    active_tab="lead"
    
class PsychometricFAQListView(BaseListView):
    template_name = "topteenadmin/psychometricfaq_list.html"
    title="PsychometricFaq"
    active_tab="psychometric"
    model = PsychometricFAQ
    filterset_class = PsychometricFaqFilter
    
class PsychometricFAQCreate(BaseCreateView):
    template_name = "topteenadmin/psychometricfaq_form.html"
    model=PsychometricFAQ
    form_class=PsychometricFaqModelForm
    success_url=reverse_lazy('topteenadminmanaged:psychometricfaqlist')
    title="PsychometricFaq"
    active_tab="psychometric"
    success_message="PsychometricFaq created successfully."
    
class PsychometricUpdateView(BaseUpdateView):
    template_name = "topteenadmin/psychometricfaq_form.html"
    model=PsychometricFAQ
    form_class=PsychometricFaqModelForm
    success_url=reverse_lazy('topteenadminmanaged:psychometricfaqlist')
    title="PsychometricFaq"
    active_tab="psychometric"
    success_message="PsychometricFaq created successfully."

class PsychometricDeleteView(BaseDeleteView):
    model = PsychometricFAQ
    active_tab="pyschometric"
    success_url = reverse_lazy('topteenadminmanaged:psychometricfaqlist')
    success_message="Psychometric deleted successfully."
    
class PsychometricDetailView(BaseDetailView):
    template_name = "topteenadmin/psychomtric_detail.html"
    model = PsychometricFAQ
    title="PyschometricFaq"
    form_class=PsychometricFaqModelForm
    success_url=reverse_lazy('topteenadminmanaged:psychometricfaqlist')
    active_tab="psychometric"


@method_decorator(login_required, name='dispatch')
class StudentListView(TemplateView):
    template_name = "topteenadmin/student_list.html"
    
    def get_context_data(self, **kwargs):
        """Return minimal initial data - heavy processing moved to AJAX endpoints"""
        context = super().get_context_data(**kwargs)
        context['active_tab'] = 'students'
        context['title'] = 'Students'
        context['meta_title'] = 'Students'
        
        # Return minimal context - statistics and school data loaded via AJAX
        context.update({
            'breadcrumb': build_admin_breadcrumb([{'title': 'Students', 'text': 'Students', 'url': reverse('topteenadminmanaged:studentlist')}]),
        })
        
        return context


@method_decorator(login_required, name='dispatch')
class StudentListStatsAPIView(View):
    """API endpoint to return student statistics via AJAX - optimized with bulk queries"""
    
    def get(self, request):
        # Load all_tests once
        all_tests = Test.objects.filter(is_active=True)
        total_tests_count = all_tests.count()
        
        # Bulk fetch TestSession exists for all users
        user_ids = list(User.objects.values_list('id', flat=True))
        
        # Prefetch all related data in bulk
        all_users = User.objects.select_related('user_profile').prefetch_related(
            Prefetch('student_management', queryset=StudentManagement.objects.select_related('class_and_section')),
            Prefetch('test_sessions', queryset=TestSession.objects.filter(is_completed=True)),
            Prefetch('testcompletion_set', queryset=TestCompletion.objects.all())
        )
        users_with_test_sessions = set(
            TestSession.objects.filter(user_id__in=user_ids).values_list('user_id', flat=True).distinct()
        )
        users_with_test_completion = set(
            TestCompletion.objects.filter(user_id__in=user_ids).values_list('user_id', flat=True).distinct()
        )
        
        # Bulk fetch completion counts for Class 12
        completed_sessions_count = {}
        if user_ids:
            completed_counts = TestSession.objects.filter(
                user_id__in=user_ids,
                is_completed=True
            ).values('user_id').annotate(count=Count('id'))
            completed_sessions_count = {item['user_id']: item['count'] for item in completed_counts}
        
        # Bulk fetch TestCompletion data
        test_completion_map = {}
        for tc in TestCompletion.objects.filter(user_id__in=user_ids).select_related('user'):
            test_completion_map[tc.user_id] = tc
        
        # Classify users using bulk data
        class12_user_ids = []
        class10_user_ids = []
        
        # Build student management map
        student_management_map = {}
        for sm in StudentManagement.objects.filter(student_id__in=user_ids).select_related('class_and_section'):
            student_management_map[sm.student_id] = sm
        
        for user_id in user_ids:
            # Check StudentManagement first
            student_class = None
            if user_id in student_management_map:
                sm = student_management_map[user_id]
                if sm.class_and_section:
                    class_name = sm.class_and_section.class_and_section
                    if class_name:
                        class_prefix = class_name[:2].strip()
                        if class_prefix in ["11", "12"]:
                            student_class = "class12"
                        elif class_prefix == "10":
                            student_class = "class10"
            
            # Fallback to test sessions/completion
            if not student_class:
                if user_id in users_with_test_sessions:
                    student_class = "class12"
                elif user_id in users_with_test_completion:
                    student_class = "class10"
            
            if student_class == "class12":
                class12_user_ids.append(user_id)
            elif student_class == "class10":
                class10_user_ids.append(user_id)
        
        # Calculate Class 12 statistics using bulk queries
        class12_total = len(class12_user_ids)
        class12_completed_count = 0
        class12_pending_count = 0
        
        if class12_user_ids and total_tests_count > 0:
            for user_id in class12_user_ids:
                completed_count = completed_sessions_count.get(user_id, 0)
                if completed_count == total_tests_count:
                    class12_completed_count += 1
                elif user_id in users_with_test_sessions:
                    class12_pending_count += 1
        
        class12_not_started_count = class12_total - class12_completed_count - class12_pending_count
        
        # Calculate Class 10 statistics using bulk queries
        class10_total = len(class10_user_ids)
        class10_completed_count = 0
        class10_pending_count = 0
        
        if class10_user_ids:
            for user_id in class10_user_ids:
                if user_id in test_completion_map:
                    tc = test_completion_map[user_id]
                    if tc.are_all_primary_tests_completed():
                        class10_completed_count += 1
                    else:
                        class10_pending_count += 1
        
        class10_not_started_count = class10_total - class10_completed_count - class10_pending_count
        
        return JsonResponse({
            'class12_total': class12_total,
            'class12_completed': class12_completed_count,
            'class12_pending': class12_pending_count,
            'class12_not_started': class12_not_started_count,
            'class10_total': class10_total,
            'class10_completed': class10_completed_count,
            'class10_pending': class10_pending_count,
            'class10_not_started': class10_not_started_count,
        })


@method_decorator(login_required, name='dispatch')
class StudentListSchoolsAPIView(View):
    """API endpoint to return school-wise student data via AJAX - optimized with bulk queries"""
    
    def get(self, request):
        # Load all_tests once
        all_tests = Test.objects.filter(is_active=True)
        total_tests_count = all_tests.count()
        
        # Bulk fetch user IDs first
        user_ids = list(User.objects.values_list('id', flat=True))
        
        # Prefetch all related data in bulk
        all_users = User.objects.select_related('user_profile').prefetch_related(
            Prefetch('student_management', queryset=StudentManagement.objects.select_related('class_and_section')),
            Prefetch('test_sessions', queryset=TestSession.objects.all()),
            Prefetch('testcompletion_set', queryset=TestCompletion.objects.all())
        )
        
        # Bulk fetch all data
        users_with_test_sessions = set(
            TestSession.objects.filter(user_id__in=user_ids).values_list('user_id', flat=True).distinct()
        )
        users_with_test_completion = set(
            TestCompletion.objects.filter(user_id__in=user_ids).values_list('user_id', flat=True).distinct()
        )
        
        # Bulk fetch completion counts
        completed_sessions_count = {}
        if user_ids:
            completed_counts = TestSession.objects.filter(
                user_id__in=user_ids,
                is_completed=True
            ).values('user_id').annotate(count=Count('id'))
            completed_sessions_count = {item['user_id']: item['count'] for item in completed_counts}
        
        # Bulk fetch TestCompletion
        test_completion_map = {}
        for tc in TestCompletion.objects.filter(user_id__in=user_ids).select_related('user'):
            test_completion_map[tc.user_id] = tc
        
        # Build student management map
        student_management_map = {}
        for sm in StudentManagement.objects.filter(student_id__in=user_ids).select_related('class_and_section'):
            student_management_map[sm.student_id] = sm
        
        # Build user profile map for school names
        user_profile_map = {}
        for up in UserProfile.objects.filter(user_id__in=user_ids):
            user_profile_map[up.user_id] = up.schoolname or "Unknown"
        
        # Classify and group by school
        class12_schools = {}
        class10_schools = {}
        
        for user in all_users:
            user_id = user.id
            school_name = user_profile_map.get(user_id, "Unknown")
            
            # Determine student class
            student_class = None
            if user_id in student_management_map:
                sm = student_management_map[user_id]
                if sm.class_and_section:
                    class_name = sm.class_and_section.class_and_section
                    if class_name:
                        class_prefix = class_name[:2].strip()
                        if class_prefix in ["11", "12"]:
                            student_class = "class12"
                        elif class_prefix == "10":
                            student_class = "class10"
            
            if not student_class:
                if user_id in users_with_test_sessions:
                    student_class = "class12"
                elif user_id in users_with_test_completion:
                    student_class = "class10"
            
            if student_class == "class12":
                if school_name not in class12_schools:
                    class12_schools[school_name] = {
                        'name': school_name,
                        'total_students': 0,
                        'completed': 0,
                        'pending': 0,
                        'not_started': 0
                    }
                class12_schools[school_name]['total_students'] += 1
                
                # Check completion status
                completed_count = completed_sessions_count.get(user_id, 0)
                if completed_count == total_tests_count and total_tests_count > 0:
                    class12_schools[school_name]['completed'] += 1
                elif user_id in users_with_test_sessions:
                    class12_schools[school_name]['pending'] += 1
                else:
                    class12_schools[school_name]['not_started'] += 1
                    
            elif student_class == "class10":
                if school_name not in class10_schools:
                    class10_schools[school_name] = {
                        'name': school_name,
                        'total_students': 0,
                        'completed': 0,
                        'pending': 0,
                        'not_started': 0
                    }
                class10_schools[school_name]['total_students'] += 1
                
                # Check completion status
                if user_id in test_completion_map:
                    tc = test_completion_map[user_id]
                    if tc.are_all_primary_tests_completed():
                        class10_schools[school_name]['completed'] += 1
                    else:
                        class10_schools[school_name]['pending'] += 1
                else:
                    class10_schools[school_name]['not_started'] += 1
        
        # Sort schools by name
        class12_schools_list = sorted(class12_schools.values(), key=lambda x: x['name'])
        class10_schools_list = sorted(class10_schools.values(), key=lambda x: x['name'])
        
        return JsonResponse({
            'class12_schools': class12_schools_list,
            'class10_schools': class10_schools_list,
        })


@method_decorator(login_required, name='dispatch')
class StudentTestHistoryView(TemplateView):
    template_name = "topteenadmin/student_test_history.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_id = kwargs.get('user_id')
        
        try:
            student = User.objects.get(id=user_id)
        except User.DoesNotExist:
            context['error'] = 'Student not found'
            return context
        
        context['active_tab'] = 'students'
        context['title'] = f'{student.name} - Test History'
        context['meta_title'] = f'{student.name} - Test History'
        
        # Get all test sessions for this student
        test_sessions = TestSession.objects.filter(user=student).select_related('test').order_by('-created_at')
        
        # Get all tests to show which ones haven't been started
        all_tests = Test.objects.filter(is_active=True).order_by('title')
        started_test_ids = set(test_sessions.values_list('test_id', flat=True))
        
        # Create comprehensive test data
        test_data = []
        for test in all_tests:
            session = test_sessions.filter(test=test).first()
            if session:
                # Get test result if available (OneToOne relationship)
                try:
                    result = session.result
                    score = result.score
                    grade = result.grade
                    result_data = result.result_data
                except AttributeError:
                    score = None
                    grade = None
                    result_data = None
                
                # Calculate time taken
                time_taken = None
                if session.start_time and session.end_time:
                    delta = session.end_time - session.start_time
                    time_taken = delta.total_seconds()
                
                test_data.append({
                    'test': test,
                    'is_started': True,
                    'is_completed': session.is_completed,
                    'start_time': session.start_time,
                    'end_time': session.end_time,
                    'time_taken': time_taken,
                    'attempt_count': session.attempt_count,
                    'score': score,
                    'grade': grade,
                    'result_data': result_data,
                })
            else:
                test_data.append({
                    'test': test,
                    'is_started': False,
                    'is_completed': False,
                    'not_started': True,
                    'start_time': None,
                    'end_time': None,
                    'time_taken': None,
                    'attempt_count': 0,
                    'score': None,
                    'grade': None,
                })
        
        # Get statistics
        total_tests = all_tests.count()
        completed_tests = sum(1 for data in test_data if data.get('is_completed'))
        pending_tests = sum(1 for data in test_data if data.get('is_started') and not data.get('is_completed'))
        not_started_tests = sum(1 for data in test_data if data.get('not_started'))
        
        context.update({
            'student': student,
            'test_data': test_data,
            'total_tests': total_tests,
            'completed_tests': completed_tests,
            'pending_tests': pending_tests,
            'not_started_tests': not_started_tests,
            'breadcrumb': build_admin_breadcrumb([
                {'title': 'Students', 'text': 'Students', 'url': reverse('topteenadminmanaged:studentlist')},
                {'title': student.name, 'text': student.name, 'url': '#'}
            ]),
        })
        
        return context


def _staff_required(view_func):
    """Require login and staff/superuser for admin-only actions."""
    def wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect(settings.LOGIN_URL)
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'You do not have permission to perform this action.')
            return redirect('topteenadmin:topteendashboard')
        return view_func(request, *args, **kwargs)
    return wrapped


@method_decorator(login_required, name='dispatch')
class AdminResetUserPasswordView(TemplateView):
    """Allow staff/superuser to set a new password for any user (no current password required)."""
    template_name = 'topteenadmin/admin_reset_password.html'

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'You do not have permission to reset user passwords.')
            return redirect('topteenadmin:topteendashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_id = kwargs.get('user_id')
        try:
            target_user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            context['error'] = 'User not found'
            return context
        context['target_user'] = target_user
        context['active_tab'] = 'students'
        context['breadcrumb'] = build_admin_breadcrumb([
            {'title': 'Students', 'text': 'Students', 'url': reverse('topteenadminmanaged:studentlist')},
            {'title': 'User list', 'text': 'User list', 'url': reverse('topteenadminmanaged:userlist')},
            {'title': 'Reset password', 'text': 'Reset password', 'url': '#'},
        ])
        return context

    def post(self, request, *args, **kwargs):
        user_id = kwargs.get('user_id')
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'Permission denied.')
            return redirect('topteenadmin:topteendashboard')
        try:
            target_user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            messages.error(request, 'User not found.')
            return redirect(reverse('topteenadminmanaged:userlist'))
        password1 = request.POST.get('new_password1', '').strip()
        password2 = request.POST.get('new_password2', '').strip()
        if not password1:
            messages.error(request, 'Please enter a new password.')
            return redirect(reverse('topteenadminmanaged:admin_reset_user_password', kwargs={'user_id': user_id}))
        if password1 != password2:
            messages.error(request, 'The two password fields did not match.')
            return redirect(reverse('topteenadminmanaged:admin_reset_user_password', kwargs={'user_id': user_id}))
        if len(password1) < 8:
            messages.error(request, 'Password must be at least 8 characters.')
            return redirect(reverse('topteenadminmanaged:admin_reset_user_password', kwargs={'user_id': user_id}))
        target_user.set_password(password1)
        target_user.save()
        messages.success(request, f'Password for {getattr(target_user, "name", None) or target_user.email} has been reset successfully.')
        next_url = request.POST.get('next') or request.GET.get('next')
        if next_url:
            return redirect(next_url)
        return redirect(reverse('topteenadminmanaged:userlist'))


@method_decorator(login_required, name='dispatch')
class UserListView(TemplateView):
    """List all users with actions: Test history, Reset password. Staff/superuser only."""
    template_name = 'topteenadmin/user_list.html'

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'You do not have permission to view the user list.')
            return redirect('topteenadmin:topteendashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        users = User.objects.select_related('user_profile').order_by('-created')
        context['user_list'] = users
        context['active_tab'] = 'students'
        context['title'] = 'User list'
        context['meta_title'] = 'User list'
        context['breadcrumb'] = build_admin_breadcrumb([
            {'title': 'Students', 'text': 'Students', 'url': reverse('topteenadminmanaged:studentlist')},
            {'title': 'User list', 'text': 'User list', 'url': '#'},
        ])
        return context


@method_decorator(login_required, name='dispatch')
class ExportAllClass12ResultsView(View):
    """
    Export comprehensive student results to Excel for ALL Class 12 students
    Following the exact format from EXPORT_FUNCTIONALITY_DOCUMENTATION.md
    """
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import json
        from datetime import datetime, timezone
        
        # Get all Class 12 students
        class12_users = []
        all_users = User.objects.select_related('user_profile').prefetch_related(
            'test_sessions', 'test_sessions__test'
        )
        
        for user in all_users:
            student_class = self.get_student_class(user)
            if student_class == "class12":
                class12_users.append(user)
        
        # Get tests by title (matching documentation format)
        personality_test = Test.objects.filter(title='Personality Assessment').first()
        motivation_test = Test.objects.filter(title='Motivation Assessment').first()
        interest_test = Test.objects.filter(title='Career Interest Inventory').first()
        aptitude_test = Test.objects.filter(title='Aptitude Assessment').first()
        
        # Get questions for each test
        personality_questions = Question.objects.filter(test=personality_test).order_by('order') if personality_test else []
        motivation_questions = Question.objects.filter(test=motivation_test).order_by('order') if motivation_test else []
        interest_questions = Question.objects.filter(test=interest_test).order_by('order') if interest_test else []
        
        # Get aptitude sections and questions
        aptitude_sections = []
        if aptitude_test:
            sections = Sections.objects.filter(test=aptitude_test).order_by('order')
            for section in sections:
                questions = Question.objects.filter(section=section).order_by('order')
                aptitude_sections.append({
                    'section': section,
                    'questions': questions
                })
        
        # Create headers following documentation format
        headers = [
            'Name', 'Gender', 'GRADE', 'Test Completion Status', 'Stream',
            'Personality Time Taken', 'Personality Time (Minutes)'
        ]
        
        # Add personality question headers
        for q in personality_questions:
            headers.append(f'Personality Q{q.order}')
        
        # Add motivation headers
        headers.extend(['Motivation Time', 'Motivation Time (Minutes)'])
        for q in motivation_questions:
            headers.append(f'Motivation Q{q.order}')
        
        # Add interest headers
        headers.extend(['Interest Time', 'Interest Time (Minutes)'])
        for q in interest_questions:
            headers.append(f'Interest Q{q.order}')
        
        # Add aptitude headers
        for section_data in aptitude_sections:
            section = section_data['section']
            questions = section_data['questions']
            headers.append(f'{section.title} Time')
            headers.append(f'{section.title} Time (Minutes)')
            headers.append(f'{section.title} Start Time')
            headers.append(f'{section.title} End Time')
            headers.append(f'{section.title} Manually Fixed')
            for q in questions:
                headers.append(f'{section.title} Q{q.order}')
        
        # Add aptitude test result headers
        headers.extend([
            'Aptitude Above Average',
            'Aptitude Average', 
            'Aptitude Below Average'
        ])
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Student Results"
        
        # Write headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Calculate statistics for verification (same as display)
        export_total = len(class12_users)
        export_completed = 0
        export_pending = 0
        export_not_started = 0
        
        # Process each Class 12 student
        row = 2
        for user in class12_users:
            # Basic user info
            name = user.name or user.username or "N/A"
            
            # Get student class for this user (already verified as class12)
            student_class = self.get_student_class(user)
            
            # Get Gender, Grade (Class), Stream from UserProfile and StudentManagement
            gender = "Unknown"
            grade = ""  # Will contain student class (e.g., "Class 12", "Class 11")
            test_completion_status = "Not Started"
            stream = ""
            
            try:
                if hasattr(user, 'user_profile') and user.user_profile:
                    if user.user_profile.gender:
                        gender_choices = dict(user.user_profile._meta.get_field('gender').choices)
                        gender = gender_choices.get(user.user_profile.gender, "Unknown")
                
                # Get student class from StudentManagement
                try:
                    student_management = StudentManagement.objects.filter(student=user).first()
                    if student_management and student_management.class_and_section:
                        class_name = student_management.class_and_section.class_and_section
                        if class_name:
                            grade = class_name  # e.g., "12-A", "11-B", "Class 12"
                        stream = student_management.class_and_section.stream or ""
                    else:
                        # Fallback: try to get from UserProfile
                        if hasattr(user, 'user_profile') and user.user_profile:
                            grade = user.user_profile.grade or ""
                    
                    # If still no grade, try to infer from test sessions (user is Class 12)
                    if not grade:
                        # Check if we can infer class from student_class determination
                        if student_class == "class12":
                            grade = "Class 12"  # Default fallback
                except:
                    pass
            except:
                pass
            
            # Get user's test sessions and calculate completion status (EXACT SAME LOGIC AS DISPLAY)
            test_sessions = TestSession.objects.filter(user=user)
            all_tests = Test.objects.filter(is_active=True)
            completed_count = test_sessions.filter(is_completed=True).count()
            total_tests_count = all_tests.count()
            
            # Use EXACT same logic as StudentListView for consistency
            if completed_count == total_tests_count and total_tests_count > 0:
                test_completion_status = "Completed"
                export_completed += 1
            elif test_sessions.exists():
                test_completion_status = f"Pending ({completed_count}/{total_tests_count})"
                export_pending += 1
            else:
                test_completion_status = "Not Started"
                export_not_started += 1
            
            # Initialize data row
            data_row = [name, gender, grade, test_completion_status, stream]
            
            # Personality Assessment data
            personality_session = test_sessions.filter(test=personality_test).first() if personality_test else None
            personality_time = ""
            personality_time_minutes = ""
            personality_answers = [""] * len(personality_questions)
            
            if personality_session:
                # Calculate time taken
                if personality_session.end_time and personality_session.start_time:
                    time_diff = personality_session.end_time - personality_session.start_time
                    if time_diff.total_seconds() >= 0:
                        personality_time = f"{int(time_diff.total_seconds())}s"
                        personality_time_minutes = f"{round(time_diff.total_seconds() / 60, 2)}"
                    else:
                        personality_time = "0s"
                        personality_time_minutes = "0"
                elif personality_session.start_time and not personality_session.end_time:
                    personality_time = "Incomplete"
                    personality_time_minutes = "Incomplete"
                else:
                    personality_time = "N/A"
                    personality_time_minutes = "N/A"
                
                # Get answers
                try:
                    user_response = UserResponse.objects.filter(session=personality_session).first()
                    if user_response and user_response.selected_answer:
                        submitted_answers = user_response.selected_answer.get('submitted_answers', {})
                        for i, q in enumerate(personality_questions):
                            answer_key = f"Question_{q.order}"
                            if answer_key in submitted_answers:
                                answer_data = submitted_answers[answer_key]
                                if isinstance(answer_data, dict):
                                    personality_answers[i] = answer_data.get('text', '')
                                else:
                                    personality_answers[i] = str(answer_data)
                except Exception as e:
                    print(f"Error processing personality answers for user {user.email}: {e}")
            
            data_row.append(personality_time)
            data_row.append(personality_time_minutes)
            data_row.extend(personality_answers)
            
            # Motivation Assessment data
            motivation_session = test_sessions.filter(test=motivation_test).first() if motivation_test else None
            motivation_time = ""
            motivation_time_minutes = ""
            motivation_answers = [""] * len(motivation_questions)
            
            if motivation_session:
                # Calculate time taken
                if motivation_session.end_time and motivation_session.start_time:
                    time_diff = motivation_session.end_time - motivation_session.start_time
                    if time_diff.total_seconds() >= 0:
                        motivation_time = f"{int(time_diff.total_seconds())}s"
                        motivation_time_minutes = f"{round(time_diff.total_seconds() / 60, 2)}"
                    else:
                        motivation_time = "0s"
                        motivation_time_minutes = "0"
                elif motivation_session.start_time and not motivation_session.end_time:
                    motivation_time = "Incomplete"
                    motivation_time_minutes = "Incomplete"
                else:
                    motivation_time = "N/A"
                    motivation_time_minutes = "N/A"
                
                # Get answers
                try:
                    user_response = UserResponse.objects.filter(session=motivation_session).first()
                    if user_response and user_response.selected_answer:
                        submitted_answers = user_response.selected_answer.get('submitted_answers', {})
                        for i, q in enumerate(motivation_questions):
                            answer_key = f"Question_{q.order}"
                            if answer_key in submitted_answers:
                                answer_data = submitted_answers[answer_key]
                                if isinstance(answer_data, dict):
                                    motivation_answers[i] = answer_data.get('text', '')
                                else:
                                    motivation_answers[i] = str(answer_data)
                except Exception as e:
                    print(f"Error processing motivation answers for user {user.email}: {e}")
            
            data_row.append(motivation_time)
            data_row.append(motivation_time_minutes)
            data_row.extend(motivation_answers)
            
            # Career Interest Inventory data
            interest_session = test_sessions.filter(test=interest_test).first() if interest_test else None
            interest_time = ""
            interest_time_minutes = ""
            interest_answers = [""] * len(interest_questions)
            
            if interest_session:
                # Calculate time taken
                if interest_session.end_time and interest_session.start_time:
                    time_diff = interest_session.end_time - interest_session.start_time
                    if time_diff.total_seconds() >= 0:
                        interest_time = f"{int(time_diff.total_seconds())}s"
                        interest_time_minutes = f"{round(time_diff.total_seconds() / 60, 2)}"
                    else:
                        interest_time = "0s"
                        interest_time_minutes = "0"
                elif interest_session.start_time and not interest_session.end_time:
                    interest_time = "Incomplete"
                    interest_time_minutes = "Incomplete"
                else:
                    interest_time = "N/A"
                    interest_time_minutes = "N/A"
                
                # Get answers
                try:
                    user_response = UserResponse.objects.filter(session=interest_session).first()
                    if user_response and user_response.selected_answer:
                        submitted_answers = user_response.selected_answer.get('submitted_answers', {})
                        for i, q in enumerate(interest_questions):
                            answer_key = f"Question_{q.order}"
                            if answer_key in submitted_answers:
                                answer_data = submitted_answers[answer_key]
                                if isinstance(answer_data, dict):
                                    interest_answers[i] = answer_data.get('text', '')
                                else:
                                    interest_answers[i] = str(answer_data)
                except Exception as e:
                    print(f"Error processing interest answers for user {user.email}: {e}")
            
            data_row.append(interest_time)
            data_row.append(interest_time_minutes)
            data_row.extend(interest_answers)
            
            # Aptitude Assessment data
            aptitude_session = test_sessions.filter(test=aptitude_test).first() if aptitude_test else None
            
            # Calculate alternative timing for aptitude tests (for tests taken before today)
            alternative_timing_data = {}
            if aptitude_session:
                try:
                    # Get all section sessions for this aptitude test, ordered by end time
                    all_section_sessions = SectionSession.objects.filter(
                        session=aptitude_session
                    ).order_by('end_time')
                    
                    # Check if this is a test taken before today (to apply alternative timing)
                    today = datetime.now(timezone.utc).date()
                    test_date = aptitude_session.start_time.date() if aptitude_session.start_time else None
                    
                    if test_date and test_date < today and all_section_sessions.exists():
                        # Apply alternative timing calculation
                        test_start_time = aptitude_session.start_time
                        
                        # Create timing data for each section
                        for i, section_session in enumerate(all_section_sessions):
                            section_name = section_session.section.title
                            
                            if i == 0:
                                # First section: use test session start time as start time
                                section_start = test_start_time
                            else:
                                # Subsequent sections: use previous section's end time as start time
                                prev_section_session = all_section_sessions[i-1]
                                section_start = prev_section_session.end_time
                            
                            section_end = section_session.end_time
                            
                            # Calculate duration
                            if section_start and section_end:
                                duration = section_end - section_start
                                alternative_timing_data[section_name] = {
                                    'start_time': section_start,
                                    'end_time': section_end,
                                    'duration': duration.total_seconds(),
                                    'duration_str': f"{int(duration.total_seconds())}s",
                                    'duration_minutes': f"{round(duration.total_seconds() / 60, 2)}",
                                    'manually_fixed': "Yes"
                                }
                            else:
                                alternative_timing_data[section_name] = {
                                    'start_time': section_start,
                                    'end_time': section_end,
                                    'duration': 0,
                                    'duration_str': "N/A",
                                    'duration_minutes': "N/A",
                                    'manually_fixed': "No"
                                }
                except Exception as e:
                    print(f"Error calculating alternative timing for user {user.email}: {e}")
            
            for section_data in aptitude_sections:
                section = section_data['section']
                questions = section_data['questions']
                section_time = ""
                section_time_minutes = ""
                section_start_time = ""
                section_end_time = ""
                section_manually_fixed = ""
                section_answers = [""] * len(questions)
                
                if aptitude_session:
                    # Get section session
                    try:
                        section_session = SectionSession.objects.filter(
                            session=aptitude_session, 
                            section=section
                        ).first()
                        
                        if section_session:
                            # Use alternative timing if available, otherwise use original logic
                            if section.title in alternative_timing_data:
                                timing_data = alternative_timing_data[section.title]
                                section_time = timing_data['duration_str']
                                section_time_minutes = timing_data['duration_minutes']
                                section_start_time = timing_data['start_time'].strftime('%Y-%m-%d %H:%M:%S') if timing_data['start_time'] else "N/A"
                                section_end_time = timing_data['end_time'].strftime('%Y-%m-%d %H:%M:%S') if timing_data['end_time'] else "N/A"
                                section_manually_fixed = timing_data['manually_fixed']
                            else:
                                # Original timing calculation (for newer tests)
                                if section_session.end_time and section_session.start_time:
                                    time_diff = section_session.end_time - section_session.start_time
                                    if time_diff.total_seconds() >= 0:
                                        section_time = f"{int(time_diff.total_seconds())}s"
                                        section_time_minutes = f"{round(time_diff.total_seconds() / 60, 2)}"
                                    else:
                                        section_time = "0s"
                                        section_time_minutes = "0"
                                elif section_session.start_time and not section_session.end_time:
                                    section_time = "Incomplete"
                                    section_time_minutes = "Incomplete"
                                else:
                                    section_time = "N/A"
                                    section_time_minutes = "N/A"
                                
                                section_start_time = section_session.start_time.strftime('%Y-%m-%d %H:%M:%S') if section_session.start_time else "N/A"
                                section_end_time = section_session.end_time.strftime('%Y-%m-%d %H:%M:%S') if section_session.end_time else "N/A"
                                section_manually_fixed = "No"
                            
                            # Get answers for this section
                            user_response = UserResponse.objects.filter(
                                session=aptitude_session,
                                session_section=section_session
                            ).first()
                            
                            if not user_response:
                                user_response = UserResponse.objects.filter(
                                    session=aptitude_session
                                ).first()
                            
                            if user_response and user_response.selected_answer:
                                response_data = user_response.selected_answer
                                
                                # Handle aptitude test response format (all sections in one response)
                                if 'sections' in response_data and section.title in response_data['sections']:
                                    section_data_resp = response_data['sections'][section.title]
                                    submitted_answers = section_data_resp.get('submitted_answers', {})
                                    
                                    for i, q in enumerate(questions):
                                        answer_key = f"Question_{q.order}"
                                        if answer_key in submitted_answers:
                                            answer_data = submitted_answers[answer_key]
                                            if isinstance(answer_data, dict):
                                                section_answers[i] = answer_data.get('selected_answer', answer_data.get('text', ''))
                                            else:
                                                section_answers[i] = str(answer_data)
                                else:
                                    # Fallback to old format
                                    submitted_answers = response_data.get('submitted_answers', {})
                                    for i, q in enumerate(questions):
                                        answer_key = f"Question_{q.order}"
                                        if answer_key in submitted_answers:
                                            answer_data = submitted_answers[answer_key]
                                            if isinstance(answer_data, dict):
                                                section_answers[i] = answer_data.get('selected_answer', answer_data.get('text', ''))
                                            else:
                                                section_answers[i] = str(answer_data)
                    except Exception as e:
                        print(f"Error processing aptitude section {section.title} for user {user.email}: {e}")
                
                data_row.append(section_time)
                data_row.append(section_time_minutes)
                data_row.append(section_start_time)
                data_row.append(section_end_time)
                data_row.append(section_manually_fixed)
                data_row.extend(section_answers)
            
            # Add aptitude test results
            aptitude_above_average = ""
            aptitude_average = ""
            aptitude_below_average = ""
            
            if aptitude_session:
                try:
                    # Get categories record
                    categories_record = TestTopCategories.objects.filter(
                        user=user,
                        test_paper=aptitude_test
                    ).first()
                    
                    if categories_record and categories_record.high_category:
                        # Parse the JSON data
                        high_categories = json.loads(categories_record.high_category)
                        
                        # Extract the categorized results
                        above_avg = high_categories.get("Above Average", [])
                        average = high_categories.get("Average", [])
                        below_avg = high_categories.get("Below Average", [])
                        
                        # Convert lists to comma-separated strings
                        aptitude_above_average = ", ".join(above_avg) if above_avg else "None"
                        aptitude_average = ", ".join(average) if average else "None"
                        aptitude_below_average = ", ".join(below_avg) if below_avg else "None"
                except Exception as e:
                    print(f"Error processing aptitude results for user {user.email}: {e}")
            
            # Add aptitude result data to row
            data_row.extend([
                aptitude_above_average,
                aptitude_average,
                aptitude_below_average
            ])
            
            # Write data row to worksheet
            for col, value in enumerate(data_row, 1):
                worksheet.cell(row=row, column=col, value=value)
            
            row += 1
        
        # Add statistics summary row
        summary_row = row + 1
        worksheet.cell(row=summary_row, column=1, value="SUMMARY STATISTICS:")
        worksheet.cell(row=summary_row, column=2, value=f"Total Students: {export_total}")
        worksheet.cell(row=summary_row, column=3, value=f"Completed: {export_completed}")
        worksheet.cell(row=summary_row, column=4, value=f"Pending: {export_pending}")
        worksheet.cell(row=summary_row, column=5, value=f"Not Started: {export_not_started}")
        
        # Format summary row
        for col in range(1, 6):
            cell = worksheet.cell(row=summary_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="class12_student_results_detailed.xlsx"'
        
        workbook.save(response)
        return response
    
    def get_student_class(self, user):
        """Determine if student is Class 12 or Class 10 - EXACT SAME LOGIC AS DISPLAY"""
        try:
            student_management = StudentManagement.objects.filter(student=user).first()
            if student_management and student_management.class_and_section:
                class_name = student_management.class_and_section.class_and_section
                if class_name:
                    class_prefix = class_name[:2].strip()
                    if class_prefix in ["11", "12"]:
                        return "class12"
                    elif class_prefix == "10":
                        return "class10"
        except:
            pass
        # Check by test sessions - if has post_matric sessions, likely class 12
        if TestSession.objects.filter(user=user).exists():
            return "class12"
        # If has TestCompletion, likely class 10
        if TestCompletion.objects.filter(user=user).exists():
            return "class10"
        return None


@method_decorator(login_required, name='dispatch')
class ExportAllClass10ResultsView(View):
    """
    Export comprehensive student results to Excel for ALL Class 10 students
    Following similar format to Class 12 export but adapted for Class 10 structure
    """
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import json
        from datetime import datetime, timezone
        
        # Import Class 10 models
        from app.models import Question as Class10Question, Answer as Class10Answer, Results as Class10Results
        
        # Get all Class 10 students
        class10_users = []
        all_users = User.objects.select_related('user_profile').prefetch_related('testcompletion_set')
        
        for user in all_users:
            student_class = self.get_student_class(user)
            if student_class == "class10":
                class10_users.append(user)
        
        # Get all unique test papers from Class 10 questions
        test_papers = Class10Question.objects.exclude(test_paper='').values_list('test_paper', flat=True).distinct().order_by('test_paper')
        
        # Get questions for each test paper
        test_papers_data = {}
        for test_paper in test_papers:
            questions = Class10Question.objects.filter(test_paper=test_paper).order_by('id')
            test_papers_data[test_paper] = questions
        
        # Create headers
        headers = [
            'Name', 'Gender', 'GRADE', 'Test Completion Status', 'Stream',
        ]
        
        # Add headers for each test paper
        for test_paper in test_papers:
            test_name = test_paper.replace('_', ' ').title() if test_paper else "Unknown"
            headers.append(f'{test_name} Time')
            headers.append(f'{test_name} Time (Minutes)')
            
            # Get question count for this test
            questions_count = len(test_papers_data.get(test_paper, []))
            for q_num in range(1, questions_count + 1):
                headers.append(f'{test_name} Q{q_num}')
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Class 10 Student Results"
        
        # Write headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Calculate statistics for verification (same as display)
        export_total = len(class10_users)
        export_completed = 0
        export_pending = 0
        export_not_started = 0
        
        # Process each Class 10 student
        row = 2
        for user in class10_users:
            # Basic user info
            name = user.name or user.username or "N/A"
            
            # Get student class for this user (already verified as class10)
            student_class = self.get_student_class(user)
            
            # Get Gender, Grade (Class), Stream from UserProfile and StudentManagement
            gender = "Unknown"
            grade = ""
            test_completion_status = "Not Started"
            stream = ""
            
            try:
                if hasattr(user, 'user_profile') and user.user_profile:
                    if user.user_profile.gender:
                        gender_choices = dict(user.user_profile._meta.get_field('gender').choices)
                        gender = gender_choices.get(user.user_profile.gender, "Unknown")
                
                # Get student class from StudentManagement
                try:
                    student_management = StudentManagement.objects.filter(student=user).first()
                    if student_management and student_management.class_and_section:
                        class_name = student_management.class_and_section.class_and_section
                        if class_name:
                            grade = class_name
                        stream = student_management.class_and_section.stream or ""
                    else:
                        # Fallback: try to get from UserProfile
                        if hasattr(user, 'user_profile') and user.user_profile:
                            grade = user.user_profile.grade or ""
                    
                    # If still no grade, try to infer from student_class determination
                    if not grade:
                        if student_class == "class10":
                            grade = "Class 10"
                except:
                    pass
            except:
                pass
            
            # Calculate completion status (EXACT SAME LOGIC AS DISPLAY)
            test_completion = Class10Results.objects.filter(user=user).exists()
            primary_tests_count = 0  # test1, test2, test3
            
            if test_completion:
                # Check TestCompletion model for primary tests
                test_completion_obj = TestCompletion.objects.filter(user=user).first()
                if test_completion_obj:
                    if test_completion_obj.test1_complete:
                        primary_tests_count += 1
                    if test_completion_obj.test2_complete:
                        primary_tests_count += 1
                    if test_completion_obj.test3_complete:
                        primary_tests_count += 1
                    
                    if test_completion_obj.are_all_primary_tests_completed():
                        test_completion_status = "Completed"
                        export_completed += 1
                    elif primary_tests_count > 0:
                        test_completion_status = f"Pending ({primary_tests_count}/3)"
                        export_pending += 1
                    else:
                        test_completion_status = "Not Started"
                        export_not_started += 1
                else:
                    test_completion_status = "Not Started"
                    export_not_started += 1
            else:
                test_completion_status = "Not Started"
                export_not_started += 1
            
            # Initialize data row
            data_row = [name, gender, grade, test_completion_status, stream]
            
            # Process each test paper
            for test_paper in test_papers:
                questions = test_papers_data.get(test_paper, [])
                test_result = Class10Results.objects.filter(user=user, test_paper=test_paper).first()
                
                # Get time (from modified field if available)
                test_time = ""
                test_time_minutes = ""
                if test_result and test_result.modified:
                    # We don't have start/end time, so use modified time as completion time
                    test_time = "Completed"
                    test_time_minutes = "N/A"  # No duration available
                else:
                    test_time = "N/A"
                    test_time_minutes = "N/A"
                
                data_row.append(test_time)
                data_row.append(test_time_minutes)
                
                # Get answers for each question
                answers_list = [""] * len(questions)
                
                if test_result and test_result.selected_answers:
                    try:
                        selected_answers_data = test_result.selected_answers
                        
                        # Check if it's the new format with category keys (like Spatial, Numerical, etc.)
                        if isinstance(selected_answers_data, dict):
                            # Check for submitted_answers format (test1, test2, test3)
                            if 'submitted_answers' in selected_answers_data:
                                submitted_answers = selected_answers_data['submitted_answers']
                                
                                for i, question in enumerate(questions):
                                    # Try different answer key formats
                                    answer_key = f"Question_{i + 1}"
                                    answer_key_alt = f"question_{i + 1}"
                                    answer_key_id = f"Question_{question.id}"
                                    
                                    answer_value = ""
                                    if answer_key in submitted_answers:
                                        answer_value = submitted_answers[answer_key]
                                    elif answer_key_alt in submitted_answers:
                                        answer_value = submitted_answers[answer_key_alt]
                                    elif answer_key_id in submitted_answers:
                                        answer_value = submitted_answers[answer_key_id]
                                    
                                    # Convert answer value to text if needed
                                    if answer_value:
                                        # If it's a number, try to get answer text from Answer model
                                        try:
                                            answer_obj = Class10Answer.objects.filter(
                                                question=question,
                                                id=int(answer_value)
                                            ).first()
                                            if answer_obj:
                                                answers_list[i] = answer_obj.answer_text or str(answer_value)
                                            else:
                                                answers_list[i] = str(answer_value)
                                        except (ValueError, TypeError):
                                            # If not a valid number, use as string
                                            answers_list[i] = str(answer_value)
                            else:
                                # Format with category keys (like {'Spatial': [...], 'Numerical': [...]})
                                # Extract answers from category lists
                                for category_key, category_questions in selected_answers_data.items():
                                    if isinstance(category_questions, list):
                                        for q_data in category_questions:
                                            if isinstance(q_data, dict):
                                                selected_answer = q_data.get('selected_answer', '')
                                                question_text = q_data.get('question_text', '')
                                                
                                                # Find matching question by text
                                                for i, question in enumerate(questions):
                                                    if question.question_text == question_text:
                                                        answers_list[i] = selected_answer
                                                        break
                    except Exception as e:
                        print(f"Error processing answers for user {user.email}, test {test_paper}: {e}")
                
                data_row.extend(answers_list)
            
            # Write data row to worksheet
            for col, value in enumerate(data_row, 1):
                worksheet.cell(row=row, column=col, value=value)
            
            row += 1
        
        # Add statistics summary row
        summary_row = row + 1
        worksheet.cell(row=summary_row, column=1, value="SUMMARY STATISTICS:")
        worksheet.cell(row=summary_row, column=2, value=f"Total Students: {export_total}")
        worksheet.cell(row=summary_row, column=3, value=f"Completed: {export_completed}")
        worksheet.cell(row=summary_row, column=4, value=f"Pending: {export_pending}")
        worksheet.cell(row=summary_row, column=5, value=f"Not Started: {export_not_started}")
        
        # Format summary row
        for col in range(1, 6):
            cell = worksheet.cell(row=summary_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="class10_student_results_detailed.xlsx"'
        
        workbook.save(response)
        return response
    
    def get_student_class(self, user):
        """Determine if student is Class 10 - EXACT SAME LOGIC AS DISPLAY"""
        try:
            student_management = StudentManagement.objects.filter(student=user).first()
            if student_management and student_management.class_and_section:
                class_name = student_management.class_and_section.class_and_section
                if class_name:
                    class_prefix = class_name[:2].strip()
                    if class_prefix == "10":
                        return "class10"
        except:
            pass
        # If has TestCompletion, likely class 10
        if TestCompletion.objects.filter(user=user).exists():
            return "class10"
        return None


@method_decorator(login_required, name='dispatch')
class ExportClass12TestQuestionsView(View):
    """
    Export Class 12 test questions and correct answers grouped by category
    """
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Class 12 Test Questions"
        
        # Headers
        headers = [
            'Question No', 'Question', 'Question Image URL', 'Correct Answer',
            'Section', 'Category', 'Difficulty', 'Time Limit (seconds)'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Get all Class 12 tests
        all_tests = Test.objects.filter(is_active=True).order_by('title')
        
        row = 2
        for test in all_tests:
            # Check if test has sections (like Aptitude)
            sections = Sections.objects.filter(test=test).order_by('order')
            
            if sections.exists():
                # Test has sections - process each section
                for section in sections:
                    questions = Question.objects.filter(section=section).order_by('order')
                    
                    for question in questions:
                        # Get correct answer
                        correct_answer = ""
                        if question.question_type == 'multiple_choice':
                            correct_answer_obj = Answer.objects.filter(
                                question=question, is_correct=True
                            ).first()
                            if correct_answer_obj:
                                correct_answer = correct_answer_obj.text
                        elif question.question_type in ['scale', 'true_false']:
                            # For scale/true_false, get the answer with highest score or check is_correct
                            answer_obj = Answer.objects.filter(question=question).first()
                            if answer_obj:
                                correct_answer = answer_obj.text
                        
                        # Get question image URL
                        question_image_url = ""
                        if question.image:
                            question_image_url = request.build_absolute_uri(question.image.url)
                        
                        # Get category (from question_dimension or answer category)
                        category = question.question_dimension or "N/A"
                        category_display = dict(question._meta.get_field('question_dimension').choices).get(category, category)
                        
                        # Get difficulty
                        difficulty = question.question_level or "N/A"
                        
                        # Write data row
                        data_row = [
                            question.order,
                            question.text,
                            question_image_url,
                            correct_answer,
                            section.title,
                            category_display,
                            difficulty,
                            "N/A"  # Time limit not stored in model
                        ]
                        
                        for col, value in enumerate(data_row, 1):
                            worksheet.cell(row=row, column=col, value=value)
                        
                        row += 1
            else:
                # Test without sections - process questions directly
                questions = Question.objects.filter(test=test).order_by('order')
                
                for question in questions:
                    # Get correct answer
                    correct_answer = ""
                    if question.question_type == 'multiple_choice':
                        correct_answer_obj = Answer.objects.filter(
                            question=question, is_correct=True
                        ).first()
                        if correct_answer_obj:
                            correct_answer = correct_answer_obj.text
                    elif question.question_type in ['scale', 'true_false']:
                        answer_obj = Answer.objects.filter(question=question).first()
                        if answer_obj:
                            correct_answer = answer_obj.text
                    
                    # Get question image URL
                    question_image_url = ""
                    if question.image:
                        question_image_url = request.build_absolute_uri(question.image.url)
                    
                    # Get category
                    category = question.question_dimension or "N/A"
                    category_display = dict(question._meta.get_field('question_dimension').choices).get(category, category)
                    
                    # Get difficulty
                    difficulty = question.question_level or "N/A"
                    
                    # Write data row
                    data_row = [
                        question.order,
                        question.text,
                        question_image_url,
                        correct_answer,
                        "N/A",  # No section
                        category_display,
                        difficulty,
                        "N/A"  # Time limit not stored in model
                    ]
                    
                    for col, value in enumerate(data_row, 1):
                        worksheet.cell(row=row, column=col, value=value)
                    
                    row += 1
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="class12_test_questions_answers.xlsx"'
        
        workbook.save(response)
        return response


@method_decorator(login_required, name='dispatch')
class ExportClass10TestQuestionsView(View):
    """
    Export Class 10 test questions and correct answers grouped by category
    """
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Import Class 10 models
        from app.models import Question as Class10Question, Answer as Class10Answer
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Class 10 Test Questions"
        
        # Headers
        headers = [
            'Question No', 'Question', 'Question Image URL', 'Correct Answer',
            'Section', 'Category', 'Difficulty', 'Time Limit (seconds)'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Get all Class 10 questions grouped by test_paper and category
        questions = Class10Question.objects.all().order_by('test_paper', 'category', 'id')
        
        row = 2
        question_no = 1
        
        # Group by test_paper
        current_test_paper = None
        for question in questions:
            # Group by test_paper (section)
            if question.test_paper != current_test_paper:
                current_test_paper = question.test_paper
                question_no = 1
            
            # Get correct answer
            correct_answer = ""
            correct_answer_obj = Class10Answer.objects.filter(
                question=question, is_correct=True
            ).first()
            if correct_answer_obj:
                correct_answer = correct_answer_obj.answer_text or ""
            
            # Get question image URL
            question_image_url = ""
            if question.question_image:
                question_image_url = request.build_absolute_uri(question.question_image.url)
            
            # Write data row
            data_row = [
                question_no,
                question.question_text or "",
                question_image_url,
                correct_answer,
                question.test_paper or "N/A",  # Use test_paper as section
                question.category or "N/A",
                "N/A",  # Difficulty not stored in Class 10 model
                "N/A"  # Time limit not stored in model
            ]
            
            for col, value in enumerate(data_row, 1):
                worksheet.cell(row=row, column=col, value=value)
            
            row += 1
            question_no += 1
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="class10_test_questions_answers.xlsx"'
        
        workbook.save(response)
        return response


@method_decorator(login_required, name='dispatch')
class ExportClass12ResultsView(View):
    """
    Export comprehensive student results to Excel for Class 12 (app_post_matric)
    """
    def get(self, request, user_id):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import json
        
        try:
            student = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return HttpResponse("Student not found", status=404)
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"{student.name} - Results"
        
        # Get all tests for class 12
        all_tests = Test.objects.filter(is_active=True).order_by('title')
        
        # Create headers
        headers = ['Name', 'Email', 'Mobile', 'Gender', 'Grade', 'Stream']
        
        # Process each test
        tests_data = []
        for test in all_tests:
            test_session = TestSession.objects.filter(
                user=student,
                test=test
            ).first()
            
            if test_session:
                # Get sections if they exist
                sections = Sections.objects.filter(test=test).order_by('order')
                
                if sections.exists():
                    # Test has sections (like Aptitude)
                    for section in sections:
                        headers.append(f'{test.title} - {section.title} Time')
                        headers.append(f'{test.title} - {section.title} Time (Minutes)')
                        
                        # Get section questions
                        section_questions = Question.objects.filter(
                            section=section
                        ).order_by('order')
                        for q in section_questions:
                            headers.append(f'{test.title} - {section.title} Q{q.order}')
                else:
                    # Test without sections (like Personality, Motivation, Interest)
                    headers.append(f'{test.title} Time')
                    headers.append(f'{test.title} Time (Minutes)')
                    
                    # Get test questions
                    test_questions = Question.objects.filter(
                        test=test
                    ).order_by('order')
                    for q in test_questions:
                        headers.append(f'{test.title} Q{q.order}')
                
                tests_data.append({
                    'test': test,
                    'session': test_session,
                    'has_sections': sections.exists()
                })
        
        # Write headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Initialize data row
        data_row = [
            student.name or "N/A",
            student.email or "N/A",
            student.mobile or "N/A",
            "",  # Gender - not available in current schema
            "",  # Grade - not available in current schema
            "",  # Stream - not available in current schema
        ]
        
        # Process each test's data
        for test_info in tests_data:
            test = test_info['test']
            session = test_info['session']
            has_sections = test_info['has_sections']
            
            if has_sections:
                # Process sections
                sections = Sections.objects.filter(test=test).order_by('order')
                
                for section in sections:
                    # Get section session
                    section_session = SectionSession.objects.filter(
                        session=session,
                        section=section
                    ).first()
                    
                    # Calculate time
                    section_time = ""
                    section_time_minutes = ""
                    
                    if section_session:
                        if section_session.end_time and section_session.start_time:
                            time_diff = section_session.end_time - section_session.start_time
                            if time_diff.total_seconds() >= 0:
                                section_time = f"{int(time_diff.total_seconds())}s"
                                section_time_minutes = f"{round(time_diff.total_seconds() / 60, 2)}"
                            else:
                                section_time = "0s"
                                section_time_minutes = "0"
                        elif section_session.start_time and not section_session.end_time:
                            section_time = "Incomplete"
                            section_time_minutes = "Incomplete"
                        else:
                            section_time = "N/A"
                            section_time_minutes = "N/A"
                        
                        # Get answers for this section
                        section_questions = Question.objects.filter(
                            section=section
                        ).order_by('order')
                        section_answers = [""] * len(section_questions)
                        
                        try:
                            user_response = UserResponse.objects.filter(
                                session=session,
                                session_section=section_session
                            ).first()
                            
                            if not user_response:
                                user_response = UserResponse.objects.filter(
                                    session=session
                                ).first()
                            
                            if user_response and user_response.selected_answer:
                                response_data = user_response.selected_answer
                                
                                # Handle section-based response format
                                if 'sections' in response_data and section.title in response_data['sections']:
                                    section_data_resp = response_data['sections'][section.title]
                                    submitted_answers = section_data_resp.get('submitted_answers', {})
                                    
                                    for i, q in enumerate(section_questions):
                                        answer_key = f"Question_{q.order}"
                                        if answer_key in submitted_answers:
                                            answer_data = submitted_answers[answer_key]
                                            if isinstance(answer_data, dict):
                                                section_answers[i] = answer_data.get('selected_answer', answer_data.get('text', ''))
                                            else:
                                                section_answers[i] = str(answer_data)
                                else:
                                    # Fallback to old format
                                    submitted_answers = response_data.get('submitted_answers', {})
                                    for i, q in enumerate(section_questions):
                                        answer_key = f"Question_{q.order}"
                                        if answer_key in submitted_answers:
                                            answer_data = submitted_answers[answer_key]
                                            if isinstance(answer_data, dict):
                                                section_answers[i] = answer_data.get('selected_answer', answer_data.get('text', ''))
                                            else:
                                                section_answers[i] = str(answer_data)
                        except Exception as e:
                            print(f"Error processing section {section.title} answers: {e}")
                        
                        data_row.append(section_time)
                        data_row.append(section_time_minutes)
                        data_row.extend(section_answers)
                    else:
                        # No section session
                        section_questions = Question.objects.filter(
                            section=section
                        ).order_by('order')
                        data_row.append("N/A")
                        data_row.append("N/A")
                        data_row.extend([""] * len(section_questions))
            else:
                # Test without sections
                # Calculate time
                test_time = ""
                test_time_minutes = ""
                
                if session.end_time and session.start_time:
                    time_diff = session.end_time - session.start_time
                    if time_diff.total_seconds() >= 0:
                        test_time = f"{int(time_diff.total_seconds())}s"
                        test_time_minutes = f"{round(time_diff.total_seconds() / 60, 2)}"
                    else:
                        test_time = "0s"
                        test_time_minutes = "0"
                elif session.start_time and not session.end_time:
                    test_time = "Incomplete"
                    test_time_minutes = "Incomplete"
                else:
                    test_time = "N/A"
                    test_time_minutes = "N/A"
                
                # Get answers
                test_questions = Question.objects.filter(
                    test=test
                ).order_by('order')
                test_answers = [""] * len(test_questions)
                
                try:
                    user_response = UserResponse.objects.filter(
                        session=session
                    ).first()
                    
                    if user_response and user_response.selected_answer:
                        submitted_answers = user_response.selected_answer.get('submitted_answers', {})
                        
                        for i, q in enumerate(test_questions):
                            answer_key = f"Question_{q.order}"
                            if answer_key in submitted_answers:
                                answer_data = submitted_answers[answer_key]
                                if isinstance(answer_data, dict):
                                    test_answers[i] = answer_data.get('text', answer_data.get('selected_answer', ''))
                                else:
                                    test_answers[i] = str(answer_data)
                except Exception as e:
                    print(f"Error processing {test.title} answers: {e}")
                
                data_row.append(test_time)
                data_row.append(test_time_minutes)
                data_row.extend(test_answers)
        
        # Write data row to worksheet
        for col, value in enumerate(data_row, 1):
            worksheet.cell(row=2, column=col, value=value)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"class12_results_{student.name.replace(' ', '_')}_{student.id}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        return response


# Media Library Views for S3 File Management
from core.s3_utils import get_s3_upload_service
from core.models import S3FileUpload
import json


@method_decorator(login_required, name='dispatch')
class MediaLibraryView(TemplateView):
    """Media Library view with FTP-like file manager"""
    template_name = "topteenadmin/media_library.html"
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['active_tab'] = 'media_library'
        ctx['meta_title'] = 'Media Library'
        ctx['html_head'] = {'title': 'Media Library', 'description': 'Manage S3 files and folders'}
        ctx['breadcrumb'] = build_admin_breadcrumb([
            {'title': 'Media Library', 'text': 'Media Library', 'url': '#'}
        ])
        
        # Get media library base folder
        media_library_base = getattr(settings, 'S3_MEDIA_LIBRARY_BASE_FOLDER', 'medialibrary')
        
        # Get current folder path (relative to medialibrary base)
        folder_path = self.request.GET.get('folder', '').strip()
        ctx['current_folder'] = folder_path
        
        # Build full S3 path (medialibrary/ + subfolder)
        if folder_path:
            full_s3_path = f"{media_library_base}/{folder_path}".strip('/')
        else:
            full_s3_path = media_library_base
        
        # Get folder breadcrumb
        breadcrumb_folders = []
        if folder_path:
            parts = folder_path.strip('/').split('/')
            current_path = ''
            for part in parts:
                current_path = f"{current_path}/{part}" if current_path else part
                breadcrumb_folders.append({
                    'name': part,
                    'path': current_path
                })
        ctx['folder_breadcrumb'] = breadcrumb_folders
        
        # Get S3 service
        s3_service = get_s3_upload_service()
        ctx['s3_enabled'] = s3_service.is_enabled()
        ctx['max_file_size_mb'] = getattr(settings, 'S3_MAX_FILE_SIZE_MB', 2)
        
        # Get folders and files (using full S3 path)
        if s3_service.is_enabled():
            result = s3_service.list_folders_and_files(full_s3_path)
            # Update folder paths to be relative to medialibrary base
            folders = []
            for folder in result['folders']:
                # Remove medialibrary base from path for display
                folder_full_path = folder['path']
                if folder_full_path.startswith(f"{media_library_base}/"):
                    rel_path = folder_full_path[len(f"{media_library_base}/"):]
                elif folder_full_path == media_library_base:
                    rel_path = ""
                else:
                    rel_path = folder_full_path
                folders.append({
                    'name': folder['name'],
                    'path': rel_path
                })
            ctx['folders'] = folders
            ctx['files'] = result['files']
        else:
            ctx['folders'] = []
            ctx['files'] = []
        
        return ctx


@method_decorator(login_required, name='dispatch')
class MediaLibraryUploadView(View):
    """Handle file upload to S3"""
    
    def post(self, request):
        try:
            folder_path = request.POST.get('folder_path', '').strip()
            uploaded_file = request.FILES.get('file')
            
            if not uploaded_file:
                return JsonResponse({
                    'success': False,
                    'error': 'No file provided'
                })
            
            s3_service = get_s3_upload_service()
            
            if not s3_service.is_enabled():
                return JsonResponse({
                    'success': False,
                    'error': 'S3 upload is disabled'
                })
            
            # Get media library base folder and build full path
            media_library_base = getattr(settings, 'S3_MEDIA_LIBRARY_BASE_FOLDER', 'medialibrary')
            if folder_path:
                full_folder_path = f"{media_library_base}/{folder_path}".strip('/')
            else:
                full_folder_path = media_library_base
            
            result = s3_service.upload_file(
                file_obj=uploaded_file,
                folder_path=full_folder_path,
                uploaded_by=request.user.username if request.user.is_authenticated else ''
            )
            
            if result['success']:
                return JsonResponse({
                    'success': True,
                    'message': 'File uploaded successfully',
                    'file': {
                        'id': result.get('upload_id'),
                        'name': uploaded_file.name,
                        'url': result.get('s3_url'),
                        'size': result.get('file_size')
                    }
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result.get('error', 'Upload failed')
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })


@method_decorator(login_required, name='dispatch')
class MediaLibraryDeleteFileView(View):
    """Handle file deletion from S3"""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            s3_key = data.get('s3_key')
            file_id = data.get('file_id')
            
            if not s3_key and not file_id:
                return JsonResponse({
                    'success': False,
                    'error': 'File key or ID required'
                })
            
            # Get s3_key from database if only ID provided
            if not s3_key and file_id:
                try:
                    file_obj = S3FileUpload.objects.get(id=file_id)
                    s3_key = file_obj.s3_key
                except S3FileUpload.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': 'File not found'
                    })
            
            s3_service = get_s3_upload_service()
            result = s3_service.delete_file(s3_key)
            
            if result['success']:
                return JsonResponse({
                    'success': True,
                    'message': 'File deleted successfully'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result.get('error', 'Delete failed')
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })


@method_decorator(login_required, name='dispatch')
class MediaLibraryCreateFolderView(View):
    """Handle folder creation in S3"""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            folder_name = data.get('folder_name', '').strip()
            parent_folder = data.get('parent_folder', '').strip()
            
            if not folder_name:
                return JsonResponse({
                    'success': False,
                    'error': 'Folder name is required'
                })
            
            # Get media library base folder
            media_library_base = getattr(settings, 'S3_MEDIA_LIBRARY_BASE_FOLDER', 'medialibrary')
            
            # Build full folder path (medialibrary/ + parent + folder_name)
            if parent_folder:
                folder_path = f"{media_library_base}/{parent_folder}/{folder_name}".strip('/')
            else:
                folder_path = f"{media_library_base}/{folder_name}".strip('/')
            
            s3_service = get_s3_upload_service()
            result = s3_service.create_folder(folder_path)
            
            if result['success']:
                # Return relative path (without medialibrary base) for display
                rel_path = folder_path.replace(f"{media_library_base}/", "", 1) if folder_path.startswith(media_library_base) else folder_path
                return JsonResponse({
                    'success': True,
                    'message': 'Folder created successfully',
                    'folder': {
                        'name': folder_name,
                        'path': rel_path
                    }
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result.get('error', 'Create folder failed')
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })


@method_decorator(login_required, name='dispatch')
class MediaLibraryDeleteFolderView(View):
    """Handle folder deletion from S3"""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            folder_path = data.get('folder_path', '').strip()
            
            if not folder_path:
                return JsonResponse({
                    'success': False,
                    'error': 'Folder path is required'
                })
            
            # Get media library base folder and build full path
            media_library_base = getattr(settings, 'S3_MEDIA_LIBRARY_BASE_FOLDER', 'medialibrary')
            if folder_path:
                full_folder_path = f"{media_library_base}/{folder_path}".strip('/')
            else:
                full_folder_path = media_library_base
            
            s3_service = get_s3_upload_service()
            result = s3_service.delete_folder(full_folder_path)
            
            if result['success']:
                return JsonResponse({
                    'success': True,
                    'message': 'Folder deleted successfully'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result.get('error', 'Delete folder failed')
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })